
import sys
import time
import os
import re
import  openpyxl

from openpyxl import worksheet 
from pprint import pprint 
import pandas as pd
from openpyxl.utils import get_column_letter 

import tkinter as tk

from openpyxl.styles import Border, Side, PatternFill, Alignment 
from openpyxl.worksheet.datavalidation import DataValidation

from .baseupf import *

class PDomainSheet(BaseSheet):
    def __init__(self,*args):
        super().__init__(*args)
        self._pddata = {}
        #self._pdnmdic = {}
        
    def update_sheet(self):
        '''
        # only during -dg option
        # addition of supply infos in PMDOMAIN table
        # addition of dropdown infos from pmempty and pmobj file
        '''

        sheet = self.get_sheet()

        supply_kw, supply_vol, supply_vss, supply_data = self.get_supply_infos()

        # find PMDOMAIN table 
        start_rowg = self.find_sheet(sheet, 'PMDOMAIN') 

        #sheet.insert_cols(2, amount = len(supply_kw) + 1)
        #print(start_rowg)
        #rgb_color = PatternFill(start_color=cell_color, end_color=cell_color).start_color.rgb
        #print("单元格A1的填充颜色RGB值:", rgb_color)
        # print(sheet.cell(start_rowg,1).fill.start_color.index)
        # print(sheet.cell(start_rowg,1).fill.end_color.index)

        for i in range(1, len(supply_kw) + 1):
            sheet.cell(start_rowg, i + 2).value = supply_kw[i-1]
        sheet.cell(start_rowg, len(supply_kw) + 3).value = 'Comment'
        self.cell_style1(sheet, [start_rowg, 3], [start_rowg, len(supply_kw) + 3])
        #print(sheet.cell(start_rowg,1).fill.start_color.index)
        #print(sheet.cell(start_rowg,1).fill.end_color.index)
        self.cell_style2(sheet, [start_rowg + 1, 3], [start_rowg + 8, len(supply_kw) + 3])

        # formulal=f'"{iodly_str},40%,50%, 60%, 70%, 80%"'
        # formula2=r'[-+]?[0-9]*\.?[0-9]+'
        # f'"{",".join(choices)}"'
        self.add_dropdown(sheet, '"PRM"', [start_rowg + 1, 3], [start_rowg + 8, len(supply_kw) + 2])
        self.add_dropdown(sheet, '"-update"', [start_rowg + 1, len(supply_kw) + 3], [start_rowg + 8, len(supply_kw) + 3])

        # find PMNETWORK table
        # SupplyPortNet	NPwellNet	InstList	MapSupplyList	Comment
        # f'"{",".join(supply_kw)}"'
        start_rowg = self.find_sheet(sheet, 'PMNETWORK')
        # self.add_dropdown(sheet, '"' + ','.join(supply_kw) + '"', [start_rowg + 1, 1], [start_rowg + 10, 1])
        # obj_data = list(self._upfdg._objfile_data.keys())
        # self.add_dropdown(sheet, '"' + ','.join(obj_data) + '"', [start_rowg + 1, 3], [start_rowg + 10, 3])
        self.get_impl_obj(sheet, start_rowg, 'PMNETWORK')

        # find PMBOUNDARY table
        # ApplyPorts	Elements	ExcludeList	DriverSupply	ReceiverSupply	Attribute	Comment
        start_rowg = self.find_sheet(sheet, 'PMBOUNDARY')
        self.add_dropdown(sheet, '"inputs,outputs,both"', [start_rowg + 1, 1], [start_rowg + 10, 1])
        #print(self.get_impl_obj(sheet, start_rowg, 'PMBOUNDARY'))
        kw_ele,kw_exd = self.get_impl_obj(sheet, start_rowg, 'PMBOUNDARY')
        if kw_ele:
            self.add_dropdown(sheet, '"' + ','.join(kw_ele) + '"', [start_rowg + 1, 2], [start_rowg + 10, 2])
        if kw_exd:  
            self.add_dropdown(sheet, '"' + ','.join(kw_exd) + '"', [start_rowg + 1, 3], [start_rowg + 10, 3])

        self.add_dropdown(sheet, '"' + ','.join(supply_kw) + '"', [start_rowg + 1, 4], [start_rowg + 10, 5])
        kw_bd = ['iso_sink', 'iso_source', 'snps_derived', 'related_supply_default_primary', 'resolved_iso_strategy']
        self.add_dropdown(sheet, '"' + ','.join(kw_bd) + '"', [start_rowg + 1, 6], [start_rowg + 10, 6])
        kw_bd = ['-feedthrough', '-unconnected', '-is_analog']
        self.add_dropdown(sheet, '"' + ','.join(kw_bd) + '"', [start_rowg + 1, 7], [start_rowg + 10, 7])        
  

    def read_data(self):
        sheet = self.get_sheet()
        self._pddata = self.get_table_contxt(sheet)
        # print('_pddata: ', self._pddata)

    def change_sheet(self):
        pass

    def dump_json(self,json_file):
        self._data = self._pddata
        self.write_json(json_file)

    def write_upf(self,mdname,blkalias,blklvl,upf_file):
        sheet = self.get_sheet()

        # mdname = self._upfdg._vfile_data['module_name']
        # blkalias = self._upfdg._hier_tree._blocks[mdname].alias
        # blklvl = self._upfdg._hier_tree._blocks[mdname].hdlevel

        #supply_kw, supply_vol, supply_vss, supply_data = self.get_supply_infos()
        # supply_kw  = supply_kw.extend(supply_vss)
        #pdkeys = list(self._pddata.keys())

        # pdomaindict = {}
        # pdomainlist = [(key, val) for key, val in self._pddata.items() if re.search(r'PMDOMAIN_Row\d+',key) and val['PDName']]
        # for k,v in pdomainlist:
        #     pdomaindict[k] = v
        # pdomainkeys = [x for x in pdomainlist[0] if re.search(r'PMDOMAIN_Row\d+',x)]
        # pdomainkeys.sort()
        pdomaindict, pdomainkeys = self.get_rows(self._pddata,'PMDOMAIN_Row','PDName', 'PDName')

        # pmnetdict = {}
        # pmnetlist = [(key,val) for key, val in self._pddata.items() if re.search(r'PMNETWORK_Row\d+',key) and val['SupplyPortNet']]
        # for k,v in pmnetlist:
        #     pmnetdict[k] = v
        # pmnetkeys = [x for x in pmnetlist[0] if re.search(r'PMNETWORK_Row\d+',x)]
        # pmnetkeys.sort()
        pmnetdict, pmnetkeys = self.get_rows(self._pddata,'PMNETWORK_Row','SupplyPortNet', 'SupplyPortNet')

        # pmbrddict = {}
        # pmbdrdlist = [(key,val) for key, val in self._pddata.items() if re.search(r'PMBOUNDARY_Row\d+',key) and val['Elements']]
        # for k,v in pmbdrdlist:
        #     pmbrddict[k] = v        
        # pmbdrdkeys = [x for x in pdkeys if re.search(r'PMBOUNDARY_Row\d+',x)]
        # pmbdrdkeys.sort()
        pmbrddict, pmbdrdkeys = self.get_rows(self._pddata,'PMBOUNDARY_Row','ApplyPorts|Elements','ApplyPorts')

        # spy_kwd = [x for x in list[self._pddata[pdomainkeys[0]].keys()] if not x in ['PDName','Elements','Comment']].extend(supply_vss)
        # print(spy_kwd)


        upf_lines = f'''
# ========================================= #
# create supply port & net & set
# ========================================= #
'''    
        upf_lines += self.crt_supply(blkalias,pmnetkeys,pmnetdict)

        upf_lines += f'''
# ========================================= #
# create power domain
# ========================================= #
'''      
        upf_lines += self.crt_domain(blkalias,pdomainkeys,pdomaindict) 

        upf_lines += f'''
# ========================================= #
# supply connection
# ========================================= #
'''      
        upf_lines += self.connect_supply(blkalias,pmnetkeys,pmnetdict)

        upf_lines += f'''
# ========================================= #
# boundary port related supply
# ========================================= #
'''      
        upf_lines += self.port_supply(blkalias,pmbdrdkeys,pmbrddict)       

        self.save_text(upf_lines,upf_file)
       
    def crt_supply(self,blkalias,prows,pdict):

        upf_lines = ''

        upf_lines += f'''
# create supply port and net
        '''

        # supply net
        for prow in prows:
            rowdict = pdict[prow]
            sportnet = rowdict['SupplyPortNet']
            if '(' in sportnet:
                sportnet = ''.join(sportnet.split('(')[0])

            if sportnet and not rowdict['InstList'] and not rowdict['MapSupplyList']:

                if re.search(r'_PSW\d+',sportnet):
                    if rowdict['Comment'] == 'PAL OUT':
                        upf_lines += f'''
# line: {prow}
create_supply_port {sportnet} -direction out
create_supply_net {sportnet} -resolve parallel
connect_supply_net {sportnet} -ports {sportnet}
'''
                    if rowdict['Comment'] == 'PAL':
                        upf_lines += f'''
# line: {prow}
# create_supply_port {sportnet} -direction out
create_supply_net {sportnet} -resolve parallel
# connect_supply_net {sportnet} -ports {sportnet}
'''
                    if rowdict['Comment'] == 'OUT':
                        upf_lines += f'''
# line: {prow}
create_supply_port {sportnet} -direction out
create_supply_net {sportnet} 
connect_supply_net {sportnet} -ports {sportnet}
'''
                else:
                    upf_lines += f'''
# line: {prow}
create_supply_port {sportnet} -direction in
create_supply_net {sportnet} 
connect_supply_net {sportnet} -ports {sportnet}
'''

        upf_lines += f'''
# create supply set
'''
        # supply set
        for prow in prows:
            rowdict = pdict[prow]
            sportnet = rowdict['SupplyPortNet'] 

            if sportnet and rowdict['NPwellNet'] and not rowdict['InstList'] and not rowdict['MapSupplyList']:

                
                if '(' in sportnet:
                    nsportnet = ''.join(sportnet.split('(')[0])
                    nvss = ''.join(sportnet.split('(')[1]).replace(')','')

                    upf_lines += f'''
# line: {prow}
create_supply_set SS_${{{blkalias}}}_{nsportnet} \
        -function   {{power {nsportnet}}} \
        -function   {{nwell {rowdict['NPwellNet']}}} \
        -function   {{pwell {nvss}}} \
        -function   {{ground {nvss}}} 
'''           
                else:
                    upf_lines += f'''
# line: {prow}
create_supply_set SS_${{{blkalias}}}_{sportnet} \
        -function   {{power {sportnet}}} \
        -function   {{nwell {rowdict['NPwellNet']}}} \
        -function   {{pwell VSS}} \
        -function   {{ground VSS}} 
''' 
        # equalent net
        for prow in prows:
            rowdict = pdict[prow]
            sportnet = rowdict['SupplyPortNet']
            cmt =  rowdict['Comment']

            if sportnet and cmt and not rowdict['NPwellNet'] and not rowdict['InstList'] and not rowdict['MapSupplyList']:
                upf_lines += f'''
# equalent net
'''          
                if len(sportnet.split('') == 2) and re.search(r'SNE\d+',cmt):
                    snets = sportnet.split(' ')
                    upf_lines += f'''
set_equivalent -nets {{f'{snets[0]} {snets[1]}}}
'''
        
        return upf_lines
    
 
    
    def crt_domain(self,blkalias,prows,pdict):
        upf_lines = '' 

        i = 0
        #pddic = {}
        for prow in prows:
            rowdict = pdict[prow]
            pdname = rowdict['PDName']
            pdele = rowdict['Elements']
            i += 1

            prmdic = [(key,val) for key,val in rowdict.items() if val == 'PRM']
            if len(prmdic) > 1:
                upf_error(f'{prow} has two or more primary power. Please check it.')
                prm = 'NOPRM'
            else:
                prm = prmdic[0][0]
            
            nprm = ''.join(prm.split('_'))
            if pdele =='.':
                pdele = '{.}'

            #self._pdnmdic[f'{nprm}'] = f'PD{i}_${{{blkalias}}}_{nprm}_{pdname}'
            
            upf_lines += f'''
# line: {prow}
create_power_domain PD{i}_${{{blkalias}}}_{nprm}_{pdname} \
                    -supply {{primary SS_${{{blkalias}}}_{prm}}} \
                    -elements   "{pdele}"
'''

        return upf_lines

    def connect_supply(self,blkalias,prows,pdict):

        upf_lines = '' 
        # supply net
        for prow in prows:
            rowdict = pdict[prow]
            sportnets = rowdict['SupplyPortNet']
            insts = rowdict['InstList']
            mappins = rowdict['MapSupplyList']

            if sportnets and mappins:
                sportnets_lst = sportnets.split(' ')
                mappins_lst = mappins.split(' ')
                if len(sportnets_lst) != len(mappins_lst):
                    upf_error(f'{prow} supply mapping number is not correct. Please check it.')
                else:
                    # nmap = ''
                    # for i,j in zip(sportnets,mappins):
                    #     nmap += f'{i} {j} '
                    # nmap = nmap.strip().split(' ')
                    # print(nmap)

                    # print('sportnets,mappins: ', sportnets,mappins)
                    sportnetsx = sportnets.split(' ')
                    mappinsx = mappins.split(' ')
                    # print('sportnetsx,mappinsx: ', sportnetsx,mappinsx)
                    for inst in insts.split(' '):
                        ninsts = f'$UPFVAR({inst},${{{blkalias}}})'
                        upf_lines += f'''
    # line: {prow}
    foreach inst {ninsts} {{
    '''
                        for i,j in zip(sportnetsx,mappinsx):
                            upf_lines += f'''
        connect_supply_net {i} -ports $inst/{j}
    '''
                        upf_lines += f'''
    }}
    '''
            else:
                upf_warn(f'{prow} SupplyPortNet or MapSupplyList is empty. Please check it.')

        return upf_lines

    def port_supply(self,blkalias,prows,pdict):
         
        upf_lines = '' 
        kwds = ['ApplyPorts','Elements','ExcludeList','DriverSupply','ReceiverSupply','Attribute','Comment']        
        # support ports, elements
        for prow in prows:
            rowdict = pdict[prow]

            cell = []
            for kwd in kwds:
                cell.append(rowdict[kwd])
            drct,eles,exd,drv,rec,attr,cmt = cell
            
            if eles =='.':
                eles  = '{.}'

            cmd_line = ''
            if drct:
                cmd_line += f'-applies_to {drct} '
            if eles:
                cmd_line += f'-ports $UPFVAR({eles},${{{blkalias}}}) '
            if exd:
                cmd_line += f'-exclude_ports $UPFVAR({exd},${{{blkalias}}}) '
            if drv:
                cmd_line += f'-driver_supply SS_${{{blkalias}}}_{drv} '
            if rec:
                cmd_line += f'-receiver_supply SS_${{{blkalias}}}_{rec} '
            if attr:
                cmd_line += f'{attr} '

            if (drct or eles or exd) and (drv or rec):
                upf_lines += f'''
# line: {prow}
set_port_attributes {cmd_line}
'''      
            
        return upf_lines
        


    def check_sheet(self):
        pass   





