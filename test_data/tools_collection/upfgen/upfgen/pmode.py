import sys
import time
import os
import re
import  openpyxl

from openpyxl import worksheet 
from pprint import pprint 
import pandas as pd
from openpyxl.utils import get_column_letter 

import tkinter as tk

from openpyxl.styles import Border, Side, PatternFill, Alignment 
from openpyxl.worksheet.datavalidation import DataValidation

from .baseupf import *

class PModeSheet(BaseSheet):
    def __init__(self,*args):
        super().__init__(*args)
        self._pmdata = {}
        


    def update_sheet(self):
        
        sheet = self.get_sheet()

        supply_kw, supply_vol, supply_vss = self.get_supply_infos()

        # find PMMODE table 
        start_rowg = self.find_sheet(sheet, 'PMMODE') 
        for i in range(1, len(supply_kw) + 1):
            sheet.cell(start_rowg, i + 1).value = supply_kw[i-1]
        sheet.cell(start_rowg, len(supply_kw) + 2).value = 'Comment'
        self.cell_style1(sheet, [start_rowg, 2], [start_rowg, len(supply_kw) + 2])

        self.cell_style2(sheet, [start_rowg + 1, 2], [start_rowg + 10, len(supply_kw) + 2])

        self.add_dropdown(sheet, '"' + ','.join(supply_vol) + '"', [start_rowg + 1, 2], [start_rowg + 10, len(supply_kw) + 2])
        self.add_dropdown(sheet, '"-update"', [start_rowg + 1, len(supply_kw) + 2], [start_rowg + 10, len(supply_kw) + 2])

    def read_data(self):
        sheet = self.get_sheet()
        self._pmdata = self.get_table_contxt(sheet)
        # print('_pmdata: ', self._pmdata)

    def dump_json(self,json_file):
        self._data = self._pmdata
        self.write_json(json_file)

    def change_sheet(self):
        pass

    def write_upf(self,mdname,blkalias,blklvl,upf_file):
        sheet = self.get_sheet()

        # mdname = self._upfdg._vfile_data['module_name']
        # blkalias = self._upfdg._hier_tree._blocks[mdname].alias
        # blklvl = self._upfdg._hier_tree._blocks[mdname].hdlevel

        supply_kw, supply_vol, supply_vss, supply_data = self.get_supply_infos()
        # supply_kw  = supply_kw.extend(supply_vss)
        # print(' supply_vss, supply_data: ',  supply_vss, supply_data)
        # supply_vss, supply_data: []
        # {'VDD_CORE': '0.75v 0.7v 0.65v',
        #  'VDD_MM_CSS': '0.65v 0.6v 0.55v PSO1',
        #  'VDDM_CLPS': '0.8v 0.75v 0.7v PSO2'}

        pmodedict, pmodekeys= self.get_rows(self._pmdata,'PMMODE_Row','PMName','PMName')
        upf_lines = f'''
# ========================================= #
# Add power state table
# ========================================= #
'''    

        for key,val in supply_data.items():
            keyg = ''.join(key.split('_'))
            if not key in supply_vss:
                if val:
                    # print('val: ',val)
                    vlgx = val.split(' ') if ' ' in val else val.split()
                    vlg = vlgx[0].replace('.','P')
                    upf_lines += f'''
# Add power state for SS_${{{blkalias}}}_{keyg}
add_power_state SS_${{{blkalias}}}_{keyg}    \\
        -state {{  {vlg}   \\
        -supply_expr {{power == {{FULL_ON {vlgx[0]}}}}} && ground == {{FULL_ON 0.0}}}} \\
        -simstate  NORMAL \\
    }}
'''
                if len(vlgx) > 1:
                    nval = vlgx[1:]
                    for nv in nval:
                        if re.search(r'PSO\d+',nv):
                            upf_lines += f'''
add_power_state SS_${{{blkalias}}}_{keyg}    \\
        -state {{  OFF      \\
        -supply_expr {{power == {{OFF}} && ground == {{FULL_ON 0.0}}}}  \\
        -simstate CORRUPT  \\
    }}   \\
    -update
'''
                        else:
                            vlg = nv.replace('.','P')
                            upf_lines += f'''
add_power_state SS_${{{blkalias}}}_{keyg}    \\
        -state {{  {vlg}   \\
        -supply_expr {{power == {{FULL_ON {nv}}}}} && ground == {{FULL_ON 0.0}}}} \\
        -simstate  NORMAL \\
    }}  \\
    -update
'''
        
        upf_lines += self.add_pmode(blkalias,pmodedict,pmodekeys)      

        self.save_text(upf_lines,upf_file)

    def add_pmode(self,blkalias,pdict,prows):
        upf_lines = ''
        i = 0

        # pdsht = self._upfdg._sheets['PDomain']
        pdsht = self._upfdg._wb['PDomain']
        # print('pdsht: ', pdsht)
        pdname = self.get_pdname(blkalias,pdsht)
        # print('pdname: ', pdname)
        for vl in pdname.values():
            if re.search(r'PD1_',vl):
                pdnm = vl

        upf_lines += f'''
add_power_state     {pdnm}    \\'''
        # upf_lines.strip()
        # arg_line = ''
        for prow in prows:
            rowdict = pdict[prow]
            # print('rowdict: ', rowdict)
            i += 1
            pmnm = f'PM{i}_${{{blkalias}}}_' + f'{rowdict["PMName"]}'
            upf_lines += f'''
                    -state {pmnm}  -logic_expr {{
'''
            for key,val in rowdict.items():
                if key not in ('PMName', 'Comment'):
                    keyg = ''.join(key.split('_'))                    
                    pss = f'SS_${{{blkalias}}}_{keyg}'
                    if val not in ['off','0v']:
                        valg = val.replace('.','P')
                    else:
                        valg = 'OFF'
                    upf_lines += f'  {pss} == {valg} && '
            #i += 1

            if len(prows) == i - 1:
                upf_lines = upf_lines.rstrip('&&') + f'}} '
            else:
                upf_lines = upf_lines.rstrip('&&') + f'}} \\'

        # arg_line.lstrip()
        # upf_lines += arg_line
        
        return upf_lines



    def check_sheet(self):
        pass 



