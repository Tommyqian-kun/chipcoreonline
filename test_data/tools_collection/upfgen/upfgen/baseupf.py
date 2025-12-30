

import json
import os
import re
import sys
from os.path import dirname
import time

import yaml
#from itertools import chain

from openpyxl import worksheet 
from pprint import pprint 
import pandas as pd
from openpyxl.utils import get_column_letter 

import tkinter as tk

from openpyxl.styles import Border, Side, PatternFill, Font, Alignment 
from openpyxl.worksheet.datavalidation import DataValidation

from com.base import *

'''
@ define BaseSheet for design guide sheets
@ define BaseInputs to parse vfile, pmfile and objfile 
@ print message severity
'''


# upfdg is XupfDesignGuide object
class BaseSheet(object):
    def __init__(self, upfdg, sheetname):
        self._upfdg = upfdg
        self._sheetname = sheetname
        self._data = []
        self._vardef = {}
        self._pdnmdict = {}
    
    def get_sheet(self):
        return self._upfdg._wb[self._sheetname]

    def read_data(self):
        raise NotImplementedError(self.__class__.__name__ + ' raad_data not implemented yet')

    def write_json(self, filepath):
        os.makedirs(dirname(filepath), exist_ok=True)
        jsonstr = json.dumps(self._data, indent=4)
        with open(filepath,'w') as fw:
            print(jsonstr, file=fw)

    def find_sheet(self, sheet, skw):
        start_rowg = 1
        # TABCONST = ['PMVAR','PMCELL','PMHIER','PMDOMAIN','PMNETWORK','PMBOUNDARY','PMISO','PMLS','PMRET','PMPSW','PMRPT','PMMODE']
        TABCONST = ['PMVAR','PMCELL','PMDOMAIN','PMNETWORK','PMBOUNDARY','PMISO','PMLS','PMRET','PMPSW','PMMODE']
        for i in range(1,sheet.max_row+1):
            if skw in TABCONST and sheet.cell(i,1).value == skw:
                start_rowg = i + 1
                break  
        return  start_rowg 

    def get_vardef_value(self, sheet):

        start_rowg = self.find_sheet(sheet, 'PMVAR')
        # end_rowg = self.find_sheet(sheet, 'PMHIER')
        end_rowg = self.find_sheet(sheet, 'PMCELL')
        for i in range(start_rowg + 1, end_rowg-1):
            key = sheet.cell(row=i, column=1).value
            val = sheet.cell(row=i, column=2).value
            self._vardef[key] = val

        self._vardef['UPF_DIR'] = self._upfdg._upfdir
        self._vardef['COM_DIR'] = self._upfdg._upfdir
        self._vardef['UPF_VERSION'] = '2.1'
        # self._vardef['BOUNDARY_MODE'] = 'lower'
        # self._vardef['HD_MOD_NAME'] = self._upfdg._mdname
        # self._vardef['HD_PROCESS'] = ''
        self._vardef['SS_MODE'] = 'full'
        self._vardef['SCOPE_TYPE'] = 'parent; # self'

        return self._vardef   

    def set_name_style(self, kw):
        #time_stamp = time.strftime("%Y%m%d%H%M%S", time.localtime())
        #CONST = f'Generic_Xupf_{time_stamp}'
        CONST = f'Generic_XUPF'
        return kw + '_' + CONST

    # def change_space(self, dgfile):

    #     #sheet = self.get_sheet()
    #     shtname = self._sheetname
    #     sheet = self._upfdg._wb[shtname]

    #     # Find variable start row num, below "Variable" header 
    #     start =0       
    #     df = pd.read_excel(dgfile, sheet_name=shtname, engine='openpyxl') 
    #     df.loc[len(df)] = list(df.columns)
    #     for col in df.columns:
    #         index = list(df.columns).index(col)
    #         letter = get_column_letter(index + 1)
    #         collen = df[col].apply(lambda x: len(self.max_str(str(x)).encode())).max()
    #         # sheet.column dimensions[letter).width = collen*o.9 
    #         sheet.column_dimensions[letter].width = collen * 1.05

    def max_str(self, li):
        max = 0
        max_str = ''
        try:
            for i in li.split('\n'):
                if len(i) > max:
                    max =len(i)
                    max_str=i 
            return max_str 
        except:
            return li
        


    # showErrorMessage=False,showDropDown=True
    def add_dropdown(self, sheet, options, start, end):       
        dv = DataValidation(type="list", formula1=options, showErrorMessage=False)
        sheet.add_data_validation(dv)
        if len(start) == 2 and len(end) == 2:
            for i in range(start[0], end[0] + 1):
                for j in range(start[1], end[1] + 1):
                    dv.add(sheet.cell(i,j))
        if len(start) == 1 and len(end) == 1:
            dv.add(sheet.cell(start[0],end[0]))


    def cell_style1(self, sheet, start, end):
        border=Border(left=Side(border_style='thin', color='000000'),
                      right=Side(border_style='thin', color='000000'),
                      top=Side(border_style='thin', color='000000'),
                      bottom=Side(border_style='thin', color='000000'))
        #bgfill = PatternFill(fill_type='solid', start_color='fff2cc', end_color='fff2cc') 
        #bgfill = PatternFill(fill_type = 'solid', start_color='197e00',end_color='197e00')
        bgfill = PatternFill(fill_type = 'solid', start_color='FF385724',end_color='FF333300')
        font = Font(name='等线', size=11, color='FFFFFF')
        for i in range(start[0], end[0] + 1):
            for j in range(start[1], end[1] + 1):
                sheet.cell(i,j).border=border 
                sheet.cell(i,j).fill=bgfill
                sheet.cell(i,j).font=font
                sheet.cell(i,j).alignment = Alignment(horizontal='left', vertical='center',wrapText=True) 
                sheet.cell(i,j).alignment = Alignment(horizontal='left', vertical='center',wrapText=True) 
                sheet.cell(i,j).alignment=Alignment(horizontal='left', vertical='center') 

    def cell_style2(self, sheet, start, end):
        border=Border(left=Side(border_style='thin', color='000000'),
                      right=Side(border_style='thin', color='000000'),
                      top=Side(border_style='thin', color='000000'),
                      bottom=Side(border_style='thin', color='000000'))
        #bgfill = PatternFill(fill_type='solid', start_color='fff2cc', end_color='fff2cc') 
        #bgfill = PatternFill(fill_type = 'solid', start_color='197e00',end_color='197e00')
        bgfill = PatternFill(fill_type = 'solid', start_color='FFFFFF',end_color='FFFFFF')
        #font = Font(name='等线', size=11, color='FFFFFF')
        for i in range(start[0], end[0] + 1):
            for j in range(start[1], end[1] + 1):
                sheet.cell(i,j).border=border 
                sheet.cell(i,j).fill=bgfill
                #sheet.cell(i,j).font=font
                sheet.cell(i,j).alignment = Alignment(horizontal='left', vertical='center',wrapText=True) 
                sheet.cell(i,j).alignment = Alignment(horizontal='left', vertical='center',wrapText=True) 
                sheet.cell(i,j).alignment=Alignment(horizontal='left', vertical='center',wrapText=True)

    def get_supply_infos(self):

        delkeys = ['module_name', 'ISO_CTRL', 'RET_SAVE', 'RET_RES', 'PSO_CTRL', 'PSO_ACK']
        supply_datag = self._upfdg._vfile_data
        supply_data = {}

        for ky,vl in supply_datag.items():
            if ky not in delkeys:
                supply_data[ky] = vl

        # print('supply_data: ', supply_data)
        supply_kw = []
        supply_val = []
        supply_vss = []
        supply_tmp = ''
        for key,val in supply_data.items():
            if '0v' in val or '0.0v' in val:
                supply_vss.append(key)
                upf_info(f'Ground pin is {key}')
            elif 'PSO' in val:
                supply_kw.append(key)
                for i in range(1, int(val[-1]) + 1):
                    supply_kw.append(key + f'_PSW{i}')
                supply_tmp = supply_tmp + ' ' + supply_data[key].split('PSO')[0].strip()
            else:
                supply_kw.append(key)
                supply_tmp = supply_tmp + ' ' + supply_data[key].strip()

        #print(supply_tmp.strip().split(','))
        float_list = [float(x.strip('v')) for x in supply_tmp.strip().split()]
        unique_floats = set(float_list)
        sorted_floats = sorted(unique_floats, reverse=True)
        supply_val = [str(x) + 'v' for x in sorted_floats]
        supply_val.append('off')
        supply_val.append('0v')

        return supply_kw,supply_val,supply_vss,supply_data
    

    def get_ctl_sig(self, ctsig):
        ctrl = []
        for ct in ctsig:
            if re.search(r'\[\d+:\d+\]', ct):
                sig = ct.split('[')[0].strip()
                st = int(ct.split(':')[0].strip()[-1])
                ed = int(ct.split(':')[1].strip()[0])
                for i in range(ed,st+1):
                    ctrl.append(sig + '[' + str(i) + ']')
            else:
                ctrl.append(ct)
        return ctrl

    def get_table_loc(self,sheet, shnm='') -> dict:

        #TABCONSTT = ['PMVAR','PMCELL','PMHIER','PMDOMAIN','PMNETWORK','PMBOUNDARY','PMISO','PMLS','PMPSW','PMRET','PMRPT','PMMODE']

        if not shnm:
            sheetname = self._sheetname
        else:
            sheetname = shnm

        if sheetname == 'VarDef':
            TABCONST = ['PMVAR','PMCELL']
        if sheetname == 'PDomain':
            TABCONST = ['PMDOMAIN','PMNETWORK','PMBOUNDARY']
        if sheetname == 'PStrategy':
            TABCONST = ['PMISO','PMLS','PMPSW','PMRET']
        if sheetname == 'PMode':
            TABCONST = ['PMMODE']

        # row_start max_col              
        row_start = ''
        max_row = ''
        max_col = ''
        # row_start max_col max_row
        table_row_loc = {}
        for kw in TABCONST:
            strow = self.find_sheet(sheet, kw)
            row_start = str(strow)
            for i in range(1,sheet.max_column + 1):
                if sheet.cell(strow,i).value == 'Comment':
                    max_col = str(i)
                    break

            if kw in ['PMCELL','PMBOUNDARY','PMRET','PMMODE']:
                #table_row_loc[kw] = row_start[kw] + ' ' + str(int(row_start[kw].split()[0]) + 20)
                table_row_loc[kw] = row_start + ' ' + str(sheet.max_row + 2) + ' ' + max_col
            else:
                idx = TABCONST.index(kw) + 1
                # print('dfsg: ', TABCONST,idx)
                max_row = self.find_sheet(sheet,TABCONST[idx]) - 1
                #table_row_loc[kw] = row_start[kw] + ' ' + str(int(row_start[TABCONST[idx]].split()[0]) - 2)
                table_row_loc[kw] = row_start + ' ' + str(max_row) + ' ' + max_col

        return table_row_loc
  
    def get_table_contxt(self,sheet, shnm='', tabnm=[]) -> dict:
        # row_start max_col max_row
        tab_loc = self.get_table_loc(sheet,shnm)
        print('sheet: tab_loc: ', tab_loc)

        if not shnm:
            sheetname = self._sheetname
        else:
            sheetname = shnm

        #TABCONST = ['PMVAR','PMCELL','PMHIER','PMDOMAIN','PMNETWORK','PMBOUNDARY','PMISO','PMLS','PMRET','PMPSW','PMRPT','PMMODE']
        if sheetname == 'VarDef':
            # TABCONST = ['PMVAR','PMCELL','PMHIER']
            if not tabnm:
                TABCONST = ['PMVAR','PMCELL']
            else:
                TABCONST = tabnm
        if sheetname == 'PDomain':
            if not tabnm:
                TABCONST = ['PMDOMAIN','PMNETWORK','PMBOUNDARY']
            else:
                TABCONST = tabnm
        if sheetname == 'PStrategy':
            # TABCONST = ['PMISO','PMLS','PMRET','PMPSW','PMRPT']
            if not tabnm:
                TABCONST = ['PMISO','PMLS','PMRET','PMPSW']
            else:
                TABCONST = tabnm
        if sheetname == 'PMode':
            if not tabnm:
                TABCONST = ['PMMODE']
            else:
                TABCONST = tabnm

        table_contxt = {}
        #row_contxt = {}
        if TABCONST:
            for kw in TABCONST:
                start_row = int(tab_loc[kw].split(' ')[0])
                end_row = int(tab_loc[kw].split(' ')[1])
                end_col = int(tab_loc[kw].split(' ')[2])
                if kw == 'PMVAR':
                    for i in range(start_row,end_row+1):
                        key = sheet.cell(i+1,1).value
                        val = str(sheet.cell(i+1,2).value)
                        if key:
                            table_contxt[key] = val.strip()
                        # print('PMVARdfd: ', table_contxt)
                        # if key and val:
                        #     table_contxt[key] = val
                else:
                    table_contxt.update(self.get_row_txt(sheet,kw,start_row,end_row,end_col))

        return table_contxt

    def get_row_txt(self, sheet, kw, start_row, end_row, end_col):
        row_contxt = {}
        table_contxt = {}
        for i in range(1, end_row - start_row):
            for j in range(1, end_col + 1):
                key = sheet.cell(start_row, j).value
                val = sheet.cell(start_row + i, j).value
                val_col1 = sheet.cell(start_row + i, 1).value
                if val_col1:
                    if re.search(r'^#', val_col1.strip()):
                        continue
                if key:     key = str(key).strip()
                if val:     val = str(val).strip()
                row_contxt[key] = val
                # if key and val:
                #     row_contxt[key] = val

            all_none = all(ele is None for ele in list(row_contxt.values()))
            if not all_none and row_contxt:
                table_contxt[f'{kw}_Row{start_row + i}'] = row_contxt
            row_contxt = {}
            # for key in table_contxt.keys():
            #     if 'TMCLK' in key:
            #         print(table_contxt)

        return table_contxt


    def get_impl_obj(self, sheet, start_rowg, kwd):
        keywds = list(self._upfdg._objfile_data.keys())
        TABCONST = ['PMVAR','PMCELL','PMHIER','PMDOMAIN','PMNETWORK','PMBOUNDARY','PMISO','PMLS','PMPSW','PMRET','PMRPT','PMMODE']        
        
        supply_kw, supply_vol,supply_vss, supply_data = self.get_supply_infos()
        supply_kw = supply_kw.extend(supply_vss)
        notvss = [x for x in supply_vss if not x in ['VSS']]

        #print(keywds)

        if kwd == 'PMNETWORK':
            # SupplyPortNet	NPwellNet	InstList	MapSupplyList	Comment
            row_tmp = start_rowg
            row_tot = start_rowg + 10
            if len(supply_kw) > 8:
                sheet.insert_rows(start_rowg + 3, 8)
                row_tot += 8
                #self.cell_style2(sheet,[start_rowg + 3,1], [start_rowg + 12,5])
            if len(keywds) > 4:
                sheet.insert_rows(start_rowg + 11, 6)
                row_tot += 6
            self.cell_style2(sheet,[start_rowg + 1,1], [row_tot + 1,5])

            virpwr = [x for x in supply_kw if re.search(r'_PSW\d+',x)]
            relpwr = [x for x in supply_kw if not re.search(r'_PSW\d+',x)]
            cmt = ['PAL','PAL OUT','OUT','SNE1','SNE2','SNE3','SNE4','SNE5']
            if len(notvss) > 0:
                ncmt = notvss.extend(cmt)
            else:
                ncmt = cmt
            #print(cmt)
            self.add_dropdown(sheet, '"' + ','.join(supply_kw) + '"', [start_rowg + 1, 1], [row_tot + 1 , 1])
            self.add_dropdown(sheet, '"' + ','.join(relpwr) + '"', [start_rowg + 1, 2], [row_tot + 1 , 2])
            self.add_dropdown(sheet, '"' + ','.join(ncmt) + '"', [start_rowg + 1, 5], [row_tot + 1, 5])
            
            for i in range(1, len(supply_kw) + 1):               
                sheet.cell(start_rowg + i,1).value = supply_kw[i-1]
                if re.search(r'_PSW\d+',supply_kw[i-1]):
                    nwell = supply_kw[i-1].split('_PSW')[0].strip()
                    sheet.cell(start_rowg + i,2).value = nwell
                else:
                    sheet.cell(start_rowg + i,2).value = supply_kw[i-1]
                row_tmp += 1
            #print(row_tmp)

            for ky in keywds:
                #row_tmp += 1
                if '_conspy_insts' in ky:
                    row_tmp += 1
                    nky = ky.split('_conspy_insts')[0].strip()
                    sheet.cell(row_tmp,1).value = self._upfdg._objfile_data[nky + '_outer_spy']
                    sheet.cell(row_tmp,3).value = nky + '_conspy_insts' #self._upfdg._objfile_data[nky + '_conspy_insts']
                    #self.add_dropdown(sheet,'"' + ky + '"',)
                    sheet.cell(row_tmp,4).value = self._upfdg._objfile_data[nky + '_inner_spy']
                if '_conspy_hinsts' in ky:
                    row_tmp += 1
                    nky = ky.split('_conspy_hinsts')[0].strip()
                    sheet.cell(row_tmp,1).value = self._upfdg._objfile_data[nky + '_outer_spy']
                    #sheet.cell(row_tmp,3).value = self._upfdg._objfile_data[nky + '_conspy_hinsts']
                    sheet.cell(row_tmp,3).value = nky + '_conspy_hinsts'
                    sheet.cell(row_tmp,4).value = self._upfdg._objfile_data[nky + '_inner_spy']
            
            row_tmp = 0

        if kwd == 'PMBOUNDARY':
            bd_ele = []
            bd_exd = []
            flg_ele = 0
            flg_exd = 0
            for ky in keywds:
                if re.search(r'_spa_inport|_spa_outport|_spa_inhpin|_spa_outhpin', ky):
                    bd_ele.append(ky)
                    flg_ele = 1
                if re.search(r'_exdspa_inport|_exdspa_outport|_exdspa_inhpin|_exdspa_outhpin', ky):
                    bd_exd.append(ky)
                    flg_exd = 1
                if re.search(r'_spa_in$|_spa_out$|_spa_input|_spa_output|_spa_inpin|_spa_outpin', ky):
                    bd_ele.append(ky)
                    flg_ele = 1
                if re.search(r'_exdspa_in$|_exdspa_out$|_exdspa_input|_exdspa_output|_exdspa_inpin|_exdspa_outpin', ky):
                    bd_exd.append(ky)
                    flg_exd = 1                    
            if not flg_ele:     bd_ele = None
            if not flg_exd:     bd_exd = None

            return bd_ele,bd_exd

        if kwd == 'PMISO':
            iso_ele = []
            iso_exd = []
            iso_no = []
            flg_ele = 0
            flg_exd = 0
            flg_no = 0
            for ky in keywds:
                if re.search(r'_iso_inport|_iso_outport|_iso_inhpin|_iso_outhpin|_ctliso_inport|_ctliso_inhpin|_fdthiso_inport|_fdthiso_outport|_fdthiso_inhpin|_fdthiso_outhpin', ky):
                    iso_ele.append(ky)
                    flg_ele = 1
                if re.search(r'_exdiso_inport|_exdiso_outport|_exdiso_inhpin|_exdiso_outhpin', ky):
                    iso_exd.append(ky)
                    flg_exd = 1
                if re.search(r'_noiso_inport|_noiso_outport|_noiso_inhpin|_noiso_outhpin', ky):
                    iso_no.append(ky)
                    flg_no = 1

                if re.search(r'_iso_in$|_iso_out$|_ctliso_in$|_fdthiso_in$|_fdthiso_out$', ky):
                    iso_ele.append(ky)
                    flg_ele = 1
                if re.search(r'_exdiso_in$|_exdiso_out$', ky):
                    iso_exd.append(ky)
                    flg_exd = 1
                if re.search(r'_noiso_in$|_noiso_out$', ky):
                    iso_no.append(ky)
                    flg_no = 1

                if re.search(r'_iso_input|_iso_output|_ctliso_input|_fdthiso_input|_fdthiso_output', ky):
                    iso_ele.append(ky)
                    flg_ele = 1
                if re.search(r'_iso_inpin|_iso_outpin|_ctliso_inpin|_fdthiso_inpin|_fdthiso_outpin', ky):
                    iso_ele.append(ky)
                    flg_ele = 1                    
                if re.search(r'_exdiso_input|_exdiso_output|_exdiso_outpin', ky):
                    iso_exd.append(ky)
                    flg_exd = 1
                if re.search(r'_noiso_input|_noiso_output|_noiso_outpin', ky):
                    iso_no.append(ky)
                    flg_no = 1

            if not flg_ele:     iso_ele = None
            if not flg_exd:     iso_exd = None
            if not flg_no:      iso_no = None

            return iso_ele,iso_exd,iso_no
                

        if kwd == 'PMLS':
            ls_ele = []
            ls_exd = []
            ls_no = []
            flg_ele = 0
            flg_exd = 0
            flg_no = 0                
            for ky in keywds:
                if re.search(r'_ls_inport|_ls_outport|_ls_inhpin|_ls_outhpin|_fdthls_inport|_fdthls_outport|_fdthls_inhpin|_fdthls_outhpin', ky):
                    ls_ele.append(ky)
                    flg_ele = 1
                if re.search(r'_exdls_inport|_exdls_outport|_exdls_inhpin|_exdls_outhpin', ky):
                    ls_exd.append(ky)
                    flg_exd = 1
                if re.search(r'_nols_inport|_nols_outport|_nols_inhpin|_nols_outhpin', ky):
                    ls_no.append(ky) 
                    flg_no = 1

                if re.search(r'_ls_in$|_ls_out$|_fdthls_in$|_fdthls_out$|_ls_input|_ls_output|_fdthls_input|_fdthls_output|_ls_inpin|_ls_outpin|_fdthls_inpin|_fdthls_outpin', ky):
                    ls_ele.append(ky)
                    flg_ele = 1
                if re.search(r'_exdls_in$|_exdls_out$|_exdls_input|_exdls_output|_exdls_inpin|_exdls_outpin', ky):
                    ls_exd.append(ky)
                    flg_exd = 1
                if re.search(r'_nols_in$|_nols_out$|_nols_input|_nols_output|_nols_inpin|_nols_outpin', ky):
                    ls_no.append(ky) 
                    flg_no = 1                     
            if not flg_ele:     iso_ele = None
            if not flg_exd:     iso_exd = None
            if not flg_no:      iso_no = None

            return ls_ele,ls_exd,ls_no
        
        if kwd == 'PMPSW':
            psw_ele = []
            psw_exd = []
            flg_ele = 0
            flg_exd = 0
            for ky in keywds:
                if re.search(r'_ctlpsw_inport|_ctlpsw_inhpin', ky):
                    psw_ele.append(ky)
                    flg_ele = 1
                if re.search(r'_ackpsw_inport|_ackpsw_inhpin', ky):
                    psw_exd.append(ky)
                    flg_exd = 1

                if re.search(r'_ctlpsw_in$|_ctlpsw_input|_ctlpsw_inpin', ky):
                    psw_ele.append(ky)
                    flg_ele = 1
                if re.search(r'_ackpsw_in$|_ackpsw_input|_ackpsw_inpin', ky):
                    psw_exd.append(ky)
                    flg_exd = 1
            if not flg_ele:     psw_ele = None
            if not flg_exd:     psw_exd = None    

            return psw_ele,psw_exd

        if kwd == 'PMRET':
            ret_save = []
            ret_res = []
            ret_ele = []
            ret_exd = []
            ret_no = []
            flg_save = 0
            flg_res = 0
            flg_ele = 0
            flg_exd = 0
            flg_no = 0                
            for ky in keywds:
                if re.search(r'_saveret_inport|_saveret_inhpin', ky):
                    ret_save.append(ky)
                    flg_save = 1
                if re.search(r'_resret_inport|_resret_inhpin', ky):
                    ret_res.append(ky)
                    flg_res = 1
                if re.search(r'_ret_insts|_ret_hinsts', ky):
                    ret_ele.append(ky)
                    flg_ele = 1
                if re.search(r'_exdret_insts|_exdret_hinsts', ky):
                    ret_exd.append(ky)
                    flg_exd = 1
                if re.search(r'_noret_insts|_noret_hinsts', ky):
                    ret_no.append(ky)
                    flg_no = 1

                if re.search(r'_saveret_in$|_saveret_input|_saveret_inpin', ky):
                    ret_save.append(ky)
                    flg_save = 1
                if re.search(r'_resret_in$|_resret_input|_resret_inpin', ky):
                    ret_res.append(ky)
                    flg_res = 1


            if not flg_save:    ret_save = None
            if not flg_res:     ret_res = None 
            if not flg_ele:     ret_ele = None
            if not flg_exd:     ret_exd = None
            if not flg_no:      ret_no = None

            return ret_save,ret_res,ret_ele,ret_exd,ret_no

        if kwd == 'PMRPT':
            rpt_ele = []
            rpt_exd = []
            flg_ele = 0
            flg_exd = 0
            for ky in keywds:
                if re.search(r'_rpt_inport|_rpt_outport|_rpt_inhpin|_rpt_outhpin', ky):
                    rpt_ele.append(ky)
                    flg_ele = 1
                if re.search(r'_exdrpt_inport|_exdrpt_outport|_exdrpt_inhpin|_exdrpt_outhpin', ky):
                    rpt_exd.append(ky)
                    flg_exd = 1

            for ky in keywds:
                if re.search(r'_rpt_in$|_rpt_out$|_rpt_input|_rpt_output|_rpt_inpin|_rpt_outpin', ky):
                    rpt_ele.append(ky)
                    flg_ele = 1
                if re.search(r'_exdrpt_in$|_exdrpt_out$|_exdrpt_input|_exdrpt_output|_exdrpt_inpin|_exdrpt_outpin', ky):
                    rpt_exd.append(ky)
                    flg_exd = 1

            if not flg_ele:     rpt_ele = None
            if not flg_exd:     rpt_exd = None     

            return rpt_ele,rpt_exd
        
    def save_text(self, context,file):
        with open(file, 'w') as fw:
            fw.write(context)

    def get_rows(self,pmdata,keyrow,kwd,ckwd):
        pmdict = {}
        pmkeys = []
        pmlist = []
        tab = keyrow.split('_')[0]
        # print('pmdata: ', pmdata)

        # pmlist = [(key, val) for key, val in pmdata.items() if keyrow in key and not re.search(r'^#',val[f'{ckwd}'].strip()) and val[f'{kwd}']]
        for key, val in pmdata.items():
            if keyrow in key and not re.search(r'^#', val[f'{ckwd}'].strip()):
                if '|' in kwd:
                    kwlst = kwd.split('|')
                    kflg = False
                    for kl in kwlst:
                        if val[f'{kl}']: kflg = True
                    if kflg:
                        pmlist.append((key, val))
                else:
                    if val[f'{kwd}']:
                        pmlist.append((key, val))

        if pmlist:
            for k,v in pmlist:
                pmdict[k] = v
                pmkeys.append(k)
            pmkeys.sort(key=lambda x: int(x.split("Row")[1]))
        else:
            print(f'The table {tab} is empty.')

        # print(f'real tabale {tab}', pmkeys,pmdict)

        return pmdict, pmkeys     

    def get_pdname(self,blkalias,sheet):
        i = 0
        pddic = {}

        # print('get_pdname--sheet: ', sheet)
        contxt = self.get_table_contxt(sheet,'PDomain',['PMDOMAIN'])
        pdkeys = list(contxt.keys())
        # print('pdfsmcxc: ', contxt)
        pdict, prows = self.get_rows(contxt,'PMDOMAIN_Row','PDName','PDName')
        # print('pdict: prows', pdict, prows)
        if prows:
            for prow in prows:
                rowdict = pdict[prow]
                pdname = rowdict['PDName']
                i += 1

                prmdic = [(key,val) for key,val in rowdict.items() if val == 'PRM']
                if len(prmdic) > 1:
                    upf_error(f'{prow} has two or more primary power. Please check it.')
                    prm = 'NOPRM'
                else:
                    prm = prmdic[0][0]

                nprm = ''.join(prm.split('_'))
                self._pdnmdict[f'{pdname}'] = f'PD{i}_${{{blkalias}}}_{nprm}_{pdname}'
        else:
            upf_error(f'Can not find power domain definition. Please check it.')

        return self._pdnmdict




class BaseInputs(object):
    def __init__(self):
        self.vfile_data = {}
        self.pmfile_data = {}
        self.objfile_data = {}
               
    def read_vfile(self, vfile) -> dict:

        lines = self.read_text(vfile)
        for line in lines:
            line = line.replace('\n','').replace('\r','').replace('\t',' ').strip()
            if re.search(r'^\/\/', line):
                continue
            if re.search(r'^module', line):
                self.vfile_data['module_name']= re.split(' +',line)[1].strip().replace('(','')
            if re.search(r'\/\/#(\d+[vV]#|\d+.\d+[vV]#|\d+.\d+[vV]:)', line):
                #print(line)
                port_pwrg = re.split('//#',line)[0].strip().replace(',', '')
                port_pwrg = [x for x in re.split(' ', port_pwrg) if x]
                # port_pwr = re.split(' ', port_pwrg)[1].strip()
                port_pwr = port_pwrg[1]
                port_volx = re.split('//#',line)[1].strip() # maybe include 'PSO'
                if re.search(r'#PSO \d+#', port_volx):                    
                    #port_volg = ' PSO' + re.split(r'#PSO ',port_volx)[1].replace('#','')
                    port_volg = ' ' + re.split(r'#',port_volx)[1].strip().replace(' ','')
                    port_volx = re.split(r'#PSO ', port_volx)[0].strip()
                    #print(port_volg)
                else:
                    port_volg = ''

                if re.search(r':', port_volx):
                    port_vol = ' '.join(re.split(':',port_volx)).strip().lower().replace('#', '') + port_volg
                else:
                    port_vol = port_volx.lower().replace('#', '') + port_volg
                self.vfile_data[port_pwr] = port_vol
       
        self.vfile_data['ISO_CTRL'] = self.get_ctrl_ports(lines, 'ISO_CTRL')
        self.vfile_data['RET_SAVE'] = self.get_ctrl_ports(lines, 'RET_SAVE')
        self.vfile_data['RET_RES'] = self.get_ctrl_ports(lines, 'RET_RES')
        self.vfile_data['PSO_CTRL'] = self.get_ctrl_ports(lines, 'PSO_CTRL')
        self.vfile_data['PSO_ACK'] = self.get_ctrl_ports(lines, 'PSO_ACK')

        return self.vfile_data



    def read_pmfile(self, pmfile) -> dict:

        pm_data = self.read_yaml(pmfile)

        # ISOROW1, ELSROW1,
        isocells = self.get_pmcell_info(pm_data, 'ISOPowerMCell')
        lscells = self.get_pmcell_info(pm_data, 'LSPowerMCell')
        elscells = self.get_pmcell_info(pm_data, 'ELSPowerMCell')
        retcells = self.get_pmcell_info(pm_data, 'RETPowerMCell')
        pswcells = self.get_pmcell_info(pm_data, 'PSWPowerMCell')

        isocells.update(lscells)
        isocells.update(elscells)
        isocells.update(retcells)
        isocells.update(pswcells)
        self.pmfile_data = isocells

        return self.pmfile_data



    def read_objfile(self, objfile) -> dict:
        
        obj_lines = self.read_text(objfile)

        objfile_data = {}
        var_list = []
        var_nline = ''

        for line in obj_lines:
            if re.search(r'^#', line):
                var_nline += f'\n{line}'
                #var_list += line
            elif re.search(r'^set\s+', line):
                var_list.append(re.split(' +', line)[1].strip())
                if re.search(r'[|]',line):
                    nline = line.replace('[','\\[')
                    nline = nline.replace(']','\\]')
                else:
                    nline = line
                var_nline += f'\n{nline}'
            elif re.search(r'\w+[|\d+]',line):
                nline = line.replace('[','\\[')
                nline = nline.replace(']','\\]')
                var_nline += f'\n{nline}'              
            else:
                var_nline += f'\n{line}'
        #var_list = [re.split(' +', line)[1].strip() for line in obj_lines if re.search(r'^set\s+', line)]
        #print(f'objfile var list {var_list}')

        tmpfile = dirname(objfile) + '/objtmp.tcl'
        with open(tmpfile, 'w') as fw:
            fw.write(var_nline)        
        #print(var_nline)

        tcl_intp = tk.Tcl()
        tcl_intp.eval(f'source {tmpfile}')
        #tcl_intp.eval(f'source {objfile}')
        #tcl_intp.eval(' '.join(var_nline))
        
        for var in var_list:
            nvar = tcl_intp.getvar(var).strip()
            if re.search(r'\[|\]',nvar):
                nvar = nvar.replace('\\[','[')
                nvar = nvar.replace('\\]',']')
            self.objfile_data[var] = nvar

        if os.path.exists(tmpfile):
            os.system(f'rm -f {tmpfile}')

        return self.objfile_data,var_list
        


    def read_yaml(self, yaml_file):

        yaml_data = {}
        if not os.path.exists(yaml_file):
            raise FileExistsError(f'{yaml_file} does not exists')
        with open(yaml_file, 'r') as fh:
            yaml_data = yaml.load(fh, yaml.FullLoader)

        return yaml_data
        
    def get_pmcell_info(self, pmdata, kwd) -> dict:

        pm_num = len(pmdata[kwd].keys())
        kws = kwd.replace('PowerMCell', '')

        pm_data = {}
        if pm_num == 0:
            upf_warn(f'{kws} power cell not found.')
        elif pm_num == 1:
            kws = kws + 'Row1'
            pm_data[kws] = pmdata[kwd]['TypeIndex1']
        else:
            for num in range(1,pm_num + 1):
                kwn = kws + f'Row{num}'
                pm_data[kwn] = pmdata[kwd][f'TypeIndex{num}']

        return pm_data
    


    def read_text(self, file):
        if not os.path.exists(file):
            raise FileExistsError(f'{file} does not exists')
            # upf_error(f'{file} not exist. Please check it.')
            # exit(1)
        else:
            txt_list = []
            with open(file,'r') as fh:
                for line in fh.readlines():
                    #pattern = r"\[\s*(\d+)\s*:\s*(\d+)\]"
                    if line.strip() == "":
                        continue
                    # if line.strip().startswith("//"):
                    #     continue                     
                    line = re.sub(r"\[\s*(\d+)\s*:\s*(\d+)\s*\]", r"[\1:\2]", line)
                    txt_list.append(line.strip())
        
            return txt_list

    def get_ctrl_ports(self, lnlist, kwd) -> list:

        full_list = []
        stwp = ''
        stw = []

        nkw = r'//#' + f'{kwd}' + '#'
        #print(nkw)
        for line in lnlist:
            if re.search(r'^\/\/', line):
                continue
            if re.search(f'{nkw}', line):
                # kw_loc = re.split(' +',line).index(r'//#ISO_CTRL#')
                # iso_ctrl.append(re.split(' +',line)[kw_loc - 1].strip().replace(',', ' ').strip())
                #stw = re.split('//#ISO_CTRL',line)[-2].strip()
                stwp = re.split(f'{nkw}',line)[-2].strip()
                if re.search(r'\/\/$', stwp):
                    stwp.replace('//','')
                elif re.search(r'\/\/\w*\s*', stwp):
                    stwp = re.split('//', stwp)[-2].strip()

                #patn = r'\[(\d+:\d+)|(\d+\s+:\d+)|(\d+:\s+\d+)|(\s+\d+:\d+)\]'    
                if re.search(r'\[\d+:\d+\]', stwp):
                    portnum = ''.join(re.findall(r'\[\d+:\d+\]', stwp)).strip()
                    stwp = re.split(r'\[\d+:\d+\]', stwp)[-1].strip()
                    #stw = re.split('wire|logic|byte|bit|reg', stw)[-1].strip()
                    #stw = re.split('input|output', stw)[-1].strip()                 
                    stw = re.split(',', stwp)[:-1] 
                    stw = [st + portnum for st in stw]
                elif re.search(r'wire|logic|byte|bit|reg', stwp):
                    stwp = re.split('wire|logic|byte|bit|reg', stwp)[-1].strip()
                    stw = re.split(',', stwp)[:-1]
                elif re.search(r'input|output', stwp):
                    stwp = re.split('input|output', stwp)[-1].strip()
                    stw = re.split(',', stwp)[:-1].strip()

                #nested_list.append(stw)
                full_list.extend(stw)

        #return list(chain(*nested_list))
        return full_list


    # def find_start_cell_location(self,sheet,var):
    #     start = 0
    #     for i in range(1, sheet.max_row+1):
    #         if sheet.cell(i,1).value == var:
    #             start = i + 1
    #             break
    #     for i in range(start, sheet.max_row+1):
    #         key = sheet.cell(row=i, column=1).value
    #         val = sheet.cell(row=i, column=2).value
    #         self._vardef[key] = val

# def modify_line_in_file(file_path, search_pattern, replacement):
#     # 打开文件并逐行读取内容
#     with open(file_path, 'r') as file:
#         lines = file.readlines()

#     # 遍历每一行并进行匹配和替换
#     modified_lines = []
#     for line in lines:
#         if search_pattern in line:
#             modified_line = line.replace(search_pattern, replacement)
#             modified_lines.append(modified_line)
#         else:
#             modified_lines.append(line)

#     # 将修改后的内容写回文件
#     with open(file_path, 'w') as file:
#         file.writelines(modified_lines)

# #EXIT_ON_FATAL = True
# message_list =  []

# def upf_log(level, msg, out=sys.stdout):
#     print(f'{level.upper()}: {msg}', flush=True, file=out)

# def upf_info(msg):
#     if msg not in message_list:
#         message_list.append(msg)
#         upf_log('upf_info', msg)

# def upf_warn(msg):
#     if msg not in message_list:
#          message_list.append(msg)
#          print(f'\033[0:31mUPF_WARN\033[0m: {msg}', flush=True)

# def upf_error(msg):
#     if msg not in message_list:
#          message_list.append(msg)
#          print(f'\033[0:31mUPF_ERROR\033[0m: {msg}', flush=True)

# def upf_fatal(msg):
#     if msg not in message_list:
#          message_list.append(msg)
#          print(f'\033[0:31mUPF_FATAL\033[0m: {msg}', flush=True)    
           
#     sys.exit(1) 


        