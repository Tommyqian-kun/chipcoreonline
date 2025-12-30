
import sys
import time
import os
import re

from os.path import dirname, abspath, basename

import  openpyxl

from .baseupf import *

class VarDefSheet(BaseSheet):
    def __init__(self,*args):
        super().__init__(*args)  
        self._vardata = {}     

    def update_sheet(self):
        '''
        # only during -dg option
        # addition of module name value from vfile
        # addition of user_defined variables
        # addition of pmcell table from pmfile, vfile
        # addition of block hier tree expanded table from hier yaml
        '''
        sheet = self.get_sheet()

        hiertree = self._upfdg._hier_tree

        # find PMVAR table
        start_rowg = self.find_sheet(sheet, 'PMVAR')
        
        mdname = self._upfdg._vfile_data['module_name']
        if mdname:
            sheet.cell(start_rowg + 1, 2).value = mdname
        
        vardef = self.get_vardef_value(sheet)

        varlist = ['T28','T16','T7','T4']
        self.add_dropdown(sheet, '"' + ','.join(varlist) + '"', [start_rowg + 2], [2])
        varlist = ['RTL','SYN','DFT_SYN','SIM','PLA','CTS','PnR','SIGNOFF']
        self.add_dropdown(sheet, '"' + ','.join(varlist) + '"', [start_rowg + 3], [2])
        varlist = ['1','0']
        self.add_dropdown(sheet, '"' + ','.join(varlist) + '"', [start_rowg + 4], [2])
        varlist = ['full','fast']
        self.add_dropdown(sheet, '"' + ','.join(varlist) + '"', [start_rowg + 5], [2])
        varlist = ['parent','self']
        self.add_dropdown(sheet, '"' + ','.join(varlist) + '"', [start_rowg + 6], [2])
        varlist = ['lower','higher']
        self.add_dropdown(sheet, '"' + ','.join(varlist) + '"', [start_rowg + 7], [2])
        varlist = ['DC','FC','PT','VCS','ICC2','GNS','INN','TPS']
        self.add_dropdown(sheet, '"' + ','.join(varlist) + '"', [start_rowg + 8], [2])
        varlist = ['2.0','2.1','3.0']
        self.add_dropdown(sheet, '"' + ','.join(varlist) + '"', [start_rowg + 11], [2])
        varlist = ['full','local']
        self.add_dropdown(sheet, '"' + ','.join(varlist) + '"', [start_rowg + 12], [2])

        block = hiertree.get_block_by_name(mdname)
        for i in range(start_rowg, sheet.max_row + 1):
            if sheet.cell(i, 1).value == 'UPF_DIR':
                sheet.cell(i, 2).value = block.constr_dir
        
        for i in range(1,start_rowg+11):
            for j in range(1,4):
                sheet.cell(start_rowg+i,j).alignment = Alignment(horizontal='left',vertical='center',wrapText=True)


        # find PMCELL table
        start_rowg = self.find_sheet(sheet, 'PMCELL')
        # PMType	PMCtrlSig	PMCell	PMSupplyPin	PMCtrlPin	PDFunction	PathType	NameFormat Comment
        # ISO         +          +        +           +            +           *            + 
        # LS          -          +        +           -            +           *            +      
        # ELS         +          +        +           +            +           *            +
        # RET         +          +        +           +            +           *            
        # PSO         +          +        +           +            +           *            

        # iso
        isoctl = ' '.join(self._upfdg._vfile_data['ISO_CTRL'])
        #print(self._upfdg._pmfile_data.keys())
        isokw = re.findall(r'ISORow\d+', ' '.join(self._upfdg._pmfile_data.keys()))
        for i in range(1, len(isokw) + 1):
            sheet.cell(start_rowg + i, 1).value = self._upfdg._pmfile_data[f'ISORow{i}']['CellType']
            #print(self._upfdg._pmfile_data[f'ISORow{i}']['CellType'])
            sheet.cell(start_rowg + i, 2).value = isoctl
            sheet.cell(start_rowg + i, 3).value = self._upfdg._pmfile_data[f'ISORow{i}']['CellPatn']
            supplypin = self._upfdg._pmfile_data[f'ISORow{i}']['PrimaryPower'] + ' ' + self._upfdg._pmfile_data[f'ISORow{i}']['BackupPower'] + ' ' + self._upfdg._pmfile_data[f'ISORow{i}']['NwellPower'] + ' ' + self._upfdg._pmfile_data[f'ISORow{i}']['PwellPower'] + ' ' + self._upfdg._pmfile_data[f'ISORow{i}']['GroundPin']
            sheet.cell(start_rowg + i, 4).value = supplypin
            sheet.cell(start_rowg + i, 5).value = self._upfdg._pmfile_data[f'ISORow{i}']['CtrlPin']
            sheet.cell(start_rowg + i, 6).value = self._upfdg._pmfile_data[f'ISORow{i}']['PDFunction']
            sheet.cell(start_rowg + i, 7).value = self._upfdg._pmfile_data[f'ISORow{i}']['PathType']
            sheet.cell(start_rowg + i, 8).value = self.set_name_style('ISO')  # 'ISO_Generic_Xconst_09292003'
        # ls
        start_rowg = start_rowg  + len(isokw)
        #isoctl = ' '.join(self._upfdg._vfile_data['ISO_CTRL'])
        #print(self._upfdg._pmfile_data.keys())
        lskw = re.findall(r'\s+LSRow\d+', ' '.join(self._upfdg._pmfile_data.keys()))
        #print(lskw)
        for i in range(1, len(lskw) + 1):
            sheet.cell(start_rowg + i, 1).value = self._upfdg._pmfile_data[f'LSRow{i}']['CellType']
            #sheet.cell(start_rowg + i, 2).value = isoctl
            sheet.cell(start_rowg + i, 3).value = self._upfdg._pmfile_data[f'LSRow{i}']['CellPatn']
            supplypin = self._upfdg._pmfile_data[f'LSRow{i}']['PrimaryPower'] + ' ' + self._upfdg._pmfile_data[f'LSRow{i}']['BackupPower'] + ' ' + self._upfdg._pmfile_data[f'LSRow{i}']['NwellPower'] + ' ' + self._upfdg._pmfile_data[f'LSRow{i}']['PwellPower'] + ' ' + self._upfdg._pmfile_data[f'LSRow{i}']['GroundPin']
            sheet.cell(start_rowg + i, 4).value = supplypin
            #sheet.cell(start_rowg + i, 5).value = self._upfdg._pmfile_data[f'LSRow{i}']['CtrlPin']
            sheet.cell(start_rowg + i, 6).value = self._upfdg._pmfile_data[f'LSRow{i}']['PDFunction']
            sheet.cell(start_rowg + i, 7).value = self._upfdg._pmfile_data[f'LSRow{i}']['PathType']
            sheet.cell(start_rowg + i, 8).value = self.set_name_style('LS')  # 'LS_Generic_Xconst_09292003'      

        # els
        start_rowg = start_rowg + len(lskw)
        elsctl = ' '.join(self._upfdg._vfile_data['ISO_CTRL'])
        elskw = re.findall(r'ELSRow\d+', ' '.join(self._upfdg._pmfile_data.keys()))
        for i in range(1, len(elskw) + 1):
            sheet.cell(start_rowg + i, 1).value = self._upfdg._pmfile_data[f'ELSRow{i}']['CellType']
            sheet.cell(start_rowg + i, 2).value = elsctl
            sheet.cell(start_rowg + i, 3).value = self._upfdg._pmfile_data[f'ELSRow{i}']['CellPatn']
            supplypin = self._upfdg._pmfile_data[f'ELSRow{i}']['PrimaryPower'] + ' ' + self._upfdg._pmfile_data[f'ELSRow{i}']['BackupPower'] + ' ' + self._upfdg._pmfile_data[f'ELSRow{i}']['NwellPower'] + ' ' + self._upfdg._pmfile_data[f'ELSRow{i}']['PwellPower'] + ' ' + self._upfdg._pmfile_data[f'ELSRow{i}']['GroundPin']
            sheet.cell(start_rowg + i, 4).value = supplypin
            sheet.cell(start_rowg + i, 5).value = self._upfdg._pmfile_data[f'ELSRow{i}']['CtrlPin']
            sheet.cell(start_rowg + i, 6).value = self._upfdg._pmfile_data[f'ELSRow{i}']['PDFunction']
            sheet.cell(start_rowg + i, 7).value = self._upfdg._pmfile_data[f'ELSRow{i}']['PathType']
            sheet.cell(start_rowg + i, 8).value = self.set_name_style('ELS')  # 'ELS_Generic_Xconst_09292003'

        # ret
        start_rowg = start_rowg + len(elskw)
        retsave = ' '.join(self._upfdg._vfile_data['RET_SAVE'])
        retres = ' '.join(self._upfdg._vfile_data['RET_RES'])
        retctl = retsave + ' | ' + retsave
        retkw = re.findall(r'RETRow\d+', ' '.join(self._upfdg._pmfile_data.keys()))
        for i in range(1, len(retkw) + 1):
            sheet.cell(start_rowg + i, 1).value = self._upfdg._pmfile_data[f'RETRow{i}']['CellType']
            sheet.cell(start_rowg + i, 2).value = retctl
            sheet.cell(start_rowg + i, 3).value = self._upfdg._pmfile_data[f'RETRow{i}']['CellPatn']
            supplypin = self._upfdg._pmfile_data[f'RETRow{i}']['PrimaryPower'] + ' ' + self._upfdg._pmfile_data[f'RETRow{i}']['BackupPower'] + ' ' + self._upfdg._pmfile_data[f'RETRow{i}']['NwellPower'] + ' ' + self._upfdg._pmfile_data[f'RETRow{i}']['PwellPower'] + ' ' + self._upfdg._pmfile_data[f'RETRow{i}']['GroundPin']
            sheet.cell(start_rowg + i, 4).value = supplypin
            sheet.cell(start_rowg + i, 5).value = self._upfdg._pmfile_data[f'RETRow{i}']['SavePin'] + ' | ' + self._upfdg._pmfile_data[f'RETRow{i}']['ResPin']
            sheet.cell(start_rowg + i, 6).value = self._upfdg._pmfile_data[f'RETRow{i}']['PDFunction']
            #sheet.cell(start_rowg + i, 7).value = self._upfdg._pmfile_data[f'RETRow{i}']['PathType']
            #sheet.cell(start_rowg + i, 8).value = self.set_name_style('RET')  # 'RET_Generic_Xconst_09292003'


        # pso
        start_rowg = start_rowg + len(retkw)
        psoctl = ' '.join(self._upfdg._vfile_data['PSO_CTRL'])
        psoack = ' '.join(self._upfdg._vfile_data['PSO_ACK'])
        psoctl = psoctl + ' | ' + psoack
        psokw = re.findall(r'PSWRow\d+', ' '.join(self._upfdg._pmfile_data.keys()))
        #print(psokw)
        for i in range(1, len(psokw) + 1):
            sheet.cell(start_rowg + i, 1).value = self._upfdg._pmfile_data[f'PSWRow{i}']['CellType']
            sheet.cell(start_rowg + i, 2).value = psoctl
            sheet.cell(start_rowg + i, 3).value = self._upfdg._pmfile_data[f'PSWRow{i}']['CellPatn']
            supplypin = self._upfdg._pmfile_data[f'PSWRow{i}']['InputPower'] + ' ' + self._upfdg._pmfile_data[f'PSWRow{i}']['OutputPower']  + ' ' + self._upfdg._pmfile_data[f'PSWRow{i}']['GroundPin']
            sheet.cell(start_rowg + i, 4).value = supplypin
            sheet.cell(start_rowg + i, 5).value = self._upfdg._pmfile_data[f'PSWRow{i}']['CtrlPin'] + ' | ' + self._upfdg._pmfile_data[f'PSWRow{i}']['AckPin']
            sheet.cell(start_rowg + i, 6).value = self._upfdg._pmfile_data[f'PSWRow{i}']['PDFunction']
            #sheet.cell(start_rowg + i, 7).value = self._upfdg._pmfile_data[f'PSWRow{i}']['PathType']
            #sheet.cell(start_rowg + i, 8).value = 'PSO_Generic_Xconst_09292003'

        # repeater

        # find PMHIER table
        start_rowg = self.find_sheet(sheet, 'PMHIER')

        expd_style = vardef['HIER_EXPD_STYLE']
        
        # get_alltrees_by_name(self, name, blktrees={}, valstyle=None, outtype='blks/trees')
        #alltrees = hiertree.get_alltrees_by_name(self._upfdg._vfile_data['module_name'],outtype='blks')
        hierblks = hiertree.get_hierblks(mdname)
        print(hierblks)

        hiertrees = hiertree.get_hiertrees(mdname)
        print(hiertrees)

        hierdepth = hiertree.get_hierdepth(hiertrees,mdname)
        print(hierdepth)

        # outtype is hd/lib/soft
        hdblks = hiertree.get_lvlblks(mdname,outtype='hd')
        print(hdblks)
        macblks = hiertree.get_lvlblks(mdname,outtype='lib')
        print(macblks)
        digblks = hiertree.get_lvlblks(mdname,outtype='soft')
        print(digblks)

        sheet.cell(start_rowg, 1).value = 'Item'
        for i in range(2, hierdepth + 2):
            sheet.cell(start_rowg, i).value = f'BlkLevel{i-1}'
        sheet.cell(start_rowg, hierdepth + 2).value = 'Comment'
        for i in range(1, 12):
            sheet.cell(start_rowg+i, 1).value = f'Index{i}'

    # sheet.column_dimensions[get_column_letter(col)].width  
    # sheet.row_dimensions[row].height
    #worksheet['A1'].alignment = openpyxl.styles.Alignment(wrapText=True)      
        #self.cell_style1(sheet, [start_rowg, 1], [start_rowg, hierdepth + 2])
        # print(sheet.cell(start_rowg,1).fill.start_color.index)
        # print(sheet.cell(start_rowg,1).fill.end_color.index)
        self.cell_style2(sheet, [start_rowg, 1], [start_rowg + 15, hierdepth + 2])
        # for i in range(1,start_rowg+15):
        #     for j in range(1,hierdepth+1):
        #         sheet.cell(start_rowg+i,j).alignment = Alignment(horizontal='left',vertical='center',wrapText=True)
        
        # sheet.column_dimensions[get_column_letter(1)].width = 20
        # for row in range(start_rowg+1,start_rowg+15):
        #     sheet.row_dimensions[row].height = 25
          
        # for col in range(2,hierdepth+1):
        #     sheet.column_dimensions[get_column_letter(col)].width = 35

        blksinfos = hiertree.get_hierblks_infos(mdname)
        
        # BlKLevel1
        sheet.cell(start_rowg+1, 2).value = blksinfos[mdname]
        # BlkLevel2
        uniq_blk2 = list(set(hiertrees[mdname]))
        self.fill_pmhier_col(sheet,start_rowg,hdblks,macblks,digblks,uniq_blk2,blksinfos,3)
        #print(uniq_blk2)

        # blk3/blk4/blk5/...
        if expd_style == 'full':
            # blk3
            # for i in range(3,hierdepth+1):
            #     varnm = f'uniq_blk{i-1}'
            #     exec(f'{varnm}')
            #     uniq_blk= self.get_uniq_blks(varnm,hiertrees)
            #     self.fill_pmhier_col(sheet,start_rowg,hdblks,macblks,digblks,uniq_blk,blksinfos)
            #     print(uniq_blk)
            if hierdepth > 2:
                uniq_blk3= self.get_uniq_blks(uniq_blk2,hiertrees)
                self.fill_pmhier_col(sheet,start_rowg,hdblks,macblks,digblks,uniq_blk3,blksinfos,4)
            
            if hierdepth > 3:
                uniq_blk4= self.get_uniq_blks(uniq_blk3,hiertrees)
                self.fill_pmhier_col(sheet,start_rowg,hdblks,macblks,digblks,uniq_blk4,blksinfos,5)

            if hierdepth > 4:
                uniq_blk5= self.get_uniq_blks(uniq_blk4,hiertrees)
                self.fill_pmhier_col(sheet,start_rowg,hdblks,macblks,digblks,uniq_blk5,blksinfos,6)    

            if hierdepth > 5:
                uniq_blk6= self.get_uniq_blks(uniq_blk5,hiertrees)
                self.fill_pmhier_col(sheet,start_rowg,hdblks,macblks,digblks,uniq_blk6,blksinfos,7) 

            if hierdepth > 6:
                uniq_blk7= self.get_uniq_blks(uniq_blk6,hiertrees)
                self.fill_pmhier_col(sheet,start_rowg,hdblks,macblks,digblks,uniq_blk7,blksinfos,8)

        hierinfogs = ' '.join([f'{key}: {val}' for key,val in hiertrees.items()])
        sheet.cell(start_rowg+12, 1).value = hierinfogs
        sheet.merge_cells(start_row=start_rowg + 12, end_row=start_rowg + 15,start_column=1,end_column=hierdepth+2)

        self.cell_style1(sheet, [start_rowg, 1], [start_rowg, hierdepth + 2])
        #self.cell_style2(sheet, [start_rowg, 1], [start_rowg + 15, hierdepth + 2])
    # sheet.column_dimensions[get_column_letter(col)].width  
    # sheet.row_dimensions[row].height
    #worksheet['A1'].alignment = openpyxl.styles.Alignment(wrapText=True)
        for i in range(1,start_rowg+15):
            for j in range(1,hierdepth+1):
                sheet.cell(start_rowg+i,j).alignment = Alignment(horizontal='left',vertical='center',wrapText=True)
        
        sheet.column_dimensions[get_column_letter(1)].width = 20
        for row in range(start_rowg+1,start_rowg+15):
            sheet.row_dimensions[row].height = 25
          
        for col in range(2,hierdepth+1):
            sheet.column_dimensions[get_column_letter(col)].width = 35

    def get_uniq_blks(self,blks,hiertrees):
        uniq_blk3g = []
        for blk3 in blks:
            if blk3 in hiertrees:
                uniq_blk3g.append([x for x in hiertrees[blk3]])
            # else:
            #     uniq_blk3g.append(blk3.split())
        uniq_blk3 = list(set(item for sublist in uniq_blk3g for item in sublist))
        
        return uniq_blk3


    def fill_pmhier_col(self,sheet,start_rowg,hdblks,macblks,digblks,uniq_blk,blksinfos,col):

        hd_blk2 = [x for x in uniq_blk if x in hdblks ]
        mac_blk2 = [x for x in uniq_blk if x in macblks ]
        dig_blk2 = [x for x in uniq_blk if x in digblks ]

        hd_num = len(hd_blk2)
        mac_num = len(mac_blk2)
        dig_num = len(dig_blk2)
        chg_row = start_rowg
        if hd_num > 0:
            for i in range(1, hd_num + 1):
                sheet.cell(start_rowg+i, col).value = blksinfos[hd_blk2[i-1]]
            chg_row = start_rowg + hd_num
        if mac_num > 0:
            for i in range(1, mac_num + 1):
                sheet.cell(chg_row+i, col).value = blksinfos[mac_blk2[i-1]]
            chg_row = chg_row + mac_num
        if dig_num > 0:
            for i in range(1, dig_num + 1):
                sheet.cell(chg_row+i, col).value = blksinfos[dig_blk2[i-1]]
            chg_row = chg_row + mac_num           

    def read_data(self):
        sheet = self.get_sheet()
        self._vardata = self.get_table_contxt(sheet)
        # nvaldata = {}
        # nvaldata["TMVAR_Row16"] = {
        #     "Variable": "UPF_DIR",
        #     "Value": f'{self._upfdg._upfdir}',
        #     "Comment": ''
        # }
        # nvaldata["TMVAR_Row17"] = {
        #     "Variable": "COM_DIR",
        #     "Value": f'{self._upfdg._upfdir}',
        #     "Comment": ''
        # }
        # nvaldata["TMVAR_Row18"] = {
        #     "Variable": "UPF_VERSION",
        #     "Value": '2.1',
        #     "Comment": ''
        # }
        # nvaldata["TMVAR_Row19"] = {
        #     "Variable": "HD_MOD_NAME",
        #     "Value": f'{self._upfdg._mdname}',
        #     "Comment": ''
        # }
        # nvaldata["TMVAR_Row20"] = {
        #     "Variable": "HD_PROCESS",
        #     "Value": '',
        #     "Comment": ''
        # }
        # nvaldata["TMVAR_Row21"] = {
        #     "Variable": "SCOPE_TYPE",
        #     "Value": 'parent; # self',
        #     "Comment": ''
        # }
        # self._vardata.update(nvaldata)
        # print('_vardata: ', self._vardata)

    def change_sheet(self):
        pass

    def dump_json(self,json_file):
        self._data = self._vardata
        self.write_json(json_file)

    def write_upf(self,mdname,blkalias,blklvl,upf_file):
        sheet = self.get_sheet()
        vardef = self.get_vardef_value(sheet)

        # mdname = self._upfdg._vfile_data['module_name']
        # blkalias = self._upfdg._hier_tree._blocks[mdname].alias
        # blklvl = self._upfdg._hier_tree._blocks[mdname].hdlevel

        upf_lines = f'''
############################################################
## Variable definition from vardef table
############################################################

if {{${{{blkalias}}} == ""}} {{
    echo "UPF_ERROR: ${{{blkalias}}} value is not set. Please check it."
}}
''' 
        upf_lines += self.com_upfout1(blklvl,blkalias,vardef)

        upf_lines += f'''
if {{$UPFVAR(BIAS_ENABLE,${{{blkalias}}})}} {{
        set_design_attributes -elements {{.}}  -attribute enable_bias true
}}

if {{$UPFVAR(BOUNDARY_MODE,${{{blkalias}}}) == "lower"}} {{
	set_design_attributes -elements {{.}}  -attribute lower_domain_boundary true
}} else {{
	set_design_attributes -elements {{.}}  -attribute lower_domain_boundary false
}}

if {{[info_exists UPF_VAR(FL_STAGE)]}} {{
}} else {{
    set UPFVAR(FL_STAGE) "SYN"
}}

if {{$UPFVAR(FL_STAGE) == ""}} {{
    echo "UPF_ERROR: UPFVAR(FL_STAGE) value is not set. Please check it."
}}
'''

        time_stamp = time.strftime("%Y%m%d%H%M%S", time.localtime())
        isofmt = self.set_name_style('ISO')
        lsfmt = self.set_name_style('LS')
        elsfmt = self.set_name_style('ELS')
        upf_lines += f'''
name_format -isolation_prefix {isofmt}_{time_stamp} \\
            -level_shift_prefix {lsfmt}_{time_stamp} \\
            -isolation_suffix "" \\
            -level_shift_suffix ""
'''        


#         upf_lines += f'''
# #########################################################################
# # Variable definition from obj file, pm cell, vfile, hier yaml file,etc.
# # Variable name conversion from obj file
# #########################################################################
#
# if {{[file exists $UPFVAR(UPF_DIR,${{{blkalias}}})inputs/$UPFVAR(HD_MOD_NAME,${{{blkalias}}}).pobj.tcl]}} {{
#     puts "UPF_INFO: Loading $UPFVAR(UPF_DIR,${{{blkalias}}}inputs/$UPFVAR(HD_MOD_NAME,${{{blkalias}}}).pobj.tcl)"
# 	load_upf $UPFVAR(UPF_DIR,${{{blkalias}}})inputs/$UPFVAR(HD_MOD_NAME,${{{blkalias}}}).pobj.tcl
# }} else {{
# 	echo "UPF_ERROR: Missing inputs/$UPFVAR(HD_MOD_NAME,${{{blkalias}}}).pobj.tcl. Please check it."
# }}
# '''
        objvar_list = self._upfdg._objfile_list
        objvar_dict = self._upfdg._objfile_data
        objfile = dirname(upf_file) + f'/{blkalias.lower()}.pobj.tcl'
        self.convert_objname(blkalias,objvar_list,objvar_dict,objfile)

        upf_lines += f'''
if {{[file exists $UPFVAR(UPF_DIR,${{{blkalias}}}){blkalias.lower()}.pobj.tcl]}} {{
    puts "UPF_INFO: Loading $UPFVAR(UPF_DIR,${{{blkalias}}}{blkalias.lower()}.pobj.tcl)"
	load_upf $UPFVAR(UPF_DIR,${{{blkalias}}}){blkalias.lower()}.pobj.tcl
}} else {{
	echo "UPF_ERROR: Missing {blkalias.lower()}.pobj.tcl. Please check it."
}}
'''

        self.save_text(upf_lines,upf_file)

    def convert_objname(self,blkalias,varlist,vardict,objfile):

        new_lines = '\n\n'
    
        for kw in varlist:
            if '^#' in kw:
                new_lines += f'\n{kw}'
            else:
                nkw = f'UPFVAR({kw},${{{blkalias}}})'
                if re.search(r'\w+\[\d+|\d+\]|.',vardict[kw]):
                    # newvar = vardict[kw].replace('[','_')
                    # newvar = newvar.replace(']','_')
                    # newvar = newvar.replace('.','_')
                    # newvarg = re.sub(r'\[','\\[',vardict[kw])
                    # newvarg = re.sub(r'\]','\\]',newvarg)
                    newvar = re.sub(r'\[|\]|\.','_',vardict[kw])
                    new_lines += f'\n\nif {{$UPFVAR(RTLSIM,${{{blkalias}}})}} {{'
                    new_lines += f'\n\tset  {nkw}  "{vardict[kw]}"'
                    new_lines += f'\n}} else {{'
                    new_lines += f'\n\tset  {nkw}   "{newvar}"'
                    new_lines += f'\n}}'
                else:
                    new_lines += f'\n\n\tset  {nkw}  "{vardict[kw]}"'
        
        self.save_text(new_lines,objfile)



    def com_upfout1(self,blklvl,alias,vardef):
        lvl_flat = ''
        val_flat = 0

        if blklvl == 'sys' or blklvl == 'top':
            lvl_flat = 'IS_CHIP'
            val_flat = 0
        if blklvl == 'blk':
            lvl_flat = 'IS_FLAT'
            val_flat = 0

        vardef_list =['HD_MOD_NAME','HD_PROCESS','LIB','PG_FLAG','BIAS_ENABLE',
                          'SS_MODE','SCOPE_TYPE','BOUNDARY_MODE','UPF_VERSION','EDA_TOOL','UPF_DIR','COM_DIR','RTLSIM']
        vardef_dic = {
            'HD_MOD_NAME'   : vardef['HD_MOD_NAME'],
            'HD_PROCESS'    : vardef['HD_PROCESS'],
            'LIB'           : '0',
            'PG_FLAG'       : '0',
            'BIAS_ENABLE'   : vardef['BIAS_ENABLE'],
            'SS_MODE'       : vardef.get('SS_MODE', 'full'),  # 默认值为 'full'
            'SCOPE_TYPE'    : vardef.get('SCOPE_TYPE', 'parent'),  # 默认值为 'parent'
            'BOUNDARY_MODE' : vardef.get('BOUNDARY_MODE', 'lower'),  # 默认值为 'lower'
            'UPF_VERSION'   : vardef.get('UPF_VERSION', '2.1'),  # 默认值为 '2.1'
            'EDA_TOOL'      : vardef['EDA_TOOL'],
            'UPF_DIR'       : vardef.get('UPF_DIR', './'),  # 默认值为当前目录
            'COM_DIR'       : vardef.get('COM_DIR', './'),  # 默认值为当前目录
            # 'HIER_EXPD_STYLE' : vardef['HIER_EXPD_STYLE'],
            'RTLSIM'        : '0'
        }
            

        upf_lines = ''
        upf_lines += f'''
 if {{[info exists UPFVAR({lvl_flat},${{{alias}}})]}} {{
}} else {{
    set UPFVAR({lvl_flat},${{{alias}}}) "{val_flat}"
}}

if {{$UPFVAR({lvl_flat},${{{alias}}}) == ""}} {{
    echo "UPF_ERROR: UPFVAR({lvl_flat},${{{alias}}}) value is not set. Please check it."
}}
'''       
        for kwd in vardef_list:
            if kwd == 'UPF_VERSION':
                upf_lines += f'''
if {{[info exists UPFVAR({kwd},${{{alias}}})]}} {{
}} else {{
    set UPFVAR({kwd},${{{alias}}}) "{vardef_dic[kwd]}"
    upf_version "{vardef_dic[kwd]}"
}}
'''
            else:
                upf_lines += f'''
if {{[info exists UPFVAR({kwd},${{{alias}}})]}} {{
}} else {{
    set UPFVAR({kwd},${{{alias}}}) "{vardef_dic[kwd]}"
}}

if {{$UPFVAR({kwd},${{{alias}}}) == ""}} {{
    echo "UPF_ERROR: UPFVAR({kwd},${{{alias}}}) value is not set. Please check it."
}}
'''
        return upf_lines

    # def save_text(self,context,file):
    #     with open(file, 'w') as fw:
    #         fw.write(context)


    def check_sheet(self):
        pass  



