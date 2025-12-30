
import sys
import time
import os
import re

from os.path import dirname, abspath, basename
import  openpyxl

from upfgen.vardef import *
from .pdomain import *
from .pstrategy import *
from .pmode import *
from com.hierpwr import HierPwrTree
from .baseupf import *
from .vardef import *

class UPFDG(object):
    def __init__(self):
        self._sheets = {}
        self._hier_tree = {}
        self._vardefdata = {}
        self._vardefpcell = {}
        self._wb = {}
        self._inputs = BaseInputs()
        # self._vardef = VarDefSheet()

        self.proj_mode = False
        self._vfile_data = None
        self._pmfile_data = None
        self._objfile_data = None
        self._objfile_list = None
        self._data = None
        self._upfdir = ''
        self._mdname = ''
        self._alias = ''
        
        # self._mdname = ''
        # self._blkalias = ''
        # self._blklvl = ''

    @property
    def hier_tree(self):
        return self._hier_tree
    
    @hier_tree.setter
    def hier_tree(self, hier_tree):
        self._hier_tree = hier_tree
        # if self._mdname and self._hier_tree._blocks[self._mdname]:
        #     self._blkalias = self._hier_tree._blocks[self._mdname].alias
        #     self._blklvl = self._hier_tree._blocks[self._mdname].hdlevel
 
    def load_design_guide(self,dg_file,kwd=''):
        # self._upfdir = abspath(dirname(dirname(dg_file)))
        self._wb = openpyxl.load_workbook(dg_file)

        valdef_sheet = self._wb['VarDef']
        # start = 0
        # for i in range(1, valdef_sheet.max_row+1):
        #     if valdef_sheet.cell(i,1).value == 'Variable':
        #         start = i + 1
        #         break
        # for i in range(start, valdef_sheet.max_row+1):
        #     key = valdef_sheet.cell(row=i, column=1).value
        #     val = valdef_sheet.cell(row=i, column=2).value
        #     self._vardef[key] = val


        self._sheets = {
            'VarDef'        : VarDefSheet(self, 'VarDef'),
            'PDomain'       : PDomainSheet(self, 'PDomain'),
            'PStrategy'     : PStrategySheet(self, 'PStrategy'),     
            'PMode'         : PModeSheet(self, 'PMode'),
        }

        self._vardefdata = self._sheets['VarDef'].get_vardef_value(valdef_sheet)
        self._vardefpcell = self._sheets['VarDef'].get_table_contxt(valdef_sheet)

        # read and convert excel to json data
        if kwd == 'json':
            for sheetname,sheet in self._sheets.items():
                json_file = dirname(dirname(dg_file)) + '/json' + f'/{sheetname.lower()}.json'
                sheet.read_data()
                sheet.dump_json(json_file)
            

    def read_vfile(self, vfile,kwd=''):
        self._upfdir = abspath(dirname(dirname(vfile)))
        self._vfile_data = self._inputs.read_vfile(vfile)
        self._mdname = self._vfile_data['module_name']
    #    print(self._vfile_data)
        if kwd == 'json':
            self._data = self._vfile_data
            json_file = dirname(dirname(vfile)) + '/json' + f'/pvlog.json'
            self.write_json(json_file)

    def read_pmfile(self, pmfile,kwd=''):
        self._pmfile_data = self._inputs.read_pmfile(pmfile)
        #print(self._pmfile_data)
        if kwd == 'json':
            self._data = self._pmfile_data
            json_file = dirname(dirname(pmfile)) + '/json' + f'/pmcell.json'
            self.write_json(json_file)
        
    def read_objfile(self, objfile,kwd=''):
        self._objfile_data, self._objfile_list = self._inputs.read_objfile(objfile)
        #print(self._objfile_data)
        #print(self._objfile_list)
        if kwd == 'json':
            self._data = self._objfile_data
            json_file = dirname(dirname(objfile)) + '/json' + f'/pobj.json'
            self.write_json(json_file)
      

    # def read_data(self):
    #     for sht in self._sheets.values():
    #         sht.read_data()       

    def update_dg(self):
        for sht in self._sheets.values():
            sht.update_sheet()

    def check_dg(self):
        for sht in self._sheets.values():
            sht.check_sheet()

    def change_dg(self,dgfile):
        for sht in self._sheets.values():
            sht.change_sheet(dgfile)

    def write_upf_files(self,prousr=False):
        mdname = self._vfile_data['module_name']
        blkalias = self._hier_tree._blocks[mdname].alias
        blklvl = self._hier_tree._blocks[mdname].hdlevel
        blk_file_path = self._upfdir

        # if not os.path.exists(blk_file_path + '/outputs/intg'):
        #     os.makedirs(blk_file_path + '/outputs/intg', exist_ok=True)
        # if not os.path.exists(blk_file_path + '/outputs/blklib'):
        #     os.makedirs(blk_file_path + '/outputs/blklib', exist_ok=True)
        # if not os.path.exists(blk_file_path + '/outputs/expd'):
        #     os.makedirs(blk_file_path + '/outputs/expd', exist_ok=True)

        # write top upf files such as ${blk}.upf and ${blk}_top.upf and ${blk}_subblk.upf and ${blk}_tune.upf
        upf_file = blk_file_path +  f'/outputs/{mdname}.pwr.upf'
        self.write_wrap_upf(mdname,blkalias,blklvl,upf_file)

        upf_file = blk_file_path +  f'/outputs/{blkalias.lower()}_top.upf'
        self.write_top_upf(mdname,blkalias,blklvl,upf_file)

        # upf_file = blk_file_path +  f'/outputs/{blkalias.lower()}_subblk.upf'
        # self.write_subblk_upf(mdname,blkalias,blklvl,upf_file)

        upf_file = blk_file_path +  f'/outputs/{blkalias.lower()}_tune.upf'
        self.write_tune_upf(mdname,blkalias,blklvl,upf_file)

        # write all other upf files from intent in dg file
        #for sht in self._sheets.values()
        for sheetname,sheet in self._sheets.items():
            upf_file = blk_file_path + f'/outputs/{blkalias.lower()}_{sheetname.lower()}.upf'
            sheet.write_upf(mdname,blkalias,blklvl,upf_file)      

    def write_wrap_upf(self,mdname,blkalias,blklvl,upf_file):

        # mdname = self._vfile_data['module_name']
        # blkalias = self._hier_tree._blocks[mdname].alias
        # blklvl = self._hier_tree._blocks[mdname].hdlevel

        upf_lines = '\nglobal UPFVAR\n'
        upf_lines += f'\nset {blkalias} {blkalias}'

        if blklvl == 'sys' or blklvl == 'top':
            upf_lines += f'\nset UPFVAR(IS_CHIP,${{{blkalias}}}) "0"'
        if blklvl == 'blk':
            upf_lines += f'\nset UPFVAR(IS_FLAT,${{{blkalias}}}) "0"'          
        upf_lines += f'\nset UPFVAR(LIB,${{{blkalias}}}) "0"'
        upf_lines += f'\nset UPFVAR(PG_FLAG,${{{blkalias}}}) "0"'
        upfver = self._vardefdata['UPF_VERSION']
        upf_lines += f'\nset UPFVAR(UPF_VERSION,${{{blkalias}}}) "{upfver}"'
        edatool = self._vardefdata['EDA_TOOL']
        upf_lines += f'\nset UPFVAR(EDA_TOOL,${{{blkalias}}}) "{edatool}"'
        upf_lines += f'\nset UPFVAR(RTLSIM,${{{blkalias}}}) "0"'

        upfdir = self._vardefdata['UPF_DIR']
        upf_lines += f'\n\nset UPFVAR(UPF_DIR,${{{blkalias}}}) "{upfdir}"'
        comdir = self._vardefdata['COM_DIR']
        upf_lines += f'\nset UPFVAR(COM_DIR,${{{blkalias}}}) "{comdir}"\n'

        upf_lines += f'\nload_upf $UPFVAR(UPF_DIR,${{{blkalias}}}){blkalias.lower()}_top.upf'

        self.save_text(upf_lines,upf_file)

    def write_top_upf(self,mdname,blkalias,blklvl,upf_file):
        # mdname = self._vfile_data['module_name']
        # blkalias = self._hier_tree._blocks[mdname].alias
        # blklvl = self._hier_tree._blocks[mdname].hdlevel

        upf_lines = f'''
############################################################
## Section 1: General Variable
############################################################

if {{${{{blkalias}}} == ""}} {{
    echo "UPF_ERROR: ${{{blkalias}}} value is not set. Please check it."
}}
'''
        if blklvl == 'sys' or blklvl == 'top':
            lev = 'Sys'
        if blklvl == 'blk':
            lev = 'Blk'

        upf_lines += f'''
if {{[info exists UPFVAR({lev}Nm,${{{blkalias}}})]}} {{
}} else {{
    set UPFVAR({lev}Nm',${{{blkalias}}}) "${{{blkalias}}}"
}}

if {{$UPFVAR({lev}Nm',${{{blkalias}}}) == ""}} {{
    echo "UPF_ERROR: UPFVAR({lev}Nm',${{{blkalias}}}) value is not set. Please check it."
}}
''' 

        upf_lines += self.com_upfout1(blklvl,blkalias)

        upf_lines += f'''
global design
if {{[info_exists UPF_VAR(FL_STAGE)]}} {{
}} else {{
    if {{[info exists design(use_mbist_rtl)] && $design(use_mbist_rtl)}} {{
        set UPFVAR(FL_STAGE) "MBIST_SYN"
    }} else {{
        set UPFVAR(FL_STAGE) "SYN"
    }}
}}

if {{$UPFVAR(FL_STAGE) == ""}} {{
    echo "UPF_ERROR: UPFVAR(FL_STAGE) value is not set. Please check it."
}}
'''
        lwalias = blkalias.lower()
#         upf_lines += f'''
# ############################################################
# ## Section 2: pglib for implementation
# ############################################################
#
# if {{$UPFVAR(IS_FLAT,${{{blkalias}}}) && $UPFVAR(LIB,${{{blkalias}}}) && $UPFVAR(PG_FLAG,${{{blkalias}}})}} {{
#     if {{[file exists $UPFVAR(UPF_DIR,${{{blkalias}}})blklib/{lwalias}_pglib.upf]}} {{
#         puts "UPF_INFO: Loading $UPFVAR(UPF_DIR,${{{blkalias}}})blklib/{lwalias}_pglib.upf"
#         load_upf $UPFVAR(UPF_DIR,${{{blkalias}}})blklib/{lwalias}_pglib.upf
#     }} else {{
#         puts "UPF_WARN: Missing $UPFVAR(UPF_DIR,${{{blkalias}}})/blklib/{lwalias}_pglib.upf. Please check it."
#     }}
# }} else {{
# ############################################################
# ## Section 3: Common UPF Proc
# ############################################################
# if {{!$UPFVAR(PG_FLAG,${{{blkalias}}})}} {{
#     if {{[file exists $UPFVAR(COM_DIR,${{{blkalias}}})pwr_proc.upf]}} {{
#         puts "UPF_INFO: Loading $UPFVAR(COM_DIR,${{{blkalias}}})pwr_proc.upf"
#         load_upf  $UPFVAR(COM_DIR,${{{blkalias}}})pwr_proc.upf
#     }} else {{
#         puts "UPF_WARN: Missing $UPFVAR(COM_DIR,${{{blkalias}}})pwr_proc.upf. Please check it."
#     }}
# }}
# '''
        upf_lines += self.com_upfout2(blkalias)

        self.save_text(upf_lines,upf_file)


    def com_upfout1(self,blklvl,alias):
        if blklvl == 'sys' or blklvl == 'top':
            keylist = ['IS_CHIP', 'LIB', 'PG_FLAG', 'UPF_DIR', 'COM_DIR']
            upfdir = self._vardefdata['UPF_DIR']
            comdir = self._vardefdata['COM_DIR']
            keydic = {
                'IS_CHIP'       : '0',
                'LIB'           : '0',
                'PG_FLAG'       : '0',
                'UPF_DIR'       : upfdir,
                'COM_DIR'       : comdir
            }
        if blklvl == 'blk':
            keywds = ['IS_FLAT', 'LIB', 'PG_FLAG', 'UPF_DIR', 'COM_DIR']
            upfdir = self._vardefdata['UPF_DIR']
            comdir = self._vardefdata['COM_DIR']
            keydic = {
                'IS_FLAT'       : '0',
                'LIB'           : '0',
                'PG_FLAG'       : '0',
                'UPF_DIR'       : upfdir,
                'COM_DIR'       : comdir
            }
        upf_lines = ''
        for kwd in keywds:
            upf_lines += f'''
if {{[info exists UPFVAR({kwd},${{{alias}}})]}} {{
}} else {{
    set UPFVAR({kwd},${{{alias}}}) "{keydic[kwd]}"
}}

if {{$UPFVAR({kwd},${{{alias}}}) == ""}} {{
    echo "UPF_ERROR: UPFVAR({kwd},${{{alias}}}) value is not set. Please check it."
}}
'''
        return upf_lines
    
    def com_upfout2(self,alias):

        balias = alias.lower()
        kwd_dic = {
            f'{balias}_vardef'     :  '2: Variables Definition',
            f'{balias}_pdomain'    :  '3: Power Domain & Supply Network',
            f'{balias}_pstrategy'  :  '4: Power Strategy',
            f'{balias}_tune'       :  '5: User Tune',
            f'{balias}_pmode'      :  '6: Power State Table'
            # f'{balias}_subblk'     :  '9: Nested Block&IP&Macro power intent'
            
        }

        upf_lines = ''
        kwd_list = [f'{balias}_vardef', f'{balias}_pdomain', f'{balias}_pstrategy', f'{balias}_tune', f'{balias}_pmode'] #, f'{balias}_subblk']
        for kwd in kwd_list: 
            #sec = str(kwd_list[kwd][0])         
            upf_lines += f'''
############################################################
## Section {kwd_dic[kwd]}
############################################################
if {{[file exists $UPFVAR(UPF_DIR,${{{alias}}}){kwd}.upf]}} {{
    puts "UPF_INFO: Loading $UPFVAR(UPF_DIR,${{{alias}}}){kwd}.upf"
    load_upf  $UPFVAR(UPF_DIR,${{{alias}}}){kwd}.upf
}} else {{
    puts "UPF_WARN: Missing $UPFVAR(UPF_DIR,${{{alias}}}){kwd}.upf. Please check it."
}}
'''
        return upf_lines



    def write_subblk_upf(self,mdname,blkalias,blklvl,upf_file):
        pass

    def write_tune_upf(self,mdname,blkalias,blklvl,upf_file):
        pass

    def write_json(self, filepath):
        os.makedirs(dirname(filepath), exist_ok=True)
        jsonstr = json.dumps(self._data, indent=4)
        with open(filepath,'w') as fw:
            print(jsonstr, file=fw)    

    def save_text(self,context,file):
        with open(file, 'w') as fw:
            fw.write(context)

    def save_workbook(self,output):
        self._wb.save(output)



    # def excel_to_json(self, excel_file, sheet_name,json_file):

    #     data_frame = pd.read_excel(excel_file, sheet_name=sheet_name)
    #     json_data = data_frame.to_dict(orient='records')
    #     if sheet_name == 'VarDef':
    #         print(json_data)        
    #     json_str = json.dumps(json_data, ensure_ascii=False)

    #     self.save_text(json_str,json_file)



#      sbclk_lines += f'''
# foreach {{ALIAS_VAL HIER_VAL}} ${key}_PAT {{
#     set {key} $ALIAS_VAL
#     if {{$SDCVAR({nsubblk_is_flat},${{{key}}}}) && !$SDCVAR(LIB,${{{key}}})}} {{
#         if {{[file exists $SDCVAR(SDC_DIR,${{{key}}})intg/{file_name}]}} {{
#             puts "SDC_INFO: Sourcing $SDCVAR(SDC_DIR,${{{key}}})intg/{file_name}"
#             source -echo -verbose  $SDCVAR(SDC_DIR,${{{key}}})intg/{file_name}
#         }} else {{
#             puts "SDC_WARN: Missing  $SDCVAR(SDC_DIR,${{{key}}})intg/{file_name}. Please check it."
#         }}
#     }}
# }}                       
# '''  

# dcdc_lines += f'''
# if {{[info exists {varname}]}} {{
# }} else {{
#     set {varname} "$SDCVAR(HIER,{self.name_flg},{self.blk_name}){varvalue}"
# }}
# '''
#             else:
#                 dcdc_lines += f'''
# if {{[info exists {varname}]}} {{
# }} else {{
#     set {varname} "{varvalue}"
# }}
# '''

#             dcdc_lines += f'''
# if {{[info exists SDCVAR{varname}]}} {{
#     unset SDCVAR({varname})
#     set SDCVAR({varname}) "{varvalue}"
# }} else {{
#     set SDCVAR({varname}) "{varvalue}"
# }}
# '''

#                 nsubblk_lines += f'''
# \tif {{[info exists SDCVAR({nsubblk_is_flat},${{{key}}})]}} {{
# \t}} else {{
# \tset SDCVAR({nsubblk_is_flat},${{{key}}})  "1"
# \t}}
# \tif {{[info exists SDCVAR(LIB,${{{key}}})]}} {{
# \t}} else {{
# \tset SDCVAR(LIB,${{{key}}})  "0"
# \t}}
# '''



