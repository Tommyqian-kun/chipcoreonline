

import json
import os
import re
import sys
from os.path import dirname
import time

import yaml
#from itertools import chain

from openpyxl import worksheet 
from pprint import pprint 
import pandas as pd
from openpyxl.utils import get_column_letter 

import tkinter as tk

from openpyxl.styles import Border, Side, PatternFill, Font, Alignment 
from openpyxl.worksheet.datavalidation import DataValidation

from com.base import *

'''
@ define BaseSheet for design guide sheets
@ define BaseInputs to parse vfile
@ print message severity
'''


# sdcdg is XsdcDesignGuide object
class BaseSheet(object):
    def __init__(self, sdcdg, sheetname):
        self._sdcdg = sdcdg
        self._sheetname = sheetname
        self._data = []
        #self._vardef = {}
        self._pdnmdict = {}
    
    def get_sheet(self):
        return self._sdcdg._wb[self._sheetname]

    def read_data(self):
        raise NotImplementedError(self.__class__.__name__ + ' read_data not implemented yet')

    def write_json(self, filepath):
        os.makedirs(dirname(filepath), exist_ok=True)
        jsonstr = json.dumps(self._data, indent=4)
        with open(filepath,'w') as fw:
            print(jsonstr, file=fw)

    def find_sheet(self, sheet, skw):
        start_rowg = 1
        # TABCONST = ['TMVAR','TMHIER','TMCLK','TMIODLY','TMIOEXP','TMINOUT','TMINTEXP','TMSTPGATE']
        TABCONST = ['TMVAR','TMCLK','TMIODLY','TMIOEXP','TMINOUT','TMINTEXP','TMSTPGATE']
        #print(skw,sheet)
        for i in range(1,sheet.max_row+1):
            if skw in TABCONST and sheet.cell(i,1).value == skw:
                start_rowg = i + 1
                break  
        return  start_rowg 

    def get_vardef_value(self, sheet):
        vardef = {}
        start_rowg = self.find_sheet(sheet, 'TMVAR')
        # end_rowg = self.find_sheet(sheet, 'TMHIER')
        for i in range(start_rowg + 1, start_rowg + 15):
            key = sheet.cell(row=i, column=1).value
            val = sheet.cell(row=i, column=2).value
            vardef[key] = val

        vardef['SDC_DIR'] = self._sdcdg._sdcdir
        vardef['COM_DIR'] = self._sdcdg._sdcdir
        vardef['DFT_DIR'] = ''
        # vardef['HD_MOD_NAME'] = self._sdcdg._mdname
        vardef['HD_PROCESS'] = ''
        vardef['CYCLE_LIST'] = '[list CYCLE500M]'

        # print('vardef:', vardef)
        return vardef   

    def set_name_style(self, kw):
        #time_stamp = time.strftime("%Y%m%d%H%M%S", time.localtime())
        #CONST = f'Generic_Xsdc_{time_stamp}'
        CONST = f'Generic_XSDC'
        return kw + '_' + CONST

    # showErrorMessage=False,showDropDown=True
    def add_dropdown(self, sheet, options, start, end):       
        dv = DataValidation(type="list", formula1=options, showErrorMessage=False)
        sheet.add_data_validation(dv)
        if len(start) == 2 and len(end) == 2:
            for i in range(start[0], end[0] + 1):
                for j in range(start[1], end[1] + 1):
                    dv.add(sheet.cell(i,j))
        if len(start) == 1 and len(end) == 1:
            dv.add(sheet.cell(start[0],end[0]))


    def cell_style1(self, sheet, start, end):
        border=Border(left=Side(border_style='thin', color='000000'),
                      right=Side(border_style='thin', color='000000'),
                      top=Side(border_style='thin', color='000000'),
                      bottom=Side(border_style='thin', color='000000'))
        #bgfill = PatternFill(fill_type='solid', start_color='fff2cc', end_color='fff2cc') 
        #bgfill = PatternFill(fill_type = 'solid', start_color='197e00',end_color='197e00')
        bgfill = PatternFill(fill_type = 'solid', start_color='FF385724',end_color='FF333300')
        font = Font(name='等线', size=11, color='FFFFFF')
        for i in range(start[0], end[0] + 1):
            for j in range(start[1], end[1] + 1):
                sheet.cell(i,j).border=border 
                sheet.cell(i,j).fill=bgfill
                sheet.cell(i,j).font=font
                sheet.cell(i,j).alignment = Alignment(horizontal='left', vertical='center',wrapText=True) 
                sheet.cell(i,j).alignment = Alignment(horizontal='left', vertical='center',wrapText=True) 
                sheet.cell(i,j).alignment=Alignment(horizontal='left', vertical='center') 

    def cell_style2(self, sheet, start, end):
        border=Border(left=Side(border_style='thin', color='000000'),
                      right=Side(border_style='thin', color='000000'),
                      top=Side(border_style='thin', color='000000'),
                      bottom=Side(border_style='thin', color='000000'))
        #bgfill = PatternFill(fill_type='solid', start_color='fff2cc', end_color='fff2cc') 
        #bgfill = PatternFill(fill_type = 'solid', start_color='197e00',end_color='197e00')
        bgfill = PatternFill(fill_type = 'solid', start_color='FFFFFF',end_color='FFFFFF')
        #font = Font(name='等线', size=11, color='FFFFFF')
        for i in range(start[0], end[0] + 1):
            for j in range(start[1], end[1] + 1):
                sheet.cell(i,j).border=border 
                sheet.cell(i,j).fill=bgfill
                #sheet.cell(i,j).font=font
                sheet.cell(i,j).alignment = Alignment(horizontal='left', vertical='center',wrapText=True) 
                sheet.cell(i,j).alignment = Alignment(horizontal='left', vertical='center',wrapText=True) 
                sheet.cell(i,j).alignment=Alignment(horizontal='left', vertical='center',wrapText=True)

    def get_supply_infos(self):

        delkeys = ['module_name', 'ISO_CTRL', 'RET_SAVE', 'RET_RES', 'PSO_CTRL', 'PSO_ACK']
        supply_datag = self._sdcdg.vfile_data
        supply_data = {}

        for ky,vl in supply_datag.items():
            if ky not in delkeys:
                supply_data[ky] = vl

        supply_kw = []
        supply_val = []
        supply_vss = []
        supply_tmp = ''
        for key,val in supply_data.items():
            if '0v' in val or '0.0v' in val:
                supply_vss.append(key)
                sdc_info(f'Ground pin is {key}')
            elif 'PSO' in val:
                supply_kw.append(key)
                for i in range(1, int(val[-1]) + 1):
                    supply_kw.append(key + f'_PSW{i}')
                supply_tmp = supply_tmp + ' ' + supply_data[key].split('PSO')[0].strip()
            else:
                supply_kw.append(key)
                supply_tmp = supply_tmp + ' ' + supply_data[key].strip()

        #print(supply_tmp.strip().split(','))
        float_list = [float(x.strip('v')) for x in supply_tmp.strip().split()]
        unique_floats = set(float_list)
        sorted_floats = sorted(unique_floats, reverse=True)
        supply_val = [str(x) + 'v' for x in sorted_floats]
        supply_val.append('off')
        supply_val.append('0v')

        return supply_kw,supply_val,supply_vss,supply_data
    

    def get_ctl_sig(self, ctsig):
        ctrl = []
        for ct in ctsig:
            if re.search(r'\[\d+:\d+\]', ct):
                sig = ct.split('[')[0].strip()
                st = int(ct.split(':')[0].strip()[-1])
                ed = int(ct.split(':')[1].strip()[0])
                for i in range(ed,st+1):
                    ctrl.append(sig + '[' + str(i) + ']')
            else:
                ctrl.append(ct)
        return ctrl

    def get_table_loc(self,sheet) -> dict:

        if self._sheetname == 'VarDef':
            TABCONST = ['TMVAR']
        if self._sheetname == 'ClkDef':
            TABCONST = ['TMCLK']
        if self._sheetname == 'IODly':
            TABCONST = ['TMIODLY']
        # if self._sheetname == 'IOExp':
        #     TABCONST = ['TMIOEXP','TMINOUT']
        # if self._sheetname == 'IntExp':
        #     TABCONST = ['TMINTEXP','TMSTPGATE']
        if self._sheetname == 'Exp':
            TABCONST = ['TMIOEXP','TMINOUT','TMINTEXP','TMSTPGATE']

        # row_start max_col              
        row_start = ''        
        max_row = ''
        max_col = ''
        # row_start max_col max_row
        table_row_loc = {}
        for kw in TABCONST:
            strow = self.find_sheet(sheet, kw)
            row_start = str(strow)
            for i in range(1,sheet.max_column + 1):
                if sheet.cell(strow,i).value == 'Comment':
                    max_col = str(i)
                    #print(row_start[kw])
                    break              

            print(kw,row_start)
            # if kw in ['TMHIER','TMCLK','TMIODLY','TMINOUT','TMSTPGATE']:
            if kw in ['TMVAR','TMCLK','TMIODLY','TMSTPGATE']:
                #table_row_loc[kw] = row_start[kw] + ' ' + str(int(row_start[kw].split()[0]) + 20)
                table_row_loc[kw] = row_start + ' ' + str(sheet.max_row + 2) + ' ' + max_col
            else:
                idx = TABCONST.index(kw) + 1
                max_row = self.find_sheet(sheet,TABCONST[idx]) - 1
                # print(kw,idx,max_row)
                #table_row_loc[kw] = row_start[kw] + ' ' + str(int(row_start[TABCONST[idx]].split()[0]) - 2)
                table_row_loc[kw] = row_start + ' ' + str(max_row) + ' ' + max_col

        #print(table_row_loc)
        return table_row_loc
  
    def get_table_contxt(self,sheet) -> dict:
        # row_start max_col max_row
        tab_loc = self.get_table_loc(sheet)
        print(tab_loc)

        TABCONST = []
        if self._sheetname == 'VarDef':
            TABCONST = ['TMVAR']
        if self._sheetname == 'ClkDef':
            TABCONST = ['TMCLK']
        if self._sheetname == 'IODly':
            TABCONST = ['TMIODLY']
        # if self._sheetname == 'IOExp':
        #     TABCONST = ['TMIOEXP','TMINOUT']
        # if self._sheetname == 'IntExp':
        #     TABCONST = ['TMINTEXP','TMSTPGATE']
        if self._sheetname == 'Exp':
            TABCONST = ['TMIOEXP','TMINOUT','TMINTEXP','TMSTPGATE']

        table_contxt = {}
        #row_contxt = {}
        if TABCONST:
            for kw in TABCONST:
                start_row = int(tab_loc[kw].split(' ')[0])
                end_row = int(tab_loc[kw].split(' ')[1])
                end_col = int(tab_loc[kw].split(' ')[2])
                # if kw == 'TMSTPGATE':
                #     print('TMSTPGATE:',start_row,end_row,end_col)
                if kw == 'PMVAR':
                    for i in range(start_row, end_row + 1):
                        key = sheet.cell(i + 1, 1).value
                        val = str(sheet.cell(i + 1, 2).value)
                        if key:
                            table_contxt[key] = val.strip()
                        # print('PMVARdfd: ', table_contxt)
                        # if key and val:
                        #     table_contxt[key] = val
                else:
                    table_contxt.update(self.get_row_txt(sheet,kw,start_row,end_row,end_col))
                # if kw == 'TMSTPGATE':
                #     print('TMSTPGATE:',table_contxt)

        return table_contxt

    def get_row_txt(self,sheet,kw,start_row,end_row,end_col):
        row_contxt = {}
        table_contxt = {}
        for i in range(1,end_row-start_row):
            for j in range(1,end_col+1):
                key = sheet.cell(start_row,j).value
                val = sheet.cell(start_row+i,j).value
                val_col1 = sheet.cell(start_row+i,1).value
                if val_col1:
                    if re.search(r'^#',val_col1.strip()):
                        continue
                if key:     key = str(key).strip()
                if val:     val = str(val).strip()
                row_contxt[key] = val
                # if key and val:
                #     row_contxt[key] = val
            all_none = all(ele is None for ele in list(row_contxt.values()))
            if not all_none and row_contxt:
                table_contxt[f'{kw}_Row{start_row+i}'] = row_contxt
            row_contxt = {}
            # for key in table_contxt.keys():
            #     if 'TMCLK' in key:
            #         print(table_contxt)
        
        return table_contxt

        
    def save_text(self, context,file):
        with open(file, 'w') as fw:
            fw.write(context)

    def get_rows(self,pmdata,keyrow,kwd,ckwd):
        pmdict = {}
        pmlist = [(key, val) for key, val in pmdata.items() if re.search(r'{keyrow}\d+',key) and not re.search(r'^#',val[f'{ckwd}'].strip()) and val[f'{kwd}']]
        for k,v in pmlist:
            pmdict[k] = v
        pmkeys = [x for x in pmlist if re.search(r'{keyrow}\d+',x)]
        pmkeys.sort() 

        return pmdict, pmkeys     


class BaseInputs(object):
    def __init__(self):
        self.vfile_data = {}
        self.vfile_list = []
               
    def read_vfile(self, vfile) -> dict:

        lines = self.read_text(vfile)

        relclknum = 0
        for line in lines:
            line = line.replace('\n','').replace('\r','').replace('\t',' ').strip()
            if re.search(r'^\/\/', line) and '#RelClock:' not in line:
                continue

            if re.search(r'^module', line):
                self.vfile_data['module_name']= re.split(' +',line)[1].strip().replace('(','')
                self.vfile_list.append('module_name')
                continue

            if '#RelClock:' in line:
                relclknum += 1
                relclk = line.split('#RelClock:')[1].strip().replace('#','')
                self.vfile_list.append(f'RelClock{relclknum}')
                self.vfile_data[f'RelClock{relclknum}'] = relclk
                continue
                
            if re.search(r'^\);$',line):
                break

            dirc = ''
            portnum = ''
            kwd = ''
            if re.search(r'^input|^output|^inout',line):
                tline = line.split(' ')
                sline = [x for x in tline if x != '']
                dirc = sline[0]
                dircg = dirc
                
                if re.search(r'wire|logic|byte|bit|reg|tri1|tri0',line):               
                    if re.search(r'\[\d+:\d+\]',line):
                        lineg = ' '.join(sline[3:])
                        portnum = sline[2]
                    else:
                        lineg = ' '.join(sline[2:])
                        portnum = '1'             
                else:
                    if re.search(r'\[\d+:\d+\]',line):
                        lineg = ' '.join(sline[2:])
                        portnum = sline[1]
                    else:
                        lineg = ' '.join(sline[1:])
                        portnum = '1'
                portnumg = portnum
                

                sigchar = lineg.replace(' ','')
                # print(sigchar)
                if re.search(r'\/\/#\w+#',sigchar):
                    #kwdg = ''.join(re.findall(r'\/\/(#\w+#)+', sigchar)).strip()
                    if '##' in sigchar:
                        kwd = sigchar.replace('##',' ').split('#')[1]
                    else:
                        kwd = sigchar.split('#')[1]
                    
                    if ',' in sigchar:
                        sigcharg = sigchar.split(',')
                        for ich in sigcharg:
                            if r'#\w+#' not in ich and '//' not in ich:
                                self.vfile_data[ich] = [dircg,portnumg,kwd]
                                self.vfile_list.append(ich)
                    else:
                        sdc_warn(f'{sigchar} not found , symbol ...')
                        ish = sigchar.split(r'//')[0].strip()
                        self.vfile_data[ish] = [dircg,portnumg,kwd]
                        self.vfile_list.append(ish)
                else:
                    kwd = 'None'
                    if ',' in sigchar:
                        sigcharg = sigchar.split(',')                        
                        for ich in sigcharg:
                            if '//' not in ich and ich != '':
                                self.vfile_data[ich] = [dircg,portnumg,kwd]
                                self.vfile_list.append(ich)
                    else:
                        sdc_warn(f'{sigchar} not found , symbol ...')
                        if '//' in sigchar:
                            ish = sigchar.split(r'//')[0].strip()
                        else:
                            ish = sigchar.strip()
                            self.vfile_data[ish] = [dircg,portnumg,kwd]
                            self.vfile_list.append(ish)

            else:
                if re.search(r'^\S+,$',line) and '//' not in line:
                    sline = line.split(',')
                    if re.search(r'\/\/#\w+#',line):
                        #kwd = ''.join(re.findall(r'\/\/#\w+#', line)).strip().split('#')[1]
                        if '##' in line:
                            kwd = line.replace('##',' ').split('#')[1]
                        else:
                            kwd = line.split('#')[1]
                        
                        for ich in sline:
                            if r'#\w+#' not in ich and '//' not in ich:
                                self.vfile_data[ich] = [dircg,portnumg,kwd]
                                self.vfile_list.append(ich)
                    else:
                        kwd = 'None'
                        for ich in sline:
                            if '//' not in ich and ich != '':
                                self.vfile_data[ich] = [dircg,portnumg,kwd]
                                self.vfile_list.append(ich)                       
                else:
                    #sline = line.split(' +')
                    tline = line.split(' ')
                    sline = [x for x in tline if x != '']
                    if re.search(r'wire|logic|byte|bit|reg|tri1|tri0',line):               
                        if re.search(r'\[\d+:\d+\]',line):
                            lineg = ' '.join(sline[2:])
                            portnumg = sline[1]
                        else:
                            lineg = ' '.join(sline[1:])
                            #portnum = '1'             
                    else:
                        if re.search(r'\[\d+:\d+\]',line):
                            lineg = ' '.join(sline[1:])
                            portnumg = sline[0]
                        else:
                            lineg = ' '.join(line[0:])
                            #portnum = '1'

                    sigchar = lineg.replace(' ','')
                    if re.search(r'\/\/#\w+#',sigchar):
                        #kwd = ''.join(re.findall(r'\/\/#\w+#', sigchar)).strip().split('#')[1]
                        #kwdg = ''.join(re.findall(r'\/\/(#\w+#)+', sigchar)).strip()
                        if '##' in sigchar:
                            kwd = sigchar.replace('##',' ').split('#')[1]
                        else:
                            kwd = sigchar.split('#')[1]
                        
                        if ',' in sigchar:
                            sigcharg = sigchar.split(',')
                            #kwd = ''.join(re.findall(r'\/\/#\w+#', sigchar)).strip().split('#')[1]
                            for ich in sigcharg:
                                if r'#\w+#' not in ich and '//' not in ich:
                                    self.vfile_data[ich] = [dircg,portnumg,kwd]
                                    self.vfile_list.append(ich)
                        else:
                            sdc_warn(f'{sigchar} not found , symbol ...')
                            ish = sigchar.split('//')[0].strip()
                            self.vfile_data[ish] = [dircg,portnumg,kwd]
                            self.vfile_list.append(ish)
                    else:
                        kwd = 'None'
                        if ',' in sigchar:
                            sigcharg = sigchar.split(',')                           
                            for ich in sigcharg:
                                if '//' not in ich and ich != '':
                                    self.vfile_data[ich] = [dircg,portnumg,kwd]
                                    self.vfile_list.append(ich)
                        else:
                            sdc_warn(f'{sigchar} not found , symbol ...')
                            if '//' in sigchar:
                                ish = sigchar.split('//')[0].strip()
                            else:
                                ish = sigchar.strip()
                            self.vfile_data[ish] = [dircg,portnumg,kwd]
                            self.vfile_list.append(ish)

        return self.vfile_list, self.vfile_data

        


    def read_yaml(self, yaml_file):

        yaml_data = {}
        if not os.path.exists(yaml_file):
            raise FileExistsError(f'{yaml_file} does not exists')
        with open(yaml_file, 'r') as fh:
            yaml_data = yaml.load(fh, yaml.FullLoader)

        return yaml_data
    


    def read_text(self, file):
        if not os.path.exists(file):
            raise FileExistsError(f'{file} does not exists')
            # sdc_error(f'{file} not exist. Please check it.')
            # exit(1)
        else:
            txt_list = []
            with open(file,'r') as fh:
                for line in fh.readlines():
                    if line.strip() == "":
                        continue
                    if line.strip().startswith("//") and '#RelClock:' not in line.strip():
                         continue   
                    line = re.sub(r"\[\s*(\d+)\s*:\s*(\d+)\s*\]", r"[\1:\2]", line)
                    txt_list.append(line.strip())
        
            return txt_list

 



 