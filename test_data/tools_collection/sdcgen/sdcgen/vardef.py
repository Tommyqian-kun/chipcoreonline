

import sys
import time
import os
import re

from os.path import dirname, abspath, basename

import  openpyxl

from .basesdc import *
from com.base import *
from .clkdef import *



class VarDefSheet(BaseSheet):
    def __init__(self,*args):
        super().__init__(*args)  
        self._valdata = {}  
        self._vardata = {}
        #self._clkdata = self.get_table_contxt(self._sdcdg._wb['ClkDef'])
        #self._clkdef = None

        self._vfdata = self._sdcdg._vfile_data
        self._hiertree = self._sdcdg._hier_tree
        self._sdcdir = self._sdcdg._sdcdir
        self._mdname = self._sdcdg._mdname

    def update_sheet(self):
        '''
        # only during -dg option
        # addition of module name value from vfile
        # addition of user_defined variables
        # addition of block hier tree expanded table from hier yaml
        '''
        sheet = self.get_sheet()

        hiertree = self._sdcdg._hier_tree

        # find TMVAR table
        start_rowg = self.find_sheet(sheet, 'TMVAR')
        
        mdname = self._sdcdg._vfile_data['module_name']
        if mdname:
            sheet.cell(start_rowg + 1, 2).value = mdname

        vardef = self.get_vardef_value(sheet)
     
        # varlist = ['T28','T16','T7','T4']
        # self.add_dropdown(sheet, '"' + ','.join(varlist) + '"', [start_rowg + 2], [2])
        # varlist = ['RTL','SYN','DFT_SYN','SIM','PLA','CTS','PnR','SIGNOFF']
        # self.add_dropdown(sheet, '"' + ','.join(varlist) + '"', [start_rowg + 4], [2])
        # varlist = ['DC WLM','DC SPG','GNS PLE','GNS ISP']
        # self.add_dropdown(sheet, '"' + ','.join(varlist) + '"', [start_rowg + 5], [2])
        # varlist = ['full','local']
        # self.add_dropdown(sheet, '"' + ','.join(varlist) + '"', [start_rowg + 9], [2])
        varlist = ['70%','60%','50%','40%','30%','0','-10%']
        self.add_dropdown(sheet, '"' + ','.join(varlist) + '"', [start_rowg + 10], [2])
        varlist = ['20%','10%','0','-10%','20%']
        self.add_dropdown(sheet, '"' + ','.join(varlist) + '"', [start_rowg + 11], [2])

        block = hiertree.get_block_by_name(mdname)
        for i in range(start_rowg, sheet.max_row + 1):
            if sheet.cell(i, 1).value == 'SDC_DIR':
                sheet.cell(i, 2).value = block.constr_dir
        
        for i in range(1,start_rowg+11):
            for j in range(1,4):
                sheet.cell(start_rowg+i,j).alignment = Alignment(horizontal='left',vertical='center',wrapText=True)


        # find TMHIER table
        # start_rowg = self.find_sheet(sheet, 'TMHIER')
        #
        # expd_style = vardef['HIER_EXPD_STYLE']
        #
        # hierblks = hiertree.get_hierblks(mdname)
        # #print(hierblks)
        #
        # hiertrees = hiertree.get_hiertrees(mdname)
        # #print(hiertrees)
        #
        # hierdepth = hiertree.get_hierdepth(hiertrees,mdname)
        # #print(hierdepth)
        #
        # # outtype is hd/lib/soft
        # hdblks = hiertree.get_hierlvlblks(mdname,outtype='hd')
        # #print(hdblks)
        # macblks = hiertree.get_hierlvlblks(mdname,outtype='lib')
        # #print(macblks)
        # digblks = hiertree.get_hierlvlblks(mdname,outtype='soft')
        # #print(digblks)
        #
        # sheet.cell(start_rowg, 1).value = 'Item'
        # for i in range(2, hierdepth + 2):
        #     sheet.cell(start_rowg, i).value = f'BlkLevel{i-1}'
        # sheet.cell(start_rowg, hierdepth + 2).value = 'Comment'
        # for i in range(1, 12):
        #     sheet.cell(start_rowg+i, 1).value = f'Index{i}'
        #
        # self.cell_style2(sheet, [start_rowg, 1], [start_rowg + 15, hierdepth + 2])
        #
        # blksinfos = hiertree.get_hierblks_infos(mdname)
        #
        # # BlKLevel1
        # sheet.cell(start_rowg+1, 2).value = blksinfos[mdname]
        # # BlkLevel2
        # uniq_blk2 = list(set(hiertrees[mdname]))
        # self.fill_pmhier_col(sheet,start_rowg,hdblks,macblks,digblks,uniq_blk2,blksinfos,3)
        # #print(uniq_blk2)
        #
        # # blk3/blk4/blk5/...
        # if expd_style == 'full':
        #     # blk3
        #     # for i in range(3,hierdepth+1):
        #     #     varnm = f'uniq_blk{i-1}'
        #     #     exec(f'{varnm}')
        #     #     uniq_blk= self.get_uniq_blks(varnm,hiertrees)
        #     #     self.fill_pmhier_col(sheet,start_rowg,hdblks,macblks,digblks,uniq_blk,blksinfos)
        #     #     print(uniq_blk)
        #     if hierdepth > 2:
        #         uniq_blk3= self.get_uniq_blks(uniq_blk2,hiertrees)
        #         self.fill_pmhier_col(sheet,start_rowg,hdblks,macblks,digblks,uniq_blk3,blksinfos,4)
        #
        #     if hierdepth > 3:
        #         uniq_blk4= self.get_uniq_blks(uniq_blk3,hiertrees)
        #         self.fill_pmhier_col(sheet,start_rowg,hdblks,macblks,digblks,uniq_blk4,blksinfos,5)
        #
        #     if hierdepth > 4:
        #         uniq_blk5= self.get_uniq_blks(uniq_blk4,hiertrees)
        #         self.fill_pmhier_col(sheet,start_rowg,hdblks,macblks,digblks,uniq_blk5,blksinfos,6)
        #
        #     if hierdepth > 5:
        #         uniq_blk6= self.get_uniq_blks(uniq_blk5,hiertrees)
        #         self.fill_pmhier_col(sheet,start_rowg,hdblks,macblks,digblks,uniq_blk6,blksinfos,7)
        #
        #     if hierdepth > 6:
        #         uniq_blk7= self.get_uniq_blks(uniq_blk6,hiertrees)
        #         self.fill_pmhier_col(sheet,start_rowg,hdblks,macblks,digblks,uniq_blk7,blksinfos,8)
        #
        # hierinfogs = ' '.join([f'{key}: {val}' for key,val in hiertrees.items()])
        # sheet.cell(start_rowg+12, 1).value = hierinfogs
        # sheet.merge_cells(start_row=start_rowg + 12, end_row=start_rowg + 15,start_column=1,end_column=hierdepth+2)
        #
        # self.cell_style1(sheet, [start_rowg, 1], [start_rowg, hierdepth + 2])
        # for i in range(1,start_rowg+15):
        #     for j in range(1,hierdepth+1):
        #         sheet.cell(start_rowg+i,j).alignment = Alignment(horizontal='left',vertical='center',wrapText=True)
        #
        # sheet.column_dimensions[get_column_letter(1)].width = 20
        # for row in range(start_rowg+1,start_rowg+15):
        #     sheet.row_dimensions[row].height = 25
        #
        # for col in range(2,hierdepth+1):
        #     sheet.column_dimensions[get_column_letter(col)].width = 35

    # def get_uniq_blks(self,blks,hiertrees):
    #     uniq_blk3g = []
    #     for blk3 in blks:
    #         if blk3 in hiertrees:
    #             uniq_blk3g.append([x for x in hiertrees[blk3]])
    #         # else:
    #         #     uniq_blk3g.append(blk3.split())
    #     uniq_blk3 = list(set(item for sublist in uniq_blk3g for item in sublist))
    #
    #     return uniq_blk3


    # def fill_pmhier_col(self,sheet,start_rowg,hdblks,macblks,digblks,uniq_blk,blksinfos,col):
    #
    #     hd_blk2 = [x for x in uniq_blk if x in hdblks ]
    #     mac_blk2 = [x for x in uniq_blk if x in macblks ]
    #     dig_blk2 = [x for x in uniq_blk if x in digblks ]
    #
    #     hd_num = len(hd_blk2)
    #     mac_num = len(mac_blk2)
    #     dig_num = len(dig_blk2)
    #     chg_row = start_rowg
    #     if hd_num > 0:
    #         for i in range(1, hd_num + 1):
    #             sheet.cell(start_rowg+i, col).value = blksinfos[hd_blk2[i-1]]
    #         chg_row = start_rowg + hd_num
    #     if mac_num > 0:
    #         for i in range(1, mac_num + 1):
    #             sheet.cell(chg_row+i, col).value = blksinfos[mac_blk2[i-1]]
    #         chg_row = chg_row + mac_num
    #     if dig_num > 0:
    #         for i in range(1, dig_num + 1):
    #             sheet.cell(chg_row+i, col).value = blksinfos[dig_blk2[i-1]]
    #         chg_row = chg_row + mac_num


###########################################################

    def read_data(self):
        sheet = self.get_sheet()
        self._valdata = self.get_table_contxt(sheet)
        nvaldata = {}
        nvaldata["TMVAR_Row14"] = {
            "Variable": "SDC_DIR",
            "Value": f'{self._sdcdir}',
            "Comment": ''
        }
        nvaldata["TMVAR_Row15"] = {
            "Variable": "COM_DIR",
            "Value": f'{self._sdcdir}',
            "Comment": ''
        }
        nvaldata["TMVAR_Row16"] = {
            "Variable": "DFT_DIR",
            "Value": '',
            "Comment": ''
        }
        # nvaldata["TMVAR_Row17"] = {
        #     "Variable": "HD_MOD_NAME",
        #     "Value": f'{self._mdname}',
        #     "Comment": ''
        # }
        nvaldata["TMVAR_Row17"] = {
            "Variable": "HD_PROCESS",
            "Value": '',
            "Comment": ''
        }
        nvaldata["TMVAR_Row18"] = {
            "Variable": "CYCLE_LIST",
            "Value": '[list CYCLE500M]',
            "Comment": ''
        }
        self._valdata.update(nvaldata)

    def check_sheet(self):
        pass

    def change_sheet(self):
        pass

    def dump_json(self,json_file):
        self._data = self._valdata
        self.write_json(json_file)

    def write_sdc(self,sdc_dir,prousr=False):
        sheet = self.get_sheet()
        self._vardata = self.get_vardef_value(sheet)
        mdname = self._sdcdg._vfile_data['module_name']
        alias = self._sdcdg._hier_tree._blocks[mdname].alias
        hdlvl = self._sdcdg._hier_tree._blocks[mdname].hdlevel
        pwr = self._sdcdg._hier_tree._blocks[mdname].prime_pwr
        if hdlvl == 'sys':
            lvl = 'sys'
        else:
            lvl = 'blk'

        sdc_file = sdc_dir + f'{alias.lower()}_{lvl}var.sdc'
        
        # mdname = self._sdcdg._vfile_data['module_name']
        # alias = self._sdcdg._hier_tree._blocks[mdname].alias
        # blklvl = self._sdcdg._hier_tree._blocks[mdname].hdlevel
        #dcdcnm = self._sdcdg._hier_tree._blocks[mdname].prime_pwr.split(' ')[0]
        #dcdcvl = self._sdcdg._hier_tree._blocks[mdname].prime_pwr.split(' ')[1:]
        hierdcdc = self._sdcdg._hier_tree.get_hier_dcdc(mdname,hflg=False)
        #self._clkdata = self.get_table_contxt(self._sdcdg._wb['ClkDef'])
        var_lines = self.get_vars(mdname,alias,lvl,pwr,hierdcdc,prousr,False)
        self.save_text(var_lines,sdc_file)

        #hierdcdc = self._sdcdg._hier_tree.get_hier_dcdc(mdname,hflg=True)
        varintg_lines = self.get_vars(mdname,alias,lvl,pwr,hierdcdc,prousr,True)
        sdc_file = sdc_dir + f'../intg/{alias.lower()}_{lvl}var_intg.sdc'
        self.save_text(varintg_lines,sdc_file)
        
    def get_vars(self,mdname,alias,lvl,pwr,hierdcdc,prousr=False,fintg=False):
        cur_clkdef = self._sdcdg._sheets['ClkDef']
        var_lines = ''

        var_lines += f'''
############################################################
## Variables Definition 
############################################################

if {{${{{alias}}} == ""}} {{
    echo "SDC_ERROR: ${{{alias}}} value is not set. Please check it."
}}
'''

        for dcdc in hierdcdc:
            dcnm = dcdc.split(' ')[0]
            #dcvl = dcdc.split(' ')[1]
            var_lines += f'''
if {{${{{dcnm}}} == ""}} {{
    echo "SDC_ERROR: ${{{dcnm}}} value is not set. Please check it."
}}
'''
        # var_lines += self.var_out1(f'HD_MODE_NAME,${{{alias}}}',self._vardata['HD_MOD_NAME'])
        var_lines += self.var_out1(f'HD_PROCESS,${{{alias}}}',self._vardata['HD_PROCESS'])

        if lvl == 'sys':
            var_lines += self.var_out1(f'SysNm,${{{alias}}}',alias)
            var_lines += self.var_out1(f'HIER,SYS,${{{alias}}}','',judge_fg=0)
            var_lines += self.var_out1(f'IS_CHIP,${{{alias}}}','0')
            var_lines += self.var_out1(f'LIB,${{{alias}}}','0')
        if lvl == 'blk':
            var_lines += self.var_out1(f'BlkNm,${{{alias}}}',alias)
            var_lines += self.var_out1(f'HIER,BLK,${{{alias}}}','',judge_fg=0)
            var_lines += self.var_out1(f'IS_FLAT,${{{alias}}}','0')  
            var_lines += self.var_out1(f'LIB,${{{alias}}}','0')

        var_lines += self.var_out1(f'SYN_WLM_SEL',self._vardata['SYN_WLM_SEL']) 
        var_lines += self.var_out1(f'FL_STAGE',self._vardata['FL_STAGE'])
        

        for key,val in self._vardata.items():
            if key and key not in ['HD_MOD_NAME','HD_PROCESS','SYN_WLM_SEL','FL_STAGE','CYCLE_LIST']:
                var_lines += self.var_out1(f'{key},${{{alias}}}',val)  
                # if re.search(r'SDC_DIR',key):
                #     var_lines += self.var_out1(f'{key},{alias}',f'{val}/sdc/outputs/')
                # if re.search(r'DFT_DIR|COM_DIR',key):
                #     var_lines += self.var_out1(f'{key},{alias}',val)
                if re.search(r'HIER_BLK|HIER_SYS',key):
                    var_lines += f'''
set {key} "{val}"
'''
        crgipclk_info = {}
        if cur_clkdef._crgalsiptval:
            crgipclk_info.update(cur_clkdef._crgalsiptval)
        if cur_clkdef._ipalsiptval:
            crgipclk_info.update(cur_clkdef._ipalsiptval)
        if crgipclk_info:
            #print('crgipclk_info+++++++',crgipclk_info.keys())
            #print('_crgals+++++++++++++++',cur_clkdef._crgals,cur_clkdef._ipals)
            for ky,vl in crgipclk_info.items():
                if vl:
                    val = ky.split('_')[1]
                    var = ky.split('_')[2]
                    kw = ky.split('_')[3]
                    #cival = vl[0][0].split('_')[1]
                    if 'CRG' in kw or 'PLL' in kw:
                        civar = cur_clkdef._crgals[ky]
                    if 'MACLIB' in kw or 'DIGSOFT' in kw:
                        civar = cur_clkdef._ipals[ky]
                    #print('civar++++++++++++++:',civar,var)
                    if not civar in var:
                        var_lines += f'''
# {ky} variable in hier yaml: {var}
# {ky} value in hier yaml: {val}
# {ky} variable in sdc: {civar}
if {{![info exists {var}]}} {{
    echo "SDC_ERROR: ${{{var}}} value is not set. please check it"
}} else {{
    set {civar} "${{{var}}}"
}}

if {{${civar} == ""}} {{
    echo "SDC_ERROR: ${{{civar}}} value is not set. please check it"
}}
'''
                    else:
                        sdc_error(f'Mismatch crgip alias value bet hier yaml and crgip sdc for {kw}')
                else:
                    sdc_warn(f'Missing {ky} header info. or related sdc file.')

        var_lines += f'''
###########################################################
## CP/Q pin, need to follow change name in gtech file
set DataPin "D"; # for STA/PLA/CTS/PnR/SIGNOFF
set ClkPin "CP"; # for STA/PLA/CTS/PnR/SIGNOFF
if {{[lsearch $SDCVAR(SYN_WLM_SEL) "Genus"] != -1 && ($SDCVAR(FL_STAGE) == "RTL" || $SDCVAR(FL_STAGE) == "SYN" || $SDCVAR(FL_STAGE) == "MBIST_SYN")}} {{
        set DataPin "D"
        set ClkPin "CP"
}}
if {{[lsearch $SDCVAR(SYN_WLM_SEL) "DC"] != -1 && ($SDCVAR(FL_STAGE) == "RTL" || $SDCVAR(FL_STAGE) == "SYN" || $SDCVAR(FL_STAGE) == "MBIST_SYN")}} {{
        set DataPin "next_state"
        set ClkPin "clocked_on"
}}
if {{[lsearch $SDCVAR(SYN_WLM_SEL) "FC"] != -1 && ($SDCVAR(FL_STAGE) == "RTL" || $SDCVAR(FL_STAGE) == "SYN" || $SDCVAR(FL_STAGE) == "MBIST_SYN")}} {{
        set DataPin "next_state"
        set ClkPin "clocked_on"
}}

'''
        for dcdc in hierdcdc:
            #print(hierdcdc)
            dcnm = dcdc.split(' ')[0]
            dcvl = dcdc.split(' ')[1]
            var_lines += self.var_out1(f'DCDC_VL,${{{dcnm}}}',dcvl)

        
        if not cur_clkdef._clknmlst:
            cur_clkdef.get_clkdata_by_clkname(cur_clkdef._clkdata)
        clknmlst = cur_clkdef._clknmlst
        clknmdata = cur_clkdef._clknmdata   
        #print('clkdef',self._clknmlst)     
        cur_clkdef.get_cycle_from_clkdef()
        cur_clkdef.get_cycle_from_crgip()
        #print('_cycle_clkdeflst',cur_clkdef._cycle_clkdeflst)
        #print('_cycle_crgiplst',cur_clkdef._cycle_crgiplst)
        cycles = cur_clkdef._cycle_clkdeflst + cur_clkdef._cycle_crgiplst
        #print(cycles)
        if cycles:
            cyclst = list(set(cycles))
        cycvarlst =[x for x in self._vardata['CYCLE_LIST'].split(' ')[1:] if x]
        allcycs = cyclst + cycvarlst
        cycle_str = '[list '
        for cyc in allcycs:
            cycle_str = cycle_str + f'{cyc} '
        if ']' not in cycle_str:
            cycle_str = cycle_str.strip() + ']'
        cycle_str = cycle_str.rstrip()

        #print(cycle_str)
        var_lines += f'set SDCVAR(CYCLE_LIST) "{cycle_str}"; ### set "CYCLE_LIST"'
        var_lines += f'''
period_def $SDCVAR(CYCLE_LIST)
'''

        var_lines += f'''
############################################################
## Clock Definition From ClkDef
############################################################
'''

#################################################################
    # clkintg: all clk vars & definition
    #   -CRGIN/IPIN/HDIN    ----> blkvar
    #   -CRGOUT             ---> updt
    #   -IPOUT              ---> subblk
    #   -HDOUT              ----> updt or subblk(nocrg)
    # outclktype from crg: only mstclk/srcpin/grpnm
    #   -CRGIN/HDIN/IPIN    -----> updt
    # outclktype from ip: only mstclk/srcpin/grpnm
    #   -CRGIN/HDIN/IPIN    -----> subblk    
#################################################################
        #clkdata = self.get_table_contxt(self._sdcdg._wb['ClkDef'])
        #cur_clkdef.get_clkdata_by_clkname(clkdata)   
        nclknmlst = []
        nclknmdata = {}
        for clknm in clknmlst:
            if not cur_clkdef.is_crgiphd_goutclk(clknm,'CRGOUT') and not cur_clkdef.is_crgiphd_goutclk(clknm,'IPOUT') and not cur_clkdef.is_crgiphd_goutclk(clknm,'HDOUT'):
                nclknmlst.append(clknm)
                nclknmdata[clknm] = clknmdata[clknm]     
        var_lines += cur_clkdef.get_clkvar(mdname,alias,pwr,nclknmlst,nclknmdata,prousr,fintg,'CIHIN')
        #var_lines += cur_clkdef.get_clkvar(mdname,alias,pwr,clknmlst,clknmdata,fintg,'CRGIN')
        # var_lines += cur_clkdef.get_clkvar(mdname,alias,pwr,clknmlst,clknmdata,fintg,'IPIN')
        # var_lines += cur_clkdef.get_clkvar(mdname,alias,pwr,clknmlst,clknmdata,fintg,'HDIN')
        #print(var_lines)

        return  var_lines

    # def get_vars_intg(self,mdname,alias,lvl,hierdcdc):
        
    #     var_lines = ''

    #     #cur_clkdef = self.get_table_contxt(self._sdcdg._wb['ClkDef'])
    #     cur_clkdef.get_clkvar_intg(mdname,self._hiertree,self._vfdata)
    #     var_lines += cur_clkdef._clkvarlines

    #     return var_lines

##############################################

    def var_out1(self,var,value,judge_fg=1):
        flines = ''
        if 'DCDC' in var:
            varg = 'DCDC_VL: ' + var.split(',')[1]
        elif ',' in var:
            varg = var.split(',')[0]
        else:
            varg = var
        flines += f'''
# {varg}
if {{[info exists SDCVAR({var})]}} {{
}} else {{
    set SDCVAR({var})   "{value}"; ## set "{var}"
}}
'''

        if judge_fg == 1:
            flines += f'''
if {{$SDCVAR({var}) == ""}} {{
    echo "SDC_ERROR: SDCVAR({var}) value is not set. please check it"
}}

'''
        return flines
    




