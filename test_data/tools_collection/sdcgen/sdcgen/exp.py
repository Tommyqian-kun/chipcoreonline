
import sys
import time
import os
import re

from os.path import dirname, abspath, basename

import  openpyxl

from .basesdc import *
from com.base import *
from .clkdef import *


class ExpSheet(BaseSheet):
    def __init__(self ,*args):
        super().__init__(*args)
        self._expdata = {}
        self._hier_tree = self._sdcdg._hier_tree
        # self._clkdef = None
        self._vardata = self.get_vardef_value(self._sdcdg._wb['VarDef'])

        self._lvl = 'blk'
        self._flt = 'IS_FLAT'

        self._mdname = ''

    def update_sheet(self):
        '''
        # only during -dg option
        # addition of module name value from vfile
        # addition of block hier tree expanded table from hier yaml
        '''
        sheet = self.get_sheet()

        hiertree = self._hier_tree
        vdata = self._sdcdg._vfile_data
        vlist = self._sdcdg._vfile_list

        # find TMIOEXP table
        start_rowg = self.find_sheet(sheet, 'TMIOEXP')

        nvdata = {}
        nvlist = []
        for kwd in vlist:
            if 'module_name' not in kwd and 'RelClock' not in kwd:
                if 'TCLK' not in vdata[kwd][2] and re.search(r'IDEAL|CASEXP|FPEXP|MCPEXP' ,vdata[kwd][2]):
                    # 'IDEAL' in vdata[kwd][2] or 'FPEXP' in vdata[kwd][2] or 'MCPEXP' in vdata[kwd][2] or 'CASEXP' in vdata[kwd][2]:
                    nvlist.append(kwd)
                    nvdata[kwd] = vdata[kwd]

        sheet.insert_rows(start_rowg + 9, len(nvlist) + 1)
        self.cell_style2(sheet, [start_rowg + 9, 1], [start_rowg + 9 + len(nvlist) + 1, 10])

        varlist = ['Y']
        self.add_dropdown(sheet, '"' + ','.join(varlist) + '"', [start_rowg + 1 ,3], [start_rowg + len(nvlist) + 11 ,3])
        varlist = ['0' ,'1']
        self.add_dropdown(sheet, '"' + ','.join(varlist) + '"', [start_rowg + 1 ,4], [start_rowg + len(nvlist) + 11 ,4])
        varlist = ['-setup' ,'-hold' ,'all']
        self.add_dropdown(sheet, '"' + ','.join(varlist) + '"', [start_rowg + 1 ,5], [start_rowg + len(nvlist) + 11 ,5])
        varlist = ['-start 2 1' ,'-end 2 1' ,'-start NA 1' ,'-end 2 NA']
        self.add_dropdown(sheet, '"' + ','.join(varlist) + '"', [start_rowg + 1 ,6], [start_rowg + len(nvlist) + 11 ,6])

        varlist = ['pin [list ]' ,'clk [list ]']
        self.add_dropdown(sheet, '"' + ','.join(varlist) + '"', [start_rowg + 1 ,7], [start_rowg + len(nvlist) + 11 ,7])
        varlist = ['pin [list ]']
        self.add_dropdown(sheet, '"' + ','.join(varlist) + '"', [start_rowg + 1 ,8], [start_rowg + len(nvlist) + 11 ,8])
        varlist = ['pin [list ]' ,'clk [list ]']
        self.add_dropdown(sheet, '"' + ','.join(varlist) + '"', [start_rowg + 1 ,9], [start_rowg + len(nvlist) + 11 ,9])


        i = 0
        for kwd in nvlist:
            if i < len(nvlist) + 1:
                # expkw = 'module_name' not in kwd and 'RelClock' not in kwd and 'IDEAL' not in vdata[kwd][2] and 'TCLK' not in vdata[kwd][2] and 'ANA' not in vdata[kwd][2]

                if 'CASEXP'  in nvdata[kwd][2] and 'FPEXP'  in nvdata[kwd][2]:
                    sdc_warn(f'port {kwd} includes CASEXP and FPEXP.')
                elif 'MCPEXP'  in nvdata[kwd][2] and 'FPEXP'  in nvdata[kwd][2]:
                    sdc_error(f'port {kwd} includes MCPEXP and FPEXP.')
                elif 'CASEXP'  in nvdata[kwd][2] and 'MCPEXP'  in nvdata[kwd][2]:
                    sdc_warn(f'port {kwd} includes CASEXP and MCPEXP.')

                if re.search(r'IDEAL|CASEXP|FPEXP|MCPEXP' ,nvdata[kwd][2]):
                    i += 1
                    if vdata[kwd][1] != '1':
                        sheet.cell(start_rowg + i ,1).value = kwd + nvdata[kwd][1]
                    else:
                        sheet.cell(start_rowg + i ,1).value = kwd
                    sheet.cell(start_rowg + i ,2).value = nvdata[kwd][0]
                    if 'IDEAL' in nvdata[kwd][2]:
                        sheet.cell(start_rowg + i ,3).value = 'Y'


        # find TMINOUT table
        start_rowg = self.find_sheet(sheet, 'TMINOUT')

        varlist = ['70%' ,'60%' ,'50%' ,'40%' ,'30%' ,'0' ,'-10%']
        self.add_dropdown(sheet, '"' + ','.join(varlist) + '"', [start_rowg + 1 ,3], [start_rowg + 10 ,4])
        varlist = ['0.2' ,'0.1' ,'0.05']
        self.add_dropdown(sheet, '"' + ','.join(varlist) + '"', [start_rowg + 1 ,5], [start_rowg + 10 ,5])

        clkdef = self._sdcdg._sheets['ClkDef']
        # varlist = clkdef.get_clkinfo_from_crgip('IO','0')
        # varlist = clkdef.get_clkinfo_from_crgip('IO','0')
        clknm = clkdef.get_clkinfo_from_crgip('0' ,'IO')
        alsck = []
        varlist = []
        varlistg = []
        for x in clknm:
            for k ,v in x.items():
                cials = k.split(' ')[0].split('_')[1]
                nalsck = [f'{cials} {x}' for x in v]
                alsck.extend(nalsck)
        varlist.extend(alsck)
        for kwd in vlist:
            if 'RelClock' in kwd:
                varlistg.append(vdata[kwd])
        if varlistg:
            varlistg = list(set(varlistg))
            varlist.extend(varlistg)
        self.add_dropdown(sheet, '"' + ','.join(varlist) + '"', [start_rowg + 1 ,6], [start_rowg + 10 ,6])

    ######################################################
    def read_data(self):
        sheet = self.get_sheet()
        self._expdata = self.get_table_contxt(sheet)

    def check_sheet(self):
        pass

    def change_sheet(self):
        pass

    def dump_json(self ,json_file):
        self._data = self._expdata
        self.write_json(json_file)

    ######################################################
    def write_sdc(self, sdc_dir,prousr=False):
        sheet = self.get_sheet()

        mdname = self._sdcdg._vfile_data['module_name']
        alias = self._hier_tree._blocks[mdname].alias
        hdlvl = self._hier_tree._blocks[mdname].hdlevel
        pwr = self._hier_tree._blocks[mdname].prime_pwr

        self._mdname = mdname

        if hdlvl == 'sys':
            self._lvl = 'sys'
            self._flt = 'IS_CHIP'
        if hdlvl == 'blk':
            self._lvl = 'blk'
            self._flt = 'IS_FLAT'

        # sdc_file = sdc_dir + f'{alias.lower()}_exp.sdc'
        # self.write_exp(mdname, alias, sdc_file, False)

        sdc_file = sdc_dir + f'{alias.lower()}_ioexp.sdc'
        self.write_ioexp(mdname, alias, sdc_file, prousr, fintg=False)
        sdc_file = sdc_dir + f'../intg/{alias.lower()}_ioexp_intg.sdc'
        self.write_ioexp(mdname, alias, sdc_file, prousr, fintg=True)

        sdc_file = sdc_dir + f'{alias.lower()}_intexp.sdc'
        self.write_intexp(mdname, alias, sdc_file, prousr, fintg=False)
        sdc_file = sdc_dir + f'../intg/{alias.lower()}_intexp_intg.sdc'
        self.write_intexp(mdname, alias, sdc_file, prousr, fintg=True)


    def write_exp(self, mdname, alias, sdc_file, prousr=False, fintg=False):

        exp_lines = ''

        # internal expception
        intexp_lines = ''

        # io exception
        ioexp_lines = ''

        iofp_rows = {}
        iomcp_rows = {}
        ioideal_rows = {}
        iocase_rows = {}
        ioinout_rows = {}
        for kw, vl in self._expdata.items():
            num = kw.split('Row')[1]
            if re.search(r'TMIOEXP_Row', kw):
                # PortNm	Direction	Ideal	CaseVal	FP	MCP	From	Through	To	Comment
                if vl['Ideal'] == 'Y':
                    ioideal_rows[f'IDEAL_Row{num}'] = vl
                if re.search(r'0|1', str(vl['CaseVal'])):
                    iocase_rows[f'CASE_Row{num}'] = vl
                if vl['FP']:
                    iofp_rows[f'FP_Row{num}'] = vl
                if vl['MCP']:
                    iomcp_rows[f'MCP_Row{num}'] = vl

            if re.search(r'TMINOUT_Row', kw):
                # PortIn	PortOut	DlyIn	DlyOut	RealDly	ClkNm	Vol	Comment
                ioinout_rows[f'INOUT_Row{num}'] = vl

        ioexp_lines += f'''
        ################################################
        ## IO False Path
        ################################################
        '''
        ioexp_lines += self.set_iocmd(iofp_rows, 'set_false_path', alias, fintg)

        ioexp_lines += f'''
        ################################################
        ## IO Multicycle Path
        ################################################
        '''
        ioexp_lines += self.set_iocmd(iomcp_rows, 'set_multicycle_path', alias, fintg)

        ioexp_lines += f'''
        ################################################
        ## IO Case Setting
        ################################################
        '''
        ioexp_lines += self.set_iocmd(iocase_rows, 'set_case_analysis', alias, fintg)

        ioexp_lines += f'''
        ################################################
        ## IO Ideal Network
        ################################################
        '''
        ioexp_lines += self.set_iocmd(ioideal_rows, 'set_ideal_network', alias, fintg)


        # ioexp_lines += f'''
        # ################################################
        # ## INOUT Max Delay Path
        # ################################################
        # '''
        # ioexp_lines += self.set_iocmd(ioinout_rows, 'set_max_delay', alias, fintg)

        # sub harden blk
        # if not fintg:
        #     blkf = 'ioexp'
        #     ioexp_lines += self._hier_tree.set_subblk_intg(mdname, blkf)

        # self.save_text(ioexp_lines, sdc_file)
        exp_lines += ioexp_lines


        # print('_intexpdata',self._intexpdata)
        intfp_rows = {}
        intmcp_rows = {}
        intcase_rows = {}
        intstpgt_rows = {}
        for kw, vl in self._expdata.items():
            num = kw.split('Row')[1]
            if re.search(r'TMINTEXP_Row', kw):
                # FP	MCP	CaseVal	CasePin	From	Through	To	Comment
                if re.search(r'0|1', str(vl['CaseVal'])):
                    intcase_rows[f'CASE_Row{num}'] = vl
                if vl['FP']:
                    intfp_rows[f'FP_Row{num}'] = vl
                if vl['MCP']:
                    intmcp_rows[f'MCP_Row{num}'] = vl

            if re.search(r'TMSTPGATE_Row', kw):
                # StopClk	StopPin	DisClkGating	Comment
                intstpgt_rows[f'STPGATE_Row{num}'] = vl
        # print('intfp_rows:',intfp_rows)

        intexp_lines += f'''
    ################################################
    ## Internal False Path
    ################################################
    '''
        intexp_lines += self.set_intcmd(intfp_rows, 'set_false_path', alias, fintg)

        intexp_lines += f'''
    ################################################
    ## Internal Multicycle Path
    ################################################
    '''
        intexp_lines += self.set_intcmd(intmcp_rows, 'set_multicycle_path', alias, fintg)

        intexp_lines += f'''
    ################################################
    ## Internal Case Setting
    ################################################
    '''
        intexp_lines += self.set_intcmd(intcase_rows, 'set_case_analysis', alias, fintg)

        intexp_lines += f'''
    ################################################
    ## Internal STP and DisGating
    ################################################
    '''
        intexp_lines += self.set_intcmd(intstpgt_rows, 'set_sense', alias, fintg)
        intexp_lines += self.set_intcmd(intstpgt_rows, 'set_disable_clock_gating_check', alias, fintg)

        # sub harden blk
        # if not fintg:
        #     blkf = 'intexp'
        #     intexp_lines += self._hier_tree.set_subblk_intg(mdname, blkf)

        exp_lines += intexp_lines

        self.save_text(exp_lines, sdc_file)

    ######################################################
    ## write io exception
    def write_sdc_io(self, sdc_dir):
        sheet = self.get_sheet()

        mdname = self._sdcdg._vfile_data['module_name']
        alias = self._hier_tree._blocks[mdname].alias
        hdlvl = self._hier_tree._blocks[mdname].hdlevel
        pwr = self._hier_tree._blocks[mdname].prime_pwr

        self._mdname = mdname

        if hdlvl == 'sys':
            self._lvl = 'sys'
            self._flt = 'IS_CHIP'
        if hdlvl == 'blk':
            self._lvl = 'blk'
            self._flt = 'IS_FLAT'

        sdc_file = sdc_dir + f'{alias.lower()}_ioexp.sdc'
        self.write_ioexp(mdname, alias, sdc_file, fintg=False)
        sdc_file = sdc_dir + f'intg/{alias.lower()}_ioexp_intg.sdc'
        self.write_ioexp(mdname, alias, sdc_file, fintg=True)

    def write_ioexp(self, mdname, alias, sdc_file, prousr=False, fintg=False):
        # clkdata = self.get_table_contxt(self._sdcdg._wb['ClkDef'])
        # self._clkdef.get_clkdata_by_clkname(clkdata)

        clkdef = self._sdcdg._sheets['ClkDef']
        # print('self._expdata: ',self._expdata)

        ioexp_lines = ''

        iofp_rows = {}
        iomcp_rows = {}
        ioideal_rows = {}
        iocase_rows = {}
        ioinout_rows = {}
        for kw, vl in self._expdata.items():
            num = kw.split('Row')[1]
            if re.search(r'TMIOEXP_Row', kw):
                # PortNm	Direction	Ideal	CaseVal	FP	MCP	From	Through	To	Comment
                if vl['Ideal'] == 'Y':
                    ioideal_rows[f'IDEAL_Row{num}'] = vl
                if re.search(r'0|1', str(vl['CaseVal'])):
                    iocase_rows[f'CASE_Row{num}'] = vl
                if vl['FP']:
                    iofp_rows[f'FP_Row{num}'] = vl
                if vl['MCP']:
                    iomcp_rows[f'MCP_Row{num}'] = vl

            if re.search(r'TMINOUT_Row', kw):
                # PortIn	PortOut	DlyIn	DlyOut	RealDly	ClkNm	Vol	Comment
                ioinout_rows[f'INOUT_Row{num}'] = vl

        ioexp_lines += f'''
################################################
## IO False Path
################################################
'''
        ioexp_lines += self.set_iocmd(iofp_rows, 'set_false_path', alias, fintg)

        ioexp_lines += f'''
################################################
## IO Multicycle Path
################################################
'''
        ioexp_lines += self.set_iocmd(iomcp_rows, 'set_multicycle_path', alias, fintg)

        ioexp_lines += f'''
################################################
## IO Case Setting
################################################
'''
        ioexp_lines += self.set_iocmd(iocase_rows, 'set_case_analysis', alias, fintg)

        ioexp_lines += f'''
################################################
## IO Ideal Network
################################################
'''
        ioexp_lines += self.set_iocmd(ioideal_rows, 'set_ideal_network', alias, fintg)

        if prousr:
            ioexp_lines += f'''
################################################
## INOUT Max Delay Path
################################################
'''
            ioexp_lines += self.set_iocmd(ioinout_rows, 'set_max_delay', alias, fintg)

        # sub harden blk
        if not fintg:
            blkf = 'ioexp'
            ioexp_lines += self._hier_tree.set_subblk_intg(mdname, blkf)

        self.save_text(ioexp_lines, sdc_file)

    # PortNm	Direction	Ideal	CaseVal	FP	MCP	From	Through	To	Comment
    def set_iocmd(self, iodata, cmd, alias, fintg=False):
        clkdef = self._sdcdg._sheets['ClkDef']
        # cinmals = clkdef.get_crgip_clknm_alias(self._mdname)
        # print(clkdef._tclklst)
        # print('iodata: ',iodata)

        iolines = ''
        for kw, vl in iodata.items():
            rnum = kw.split('Row')[1]
            lvl = self._lvl.upper()

            if vl['Comment']:
                cmt = vl['Comment']
            else:
                cmt = 'NA'

            # case
            if cmd == 'set_case_analysis':
                if str(vl['CaseVal']):
                    caseval = vl['CaseVal']
                    rcmd = f'{cmd} {caseval}'
                    iolines += self.set_ideal_case(rnum, lvl, alias, vl, rcmd, cmt, fintg)

            # ideal
            if cmd == 'set_ideal_network' and not fintg:
                iolines += self.set_ideal_case(rnum, lvl, alias, vl, cmd, cmt, fintg)

            # fp
            if cmd == 'set_false_path':
                iolines += self.set_fp_path(rnum, lvl, alias, vl, cmd, cmt, fintg)

            # mcp
            if cmd == 'set_multicycle_path':
                iolines += self.set_mcp_path(rnum, lvl, alias, vl, cmd, cmt, fintg)

            # io maxdelay
            if cmd == 'set_max_delay' and not fintg:
                iolines += self.set_inout_maxdly(rnum, lvl, alias, vl, cmd)

        return iolines

    def set_fp_path(self, rnum, lvl, alias, vl, cmd, cmt, fintg=False):
        # clkdata = self.get_table_contxt(self._sdcdg._wb['ClkDef'])
        # self._clkdef.get_clkdata_by_clkname(clkdata)
        clkdef = self._sdcdg._sheets['ClkDef']
        # cinmals = clkdef.get_crgip_clknm_alias(self._mdname)

        if vl['PortNm']:
            portnm = vl['PortNm']
        else:
            portnm = ''
        if vl['Direction']:
            dirct = vl['Direction']
        else:
            dirct = ''
        if vl['FP']:
            fp = vl['FP']
        else:
            fp = ''
        if vl['From']:
            frm = vl['From']
        else:
            frm = ''
        if vl['Through']:
            thr = vl['Through']
        else:
            thr = ''
        if vl['To']:
            to = vl['To']
        else:
            to = ''

        iolines = ''
        if fp == 'all':
            fpval = ' '
        else:
            fpval = f'-{fp} '

        kws = self.set_frthrto_val(alias, lvl, dirct, frm, thr, to)
        psig, pstn, pedn, pnum = self.get_sig_num(portnm)

        # input
        if dirct == 'input':
            if not fintg:
                if not pnum:
                    iolines += f'''
# IOFP Row{rnum}
if {{!$SDCVAR({self._flt},${{{alias}}})}} {{
    {cmd} {fpval} -from [get_ports {portnm}] {kws} -comment "{cmt}"
}} else {{
    {cmd} {fpval} -through [get_pins $SDCVAR(HIER,{lvl},${{{alias}}}){portnm}] {kws} -comment "{cmt}"
}}
'''
                else:
                    iolines += f'''
# IOFP Row{rnum}
if {{!$SDCVAR({self._flt},${{{alias}}})}} {{
    for {{set i {pstn}}} {{$i <= {pedn}}} {{incr i}} {{
        {cmd} {fpval} -from [get_ports {psig}[i]] {kws} -comment "{cmt}"
    }}
}} else {{
    for {{set i {pstn}}} {{$i <= {pedn}}} {{incr i}} {{
        {cmd} {fpval} -through [get_pins $SDCVAR(HIER,{lvl},${{{alias}}}){psig}[i]] {kws} -comment "{cmt}"
    }}
}}
'''
            else:
                if not pnum:
                    iolines += f'''
# IOFP Row{rnum}
{cmd}  {fpval} -through [get_pins $SDCVAR(HIER,{lvl},${{{alias}}}){portnm}] {kws} -comment "{cmt}"
'''
                else:
                    iolines += f'''
# IOFP Row{rnum}
for {{set i {pstn}}} {{$i <= {pedn}}} {{incr i}} {{
    {cmd} {fpval} -through [get_pins $SDCVAR(HIER,{lvl},${{{alias}}}){psig}[i]] {kws} -comment "{cmt}"
}}
'''
                    # output
        if dirct == 'output':
            if not fintg:
                if not pnum:
                    iolines += f'''
# IOFP Row{rnum}
if {{!$SDCVAR({self._flt},${{{alias}}})}} {{
    {cmd} {fpval} {kws} -to [get_ports {portnm}] -comment "{cmt}"
}} else {{
    {cmd} {fpval} {kws} -through [get_pins $SDCVAR(HIER,{lvl},${{{alias}}}){portnm}] -comment "{cmt}"
}}
'''
                else:
                    iolines += f'''
# IOFP Row{rnum}
if {{!$SDCVAR({self._flt},${{{alias}}})}} {{
    for {{set i {pstn}}} {{$i <= {pedn}}} {{incr i}} {{
        {cmd} {fpval} {kws} -to [get_ports {psig}[i]] -comment "{cmt}"
    }}
}} else {{
    for {{set i {pstn}}} {{$i <= {pedn}}} {{incr i}} {{
        {cmd} {fpval} {kws} -through [get_pins $SDCVAR(HIER,{lvl},${{{alias}}}){psig}[i]] -comment "{cmt}"
    }}
}}
'''
            else:
                if not pnum:
                    iolines += f'''
# IOFP Row{rnum}
{cmd}  {fpval} {kws} -through [get_pins $SDCVAR(HIER,{lvl},${{{alias}}}){portnm}] -comment "{cmt}"
'''
                else:
                    iolines += f'''
# IOFP Row{rnum}
for {{set i {pstn}}} {{$i <= {pedn}}} {{incr i}} {{
    {cmd} {fpval} {kws} -through [get_pins $SDCVAR(HIER,{lvl},${{{alias}}}){psig}[i]] -comment "{cmt}"
}}
'''

        return iolines

    def set_mcp_path(self, rnum, lvl, alias, vl, cmd, cmt, fintg=False):
        # clkdata = self.get_table_contxt(self._sdcdg._wb['ClkDef'])
        # self._clkdef.get_clkdata_by_clkname(clkdata)
        # clkdef = self._sdcdg._sheets['ClkDef']
        # cinmals = clkdef.get_crgip_clknm_alias(self._mdname)

        if vl['PortNm']:
            portnm = vl['PortNm']
        else:
            portnm = ''
        if vl['Direction']:
            dirct = vl['Direction']
        else:
            dirct = ''
        if vl['MCP']:
            mcp = vl['MCP'].split(' ')
        else:
            mcp = ''
        if vl['From']:
            frm = vl['From']
        else:
            frm = ''
        if vl['Through']:
            thr = vl['Through']
        else:
            thr = ''
        if vl['To']:
            to = vl['To']
        else:
            to = ''

        iolines = ''
        # kwsetup = f'{mcp[0]} -setup {mcp[1]}'
        # kwhold = f'{mcp[0]} -hold {mcp[2]}'
        if mcp[1] != 'NA':
            kwsetup = f'-{mcp[0]} -setup {mcp[1]}'
        else:
            kwsetup = ''
        if mcp[2] != 'NA':
            kwhold = f'-{mcp[0]} -hold {mcp[2]}'
        else:
            kwhold = ''

        kws = self.set_frthrto_val(alias, lvl, dirct, frm, thr, to)
        psig, pstn, pedn, pnum = self.get_sig_num(portnm)

        # input
        if dirct == 'input':
            if not fintg:
                if not pnum:
                    iolines += f'''
# IOMCP Row{rnum}
if {{!$SDCVAR({self._flt},${{{alias}}})}} {{
'''
                    if kwsetup and kwhold:
                        iolines += f'''
    {cmd} {kwsetup} -from [get_ports {portnm}] {kws} -comment "{cmt}"
    {cmd} {kwhold} -from [get_ports {portnm}] {kws} -comment "{cmt}"
}} else {{
'''
                    if kwsetup and not kwhold:
                        iolines += f'''
    {cmd} {kwsetup} -from [get_ports {portnm}] {kws} -comment "{cmt}"
}} else {{
'''
                    if kwhold and not kwsetup:
                        iolines += f'''                  
    {cmd} {kwhold} -from [get_ports {portnm}] {kws} -comment "{cmt}"
}} else {{
'''
                    if kwsetup and kwhold:
                        iolines += f'''
    {cmd} {kwsetup} -through [get_pins $SDCVAR(HIER,{lvl},${{{alias}}}){portnm}] {kws} -comment "{cmt}"
    {cmd} {kwhold} -through [get_pins $SDCVAR(HIER,{lvl},${{{alias}}}){portnm}] {kws} -comment "{cmt}"
}}
'''
                    if kwsetup and not kwhold:
                        iolines += f'''
    {cmd} {kwsetup} -through [get_pins $SDCVAR(HIER,{lvl},${{{alias}}}){portnm}] {kws} -comment "{cmt}"
}}
'''
                    if kwhold and not kwsetup:
                        iolines += f'''                  
    {cmd} {kwhold} -through [get_pins $SDCVAR(HIER,{lvl},${{{alias}}}){portnm}] {kws} -comment "{cmt}"
}}
'''
                else:
                    iolines += f'''
# IOMCP Row{rnum}
if {{!$SDCVAR({self._flt},${{{alias}}})}} {{
    for {{set i {pstn}}} {{$i <= {pedn}}} {{incr i}} {{
'''
                    if kwsetup and kwhold:
                        iolines += f'''
        {cmd} {kwsetup} -from [get_ports {psig}[i]] {kws} -comment "{cmt}"
        {cmd} {kwhold} -from [get_ports {psig}[i]] {kws} -comment "{cmt}"
    }}
}} else {{
'''
                    if kwsetup and not kwhold:
                        iolines += f'''                    
        {cmd} {kwsetup} -from [get_ports {psig}[i]] {kws} -comment "{cmt}"
    }}
}} else {{
'''
                    if kwhold and not kwsetup:
                        iolines += f'''
        {cmd} {kwhold} -from [get_ports {psig}[i]] {kws} -comment "{cmt}"
    }}
}} else {{
'''
                    iolines += f'''
    for {{set i {pstn}}} {{$i <= {pedn}}} {{incr i}} {{
'''
                    if kwsetup and kwhold:
                        iolines += f'''                  
        {cmd} {kwsetup} -through [get_pins $SDCVAR(HIER,{lvl},${{{alias}}}){psig}[i]] {kws} -comment "{cmt}" 
        {cmd} {kwhold} -through [get_pins $SDCVAR(HIER,{lvl},${{{alias}}}){psig}[i]] {kws} -comment "{cmt}" 
    }}
}}
'''
                    if kwsetup and not kwhold:
                        iolines += f'''                  
        {cmd} {kwsetup} -through [get_pins $SDCVAR(HIER,{lvl},${{{alias}}}){psig}[i]] {kws} -comment "{cmt}"
    }}
}}
'''
                    if kwhold and not kwsetup:
                        iolines += f'''
        {cmd} {kwhold} -through [get_pins $SDCVAR(HIER,{lvl},${{{alias}}}){psig}[i]] {kws} -comment "{cmt}"
    }}
}}
'''
            else:
                if not pnum:
                    iolines += f'''
# IOMCP Row{rnum}
'''
                    if kwsetup:
                        iolines += f'''
{cmd}  {kwsetup} -through [get_pins $SDCVAR(HIER,{lvl},${{{alias}}}){portnm}] {kws} -comment "{cmt}"
'''
                    if kwhold:
                        iolines += f'''
{cmd}  {kwhold} -through [get_pins $SDCVAR(HIER,{lvl},${{{alias}}}){portnm}] {kws} -comment "{cmt}"
'''
                else:
                    iolines += f'''
# IOMCP Row{rnum}
for {{set i {pstn}}} {{$i <= {pedn}}} {{incr i}} {{
'''
                    if kwsetup:
                        iolines += f'''
    {cmd} {kwsetup} -through [get_pins $SDCVAR(HIER,{lvl},${{{alias}}}){psig}[i]] {kws} -comment "{cmt}"
'''
                    if kwhold:
                        iolines += f'''
    {cmd} {kwhold} -through [get_pins $SDCVAR(HIER,{lvl},${{{alias}}}){psig}[i]] {kws} -comment "{cmt}"
'''
                    iolines += f'''
}}
'''

        # output
        if dirct == 'output':
            if not fintg:
                if not pnum:
                    iolines += f'''
# IOMCP Row{rnum}
if {{!$SDCVAR({self._flt},${{{alias}}})}} {{
'''
                    if kwsetup and kwhold:
                        iolines += f'''
    {cmd} {kwsetup} {kws} -to [get_ports {portnm}] -comment "{cmt}"
    {cmd} {kwhold} {kws} -to [get_ports {portnm}] -comment "{cmt}"
}} else {{
'''
                    if kwsetup and not kwhold:
                        iolines += f'''
    {cmd} {kwsetup} {kws} -to [get_ports {portnm}] -comment "{cmt}"
}} else {{
'''
                    if kwhold and not kwsetup:
                        iolines += f'''
    {cmd} {kwhold} {kws} -to [get_ports {portnm}] -comment "{cmt}"
}} else {{
'''
                    if kwsetup and kwhold:
                        iolines += f'''
    {cmd} {kwsetup} {kws} -through [get_pins $SDCVAR(HIER,{lvl},${{{alias}}}){portnm}] -comment "{cmt}"
    {cmd} {kwhold} {kws} -through [get_pins $SDCVAR(HIER,{lvl},${{{alias}}}){portnm}] -comment "{cmt}"
}}
'''
                    if kwsetup and not kwhold:
                        iolines += f'''
    {cmd} {kwsetup} {kws} -through [get_pins $SDCVAR(HIER,{lvl},${{{alias}}}){portnm}] -comment "{cmt}"
}}
'''
                    if kwhold and not kwsetup:
                        iolines += f'''
    {cmd} {kwhold} {kws} -through [get_pins $SDCVAR(HIER,{lvl},${{{alias}}}){portnm}] -comment "{cmt}"
}}
'''
                else:
                    iolines += f'''
# IOMCP Row{rnum}
if {{!$SDCVAR({self._flt},${{{alias}}})}} {{
    for {{set i {pstn}}} {{$i <= {pedn}}} {{incr i}} {{
'''
                    if kwsetup and kwhold:
                        iolines += f'''
        {cmd} {kwsetup} {kws} -to [get_ports {psig}[i]] -comment "{cmt}"
        {cmd} {kwhold} {kws} -to [get_ports {psig}[i]] -comment "{cmt}"
    }}
}} else {{
'''
                    if kwsetup and not kwhold:
                        iolines += f'''
        {cmd} {kwsetup} {kws} -to [get_ports {psig}[i]] -comment "{cmt}"
    }}
}} else {{
'''
                    if kwhold and not kwsetup:
                        iolines += f'''
        {cmd} {kwhold} {kws} -to [get_ports {psig}[i]] -comment "{cmt}"
    }}
}} else {{
'''
                    iolines += f'''
    for {{set i {pstn}}} {{$i <= {pedn}}} {{incr i}} {{
'''
                    if kwsetup and kwhold:
                        iolines += f'''
        {cmd} {kwsetup} {kws} -through [get_pins $SDCVAR(HIER,{lvl},${{{alias}}}){psig}[i]] -comment "{cmt}"
        {cmd} {kwhold} {kws} -through [get_pins $SDCVAR(HIER,{lvl},${{{alias}}}){psig}[i]] -comment "{cmt}"
    }}
}}
'''
                    if kwsetup and not kwhold:
                        iolines += f'''
        {cmd} {kwsetup} {kws} -through [get_pins $SDCVAR(HIER,{lvl},${{{alias}}}){psig}[i]] -comment "{cmt}"
    }}
}}
'''
                    if kwhold and not kwsetup:
                        iolines += f'''
        {cmd} {kwhold} {kws} -through [get_pins $SDCVAR(HIER,{lvl},${{{alias}}}){psig}[i]] -comment "{cmt}"
    }}
}}
'''
            else:
                if not pnum:
                    iolines += f'''
# IOMCP Row{rnum}
'''
                    if kwsetup:
                        iolines += f'''
{cmd}  {kwsetup} {kws} -through [get_pins $SDCVAR(HIER,{lvl},${{{alias}}}){portnm}] -comment "{cmt}"
'''
                    if kwhold:
                        iolines += f'''
{cmd}  {kwhold} {kws} -through [get_pins $SDCVAR(HIER,{lvl},${{{alias}}}){portnm}] -comment "{cmt}"
'''
                else:
                    iolines += f'''
# IOMCP Row{rnum}
for {{set i {pstn}}} {{$i <= {pedn}}} {{incr i}} {{
'''
                    if kwsetup:
                        iolines += f'''
    {cmd} {kwsetup} {kws} -through [get_pins $SDCVAR(HIER,{lvl},${{{alias}}}){psig}[i]] -comment "{cmt}"
'''
                    if kwhold:
                        iolines += f'''
    {cmd} {kwhold} {kws} -through [get_pins $SDCVAR(HIER,{lvl},${{{alias}}}){psig}[i]] -comment "{cmt}"
'''
                    iolines += f'''
}}
'''

        return iolines

    def set_ideal_case(self, rnum, lvl, alias, vl, cmd, cmt, fintg=False):
        if vl['PortNm']:
            portnm = vl['PortNm']
        else:
            portnm = ''
        psig, pstn, pedn, pnum = self.get_sig_num(portnm)
        iolines = ''

        if not fintg:
            if 'set_case_analysis' in cmd:
                if not pnum:
                    iolines += f'''
# IOCASE Row{rnum}
if {{!$SDCVAR({self._flt},${{{alias}}})}} {{
    {cmd}  [get_ports {portnm}]
}} else {{
    {cmd}  [get_pins $SDCVAR(HIER,{lvl},${{{alias}}}){portnm}]
}}
'''
                else:
                    iolines += f'''
# IOCASE Row{rnum}
if {{!$SDCVAR({self._flt},${{{alias}}})}} {{
    for {{set i {pstn}}} {{$i <= {pedn}}} {{incr i}} {{
        {cmd}  [get_ports {psig}[i]]
    }}
}} else {{
    for {{set i {pstn}}} {{$i <= {pedn}}} {{incr i}} {{
        {cmd}  [get_pins $SDCVAR(HIER,{lvl},${{{alias}}}){psig}[i]]
    }}
}}
'''
            if 'set_ideal_network' in cmd:
                if not pnum:
                    iolines += f'''
# IOIDEAL Row{rnum}
if {{!$SDCVAR({self._flt},${{{alias}}})}} {{
    {cmd}  [get_ports {portnm}]
}}
'''
                else:
                    iolines += f'''
# IOIDEAL Row{rnum}
if {{!$SDCVAR({self._flt},${{{alias}}})}} {{
    for {{set i {pstn}}} {{$i <= {pedn}}} {{incr i}} {{
        {cmd}  [get_ports {psig}[i]]
    }}
}} 
'''
        else:
            if 'set_case_analysis' in cmd:
                if not pnum:
                    iolines += f'''
# IOCASE Row{rnum}
{cmd}  [get_pins $SDCVAR(HIER,{lvl},${{{alias}}}){portnm}]  
'''
                else:
                    iolines += f'''
# IOCASE Row{rnum}
for {{set i {pstn}}} {{$i <= {pedn}}} {{incr i}} {{
    {cmd} [get_pins $SDCVAR(HIER,{lvl},${{{alias}}}){psig}[i]]
}}
'''
        return iolines

    def set_inout_maxdly(self, rnum, lvl, alias, vl, cmd):
        # clkdata = self.get_table_contxt(self._sdcdg._wb['ClkDef'])
        # self._clkdef = self._sdcdg._sheets['ClkDef']
        # self._clkdef.get_clkdata_by_clkname(clkdata)
        # syn_wlm_sel = self._vardata['SYN_WLM_SEL']

        iolines = ''
        iolines += f'''
# INOUT Row{rnum}
if {{!$SDCVAR({self._flt},${{{alias}}})}} {{
    if {{!$SDCVAR(SYN_WLM_SEL) == "Genus PLE"}} {{
'''
        syn_sel = 'Genus PLE'
        iolines += self.set_inout_val(alias, vl, syn_sel, cmd)
        iolines += f'''       
    }} elseif {{!$SDCVAR(SYN_WLM_SEL) == "DC WLM"}} {{
'''
        syn_sel = 'DC WLM'
        iolines += self.set_inout_val(alias, vl, syn_sel, cmd)
        iolines += f'''     
    }} else {{
'''
        syn_sel = 'SPG ISP'
        iolines += self.set_inout_val(alias, vl, syn_sel, cmd)
        iolines += f'''
    }}
}}
'''

        return iolines

    def set_inout_val(self, alias, vl, syn_sel, cmd):
        # syn_wlm_sel = self._vardata['SYN_WLM_SEL']
        clkdef = self._sdcdg._sheets['ClkDef']
        # cinmals = clkdef.get_crgip_clknm_alias(self._mdname)

        # PortIn	PortOut	DlyIn	DlyOut	RealDly	ClkNm	Vol	Comment
        if vl['PortIn']:
            portin = vl['PortIn']
        else:
            portin = ''
        if vl['PortOut']:
            portout = vl['PortOut']
        else:
            portout = ''
        if vl['DlyIn']:
            dlyin = vl['DlyIn']
        else:
            dlyin = ''
        if vl['DlyOut']:
            dlyout = vl['DlyOut']
        else:
            dlyout = ''
        if vl['RealDly']:
            rdly = vl['RealDly']
        else:
            rdly = ''

        if vl['ClkNm']:
            nclknm = vl['ClkNm']
        else:
            nclknm = ''
        if nclknm:
            # als,cknm = clkdef.get_alval_clknm(self._mdname,alias,nclknm)
            # clknm = f'$SDCVAR(NAME,${{{als}}},{cknm})'
            cknm = nclknm
            if nclknm in clkdef._clknmlst:
                als = alias
            else:
                # avl = nclknm.split('_')[0]
                sp = nclknm.split('_')
                if 'NAME_' in nclknm:
                    avl = sp[1]
                else:
                    avl = sp[0]
                als = clkdef.get_als_var(self._mdname, avl)

        if vl['Vol']:
            vol = vl['Vol']
        else:
            vol = ''

        if vl['Comment']:
            cmt = vl[7]
        else:
            cmt = 'NA'

        insig, instn, inedn, innum = self.get_sig_num(portin)
        outsig, outstn, outedn, outnum = self.get_sig_num(portout)

        # print('inout:',vl)
        # print('innum,outnum',portout,innum,outnum)

        indly = self.get_dlymaxmin(cknm, als, dlyin)
        outdly = self.get_dlymaxmin(cknm, als, dlyout)

        if syn_sel == 'Genus PLE':
            uncer_factor = '0.1'
        elif syn_sel == 'DC WLM':
            uncer_factor = '0.25'
        else:
            uncer_factor = '0.05'

        iolines = ''
        if not innum and not outnum:
            iolines += f'''
        {cmd} [expr {indly} + {outdly} + {uncer_factor} * $SDCVAR(CYCLE,${{{als}}},{cknm}) + {rdly}] -from [get_ports {portin}] -to [get_ports {portout}] -comment "{cmt}"
'''

        if not innum and outnum:
            iolines += f'''
        for {{set i {outstn}}} {{$i <= {outedn}}} {{incr i}} {{
            {cmd} [expr {indly} + {outdly} + {uncer_factor} * $SDCVAR(CYCLE,${{{als}}},{cknm}) + {rdly}] -from [get_ports {portin}] -to [get_ports {outsig}[i]] -comment "{cmt}"
        }}
'''

        if innum and not outnum:
            iolines += f'''
        for {{set i {instn}}} {{$i <= {inedn}}} {{incr i}} {{
            {cmd} [expr {indly} + {outdly} + {uncer_factor} * $SDCVAR(CYCLE,${{{als}}},{cknm}) + {rdly}] -from [get_ports {insig}[i]] -to [get_ports {portout}] -comment "{cmt}"
        }}
'''
        if innum and outnum:
            iolines += f'''
        for {{set i {instn}}} {{$i <= {inedn}}} {{incr i}} {{
            for {{set j {outstn}}} {{$j <= {outedn}}} {{incr j}} {{
                {cmd} [expr {indly} + {outdly} + {uncer_factor} * $SDCVAR(CYCLE,${{{als}}},{cknm}) + {rdly}] -from [get_ports {insig}[i]] -to [get_ports {outsig}[j]] -comment "{cmt}"
            }}
        }}
'''

        return iolines

    def get_dlymaxmin(self, clknm, alias, ndly):

        if re.search(r'IO_DLY_M', ndly):
            # self.get_dlymaxmin(clknm,alias,str(self._vardata[ndly]))
            fdly = str(self._vardata[ndly])
        else:
            fdly = ndly

        if re.match(r'(-)*0\.\d+#$', fdly):
            return fdly.replace('#', '')

        elif re.match(r'(-)*\[expr \d+\w+\]#', fdly):
            return fdly.replace('#', '')

        elif re.match(r'(-)*0\.0$|0$', fdly):
            return f'0.0'

        elif re.match(r'(-)*0\.\d+$', fdly):
            return f'[expr $SDCVAR(CYCLE,${{{alias}}},{clknm}) * {fdly}]'

        elif re.match(r'(-)*\d+%$', fdly):
            xdly = int(fdly.split('%')[0]) / 100
            return f'[expr $SDCVAR(CYCLE,${{{alias}}},{clknm}) * {xdly}]'

        else:
            return fdly.replace('#', '')

    def set_frthrto_val(self, alias, lvl, dirct, frm, thr, to):
        kwlines = []
        # clkdef = self._sdcdg._sheets['ClkDef']
        # cinmals = clkdef.get_crgip_clknm_alias(self._mdname)
        # clkdata = clkdef.get_table_contxt(self._sdcdg._wb['ClkDef'])
        # self._clkdef.get_clkdata_by_clkname(clkdata)

        if dirct == 'input':
            if thr:
                thrvalg = self.get_frthrto_val(alias, lvl, thr)
                thrval = f'-through {thrvalg}'
                kwlines.append(thrval)
            if to:
                tovalg = self.get_frthrto_val(alias, lvl, to)
                toval = f'-to {tovalg}'
                kwlines.append(toval)

        if dirct == 'output':
            if frm:
                frmvalg = self.get_frthrto_val(alias, lvl, frm)
                frmval = f'-from {frmvalg}'
                kwlines.append(frmval)
            if thr:
                thrvalg = self.get_frthrto_val(alias, lvl, thr)
                thrval = f'-through {thrvalg}'
                kwlines.append(thrval)

        kws = ' '.join(kwlines)

        return kws

    def get_frthrto_val(self, alias, lvl, dval):
        clkdef = self._sdcdg._sheets['ClkDef']
        mdname = self._mdname
        # cinmals = clkdef.get_crgip_clknm_alias(mdname)
        rval = []

        sval = dval.strip()
        if re.search(r'\]\]$', sval):
            sval = sval.replace(']]', ']').strip().split(' ')
        else:
            sval = sval.strip(']').strip().split(' ')
        # print('ori sval',sval,sval[2:])
        if sval[0] == 'clk':
            rval.append(f'[get_clocks [list ')
            for cknm in sval[2:]:
                # ncknm = cknm.replace(']','').strip()
                if cknm:
                    # als,ncknm = clkdef.get_alval_clknm(mdname,alias,cknm)
                    # cknm = nclknm
                    if cknm in clkdef._clknmlst:
                        als = alias
                    else:
                        sp = cknm.split('_')
                        if 'NAME_' in cknm:
                            avl = sp[1]
                        else:
                            avl = sp[0]
                        als = clkdef.get_als_var(self._mdname, avl)
                    rval.append(f'$SDCVAR(NAME,${{{als}}},{cknm})')
            rval.append(f']]')

        if sval[0] == 'pin':
            rval.append(f'[get_pins [list ')
            # print('sval',sval)
            for pin in sval[2:]:
                # if re.search(r'\[\d+\]\s*]$',pin.strip()):
                #     print('deese',pin)
                #     npin = pin.strip().replace(']','').strip()
                #     #spin = self.name_chg(npin)
                #     rval.append(f'$SDCVAR(HIER,{lvl},${{{alias}}}){npin} ')
                if re.search(r'\w+\[\d+:\d+\]$', pin.strip()):
                    # print('erwqetq',pin)
                    mpins = self.get_mbit_chg(pin.strip())
                    # print('mpins',mpins)
                    for mpin in mpins:
                        rval.append(f'$SDCVAR(HIER,{lvl},${{{alias}}}){mpin}')
                elif re.search(
                        r'\/D$|\/CLK$|\/CP$|\/E$|\/Q$|\/CD$|\/CK$|\/\$ClkPin|\/\$DataPin|\/\$\{ClkPin\}|\/\$\{DataPin\}',
                        pin.strip()):
                    gpin = pin.strip().split('/')
                    if re.search(r'[\d+]', gpin[-2]):
                        tpin = gpin[-2].replace(']', '_reg_')
                        tpin = tpin.replace('[', '_')
                    else:
                        tpin = gpin[-2] + '_reg_'
                    npin = []
                    if gpin[:-2]:
                        npin.append('/'.join(gpin[:-2]))
                    npin.append(tpin)
                    npin.append(gpin[-1])
                    spin = '/'.join(npin)
                    # npin.append(gpin[:-3])
                    # npin.append(tpin)
                    # npin.append(gpin[-1])
                    # spin = '/'.join(npin)
                    rval.append(f'$SDCVAR(HIER,{lvl},${{{alias}}}){spin}')
                else:
                    rval.append(f'$SDCVAR(HIER,{lvl},${{{alias}}}){pin.strip()}')
            rval.append(f']]')
        # print('rval',rval)
        pval = ' '.join(rval)

        return pval

    # def name_chg(self,hierpin):
    #     # u_a/u_b/D
    #     # u_a/u_b[2]/D
    #     # u_a/u_b/sig[2]
    #     # u_a/u_b/sig[5:2]

    def get_mbit_chg(self, portpin):
        mportpins = []
        sig, stn, edn, num = self.get_sig_num(portpin)
        for i in range(int(stn), int(num) + 1):
            mportpins.append(f'{sig}[{i}]')

        return mportpins

    def get_sig_num(self, portnm):
        sig = ''
        stn = ''
        edn = ''
        num = ''
        iopat = re.findall(r'(\S+)(\[\d+:\d+\])', portnm)
        if iopat:
            sig = iopat[0][0]
            stn, edn, num = self.cal_num(iopat[0][1])

        iopat = re.findall(r'(\S+)(\[\d+:\d+\])(\[\d+:\d+\])', portnm)
        if iopat:
            sig = iopat[0][0]
            stn1, edn1, num1 = self.cal_num(iopat[0][1])
            stn2, edn2, num2 = self.cal_num(iopat[0][2])
            num = str(int(num1) * int(num2) - 1)
            # how to set stn/edn ??
            stn = '0'
            edn = num

        # input wire [3:0][5: 0] dat_up,  ====> dat_up[3:0][5:0]
        # input wire [3:0] datg_up[5: 0][2:0][7:0], ====> datg_up[5:0][2:0][7:0][3:0]

        iopat = re.findall(r'(\S+)(\[\d+:\d+\])(\[\d+:\d+\])(\[\d+:\d+\])', portnm)
        if iopat:
            sig = iopat[0][0]
            stn1, edn1, num1 = self.cal_num(iopat[0][1])
            stn2, edn2, num2 = self.cal_num(iopat[0][2])
            stn3, edn3, num3 = self.cal_num(iopat[0][3])
            num = str(int(num1) * int(num2) * int(num3) - 1)
            # how to set stn/edn ??
            stn = '0'
            edn = num

        iopat = re.findall(r'(\S+)(\[\d+:\d+\])(\[\d+:\d+\])(\[\d+:\d+\])(\[\d+:\d+\])', portnm)
        if iopat:
            sig = iopat[0][0]
            stn1, edn1, num1 = self.cal_num(iopat[0][1])
            stn2, edn2, num2 = self.cal_num(iopat[0][2])
            stn3, edn3, num3 = self.cal_num(iopat[0][3])
            stn4, edn4, num4 = self.cal_num(iopat[0][4])
            num = str(int(num1) * int(num2) * int(num3) * int(num4) - 1)
            # how to set stn/edn ??
            stn = '0'
            edn = num

        return sig, stn, edn, num

    def cal_num(self, npat):
        # [n:m]
        stn = npat.split(':')[1].replace(']', '')
        edn = npat.split(':')[0].replace('[', '')
        num = str(int(edn) - int(stn) + 1)
        return stn, edn, num
        # sig = portnm.split('[')[0]
        # stn = portnm.split('[')[1].split(':')[1].replace(']','')
        # edn = portnm.split('[')[1].split(':')[0]
        # print(sig,stn,edn)
        # num = int(edn) - int(stn) + 1

    ######################################################
    ## write internal exception
    def write_sdc_int(self, sdc_dir):
        sheet = self.get_sheet()

        mdname = self._sdcdg._vfile_data['module_name']
        alias = self._sdcdg._hier_tree._blocks[mdname].alias
        hdlvl = self._sdcdg._hier_tree._blocks[mdname].hdlevel
        pwr = self._sdcdg._hier_tree._blocks[mdname].prime_pwr
        self._mdname = mdname

        if hdlvl == 'sys':
            self._lvl = 'sys'
            self._flt = 'IS_CHIP'
        if hdlvl == 'blk':
            self._lvl = 'blk'
            self._flt = 'IS_FLAT'

        sdc_file = sdc_dir + f'{alias.lower()}_intexp.sdc'
        self.write_intexp(mdname, alias, sdc_file, fintg=False)
        sdc_file = sdc_dir + f'intg/{alias.lower()}_intexp_intg.sdc'
        self.write_intexp(mdname, alias, sdc_file, fintg=True)

    def write_intexp(self, mdname, alias, sdc_file, prousr=False, fintg=False):
        # clkdata = self.get_table_contxt(self._sdcdg._wb['ClkDef'])
        # self._clkdef.get_clkdata_by_clkname(clkdata)
        # self._clkdef = self._sdcdg._sheets['ClkDef']

        intexp_lines = ''

        # print('_intexpdata',self._intexpdata)
        intfp_rows = {}
        intmcp_rows = {}
        intcase_rows = {}
        intstpgt_rows = {}
        for kw, vl in self._expdata.items():
            num = kw.split('Row')[1]
            if re.search(r'TMINTEXP_Row', kw):
                # FP	MCP	CaseVal	CasePin	From	Through	To	Comment
                if re.search(r'0|1', str(vl['CaseVal'])):
                    intcase_rows[f'CASE_Row{num}'] = vl
                if vl['FP']:
                    intfp_rows[f'FP_Row{num}'] = vl
                if vl['MCP']:
                    intmcp_rows[f'MCP_Row{num}'] = vl

            if re.search(r'TMSTPGATE_Row', kw):
                # StopClk	StopPin	DisClkGating	Comment
                intstpgt_rows[f'STPGATE_Row{num}'] = vl
        # print('intfp_rows:',intfp_rows)

        intexp_lines += f'''
################################################
## Internal False Path
################################################
'''
        intexp_lines += self.set_intcmd(intfp_rows, 'set_false_path', alias, fintg)

        intexp_lines += f'''
################################################
## Internal Multicycle Path
################################################
'''
        intexp_lines += self.set_intcmd(intmcp_rows, 'set_multicycle_path', alias, fintg)

        intexp_lines += f'''
################################################
## Internal Case Setting
################################################
'''
        intexp_lines += self.set_intcmd(intcase_rows, 'set_case_analysis', alias, fintg)

        intexp_lines += f'''
################################################
## Internal STP and DisGating
################################################
'''
        intexp_lines += self.set_intcmd(intstpgt_rows, 'set_sense', alias, fintg)
        intexp_lines += self.set_intcmd(intstpgt_rows, 'set_disable_clock_gating_check', alias, fintg)

        # sub harden blk
        if not fintg:
            blkf = 'intexp'
            intexp_lines += self._hier_tree.set_subblk_intg(mdname, blkf)

        self.save_text(intexp_lines, sdc_file)

    # FP	MCP	CaseVal	CasePin	From	Through	To	Comment
    def set_intcmd(self, intdata, cmd, alias, fintg=False):
        intlines = ''
        for kw, vl in intdata.items():
            rnum = kw.split('Row')[1]
            lvl = self._lvl.upper()

            # if 'TMINTEXP' in kw and vl[7]:
            #     cmt = vl[7]
            # elif 'TMSTPGATE' in kw and vl[3]:

            if vl['Comment']:
                cmt = vl['Comment']
            else:
                cmt = 'NA'

            # case
            if cmd == 'set_case_analysis':
                if str(vl['CaseVal']):
                    caseval = vl['CaseVal']
                    rcmd = f'{cmd} {caseval}'
                    intlines += self.set_case_int(rnum, lvl, alias, vl, rcmd, cmt, fintg)

            # fp
            if cmd == 'set_false_path':
                intlines += self.set_fp_path_int(rnum, lvl, alias, vl, cmd, cmt, fintg)

            # mcp
            if cmd == 'set_multicycle_path':
                intlines += self.set_mcp_path_int(rnum, lvl, alias, vl, cmd, cmt, fintg)

            # stop clk
            if cmd == 'set_sense':
                intlines += self.set_clk_sense_gating_int(rnum, lvl, alias, vl, cmd)

            # disable clk gating
            if cmd == 'set_disable_clock_gating_check':
                intlines += self.set_clk_sense_gating_int(rnum, lvl, alias, vl, cmd)

        return intlines

    # FP	MCP	CaseVal	CasePin	From	Through	To	Comment
    def set_fp_path_int(self, rnum, lvl, alias, vl, cmd, cmt, fintg=False):
        # clkdata = self.get_table_contxt(self._sdcdg._wb['ClkDef'])
        # self._clkdef.get_clkdata_by_clkname(clkdata)

        if vl['FP']:
            fp = vl['FP']
        else:
            fp = ''
        if vl['From']:
            frm = vl['From']
        else:
            frm = ''
        if vl['Through']:
            thr = vl['Through']
        else:
            thr = ''
        if vl['To']:
            to = vl['To']
        else:
            to = ''

        intlines = ''
        if fp == 'all':
            fpval = ' '
        else:
            fpval = f'-{fp} '

        kws = self.set_frthrto_val_int(alias, lvl, frm, thr, to)
        # psig,pstn,pedn,pnum = self.cal_portpin_num(vl[0])

        intlines += f'''
# INTFP Row{rnum}
{cmd}  {fpval} {kws} -comment "{cmt}"
'''

        return intlines

    def set_mcp_path_int(self, rnum, lvl, alias, vl, cmd, cmt, fintg=False):
        # clkdata = self.get_table_contxt(self._sdcdg._wb['ClkDef'])
        # self._clkdef.get_clkdata_by_clkname(clkdata)

        if vl['MCP']:
            mcp = vl['MCP'].split(' ')
        else:
            mcp = ''
        if vl['From']:
            frm = vl['From']
        else:
            frm = ''
        if vl['Through']:
            thr = vl['Through']
        else:
            thr = ''
        if vl['To']:
            to = vl['To']
        else:
            to = ''

        kws = self.set_frthrto_val_int(alias, lvl, frm, thr, to)
        # psig,pstn,pedn,pnum = self.cal_portpin_num(vl[0])

        intlines = ''
        if mcp[1] != 'NA':
            kwsetup = f'-{mcp[0]} -setup {mcp[1]}'
        else:
            kwsetup = ''
        if mcp[2] != 'NA':
            kwhold = f'-{mcp[0]} -hold {mcp[2]}'
        else:
            kwhold = ''

        intlines += f'''
# INTMCP Row{rnum}
'''
        intlines = intlines.rstrip()
        if kwsetup:
            intlines += f'''
{cmd}  {kwsetup}  {kws} -comment "{cmt}"
'''

        if kwhold:
            intlines += f'''
{cmd}  {kwhold}  {kws} -comment "{cmt}"
'''

        return intlines

    def set_case_int(self, rnum, lvl, alias, vl, cmd, cmt, fintg=False):

        if vl['CaseVal']:
            csval = vl['CaseVal']
        else:
            csval = ''
        if vl['CasePin']:
            cspin = vl['CasePin']
        else:
            cspin = ''

        iolines = ''

        # thr = None
        # to = None
        # kws = self.get_frthrto_val(self,alias,lvl,cspin,thr,to)
        kws = self.get_frthrto_val_int(alias, lvl, cspin)
        # kwsg = kws.replace('-from','')

        iolines += f'''
# INTCASE Row{rnum}
{cmd}  {csval}  {kws}
'''

        return iolines

    def set_clk_sense_gating_int(self, rnum, lvl, alias, vl, cmd):
        # StopClk	StopPin	DisClkGating	Comment
        if vl['StopClk']:
            stopclk = vl['StopClk']
        else:
            stopclk = ''

        if vl['StopPin']:
            stoppin = vl['StopPin']
        else:
            stoppin = ''

        if vl['DisClkGating']:
            discgat = vl['DisClkGating']
        else:
            discgat = ''

        intlines = ''

        if stopclk and stoppin and cmd == 'set_sense':
            stpclk = self.get_frthrto_val_int(alias, lvl, stopclk)
            stppin = self.get_frthrto_val_int(alias, lvl, stoppin)
            intlines += f'''
# INTCLKSense Row{rnum}
{cmd} -stop_propagation -clocks {stpclk} {stppin}

'''

        if discgat and cmd == 'set_disable_clock_gating_check':
            clkgate = self.get_frthrto_val_int(alias, lvl, discgat)
            intlines += f'''
# INTCLKGating Row{rnum}            
{cmd} {clkgate}

'''

        return intlines

    def set_frthrto_val_int(self, alias, lvl, frm, thr, to):
        kwlines = []

        if frm:
            frmvalg = self.get_frthrto_val_int(alias, lvl, frm)
            frmval = f'-from {frmvalg}'
            kwlines.append(frmval)
        if thr:
            thrvalg = self.get_frthrto_val_int(alias, lvl, thr)
            thrval = f'-through {thrvalg}'
            kwlines.append(thrval)
        if to:
            tovalg = self.get_frthrto_val_int(alias, lvl, to)
            toval = f'-to {tovalg}'
            kwlines.append(toval)

        kws = ' '.join(kwlines)

        return kws

    def get_frthrto_val_int(self, alias, lvl, dval):
        clkdef = self._sdcdg._sheets['ClkDef']
        # clkdata = clkdef.get_table_contxt(self._sdcdg._wb['ClkDef'])
        # self._clkdef.get_clkdata_by_clkname(clkdata)
        rval = []

        sval = dval.strip()
        if re.search(r'\]\]$', sval):
            sval = sval.replace(']]', ']').strip().split(' ')
        else:
            sval = sval.strip(']').strip().split(' ')
        if sval[0] == 'clk':
            rval.append(f'[get_clocks [list ')
            for cknm in sval[2:]:
                if cknm:
                    # als,ncknm = clkdef.get_alval_clknm(self._mdname,alias,cknm)
                    if cknm in clkdef._clknmlst:
                        als = alias
                    else:
                        # avl = cknm.split('_')[0]
                        sp = cknm.split('_')
                        if 'NAME_' in cknm:
                            avl = sp[1]
                        else:
                            avl = sp[0]
                        als = clkdef.get_als_var(self._mdname, avl)
                    rval.append(f'$SDCVAR(NAME,${{{als}}},{cknm})')
            rval.append(f']]')

        if sval[0] == 'inst':
            rval.append(f'[get_cells [list ')
            for cell in sval[2:]:
                if re.search(r'\w+\[\d+:\d+\]', cell.strip()):
                    # ncell = cell.strip().replace(']','').strip()
                    # spin = self.name_chg(npin)
                    sig, stn, edn, num = self.get_sig_num(cell.strip())
                    for i in range(int(stn), int(edn) + 1):
                        rval.append(f'$SDCVAR(HIER,{lvl},${{{alias}}}){sig}[{i}]')
                else:
                    rval.append(f'$SDCVAR(HIER,{lvl},${{{alias}}}){cell.strip()}')
            rval.append(f']]')

        if sval[0] == 'pin':
            rval.append(f'[get_pins [list ')
            for pin in sval[2:]:
                # if re.search(r'[\d+]\s*]$',pin.strip()):
                #     npin = pin.strip().replace(']','').strip()
                #     #spin = self.name_chg(npin)
                #     rval.append(f'$SDCVAR(HIER,${lvl},${{{alias}}}){npin} ')
                if re.search(r'\w+\[\d+:\d+\]$', pin.strip()):
                    mpins = self.get_mbit_chg(pin.strip())
                    for mpin in mpins:
                        # spin = self.name_chg(mpin)
                        rval.append(f'$SDCVAR(HIER,${lvl},${{{alias}}}){mpin} ')
                elif re.search(
                        r'\/D$|\/CLK$|\/CP$|\/E$|\/Q$|\/CD$|\/CK$|\/\$ClkPin|\/\$DataPin|\/\$\{ClkPin\}|\/\$\{DataPin\}',
                        pin.strip()):
                    gpin = pin.strip().split('/')
                    if re.search(r'[\d+]', gpin[-2]):
                        tpin = gpin[-2].replace(']', '_reg_')
                        tpin = tpin.replace('[', '_')
                    else:
                        tpin = gpin[-2] + '_reg_'
                    npin = []
                    if gpin[:-2]:
                        npin.append('/'.join(gpin[:-2]))
                    npin.append(tpin)
                    npin.append(gpin[-1])
                    spin = '/'.join(npin)
                    rval.append(f'$SDCVAR(HIER,{lvl},${{{alias}}}){spin}')
                else:
                    npin = pin.strip()
                    # spin = self.name_chg(npin)
                    rval.append(f'$SDCVAR(HIER,{lvl},${{{alias}}}){npin}')
            rval.append(f']]')

        pval = ' '.join(rval)

        return pval

    def get_mbit_chg_int(self, portpin):
        mportpins = []
        sig, stn, edn, num = self.get_sig_num(portpin)
        for i in range(int(stn), int(num) + 1):
            mportpins.append(f'{sig}[{i}]')

        return mportpins

    def get_sig_num_int(self, portnm):
        sig = ''
        stn = ''
        edn = ''
        num = ''
        iopat = re.findall(r'(\S+)(\[\d+:\d+\])', portnm)
        if iopat:
            sig = iopat[0][0]
            stn, edn, num = self.cal_num(iopat[0][1])

        iopat = re.findall(r'(\S+)(\[\d+:\d+\])(\[\d+:\d+\])', portnm)
        if iopat:
            sig = iopat[0][0]
            stn1, edn1, num1 = self.cal_num(iopat[0][1])
            stn2, edn2, num2 = self.cal_num(iopat[0][2])
            num = str(int(num1) * int(num2) - 1)
            # how to set stn/edn ??
            stn = '0'
            edn = num

        # input wire [3:0][5: 0] dat_up,  ====> dat_up[3:0][5:0]
        # input wire [3:0] datg_up[5: 0][2:0][7:0], ====> datg_up[5:0][2:0][7:0][3:0]

        iopat = re.findall(r'(\S+)(\[\d+:\d+\])(\[\d+:\d+\])(\[\d+:\d+\])', portnm)
        if iopat:
            sig = iopat[0][0]
            stn1, edn1, num1 = self.cal_num(iopat[0][1])
            stn2, edn2, num2 = self.cal_num(iopat[0][2])
            stn3, edn3, num3 = self.cal_num(iopat[0][3])
            num = str(int(num1) * int(num2) * int(num3) - 1)
            # how to set stn/edn ??
            stn = '0'
            edn = num

        iopat = re.findall(r'(\S+)(\[\d+:\d+\])(\[\d+:\d+\])(\[\d+:\d+\])(\[\d+:\d+\])', portnm)
        if iopat:
            sig = iopat[0][0]
            stn1, edn1, num1 = self.cal_num(iopat[0][1])
            stn2, edn2, num2 = self.cal_num(iopat[0][2])
            stn3, edn3, num3 = self.cal_num(iopat[0][3])
            stn4, edn4, num4 = self.cal_num(iopat[0][4])
            num = str(int(num1) * int(num2) * int(num3) * int(num4) - 1)
            # how to set stn/edn ??
            stn = '0'
            edn = num

        return sig, stn, edn, num

    def cal_num_int(self, npat):
        # [n:m]
        stn = npat.split(':')[1].replace(']', '')
        edn = npat.split(':')[0].replace('[', '')
        num = str(int(edn) - int(stn) + 1)
        return stn, edn, num



