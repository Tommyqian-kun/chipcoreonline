
# _*_ coding:utf-8 _*_

import sys
import time
import os
import re

from os.path import dirname, abspath, basename
import  openpyxl
import tkinter

from .vardef import *
from .clkdef import *
from .iodly import *
# from .ioexp import *
# from .intexp import *
from .exp import *
from com.hierpwr import HierPwrTree
from .basesdc import *
from .viodly import *


class SDC_DG(object):
    def __init__(self):
        self._sheets = {}
        self._hier_tree = {}
        self._vardef = {}
        self._wb = {}
        self._sdcdir = ''
        self._mdname = ''
        self._alias = '' #self._hier_tree._blocks[self._mdname].alias
        self._hdlvl = ''
        self._pwr = ''
        self._inputs = BaseInputs()
        self._iodly = VIODly()

        #self.proj_mode = False
        self._vfile_data = None
        self._vfile_list = None
        self._data = None
        self._lvl = 'blk'
        self._flt = 'IS_FLAT'

        self._vardata = {}
        # self._sdc_dir = ''
        # self._com_dir= ''
        # self._dft_dir = ''
        # self._hd_process = ''
        # # self._hd_mod_name = self._mdname
        # self._cycle_list = '[list CYCLE500M]'


    @property
    def hier_tree(self):
        return self._hier_tree
    
    @hier_tree.setter
    def hier_tree(self, hier_tree):
        self._hier_tree = hier_tree
 
    def load_design_guide(self,dg_file,kwd=''):
        self._sdcdir = abspath(dirname(dirname(dg_file)))
        self._wb = openpyxl.load_workbook(dg_file)

        valdef_sheet = self._wb['VarDef']
        start = 0
        for i in range(1, valdef_sheet.max_row+1):
            if valdef_sheet.cell(i,1).value == 'Variable':
                start = i + 1
                break
        for i in range(start, valdef_sheet.max_row+1):
            key = valdef_sheet.cell(row=i, column=1).value
            val = valdef_sheet.cell(row=i, column=2).value
            self._vardef[key] = val

        self._sheets = {
            'VarDef'            : VarDefSheet(self, 'VarDef'),
            'ClkDef'            : ClkDefSheet(self, 'ClkDef'),
            'IODly'             : IODlySheet(self, 'IODly'),
            'Exp'               : ExpSheet(self, 'Exp'),
            # 'IOExp'             : IOExpSheet(self, 'IOExp'),
            # 'IntExp'            : IntExpSheet(self, 'IntExp'),
        }

        # read and convert excel to json data
        if kwd == 'json':
            for sheetname,sheet in self._sheets.items():
                if sheetname == 'ClkDef':
                    sheet.read_data('json')
                else:
                    sheet.read_data()            
                json_file = dirname(dirname(dg_file)) + '/json' + f'/{sheetname.lower()}.json'
                # json_file = dirname(dirname(dg_file)) + os.path.join('json', f'{sheetname.lower()}.json')
                sheet.dump_json(json_file)

            json_file = dirname(dirname(dg_file)) + '/json' + f'/hier_pwr.json'
            # json_file = dirname(dirname(dg_file)) + os.path.join('json', 'hier.json')
            self._data = self._hier_tree._yaml_data
            self.write_json(json_file)
            

    def read_vfile(self,vfile,kwd=''):
        self._vfile_list, self._vfile_data = self._inputs.read_vfile(vfile)
        self._mdname = self._vfile_data['module_name']
        # print(self._vfile_list)
        # print(self._vfile_data)

        if kwd == 'json':
            self._data = self._vfile_data
            json_file = dirname(dirname(vfile)) + '/json' + f'/vfile.json'
            self.write_json(json_file)    

    def update_dg(self):
        for sht in self._sheets.values():
            sht.update_sheet()

    def check_dg(self):
        for sht in self._sheets.values():
            sht.check_sheet()

    def change_dg(self,dgfile):
        for sht in self._sheets.values():
            sht.change_sheet(dgfile)

    def write_sdc_files(self,prousr=False):

        sdc_path = self._sdcdir
        # per_dvfs = False
        # if prousr == 'pro':
        #     sdc_info('User has profession permission.')
        #     per_dvfs = True
        # if prousr == 'fre':
        #     sdc_info('User has free permission.')
        #     per_dvfs = False

        self._alias = self._hier_tree._blocks[self._mdname].alias
        self._hdlvl = self._hier_tree._blocks[self._mdname].hdlevel
        self._pwr = self._hier_tree._blocks[self._mdname].prime_pwr

        # if not os.path.exists(sdc_path + '/outputs/intg'):
        #     os.makedirs(sdc_path + '/outputs/intg', exist_ok=True)

        # if not os.path.exists(sdc_path + '/outputs/blklib'):
        #     os.makedirs(sdc_path + '/outputs/blklib', exist_ok=True)
        # if not os.path.exists(sdc_path + '/outputs/expd'):
        #     os.makedirs(sdc_path + '/outputs/expd', exist_ok=True)

        if self._hdlvl == 'sys':
            self._lvl = 'sys'
            self._flt = 'IS_CHIP'
        else:
            self._lvl = 'blk'
            self._flt = 'IS_FLAT'

        varsht = self._sheets['VarDef'].get_sheet()
        self._vardata = self._sheets['VarDef'].get_vardef_value(varsht)
        # self._vardata['SDC_DIR'] = self._sdcdir
        # self._vardata['COM_DIR'] = self._sdcdir
        # self._vardata['DFT_DIR'] = ''
        # self._vardata['HD_MD_NAME'] = self._mdname
        # self._vardata['HD_PROCESS'] =
        # self._vardata['CYCLE_LIST'] = '[list CYCLE500M]'

        # write top sdc files such as ${blk}.sdc and ${blk}_top.sdc and ${blk}_subblk.sdc and ${blk}_tune.sdc
        # sdc_file = sdc_path + f'/outputs/{self._mdname}.func.sdc'
        # self.write_func_sdc(sdc_file)
        # sdc_file = sdc_path + f'/outputs/{self._mdname}.func.flat.sdc'
        # self.write_func_flat_sdc(sdc_file)       

        # sdc_file = sdc_path + f'/outputs/{self._mdname}.func.dvfs.sdc'
        # self.write_func_dvfs_sdc(sdc_file)
        # sdc_file = sdc_path + f'/outputs/{self._mdname}.func_dvfs.flat.sdc'
        # self.write_func_dvfs_flat_sdc(sdc_file)   

        sdc_file = sdc_path + f'/outputs/{self._alias.lower()}_{self._lvl}top.sdc'
        # sdc_file = os.path.join(f'{sdc_path}','outputs',f'{self._alias.lower()}_{self._lvl}top.sdc')
        self.write_top_sdc(sdc_file)

        # shts = ['VarDef','ClkDef','IntExp','IODly','IOExp']
        shts = ['VarDef','ClkDef','IODly','Exp']
        sdc_dir = sdc_path + '/outputs/'
        # sdc_dir = os.path.join(f'{sdc_path}','outputs')
        for sht in shts:
            #if re.search(r'VarDef|ClkDef|IODly|IOExp|IntExp',sht):
            self._sheets[sht].write_sdc(sdc_dir,prousr)

        # clkdef = self._sheets['ClkDef']
        # sdc_file = sdc_dir + f'{self._alias.lower()}_iodly.sdc'
        # self._iodly.getiodata(self._vfile_list, self._vfile_data)
        # self._data = self._iodly._iodlydata
        # json_file = sdc_path + '/json' + f'/iodly.json'
        # self.write_json(json_file)
        # self._iodly.write_sdc(self._mdname, self._alias, self._vardata, clkdef, sdc_file)

        # sdc_file = sdc_path +  f'/outputs/{self._alias.lower()}_clkcfg.sdc'
        # self.write_clkcfg_sdc(sdc_file,False)
        # sdc_file = sdc_path +  f'/outputs/intg/{self._alias.lower()}_clkcfg_intg.sdc'
        # self.write_clkcfg_sdc(sdc_file,True)

        sdc_file = sdc_path +  f'/outputs/{self._alias.lower()}_subblk.sdc'
        # sdc_file = os.path.join(f'{sdc_path}', 'outputs', f'{self._alias.lower()}_subblk.sdc')
        self.write_subblk_sdc(sdc_file,False)
        # sdc_file = sdc_path +  f'/outputs/intg/{self._alias.lower()}_subblk_intg.sdc'
        # self.write_subblk_sdc(sdc_file,True)

        sdc_file = sdc_path +  f'/outputs/{self._alias.lower()}_funproc.sdc'
        # sdc_file = os.path.join(f'{sdc_path}','outputs', f'{self._alias.lower()}_funproc.sdc')
        self.write_funcproc_sdc(sdc_file)
        sdc_file = sdc_path +  f'/outputs/{self._alias.lower()}_funcdft.sdc'
        self.write_funcdft_sdc(sdc_file)
        sdc_file = sdc_path +  f'/outputs/{self._alias.lower()}_funcom.sdc'
        self.write_funcom_sdc(sdc_file)
        sdc_file = sdc_path +  f'/outputs/{self._alias.lower()}_tune.sdc'
        self.write_tune_sdc(sdc_file)
        sdc_file = sdc_path +  f'/outputs/{self._alias.lower()}_clkgrp.sdc'
        # sdc_file = os.path.join(f'{sdc_path}', 'outputs', f'{self._alias.lower()}_clkgrp.sdc')
        self.write_clkgrp_sdc(sdc_file)

        sdc_file = sdc_path + f'/outputs/{self._mdname}.func.sdc'
        # sdc_file = os.path.join(f'{sdc_path}', 'outputs', f'{self._mdname}.func.sdc')
        self.write_func_sdc(sdc_file,False)
        # sdc_file = sdc_path + f'/outputs/{self._mdname}.func.flat.sdc'
        # self.write_func_sdc(sdc_file,True)


####################################################################
    def get_subhd_var(self,blk,fname):
        jdata = {}
        tmvars = {}
        cblk = self._hier_tree.get_block_by_name(blk)
        #bals = cblk.alias
        chdl = cblk._hdlevel
        if self._hier_tree.proj:
            hdir = cblk.constr_dir + f'sdcgen/json'
            # hdir = os.path.join(f'{cblk.constr_di}','sdc','json')
        else:                   
            hdir = self._sdcdir + f'/../../{blk}/sdcgen/json'
            # hdir = os.path.join(f'{self._sdcdir}','..','..','sdc','json')

        if chdl in ('blk','sys','top'):
            if not os.path.exists(hdir):
                sdc_warn(f'Missing directory {hdir} for intg check.')
            else:
                json_file = f'{hdir}/{fname}'
                if os.path.exists(json_file):
                    jdata = self.read_json(json_file)
                else:
                    sdc_warn(f'SDC_INFO:Missing {fname} file of {blk}')

            mvars = {k:v for k,v in jdata.items() if 'TMVAR' in k}
            #tmhier = {k:v for k,v in jdata.items() if 'TMHIER' in k}
            if mvars:
                for k,v in mvars.items():
                    tmvars[v['Variable']] = v['Value']
            else:
                sdc_error(f'Missing TMVAR info. in {blk} vardef.json')

        #print('tmvars++++++++++++:',tmvars)
        return tmvars
    
    def write_func_sdc(self,sdc_file,fintg=False):
        func_lines = ''
        als = self._alias
        name= self._mdname
        cblk = self._hier_tree.get_block_by_name(name)
        if cblk._hdlevel == 'sys':
            lvl = 'SYS'
        else:
            lvl = 'BLK'
        pwr = cblk.prime_pwr.split(' ')[0]
        flt = cblk.lvl_flat
        # sdcdir = self._sheets['VarDef']._vardata['SDC_DIR']
        # comdir = self._sheets['VarDef']._vardata['COM_DIR']
        sdcdir = self._vardata['SDC_DIR']
        comdir = self._vardata['COM_DIR']
        dftdir = self._vardata['DFT_DIR']
        iflt = "0"
        ilib = "0"
        hierblks = []
        if fintg:
            # hierblksg = self._hier_tree.get_hierlvlblks(name,outtype='hd')
            # hiercrg = self._hier_tree.get_hierlvlblks(name,outtype='crg')
            # hierblksg.extend(hiercrg)
            hierblksg = self._hier_tree.get_hierblks(name)           
            hierblks = [x for x in hierblksg if not x is name]
            iflt = "1"
            ilib = "0"
        else:
            # hierblks = cblk.get_curhd_by_name()
            # hierdig = cblk.get_curdig_by_name()
            # for bk in hierdig:
            #     sbk = self._hier_tree.get_block_by_name(bk)
            #     if sbk.hdlevel == 'crg' and not bk in hierblks:
            #         hierblks.append(bk)
            hierblks = self._hier_tree.get_curblks(name)
            # hierblksg = self._hier_tree.get_hierlvlblks(name)
            # hierblks = [x for x in hierblksg if not x is name]
            iflt = "0"
            ilib = "0"

        print('hierblks+++++++++++++++++:',hierblks)
        hierblks = list(set(hierblks))

        # curhd vars
        func_lines += f'''
       
global SDCVAR

# curhd: {als}
set {als} "{als}"
set SDCVAR({flt},${{{als}}}) "0"
set SDCVAR(LIB,${{{als}}}) "0"
set SDCVAR(HIER,{lvl},${{{als}}}) ""
set SDCVAR(SDC_DIR,${{{als}}}) "{sdcdir}"
set SDCVAR(COM_DIR,${{{als}}}) "{comdir}"
set SDCVAR(DFT_DIR,${{{als}}}) "{dftdir}"
'''
        # dcdc_vl
        dcdict = {}
        if pwr in list(self._hier_tree._pwrdata.keys()):
            dcdict[f'{pwr}'] = f'{pwr}'
        else:
            sdc_warn(f'Mismatch {name} dcdc power in hier pwr yaml')
        if hierblks:
            for hblk in hierblks:
                sbk = self._hier_tree.get_block_by_name(hblk)
                hpwr = sbk.prime_pwr.split(' ')[0]
                if hpwr in list(self._hier_tree._pwrdata.keys()):
                    dcdict[f'{hpwr}'] = f'{hpwr}'
                else:
                    #'VDDM_PLS': ['VDDM_PLS1','VDDM_PLS2']
                    varl = self._hier_tree.get_dcdc_varval_by_name(name,hblk)
                    for x in list(varl.values()):
                        if x in list(self._hier_tree._pwrdata.keys()):
                            dcdict[f'{hpwr}'] = x
                        else:
                            sdc_warn(f'Mismatch {hblk} dcdc power in hier pwr yaml')
            for k,v in dcdict.items():
                vol = self._hier_tree._pwrdata[v].split(' ')[0]
                func_lines += f'''
# dcdc: {k}:{v}
set {k} "{v}"
if {{[info exists SDCVAR(DCDC_VL,${{{k}}})]}} {{
}} else{{
    set SDCVAR(DCDC_VL,${{{k}}}) "{vol}"
}}
'''

        # subhd vars
        if hierblks:
            nomblks,mlblks,slblks = self._hier_tree.sep_mim_blks(name,hierblks)
            print('hierblkds: ', nomblks,mlblks,slblks)
        else:
            nomblks = []
            mlblks = []
            slblks = []

        if nomblks:
            for sblk in nomblks:
                sbk = self._hier_tree.get_block_by_name(sblk)
                sals = sbk.alias
                flt = sbk.lvl_flat
                slv = sbk._hdlevel
                if slv == 'sys':
                    slvl = 'SYS'
                else:
                    slvl = 'BLK'
                tbkals,tbkhier = self._hier_tree.get_hier_alias_hier(name,sblk)
                #print('hierals+++++++++++++++++++++:',sblk,tbkals)
                
                bkals = []
                bkhier = []
                if tbkals and tbkhier:
                    patval = tbkals[0] + f'_NOMIMPAT'
                    if not fintg:
                        for sa,sh in zip(tbkals,tbkhier):
                            if len(sa.split('_')) == 2:
                                bkals.append(sa)
                                bkhier.append(sh)
                    else:
                        bkals = tbkals
                        bkhier = tbkhier

                    #tmvars = self._sheets['VarDef'].get_subhd_var(sblk,'vardef.json')
                    if self._hier_tree.proj:
                        sdir = f'{sbk._constr_dir}sdcgen/outputs/'
                    else:
                        tmvars = self.get_subhd_var(sblk,'vardef.json')
                        if tmvars:
                            sdir = tmvars['SDC_DIR'] 
                        else:
                            sdir = f'{sbk._constr_dir}sdcgen/outputs/'
                            #sdc_error(f'Missing {sblk} vardef info.')              
                    if len(bkals) == 1:
                        if re.search(r'crg',slv):
                            xval = bkals[0].split('_')[-1]
                        elif re.search(r'soft|lib',slv):
                            if len(bkals[0].split('_')) >= 3:
                                xval = '_'.join(bkals[0].split('_')[-3:])
                            else:
                                xval = '_'.join(bkals[0].split('_')[-2:])
                        else:
                            xval = bkals[0]
                        func_lines += f'''
# {sblk}: {sals}
set {sals} "{xval}"
set SDCVAR({flt},${{{sals}}}) "{iflt}"
set SDCVAR(LIB,${{{sals}}}) "{ilib}"
set SDCVAR(HIER,{slvl},${{{sals}}}) "{bkhier[0]}/"
set SDCVAR(SDC_DIR,${{{sals}}}) "{sdir}"
'''
                    if len(bkals) > 1:                                                   
                        func_lines += f'''
# {sblk}: {sals}
set {patval} [list    \\
'''
                        func_lines = func_lines.rstrip()
                        for sa,sh in zip(bkals,bkhier):
                            if re.search(r'crg',slv):
                                xval = sa.split('_')[-1]
                            elif re.search(r'soft|lib',slv):
                                if len(sa.split('_')) >= 3:
                                    xval = '_'.join(sa.split('_')[-3:])
                                else:
                                    xval = '_'.join(sa.split('_')[-2:])
                            else:
                                xval = sa
                            func_lines += f'''
{xval} {sh}/   \\
'''
                            func_lines = func_lines.rstrip()

                        func_lines += f'''
]
foreach {{ALS_VAL HIER_VAL}} ${patval} {{
    set {sals} "$ALS_VAL"
    set SDCVAR({flt},${{{sals}}}) "{iflt}"
    set SDCVAR(LIB,${{{sals}}}) "{ilib}"
    set SDCVAR(HIER,{slvl},${{{sals}}}) "$HIER_VAL"
    set SDCVAR(SDC_DIR,{sals}) "{sdir}"
}}
'''
                    # else:
                    #     sdc_error(f'Missing {sblk} vardef info.')

        # patval = bkals[0] + f'_NOMIM_PAT'
        # patval = bkals[0] +  f'_MLMIM_PAT'
        if mlblks:
            for sblk in mlblks:
                sbk = self._hier_tree.get_block_by_name(sblk)
                sals = sbk.alias
                flt = sbk.lvl_flat
                slv = sbk._hdlevel
                if slv == 'sys':
                    slvl = 'SYS'
                else:
                    slvl = 'BLK'
                tbkals,tbkhier = self._hier_tree.get_hier_alias_hier(name,sblk)
                
                bkals = []
                bkhier = []
                if tbkals and tbkhier:
                    patval = tbkals[0] +  f'_MLMIMPAT'
                    if not fintg:
                        for sa,sh in zip(tbkals,tbkhier):
                            if len(sa.split('_')) == 2:
                                bkals.append(sa)
                                bkhier.append(sh)
                    else:
                        bkals = tbkals
                        bkhier = tbkhier
                    if bkhier and bkhier:
                        if self._hier_tree.proj:
                            sdir = f'{sbk._constr_dir}sdcgen/outputs/'
                        else:
                            #tmvars = self._sheets['VarDef'].get_subhd_var(sblk,'vardef.json')
                            tmvars = self.get_subhd_var(sblk,'vardef.json')
                            if tmvars:
                                sdir = tmvars['SDC_DIR']
                            else:
                                sdir =  f'{sbk._constr_dir}sdcgen/outputs/'
                                #sdc_error(f'Missing {sblk} vardef info.')                                                         
                        func_lines += f'''
# {sblk}: {sals}
set {patval} [list    \\
'''
                        func_lines = func_lines.rstrip()
                        for sa,sh in zip(bkals,bkhier):
                            if re.search(r'crg',slv):
                                xval = sa.split('_')[-1]
                            elif re.search(r'soft|lib',slv):
                                if len(sa.split('_')) >= 3:
                                    xval = '_'.join(sa.split('_')[-3:])
                                else:
                                    xval = '_'.join(sa.split('_')[-2:])
                            else:
                                xval = sa
                            func_lines += f'''
{xval} {sh}/   \\
'''
                            func_lines = func_lines.rstrip()
                        func_lines += f'''
]
foreach {{ALS_VAL HIER_VAL}} ${patval} {{
    set {sals} "$ALS_VAL"
    set SDCVAR({flt},${{{sals}}}) "{iflt}"
    set SDCVAR(LIB,${{{sals}}}) "{ilib}"
    set SDCVAR(HIER,{slvl},${{{sals}}}) "$HIER_VAL"
    set SDCVAR(SDC_DIR,{sals}) "{sdir}"
}}
'''

        if slblks:
            for sblk in slblks:
                sbk = self._hier_tree.get_block_by_name(sblk)
                sals = sbk.alias
                flt = sbk.lvl_flat
                slv = sbk._hdlevel
                if slv == 'sys':
                    slvl = 'SYS'
                else:
                    slvl = 'BLK'
                tbkals,tbkhier = self._hier_tree.get_hier_alias_hier(name,sblk)
                
                bkals = []
                bkhier = []
                if tbkals and tbkhier:
                    patval = tbkals[0] + f'_SLMIMPAT'
                    if not fintg:
                        for sa,sh in zip(tbkals,tbkhier):
                            if len(sa.split('_')) == 2:
                                bkals.append(sa)
                                bkhier.append(sh)
                    else:
                        bkals = tbkals
                        bkhier = tbkhier
                    
                    if self._hier_tree.proj:
                        sdir = f'{sbk._constr_dir}sdc/outputs/'
                    else:
                        #tmvars = self._sheets['VarDef'].get_subhd_var(sblk,'vardef.json')
                        tmvars = self.get_subhd_var(sblk,'vardef.json')
                        if tmvars:
                            sdir = tmvars['SDC_DIR'] 
                        else:
                            sdir =  f'{sbk._constr_dir}sdc/outputs/'
                            #sdc_error(f'Missing {sblk} vardef info.')
                    if bkals and bkhier:
                        #for sa,sh in zip(bkals,bkhier):
                        sval = bkals[0].split('$')[0]
                        func_lines += f'''
# {sblk}: {sals}
# set {sals} "{sval}"
set SDCVAR({flt},{sals}) "{iflt}"
set SDCVAR(LIB,{sals}) "{ilib}"
#set SDCVAR(HIER,{slvl},${{{sals}}}) "{bkhier[0]}/"
set SDCVAR(SDC_DIR,{sals}) "{sdir}"
'''

                    # else:
                    #     sdc_error(f'Missing {sblk} vardef info.')

        lals = als.lower()
        lv = lvl.lower()
        func_lines += f'''

source -echo -verbose $SDCVAR(SDC_DIR,${{{als}}}){lals}_{lv}top.sdc
'''

        self.save_text(func_lines,sdc_file)

    def write_func_flat_sdc(self,sdc_file):
        pass

    def write_func_dvfs_sdc(self,sdc_file):
        pass

    def write_func_dvfs_flat_sdc(self,sdc_file):
        pass


    def write_top_sdc(self,sdc_file):

        if self._hdlvl == 'sys':
            flt = 'IS_CHIP'
            lvl = 'sys'
            lev = 'Sys'
        if self._hdlvl == 'blk':
            flt = 'IS_FLAT'
            lvl = 'blk'
            lev = 'Blk'

        alias = self._alias.lower()
        mdname = self._mdname
        ualias = self._alias
        tclkdef = self._sheets['ClkDef']

#         sdc_lines = f'''
# #########################################################
# ## Integration of {alias} lib_based
# #########################################################
# if {{$SDCVAR({flt},${{{ualias}}}) && $SDCVAR(LIB,${{{ualias}}})}} {{
#     if {{[file exists $SDCVAR(SDC_DIR,${{{ualias}}})blklib/{alias}_{lvl}lib.sdc]}} {{
#         puts "SDC_INFO: Sourcing $SDCVAR(SDC_DIR,${{{ualias}}})blklib/{alias}_{lvl}lib.sdc."
#         source -echo -verbose $SDCVAR(SDC_DIR,${{{ualias}}})blklib/{alias}_{lvl}lib.sdc
#     }} else {{
#         puts "SDC_WARN: Missing $SDCVAR(SDC_DIR,${{{ualias}}})blklib/{alias}_{lvl}lib.sdc. Please check it."
#     }}
# }} else {{
# '''

        sdc_lines = f'''
#########################################################
## Section 1: General Variables
#########################################################
if {{[info exists SDCVAR({lev}Nm,${{{ualias}}})]}} {{
}} else {{
    set SDCVAR({lev}Nm,${{{ualias}}}) "${{{ualias}}}"
}}
if {{$SDCVAR({lev}Nm,${{{ualias}}}) == ""}} {{
    puts "SDC_ERROR: SDCVAR({lev}Nm,${{{ualias}}} is not set. Please check it."
}}

if {{[info exists SDCVAR({flt},${{{ualias}}})]}} {{
}} else {{
    set SDCVAR({flt},${{{ualias}}}) "${{{ualias}}}"
}}
if {{$SDCVAR({flt},${{{ualias}}}) == ""}} {{
    puts "SDC_ERROR: SDCVAR({flt},${{{ualias}}} is not set. Please check it."
}}

global design

if {{[info exists SDCVAR(FL_STAGE)]}} {{
}} else {{
    set SDCVAR(FL_STAGE) "RTL"
}}
if {{$SDCVAR(FL_STAGE) == ""}} {{
    puts "SDC_ERROR: SDCVAR(FL_STAGE) is not set. Please check it."
}}    

if {{$SDCVAR(FL_STAGE) !=  "SIGNOFF"}} {{
    if {{!$SDCVAR({flt},${{{ualias}}})}} {{
        if {{[array exists CLOCK_GORUP_NAME]}} {{
            array unset CLOCK_GROUP_NAME
        }}
    }}
}}

#########################################################
## Section 2: Common SDC procs
#########################################################   
if {{!$SDCVAR({flt},${{{ualias}}})}} {{ 
    if {{[file exists $SDCVAR(COM_DIR,${{{ualias}}})func_proc.sdc]}} {{
        puts "SDC_INFO: Sourcing $SDCVAR(COM_DIR,${{{ualias}}})func_proc.sdc."
        source -echo -verbose $SDCVAR(COM_DIR,${{{ualias}}})func_proc.sdc
    }} else {{
        puts "SDC_WARN: Missing $SDCVAR(COM_DIR,${{{ualias}}})func_proc.sdc. Please check it."
    }}
}}

'''

        sdc_lines += self.com_sdcout1(ualias,'3:  Block Variables',f'{alias}_{lvl}var')
        #sdc_lines += self.com_sdcout1(ualias,'3:  Clock Config',f'{alias}_clkcfg')
        fname = f'{alias}_clkcfg'
        sdc_lines += f'''
############################################################
## Section 4:  Clock Config
############################################################
if {{!$SDCVAR({flt},${{{ualias}}})}} {{
    if {{[file exists $SDCVAR(SDC_DIR,${{{ualias}}}){fname}.sdc]}} {{
        puts "SDC_INFO: Sourcing $SDCVAR(SDC_DIR,${{{ualias}}}){fname}.sdc"
        source -echo -verbose  $SDCVAR(SDC_DIR,${{{ualias}}}){fname}.sdc
    }} else {{
        puts "SDC_WARN: Missing $SDCVAR(SDC_DIR,${{{ualias}}}){fname}.sdc. Please check it."
    }}
}}

'''
        #sdc_lines += self.com_sdcout1(ualias,'4:  Block Variables',f'{alias}_{lvl}var')

        inpsdir = dirname(dirname(sdc_file)) + '/inputs'
        otpsdir = dirname(dirname(sdc_file)) + '/outputs'
        # how to solve multi autoclk order ? currently only one crg
        
        # for fname in os.listdir(inpsdir):
        #     if 'clk' in fname:
        #         autoclklst.append(f'{inpsdir}/{fname}')
        #         os.system(f'cp -rf {inpsdir}/{fname} {otpsdir}/{alias}_autoclk.sdc')
        #         autoclkflag = 1

        # 'MCUJPEG_CRG10_CRG_CRG': ['xx/xx/clk_core'];; CRG1 from inst_dig
        # 'MCUJPEG_PLL11_PLL_PLL': ['xx/xx/pll_top_wrap']
        # crgflg,crgfls = self._sheets['ClkDef'].get_macdig_info('org')
        crgflg = 0
        crgfls = 'jpeg.sdc'
        if crgflg:
            # kw : 'MCUJPEG_MCUCRG1#0_CRG_CRG' | 'MCUJPEG_MCUPLL_PLL_PLL'
            for kw,fl in crgfls.items():
                ks = kw.replace('#','').split('_')
                fn = fl.split('/')[-1].replace('.sdc','') 
                #als = kw.replace('#','')
                # als : 'MCUJPEG_MCUCRG10_CRG_CRG'
                if ks[-1] == 'CRG':
                    if '#' in kw:
                        n = ks[1][-1]
                    else:
                        n = 0
                    os.system(f'cp -rf {fl} {otpsdir}/{fn}_autoclk{n}.sdc')
                    os.system(f'cp -rf {fl} {otpsdir}/intg/{fn}_autoclk{n}_intg.sdc')
                    # crg_clkinfo = tclkdef._crgalsiptval
                    # crg_lines = self.read_text(fl)
                    # ncrg_lines = self.align_crgip_vars(kw,crg_lines,crg_clkinfo,'CRG')
                    # fcrg1 = f'{otpsdir}/{fn}_autoclk{n}.sdc'
                    # fcrg2 = f'{otpsdir}/intg/{fn}_autoclk{n}_intg.sdc'
                    # self.save_text(ncrg_lines,fcrg1)
                    # self.save_text(ncrg_lines,fcrg2)


        if crgflg:
            sdc_lines += f'''
#########################################################
## Section 5: Clock Definition
######################################################### 
if {{[file exists $SDCVAR(SDC_DIR,${{{ualias}}}){alias}_userclk.sdc]}} {{
    puts "SDC_INFO: Sourcing $SDCVAR(SDC_DIR,${{{ualias}}}){alias}_userclk.sdc."
    source -echo -verbose $SDCVAR(SDC_DIR,${{{ualias}}}){alias}_userclk.sdc
}} else {{
    puts "SDC_WARN: Missing $SDCVAR(SDC_DIR,${{{ualias}}}){alias}_userclk.sdc. Please check it."
}}
'''
            # sdc_lines += self.write_crgip_sdc(mdname,alias,crgflg,crgfls)

            sdc_lines += f'''
if {{[file exists $SDCVAR(SDC_DIR,${{{ualias}}}){alias}_updtclk.sdc]}} {{
    puts "SDC_INFO: Sourcing $SDCVAR(SDC_DIR,${{{ualias}}}){alias}_updtclk.sdc."
    source -echo -verbose $SDCVAR(SDC_DIR,${{{ualias}}}){alias}_updtclk.sdc
}} else {{
    puts "SDC_WARN: Missing $SDCVAR(SDC_DIR,${{{ualias}}}){alias}_updtclk.sdc. Please check it."
}}

'''
        else:
            sdc_lines += f'''
#########################################################
## Section 5: Clock Definition
######################################################### 
if {{[file exists $SDCVAR(SDC_DIR,${{{ualias}}}){alias}_clkdef.sdc]}} {{
    puts "SDC_INFO: Sourcing $SDCVAR(SDC_DIR,${{{ualias}}}){alias}_clkdef.sdc."
    source -echo -verbose $SDCVAR(SDC_DIR,${{{ualias}}}){alias}_clkdef.sdc
}} else {{
    puts "SDC_WARN: Missing $SDCVAR(SDC_DIR,${{{ualias}}}){alias}_clkdef.sdc. Please check it."
}}

'''
            # sdc_lines += self.write_crgip_sdc(mdname,alias,crgflg,crgfls)

        sdc_lines += self.com_sdcout1(ualias,'6: Nested Harden Block & Third_party IP',f'{alias}_subblk')
        sdc_lines += f'''
#########################################################
## Section 7: Funcdft Related Constraints
######################################################### 
if {{!$SDCVAR({flt},${{{ualias}}})}} {{
    if {{[file exists $SDCVAR(SDC_DIR,${{{ualias}}}){alias}_funcdft.sdc]}} {{
        puts "SDC_INFO: Sourcing $SDCVAR(SDC_DIR,${{{ualias}}}){alias}_funcdft.sdc."
        source -echo -verbose $SDCVAR(SDC_DIR,${{{ualias}}}){alias}_funcdft.sdc
    }} else {{
        puts "SDC_WARN: Missing $SDCVAR(SDC_DIR,${{{ualias}}}){alias}_funcdft.sdc. Please check it."
    }}
}}

'''
        sdc_lines += self.com_sdcout1(ualias,'8: Internal Exception',f'{alias}_intexp')
        sdc_lines += f'''
#########################################################
## Section 9: IO Related Constraints
######################################################### 
if {{!$SDCVAR({flt},${{{ualias}}})}} {{
    if {{[file exists $SDCVAR(SDC_DIR,${{{ualias}}}){alias}_iodly.sdc]}} {{
        puts "SDC_INFO: Sourcing $SDCVAR(SDC_DIR,${{{ualias}}}){alias}_iodly.sdc."
        source -echo -verbose $SDCVAR(SDC_DIR,${{{ualias}}}){alias}_iodly.sdc
    }} else {{
        puts "SDC_WARN: Missing $SDCVAR(SDC_DIR,${{{ualias}}}){alias}_iodly.sdc. Please check it."
    }}
}}

if {{[file exists $SDCVAR(SDC_DIR,${{{ualias}}}){alias}_ioexp.sdc]}} {{
    puts "SDC_INFO: Sourcing $SDCVAR(SDC_DIR,${{{ualias}}}){alias}_ioexp.sdc."
    source -echo -verbose $SDCVAR(SDC_DIR,${{{ualias}}}){alias}_ioexp.sdc
}} else {{
    puts "SDC_WARN: Missing $SDCVAR(SDC_DIR,${{{ualias}}}){alias}_ioexp.sdc. Please check it."
}}

'''
        
        sdc_lines += f'''
#########################################################
## Section 10: Common Constraints
######################################################### 
if {{!$SDCVAR({flt},${{{ualias}}})}} {{
    if {{[file exists $SDCVAR(SDC_DIR,${{{ualias}}})funcom.sdc]}} {{
        puts "SDC_INFO: Sourcing $SDCVAR(SDC_DIR,${{{ualias}}})funcom.sdc."
        source -echo -verbose $SDCVAR(SDC_DIR,${{{ualias}}})funcom.sdc
    }} else {{
        puts "SDC_WARN: Missing $SDCVAR(SDC_DIR,${{{ualias}}})funcom.sdc. Please check it."
    }}
}}

'''

        sdc_lines += self.com_sdcout1(ualias,'11: User Tune Constraint',f'{alias}_tune')
        sdc_lines += f'''
#########################################################
## Section 12: Clock Group
######################################################### 
if {{!$SDCVAR({flt},${{{ualias}}})}} {{
    if {{[file exists $SDCVAR(SDC_DIR,${{{ualias}}}){alias}_clkgrp.sdc]}} {{
        puts "SDC_INFO: Sourcing $SDCVAR(SDC_DIR,${{{ualias}}}){alias}_clkgrp.sdc."
        source -echo -verbose $SDCVAR(SDC_DIR,${{{ualias}}}){alias}_clkgrp.sdc
    }} else {{
        puts "SDC_WARN: Missing $SDCVAR(SDC_DIR,${{{ualias}}}){alias}_clkgrp.sdc. Please check it."
    }}
}}

'''

        self.save_text(sdc_lines,sdc_file)      

    def write_crgip_sdc(self,name,als,crgflg,crgfls):
        sdc_lines = ''
        cur_clkdef = self._sheets['ClkDef']
        scrgip = {}

        if crgflg:
            # kw : 'MCUJPEG_MCUCRG1#0_CRG_CRG' | 'MCUJPEG_MCUPLL_PLL_PLL'
            ncrg = {x:x.split('_')[1] for x,y in crgfls.items() if x.split('_')[-1] in 'CRG'}
            #npll = {x:x.split('_')[1] for x,y in crgfls.items() if x.split('_')[-1] in 'PLL'}
            scrgip.update(cur_clkdef._rowcrgdata)
            #scrgip.update(cur_clkdef._rowipdata)
            #{'MCUJPEG_MCUCRG10_MCRG_CRG': {'MCUJPEG_MCUCRG10_MCRG_CRG_CRG_Row18': {'IntgType':''}}}  
            if scrgip:
                scrg = {}
                for kx,vx in scrgip.items():
                    scrg.update({k:v for k,v in vx.items() if '_CRG_CRG_' in k})
                #sip = {k:v for k,v in scrgip.items() if '_MACLIB_IP_' in k or '_DIGSOFT_IP_' in k}
                #print('write_crgip_sdc+++++++++++++++++++:',scrg)
                ncivar = {}  
                if scrg:
                    for sk,sv in scrg.items():
                        skr = '_'.join(sk.split('_')[:-2])
                        if not skr in list(ncivar.keys()):
                            svr = sv['CRGAVR']
                            #svr = sv['IPAVR']
                            ncivar[skr] = svr
            #sdc_lines += self.write_cur_crgip_sdc(als,crgfls,npll,ncivar,'')
            sdc_lines += self.write_cur_crgip_sdc(als,crgfls,ncrg,ncivar,'')

        hblk = self._hier_tree.get_block_by_name(name)
        hdblks = hblk.get_curhd_by_name()
        if hdblks:
            for sblk in hdblks:
                #{'MCUJPEG_CR8_CR8_DIGSOFT_IP_Row17':{}}
                scrgip = cur_clkdef.get_data_from_json_intg(sblk,'_header.json','semi')
                if scrgip:
                    scrg = {k:v for k,v in scrgip.items() if '_CRG_CRG_' in k}
                    #sip = {k:v for k,v in scrgip.items() if '_MACLIB_IP_' in k or '_DIGSOFT_IP_' in k}
                    ncivar = {} 
                    if scrg:                                  
                        for sk,sv in scrg.items():
                            skr = '_'.join(sk.split('_')[:-2])
                            if not skr in list(ncivar.keys()):
                                svr = sv['CRGAVR']
                                #svr = sv['IPAVR']
                                ncivar[skr] = svr
                        #print('ncivar++++++++++++++:',ncivar)

                blk = self._hier_tree.get_block_by_name(sblk)
                sals = blk.alias      
                scrgflg,scrgfls = self._sheets['ClkDef'].get_subhd_crgip_info(sblk,'org')
                if scrgflg:
                    ncrg = {x:x.split('_')[1] for x,y in scrgfls.items() if x.split('_')[-1] in 'CRG'}
                    #npll = {x:x.split('_')[1] for x,y in scrgfls.items() if x.split('_')[-1] in 'PLL'}
                    #sdc_lines += self.write_cur_crgip_sdc(sals,scrgfls,npll,ncivar,'intg')
                    sdc_lines += self.write_cur_crgip_sdc(sals,scrgfls,ncrg,ncivar,'intg')

        return sdc_lines

    def write_cur_crgip_sdc(self,als,crgfls,ncrg,ncivar,flg='',):
        sdc_lines = ''
        if flg == 'intg':
            xf = '_intg'
        else:
            xf = ''
        if ncrg:
            # [('x_c10_c_r', '0'), ('x_p11_c_r', '1')]
            tcrg = {k:v[-1] for k,v in ncrg.items() if '#' in v}
            ycrg = {k:v for k,v in ncrg.items() if not '#' in v}
            # multi crg/ip with order
            if tcrg:
                xcrg = sorted(tcrg.items(), key = lambda x:int(x[1]))
                clst = [k for k,_ in xcrg]
                cdic = {k:v for k,v in xcrg}
                if clst:
                    for k in clst:
                        if flg == 'intg':
                            fn = crgfls[k].split('/')[-1].split('_autoclk')[0]
                        else:
                            fn = crgfls[k].split('/')[-1].replace('.sdc','')
                        #if tcrg:
                        idx = cdic[k]
                        if ncivar:
                            #kw = k.split('_')[3]
                            val = k.split('_')[1].replace('#','')
                            var = k.split('_')[2]
                            if '_CRG' in k or '_PLL' in k:
                                ks = [x for x in list(ncivar.keys()) if k.replace('#','') in x]
                                #print('TTTT+++++++:',k,ks,ncivar.keys())
                                civar = ncivar[ks[0]]
                                if not civar in var:
                                    sdc_lines += f'''
# {fn}_autoclk{idx}
# {k} variable in sdc: {civar}
set {civar} "{val}"
'''
                                else:
                                    sdc_info(f'Same crg name bet crg sdc and hier yaml.')
                        else:
                            sdc_lines += f'''
# {fn}_autoclk{idx}
'''     
                        sdc_lines = sdc_lines.rstrip()
                        sdc_lines += f'''                       
if {{[file exists $SDCVAR(SDC_DIR,${{{als}}}){flg}/{fn}_autoclk{idx}{xf}.sdc]}} {{
    puts "SDC_INFO: Sourcing $SDCVAR(SDC_DIR,${{{als}}}){flg}/{fn}_autoclk{idx}{xf}.sdc."
    source -echo -verbose $SDCVAR(SDC_DIR,${{{als}}}){flg}/{fn}_autoclk{idx}{xf}.sdc
}} else {{
    puts "SDC_WARN: Missing $SDCVAR(SDC_DIR,${{{als}}}){flg}/{fn}_autoclk{idx}{xf}.sdc. Please check it."
}}
'''
            # single crg/ip without order
            if ycrg:
                clst = list(ycrg.keys())
                if clst:
                    for k in clst:
                        if flg == 'intg':
                            fn = crgfls[k].split('/')[-1].split('_autoclk')[0]
                        else:
                            fn = crgfls[k].split('/')[-1].replace('.sdc','')
                        idx = 0
                        sdc_lines += f'''
# {fn}_autoclk{idx}
if {{[file exists $SDCVAR(SDC_DIR,${{{als}}}){flg}/{fn}_autoclk{idx}{xf}.sdc]}} {{
    puts "SDC_INFO: Sourcing $SDCVAR(SDC_DIR,${{{als}}}){flg}/{fn}_autoclk{idx}{xf}.sdc."
    source -echo -verbose $SDCVAR(SDC_DIR,${{{als}}}){flg}/{fn}_autoclk{idx}{xf}.sdc
}} else {{
    puts "SDC_WARN: Missing $SDCVAR(SDC_DIR,${{{als}}}){flg}/{fn}_autoclk{idx}{xf}.sdc. Please check it."
}}
'''        
        return sdc_lines


    def write_subblk_sdc(self,sdc_file,fintg=False):
        subblk_lines = ''
        mdname = self._mdname
        alias = self._alias
        pwr = self._hier_tree._blocks[mdname].prime_pwr
        tclkdef = self._sheets['ClkDef']

        if not fintg:
            subblk_lines += '''
#############################################################
## Nested Harden Block
#############################################################
'''
            blkf = 'subblk'            
            subblk_lines += self._hier_tree.set_subblk_intg(mdname,blkf,[],'full')

        subblk_lines += f'''
#############################################################
## Third_Party Macro IP
#############################################################
'''
        macdig_user = self._hier_tree.get_usersdc_by_name(mdname)
        macuser = {k:v for k,v in macdig_user.items() if '_MACLIB' in k}
        diguser = {k:v for k,v in macdig_user.items() if '_DIGSOFT' in k}
        if macuser:
            for ky,vl in macuser.items():
                subblk_lines += vl

        subblk_lines += f'''
#############################################################
## Third_Party Digital IP
#############################################################
'''
        if diguser:
            for ky,vl in diguser.items():
                    subblk_lines += vl               

        # if not fintg:
        #     npsdir = dirname(dirname(sdc_file)) + '/inputs'
        #     otpsdir = dirname(dirname(sdc_file)) + '/outputs'
        # else:
        #     npsdir = dirname(dirname(dirname(sdc_file))) + '/inputs'
        #     otpsdir = dirname(dirname(dirname(sdc_file))) + '/outputs'            
        # ipflg,ipfls = tclkdef.get_userip_info('org')
        # if ipflg:
        #     # kw : 'MCUJPEG_MXGPY1#0_GPY_SOFT' | 'MCUJPEG_CR8_CR8_LIB'
        #     digfls = {k:v for k,v in ipfls.items() if '_SOFT' in k}
        #     macfls = {k:v for k,v in ipfls.items() if '_LIB' in k}
        #     ip_clkinfo = tclkdef._ipalsiptval
        #     subblk_lines += self.write_digmac_sdc(alias,digfls,ip_clkinfo,'dig','Macro',otpsdir,fintg)           
        #     subblk_lines += self.write_digmac_sdc(alias,macfls,ip_clkinfo,'mac','Digital',otpsdir,fintg)


#################################################################
    # clkintg: all clk vars & definition
    #   -CRGIN/IPIN/HDIN    ----> blkvar
    #   -CRGOUT             ---> updt
    #   -IPOUT              ---> subblk
    #   -HDOUT              ----> updt or subblk(nocrg)
    # outclktype from crg: only mstclk/srcpin/grpnm
    #   -CRGIN/HDIN/IPIN    -----> updt
    # outclktype from ip: only mstclk/srcpin/grpnm
    #   -CRGIN/HDIN/IPIN    -----> subblk    
#################################################################


#         subblk_lines += f'''
# #############################################################
# ## Clk Var&Def from clkdef with IPOUT/HDOUT
# #############################################################
# '''
#         if not tclkdef._clknmlst:
#             tclkdef.get_clkdata_by_clkname(tclkdef._clkdata)
#         clknmlst = tclkdef._clknmlst
#         clknmdata = tclkdef._clknmdata
#
#         ipclknmlst = []
#         ipclknmdata = {}
#         hdclknmlst = []
#         hdclknmdata = {}
#         #subblk_lines += tclkdef.get_clkvar(mdname,alias,pwr,clknmlst,clknmdata,fintg,'OUT')
#         for clknm in clknmlst:
#             if tclkdef.is_crgiphd_goutclk(clknm,'IPOUT'):
#                 ipclknmlst.append(clknm)
#                 ipclknmdata[clknm] = clknmdata[clknm]
#             if tclkdef.is_crgiphd_goutclk(clknm,'HDOUT'):
#                 hdclknmlst.append(clknm)
#                 hdclknmdata[clknm] = clknmdata[clknm]
#         tclknmlst = []
#         if ipclknmlst:
#             subblk_lines += tclkdef.get_clkvar(mdname,alias,pwr,ipclknmlst,ipclknmdata,fintg,'IPOUT')
#             tclknmlst.extend(ipclknmlst)
#         if hdclknmlst:
#             subblk_lines += tclkdef.get_clkvar(mdname,alias,pwr,hdclknmlst,hdclknmdata,fintg,'HDOUT')
#             tclknmlst.extend(hdclknmlst)
#
#         subblk_lines += tclkdef.align_crgipoutclk_naming(mdname,'IPOUT')
#
#         if tclknmlst:
#             for clknm in tclknmlst:
#                 subblk_lines += f'''
#
# ## generated clock: {clknm}
# create_generated_clock -name $SDCVAR(NAME,${{{alias}}},{clknm}) -master_clock $SDCVAR(NAME,MST,${{{alias}}},{clknm}) -source $SDCVAR(HIER,SRC,${{{alias}}},{clknm}) $SDCVAR(DIVDEGE,${{{alias}}},{clknm}) $SDCVAR(HIER,${{{alias}}},{clknm}) -add -comment $SDCVAR(CMT,${{{alias}}},{clknm})
# '''

#         subblk_lines += tclkdef.get_clkvar(mdname,alias,pwr,clknmlst,clknmdata,fintg,'IPOUT')
#         for clknm in clknmlst:
#             if tclkdef.is_crgiphd_goutclk(clknm,'IPOUT'):
#                 subblk_lines += f'''

# ## generated clock: {clknm}
# create_generated_clock -name $SDCVAR(NAME,${{{alias}}},{clknm}) -master_clock $SDCVAR(NAME,MST,${{{alias}}},{clknm}) -source $SDCVAR(HIER,SRC,${{{alias}}},{clknm})  $SDCVAR(HIER,${{{alias}}},{clknm}) -add -comment $SDCVAR(CMT,${{{alias}}},{clknm})
# '''
#         subblk_lines += tclkdef.get_clkvar(mdname,alias,pwr,clknmlst,clknmdata,fintg,'HDOUT')
#         for clknm in clknmlst:
#             if tclkdef.is_crgiphd_goutclk(clknm,'HDOUT'):
#                 subblk_lines += f'''

# ## generated clock: {clknm}
# create_generated_clock -name $SDCVAR(NAME,${{{alias}}},{clknm}) -master_clock $SDCVAR(NAME,MST,${{{alias}}},{clknm}) -source $SDCVAR(HIER,SRC,${{{alias}}},{clknm})  $SDCVAR(HIER,${{{alias}}},{clknm}) -add -comment $SDCVAR(CMT,${{{alias}}},{clknm})
# '''

#         subblk_lines += f'''
# #############################################################
# ## Clk Variables from IP to CRGIN/IPIN/HDIN
# #############################################################
# '''
#         subblk_lines += tclkdef.set_mstsrcgrp_from_outclktype('IPOUT')

        if not fintg:
            subblk_lines += f'''
#############################################################
## Clock Group for flatten and toponly
#############################################################
'''      
            subblk_lines += self._sheets['ClkDef'].set_clkgrp(mdname,alias)
        #clkdef = self._sheets['ClkDef']
        #clkdef.get_clkdata_by_clkname(clkdef._clkdata)
        #genclk_lines = clkdef.get_genclk_lines_by_name(mdname)
        else:
            pass
        # how to deal with exclusive clk group intg
#             subblk_lines += f'''
# #############################################################
# ## Exclusively Clock Group
# #############################################################
# '''
            #subblk_lines += self._sheets['ClkDef'].set_phylog_clkgrp

        self.save_text(subblk_lines,sdc_file)

    def write_funcdft_sdc(self,sdc_file):

        mdname = self._mdname
        flt = self._flt
        lvl = self._lvl
        alias = self._alias

        funcdft_lines = f'''
#############################################################
## {mdname} funcdft constraints
#############################################################        
if {{![ info exists IS_FLAT]}} {{set IS_FLAT 0}}
if {{![ info exists HIER]}} {{set HIER ""}}
'''

        hdblksg = self._hier_tree.get_hierlvlblks(mdname,outtype='hd')
        hdblks = [x for x in hdblksg if not x is mdname]
        cblk_lib_flg = []
        for sblk in hdblks:
            blk = self._hier_tree.get_block_by_name(sblk)
            bals = blk.alias
            cblk_lib_flg.append(f'$SDCVAR(LIB,${{{bals}}}) &&')
        sblk_lib_flg = ' '.join(cblk_lib_flg).rstrip('&&')

        funcdft_lines += f'''
if {{$SDCVAR(FL_STAGE) != "RTL" && $SDCVAR(FL_STAGE) != "SYN"}} {{
    if {{!$SDCVAR({flt},${{{alias}}})}} {{
        if {{{sblk_lib_flg}}} {{
            if {{[file exists $SDCVAR(DFT_DIR,${{{alias}}}){mdname}.scanfunc.sdc]}} {{
                puts "SDC_INFO: Sourcing $SDCVAR(DFT_DIR,${{{alias}}}){mdname}.scanfunc.sdc"
                source -echo -verbose $SDCVAR(DFT_DIR,${{{alias}}}){mdname}.scanfunc.sdc
            }} else {{
                puts "SDC_WARN: Missing $SDCVAR(DFT_DIR,${{{alias}}}){mdname}.scanfunc.sdc. Please check it."
            }}
        }} else {{
            if {{[file exists $SDCVAR(DFT_DIR,${{{alias}}}){mdname}.flat.scanfunc.sdc]}} {{
                puts "SDC_INFO: Sourcing $SDCVAR(DFT_DIR,${{{alias}}}){mdname}.flat.scanfunc.sdc"
                source -echo -verbose $SDCVAR(DFT_DIR,${{{alias}}}){mdname}.flat.scanfunc.sdc
            }} else {{
                puts "SDC_WARN: Missing $SDCVAR(DFT_DIR,${{{alias}}}){mdname}.flat.scanfunc.sdc. Please check it."
            }}            
        }}
    }}
}}

if {{$SDCVAR(FL_STAGE) != "RTL" && $SDCVAR(FL_STAGE) != "SYN"}} {{
    if {{[file exists $SDCVAR(DFT_DIR,${{{alias}}}){mdname}.mbist.sdc]}} {{
        puts "SDC_INFO: Sourcing $SDCVAR(DFT_DIR,${{{alias}}}){mdname}.mbist.sdc"
        if {{$SDCVAR({flt},${{{alias}}})}} {{
            set HIER "$SDCVAR(HIER,{lvl.upper()},${{{alias}}})"
            set IS_FLAT "1"
            source -echo -verbose $SDCVAR(DFT_DIR,${{{alias}}}){mdname}.mbist.sdc
        }} else {{
            set HIER ""
            set IS_FLAT "0"
            source -echo -verbose $SDCVAR(DFT_DIR,${{{alias}}}){mdname}.mbist.sdc
        }}    
    }} else {{
        puts "SDC_WARN: Missing $SDCVAR(DFT_DIR,${{{alias}}}){mdname}.mbist.sdc. Please check it."
    }}
}}
'''
        # sub harden blk
        funcdft_lines += self._hier_tree.set_subblk_intg(mdname,'',[],'full','mbist')

        self.save_text(funcdft_lines,sdc_file)
          

    def write_funcproc_sdc(self,sdc_file):
        funcproc_lines = f'''
###########################################
proc period_def {{perd_lst}} {{
    global SDCVAR

    set cycle_lst [lsit \\
        CYCLE100M \\
        CYCLE500M \\
        CYCLE125M \\
        CYCLE152M \\
        CYCLE192M \\
        CYCLE350M \\
        CYCLE325M \\
        CYCLE650M \\
        CYCLE001M \\
        CYCLE024M \\
        CYCLE025M \\
        CYCLE026M \\
        CYCLE038M \\
        CYCLE040M \\
        CYCLE050M \\
        CYCLE052M \\
        CYCLE200M \\
        CYCLE250M \\
        CYCLE208M \\
        CYCLE333M \\
        CYCLE384M \\
        CYCLE400M \\
        CYCLE1000M \\
        CYCLE1066M \\
        CYCLE1066M7 \\
        CYCLE2000M \\
        CYCLE2400M \\
        CYCLE1200M \\
        CYCLE4266M  \\
        CYCLE33M \\
        CYCLE076M8 \\
        CYCLE800M \\
        CYCLE96M \\
        CYCLE1060M \\
        CYCLE502M \\
        CYCLE640M \\
        CYCLE032K \\
    ]

        set SDCVAR(CYCLE_LIST) [lsort -u [concat $cycle_lst $perd_lst]]

        foreach cycle $SDCVAR(CYCLE_LIST) {{
            if {{[string match *M* $cycle]}} {{
                set perid_num [string trimleft [string map {{M .}} $cycle] "CYCLE"]
                set SDCVAR($cycle) [expr (int ((1000.0/$perid_num) * 1000))/1000.0]
            }} elseif {{[string match *K* $cycle]}} {{
                set perid_num [string trimleft [string map {{K .}} $cycle] "CYCLE"]
                set SDCVAR($cycle) [expr (int ((1000.0/$perid_num) * 1000 * 1000))/1000.0]            
            }}
        }}
    }}

# proc func_mem_misc {{FL_STAGE}} {{
# }}              
 
# proc func_stop_scanclk {{}} {{
#     if {{[sizeof_coll [get_clocks -q SCAN_TEST_CLK*]] > 0}} {{
#         set occmux_inst [get_pins -hier * -filter "full_name =~ *occ*u_mux2_clk_out/u_dontouch_cell/Z" -q]
#         if {{[sizeof_coll $occmux_inst] > 0}} {{
#             set_sense -stop_propagation -clocks [get_clocks SCAN_TEST_CLK*] $occmux_list
#         }}
#     }}
# }}                            
# '''

        self.save_text(funcproc_lines,sdc_file)
    
    def write_funcom_sdc(self,sdc_file):
        funccom_lines = '''
############################################################
# setting DFT related
############################################################
#func_stop_scanclk

#############################################################
## setting memory related
#############################################################
#func_mem_misc $SDCVAR(FL_STAGE)

#############################################################
## setting Common Exception
#############################################################
# func_syncell_false
# func_false_misc; # disable clk gating, stop clk

#############################################################
## setting ideal network
#############################################################
# set ideal_ports ""
# set ideal_pins ""
# set_ideal_network $ideal_ports $ideal_pins

#############################################################
## setting DRC related
#############################################################
# set buffcell [list ]
# set_drc_syn $buffcell

#############################################################
## setting clock uncertainty
#############################################################

'''
        self.save_text(funccom_lines,sdc_file)


    def write_tune_sdc(self,sdc_file):
        sdir = f'{self._sdcdir}/outputs'
        
        sfiles = os.listdir(sdir)
        tfile = [x for x in sfiles if x.endswith('_tune.sdc')]

        tune_lines = ''
        itune_lines = ''
        uline = []

#         if tfile:
#             nfile = tfile[0].replace('_tune.sdc','')
#             tifile = f'{sdir}/intg/{nfile}_tune_intg.sdc'
#             uifile = f'{sdir}/{tfile[0]}'
#
#             sdc_info(f'Check tune sdc for intg part and regenerate tune.sdc and tune_intg.sdc.')
#             olines = self.read_text(uifile)
#             for line in olines:
#                 if '## Tune Section II' in line:
#                     break
#                 else:
#                     uline.append(line)
#             uline = ''.join(uline[:-1])
#             sline = f'''
# #########################################################################################
# ## Tune Section II: Subblk Tune Constraints Integraion
# ## This Parts with Flow handled
# ## @1: If user modify it, must rerun sdcgen flow
# #########################################################################################
# '''
#             # sub harden blk
#             blkf = 'tune'
#             tlines = self._hier_tree.set_subblk_intg(self._mdname,blkf)
#
#             os.system(f'rm -f {uifile}')
#             tune_lines = uline + sline + tlines
#             itune_lines = uline + sline
#
#             if tune_lines:
#                 self.save_text(tune_lines,sdc_file)
#             if itune_lines:
#                 self.save_text(itune_lines,tifile)
#
#             #os.system(f'cp -rf {sdir}/{tfile[0]} {sdir}/intg/{nfile}_tune_intg.sdc')
#         else:
#             sdc_error(f'Missing tune sdc and check whether delete it in mannual.')

        if not tfile:
            tune_lines = f'''
#########################################################################################
## Tune Section I: User Customized Constraints
## Must follow integration rule which covers the folllowing variables:
## SDCVAR(IS_FLAT,${{alias}}) / SDCVAR(LIB,${{alias}}) for block level
#########################################################################################
'''
            self.save_text(tune_lines, sdc_file)


    def write_subblk_clkgrp(self, alias, mflg=0):
        if not mflg:
            clkgrp_lines = f'\nforeach clk [array names SDCVAR -glob \"NAME,{alias},*\"] {{'
            clkgrp_lines += f'\n\tlappend clk_all $SDCVAR($clk)'
            clkgrp_lines += f'\n\tforeach grp_name [array name CLOCK_GROUP_NAME] {{'
            clkgrp_lines += f'\n\t\tif {{[lsearch $CLOCK_GROUP_NAME($grp_name) $SDCVAR($clk)] >= 0}} {{'
            clkgrp_lines += f'\n\t\t\tlappend clk_has_grp $SDCVAR($clk)'
            clkgrp_lines += f'\n\t\t}}'
            clkgrp_lines += f'\n\t}}'
            clkgrp_lines += f'\n}}\n'
            return clkgrp_lines
        else:
            clkgrp_lines = f'\nforeach {{ALIAS_VAL HIER_VAL}} ${alias}_PAT {{'
            clkgrp_lines += f'\n\tset {alias} $ALIAS_VAL'
            clkgrp_lines += f'\n\tforeach clk [array names SDCVAR -glob \"NAME,${{{alias}}},*\"] {{'
            clkgrp_lines += f'\n\t\tlappend clk_all $SDCVAR($clk)'
            clkgrp_lines += f'\n\t\tforeach grp_name [array name CLOCK_GROUP_NAME] {{'
            clkgrp_lines += f'\n\t\t\tif {{[lsearch $CLOCK_GROUP_NAME($grp_name) $SDCVAR($clk)] >= 0}} {{'
            clkgrp_lines += f'\n\t\t\t\tlappend clk_has_grp $SDCVAR($clk)'
            clkgrp_lines += f'\n\t\t\t}}'
            clkgrp_lines += f'\n\t\t}}'
            clkgrp_lines += f'\n\t}}\n'
            clkgrp_lines += f'\n}}\n'
            return clkgrp_lines            

    def write_clkgrp_sdc(self,sdc_file):
        mdname = self._mdname
        flt = self._flt
        lvl = self._lvl
        alias = self._alias

        clkgrp_lines = f'\nset clk_grp_f [open {mdname}_clk_group.sdc "w"]\n'
        clkgrp_lines += '\nforeach grp_name [array name CLOCK_GROUP_NAME] {'
        clkgrp_lines += '\n\tset_clock_groups -asynchronous -name $grp_name -group \"$CLOCK_GROUP_NAME($grp_name)\"'
        clkgrp_lines += '\n\tputs $clk_grp_f "set_clock_groups -asynchronous -group \\"$CLOCK_GROUP_NAME($grp_name)\\" "'
        clkgrp_lines += '\n\tputs "SDC_INFO: Setting clock group name: $grp_name"'
        clkgrp_lines += '\n}\n'

        clkgrp_lines += '\nset clk_all ""'
        clkgrp_lines += '\nset clk_has_grp ""'
        clkgrp_lines += f'\nforeach clk [array names SDCVAR -glob \"NAME,{alias},*\"] {{'
        clkgrp_lines += '\n\tlappend clk_all $SDCVAR($clk)'
        clkgrp_lines += '\n\tforeach grp_name [array name CLOCK_GROUP_NAME] {'
        clkgrp_lines += '\n\t\tif {[lsearch $CLOCK_GROUP_NAME($grp_name) $SDCVAR($clk)] >= 0} {'
        clkgrp_lines += '\n\t\t\tlappend clk_has_grp $SDCVAR($clk)'
        clkgrp_lines += '\n\t\t}'
        clkgrp_lines += '\n\t}'
        clkgrp_lines += '\n}\n'

        nomblks = []
        mlblks = []
        slblks = []
        hdblks = []
        hdblksg = self._hier_tree.get_hierlvlblks(mdname,outtype='hd')
        hdblks = [x for x in hdblksg if not x is mdname]

        for bk in hdblks:
            mlflg = None
            slflg = None
            noflg = None

            blk = self._hier_tree.get_block_by_name(bk)
            fblks = self._hier_tree.get_fblk(mdname,bk)
            for fbk in fblks:
                minfo = self._hier_tree.get_curmim_info(fbk,bk,'inst')
                if minfo == 'MLMIM':
                    mlflg = 1
                if minfo == 'SLMIM':
                    slflg = 1
                if minfo == 'NOMIM':
                    noflg = 1
            if  mlflg:
                mlblks.append(bk)
                clkgrp_lines += self.write_subblk_clkgrp(blk.alias,1)
            if slflg:
                slblks.append(bk)
                clkgrp_lines += self.write_subblk_clkgrp(blk.alias,1)
            if noflg:
                nomblks.append(bk)
                clkgrp_lines += self.write_subblk_clkgrp(blk.alias,0)

        # for nsubblk in self.subblk_hinsts:
        #     if not nsubblk.ref.name in self.multi_inst_list:
        #         nsubblk_name = nsubblk.ref.alias
        #         clkgrp_lines += self.write_subblk_clkgrp(self.get_subblk_alias(nsubblk_name))
            
        # if self.multi_inst_dict:
        #     for key in self.multi_inst_dict.keys():
        #         clkgrp_lines += self.write_subblk_clkgrp(key,multi_flag=1)

        clkgrp_lines += '\nforeach clk $clk_all {'
        clkgrp_lines += '\n\tset i "0"'
        clkgrp_lines += '\n\tforeach grp_name [array name CLOCK_GROUP_NAME] {'
        clkgrp_lines += '\n\t\tif {[lsearch $CLOCK_GROUP_NAME($grp_name) $clk] >= 0} {'
        clkgrp_lines += '\n\t\t\tset i [expr $i+1]'
        clkgrp_lines += '\n\t\t}'
        clkgrp_lines += '\n\t}'
        clkgrp_lines += '\n\tif {$i>1} {'
        clkgrp_lines += '\n\t\tputs "SDC_ERROR: $clk has been set in $i grp array"'
        clkgrp_lines += '\n\t\tputs $clk_grp_f "# clock: $clk has been set $i grp array"'
        clkgrp_lines += '\n\t} elseif {$i<1} {'
        clkgrp_lines += '\n\t\tputs \"SDC_ERROR: $clk is not set grp array\"'
        clkgrp_lines += '\n\t\tputs $clk_grp_f "# clock: $clk is not set grp array"'
        clkgrp_lines += '\n\t}'
        clkgrp_lines += '\n}'

        clkgrp_lines += '\nclose $clk_grp_f'

        self.save_text(clkgrp_lines,sdc_file)


    def write_digmac_sdc(self,alias,ipfls,ip_clkinfo,kw2,kw3,otpsdir,fintg='False'):
        subblk_lines = ''
        if ipfls:
            for kw,fl in ipfls.items():
                ks = kw.replace('#','').split('_')
                if '#' in kw:
                    n = ks[1][-1]
                else:
                    n = 0

                if fintg:
                    fi = '_intg'
                    ft = 'intg/'
                else:
                    fi = ''
                    ft = '/'
                        
                if os.path.exists(fl):
                    fn = fl.split('/')[-1].replace('.sdc','') 
                    #als = kw.replace('#','')
                    # als : 'MCUJPEG_MCUCRG10_CRG_CRG'                                         
                    ip_lines = self.read_text(fl)
                    nip_lines = self.align_crgip_vars(kw,ip_lines,ip_clkinfo,'IP')
                    fip1 = f'{otpsdir}/{fn}_{kw2}{n}.sdc'
                    fip2 = f'{otpsdir}/intg/{fn}_{kw2}{n}_intg.sdc'
                    self.save_text(nip_lines,fip1)
                    self.save_text(nip_lines,fip2)

                    subblk_lines += f'''
#############################################################
## Third_Party {kw3} IP
#############################################################
if {{[file exists $SDCVAR(SDC_DIR,${{{alias}}}){ft}{fn}_{kw2}{n}{fi}.sdc]}} {{
    puts "SDC_INFO: Sourcing $SDCVAR(SDC_DIR,${{{alias}}}){ft}{fn}_{kw2}{n}{fi}.sdc."
    source -echo -verbose $SDCVAR(SDC_DIR,${{{alias}}}){ft}{fn}_{kw2}{n}{fi}.sdc
}} else {{
    puts "SDC_WARN: Missing $SDCVAR(SDC_DIR,${{{alias}}}){ft}{fn}_{kw2}{n}{fi}.sdc. Please check it."
}}
'''
                elif re.search(r'\[file\s+exi\w+\s+(\S+)\]',fl):
                        nusr = re.findall(r'\[file\s+exi\w+\s+(\S+)\]',fl)
                        fn = nusr[0].split('/')[-1].replace('.sdc','')
                        ip_lines = self.read_text(fl)
                        nip_lines = self.align_crgip_vars(kw,ip_lines,ip_clkinfo,'IP')
                        fip1 = f'{otpsdir}/{fn}_{kw2}{n}.sdc'
                        fip2 = f'{otpsdir}/intg/{fn}_{kw2}{n}_intg.sdc'
                        self.save_text(nip_lines,fip1)
                        self.save_text(nip_lines,fip2)

                        fl = re.sub(f'{nusr[0]}', f'$SDCVAR(SDC_DIR,${{{alias}}}){ft}{fn}_{kw2}{n}{fi}.sdc', fl) 
                        subblk_lines += fl

                else:
                    sdc_warn(f'Can not find {kw} sdc file.')
        
        return subblk_lines

    def align_crgip_vars(self,kw,crg_lines,crg_clkinfo,sw):
        ncrg_lines = []
        i = 0
        svarg = ''
        evarg = ''
        sflg = False
        for n,cline in enumerate(crg_lines):
            if '## End {sw} Header' in cline:
                ncrg_lines.append(cline)
                i = n
                sflg = True
            elif sflg and n == i+2:
                #print('XXXXXnumber++++++++++++++++++:',n)
                svarg += f'''
###############################################################
## Start {sw}IN Variables Alignment for Clock Integration
###############################################################
'''
                for cky,cvl in crg_clkinfo.items():
                    #print('kw,cky+++++++++++++++:',kw,cky)
                    if kw == cky:
                        for cl in cvl:
                            mclk = '_'.join(cl[0].split('_')[1:])
                            src = '_'.join(cl[1].split('_')[1:])
                            grp = '_'.join(cl[2].split('_')[1:])
                            svarg += f'''
set {mclk}  ${{{cl[0]}}}
set {src}   ${{{cl[1]}}}
set {grp}   ${{{cl[2]}}}
'''
                svarg += f'''
###############################################################
## End {sw}IN Variables Alignment for Clock Integration
###############################################################
'''
                ncrg_lines.append(svarg)
            elif n == len(crg_lines)-1:
                #print('number++++++++++++++++++:',n,len(crg_lines))
                ncrg_lines.append(cline)
                evarg += f'''
###############################################################
## Start {sw}OUT Variables Alignment for Clock Integration
###############################################################
'''
                for cky,cvl in crg_clkinfo.items():
                    #print('kw,cky+++++++++++++++:',kw,cky)
                    if kw == cky:
                        for cl in cvl:
                            mclk = '_'.join(cl[0].split('_')[1:])
                            src = '_'.join(cl[1].split('_')[1:])
                            grp = '_'.join(cl[2].split('_')[1:])
                            evarg += f'''
set  {cl[0]}  ${{{mclk}}}
set  {cl[1]}  ${{{src}}}
set  {cl[2]}  ${{{grp}}}
'''
                evarg += f'''
###############################################################
## End {sw}OUT Variables Alignment for Clock Integration
###############################################################
'''
                ncrg_lines.append(evarg)                            
            else:
                ncrg_lines.append(cline)

        return ''.join(ncrg_lines)


    def com_sdcout1(self,alias,sec,fname):
        
        sdc_lines = f'''
############################################################
## Section {sec}
############################################################
if {{[file exists $SDCVAR(SDC_DIR,${{{alias}}}){fname}.sdc]}} {{
    puts "SDC_INFO: Sourcing $SDCVAR(SDC_DIR,${{{alias}}}){fname}.sdc"
    source -echo -verbose  $SDCVAR(SDC_DIR,${{{alias}}}){fname}.sdc
}} else {{
    puts "SDC_WARN: Missing $SDCVAR(SDC_DIR,${{{alias}}}){fname}.sdc. Please check it."
}}

'''
        return sdc_lines


    def read_json(self,file_path):
        sblk_clk_list = {}
        if os.path.exists(file_path):
            with open(file_path,'r') as fw:
                content = fw.read()
                sblk_clk_list = json.loads(content)

        #print('sblk_clk_list:',sblk_clk_list)
        return sblk_clk_list


    def write_json(self,filepath):
        os.makedirs(dirname(filepath), exist_ok=True)
        jsonstr = json.dumps(self._data, indent=4)
        with open(filepath,'w') as fw:
            print(jsonstr, file=fw) 

    def save_text(self,context,file):
        with open(file, 'w') as fw:
            fw.write(context)

    def save_workbook(self,output):
        self._wb.save(output)

    def read_text(self, file):
        if not os.path.exists(file):
            raise FileExistsError(f'{file} does not exists')
            # sdc_error(f'{file} not exist. Please check it.')
            # exit(1)
        else:
            txt_list = []
            with open(file,'r') as fh:
                for line in fh.readlines():
                    txt_list.append(line)
        
            return txt_list



