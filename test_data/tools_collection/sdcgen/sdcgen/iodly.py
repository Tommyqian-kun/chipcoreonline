
import sys
import time
import os
import re

from os.path import dirname, abspath, basename

import  openpyxl

from .basesdc import *
from com.base import *
from .clkdef import *


#     def __repr__(self):
#         return f'<{self.__class__.__name__} PortPin={self.PortPin} Direction={self.Direction}>' 
          
class IODlySheet(BaseSheet):
    def __init__(self,*args):
        super().__init__(*args)  
        self._iodlydata = {}  
        self._hiertree = self._sdcdg._hier_tree
        #self._valdata = self._sdcdg._sheets['VarDef']._valdata
        self._vardata = self.get_vardef_value(self._sdcdg._wb['VarDef'])
        
        #self._sdcdir = self._sdcdg._sdcdir
        self._mdname = self._sdcdg._mdname

        #self._clkdef = None
        self._vfdata = self._sdcdg._vfile_data
        self._hiertree = self._sdcdg._hier_tree   

    def update_sheet(self):
        '''
        # only during -dg option
        # addition of module name value from vfile
        # addition of block hier tree expanded table from hier yaml
        '''
        sheet = self.get_sheet()

        #hiertree = self._sdcdg._hier_tree

        # find TMIODLY table
        start_rowg = self.find_sheet(sheet, 'TMIODLY')

        vdata = self._sdcdg._vfile_data
        vlist = self._sdcdg._vfile_list
        
        n = 0
        for kwd in vlist:
            if 'module_name' not in kwd and 'RelClock' not in kwd:
                if re.search(r'IDEAL|CASEXP|FPEXP|ANA|TCLK',vdata[kwd][2]):
                    n += 1
        sheet.insert_rows(start_rowg + 9, len(vlist) - n)
        self.cell_style2(sheet, [start_rowg + 9, 1], [start_rowg + 9 + len(vlist) - n, 8])

        varlist = ['input','output']
        self.add_dropdown(sheet, '"' + ','.join(varlist) + '"', [start_rowg + 1,2], [start_rowg + 9 + len(vlist) - n + 1,2])
        
        #indir = self._sdcdir + '/inputs'
        clkdef = self._sdcdg._sheets['ClkDef']
        #varlist = clkdef.get_clkinfo_from_crgip('IO','0')
        clknm = clkdef.get_clkinfo_from_crgip('0','IO')
        alsck = []
        varlist = []
        varlistg = []
        for x in clknm:
            for k,v in x.items():
                cials = k.split(' ')[0].split('_')[1]
                nalsck = [f'{cials} {p}' for p in v]
                alsck.extend(nalsck)
        varlist.extend(alsck)
        for kwd in vlist:
            if 'RelClock' in kwd:
                varlistg.append(vdata[kwd])
        if varlistg:
            varlistg = list(set(varlistg))
            varlist.extend(varlistg)
        self.add_dropdown(sheet, '"' + ','.join(varlist) + '"', [start_rowg + 1,3], [start_rowg + 9 + len(vlist) - n + 1,3])
             
        varlist = ['Y']
        self.add_dropdown(sheet, '"' + ','.join(varlist) + '"', [start_rowg + 1,4], [start_rowg + 9 + len(vlist) - n + 1,4])
        
        maxlst = []
        minlst = []
        for kw in self._vardata.keys():
            if re.match(r'IO_DLY_MAX?',str(kw)):
                maxlst.append(kw)
            if re.match(r'IO_DLY_MIN?',str(kw)):
                minlst.append(kw)
        varlist = maxlst + ['70%','60%','50%','40%','30%','0','-10%']
        self.add_dropdown(sheet, '"' + ','.join(varlist) + '"', [start_rowg + 1,5], [start_rowg + 9 + len(vlist) - n + 1,5])
        varlist = minlst + ['20%','10%','0','-10%','-20%']
        self.add_dropdown(sheet, '"' + ','.join(varlist) + '"', [start_rowg + 1,6], [start_rowg + 9 + len(vlist) - n + 1,6])

        # vol
        varlist = []
        for blknm in self._hiertree.get_curblks(self._mdname):
            blk = self._hiertree.get_block_by_name(blknm)
            pwrg = blk.prime_pwr.split(' ')[0].strip()
            if f'{pwrg}' not in varlist:
                varlist.append(pwrg)
        varlistg = [f'{x}=TT0P750V' for x in varlist]
        self.add_dropdown(sheet, '"' + ','.join(varlistg) + '"', [start_rowg + 1,7], [start_rowg + 9 + len(vlist) - n + 1,7])

        i = 0
        grp_flag  = 0
        ioclknm = ''
        for kwd in vlist:
            if i < len(vlist) + 1:
                
                if 'module_name' not in kwd and 'RelClock' not in kwd:
                    if not re.search(r'IDEAL|CASEXP|FPEXP|ANA|TCLK|DFT',vdata[kwd][2]):
                        i += 1
                        if vdata[kwd][1] != '1':
                            sheet.cell(start_rowg + i,1).value = kwd + vdata[kwd][1]
                        else:
                            sheet.cell(start_rowg + i,1).value = kwd
                        sheet.cell(start_rowg + i,2).value = vdata[kwd][0]
                        sheet.cell(start_rowg + i,5).value = 'IO_DLY_MAXA'
                        sheet.cell(start_rowg + i,6).value = 'IO_DLY_MINA'

                        if grp_flag:
                            sheet.cell(start_rowg + i,3).value = ioclknm

                if 'RelClock' in kwd:
                    if not grp_flag:
                        grp_flag  = 1
                        ioclknm = vdata[kwd]
                    else:
                        grp_flag  = 0
                        ioclknm = ''

######################################################
    def read_data(self):
        sheet = self.get_sheet()
        self._iodlydata = self.get_table_contxt(sheet)
        #print(self._iodlydata)
        
    def check_sheet(self):
        pass
    
    def change_sheet(self):
        pass

    def dump_json(self,json_file):
        self._data = self._iodlydata
        self.write_json(json_file)

######################################################
    def write_sdc(self,sdc_dir,prousr=False):
        #sheet = self.get_sheet()

        mdname = self._sdcdg._vfile_data['module_name']
        alias = self._sdcdg._hier_tree._blocks[mdname].alias
        # hdlvl = self._sdcdg._hier_tree._blocks[mdname].hdlevel
        # pwr = self._sdcdg._hier_tree._blocks[mdname].prime_pwr

        sdc_file = sdc_dir +  f'{alias.lower()}_iodly.sdc'
        self.write_iodly(mdname,alias,sdc_file)

    def write_iodly(self,mdname,alias,sdc_file):
        clkdef = self._sdcdg._sheets['ClkDef']
        #cinmals = clkdef.get_crgip_clknm_alias(mdname)
        # clkdata = clkdef.get_table_contxt(self._sdcdg._wb['ClkDef'])
        # clkdef.get_clkdata_by_clkname(clkdata)
        

        #iorows = self._iodlydata.keys()
        iodly_lines = ''
        
        for key,row in self._iodlydata.items():
            #PortNm	Direction	ClkNm	ClkFall	DlyMax	DlyMin	Vol	Comment

            nrow = key.split('_')[1]
            # direction
            if row['Direction'] == 'input':
                iocmd = 'set_input_delay'

            if row['Direction'] == 'output':
                iocmd = 'set_output_delay'

            # clknm
            if row['ClkNm']:
                nclknm = row['ClkNm']
            else:
                nclknm = ''
            if nclknm:                        
                #als,cknm = clkdef.get_alval_clknm(mdname,alias,nclknm)                
                cknm = nclknm
                if nclknm in clkdef._clknmlst:
                    als = alias
                else:
                    #avl = nclknm.split('_')[0]
                    sp = nclknm.split('_')
                    if 'NAME_' in nclknm:
                        avl = sp[1]
                    else:
                        avl = sp[0]
                    als = clkdef.get_als_var(self._mdname,avl)
                clknm = f'$SDCVAR(NAME,${{{als}}},{cknm})'
            
            # clkfall
            if row['ClkFall'] == 'Y':
                clkfall = '-clock_fall'
            else:
                clkfall = ''

            # dlymax            
            if not row['DlyMax']:
                ndlymax = ''
                sdc_error(f'Missing IO dlymax value of {key}')
            else:
                ndlymax = row['DlyMax']
            if ndlymax:
                dlymax = self.get_dlymaxmin(cknm,als,ndlymax)
            else:
                 sdc_warn(f'{nrow}: {portnm} missing DlyMax value')

            # dlymin
            if not row['DlyMin']:
                ndlymin = ''
                sdc_error(f'Missing IO dlymin value of {key}')
            else:
                ndlymin = row['DlyMin']          
            if ndlymin:
                dlymin = self.get_dlymaxmin(cknm,als,ndlymin)
            else:
                sdc_warn(f'{nrow}: {portnm} missing DlyMin value')

            if row['PortNm']:
                portnm = row['PortNm']
            else:
                portnm = ''
                sdc_error(f'Missing portnm value of {key}')

            if row['Vol']:
                vol = row['Vol']
                # VDD_CORE=TT0P750V
                vlt = vol.split('=')[0].strip()
                val = vol.split('=')[1].strip()

                iodly_lines += f'''
# @IODly_{nrow}: {portnm}
if {{$SDCVAR(DCDC_VL,${{{vlt}}}) == "{val}"}} {{
'''
            else:
                iodly_lines += f'''
# @IODly_{nrow}: {portnm}
'''          
            iodly_lines = iodly_lines.rstrip()
                
            if re.search(r'\w+\[\d+:\d+\]',portnm):
                sig,stn,edn,num = self.get_sig_num(portnm)
                iodly_lines += f'''
    for {{set i {stn}}} {{i <= {edn}}} {{incr i}} {{
        {iocmd} -add_delay -clock [get_clocks {clknm}] {clkfall} -max {dlymax} [get_ports {sig}[i]]
        {iocmd} -add_delay -clock [get_clocks {clknm}] {clkfall} -min {dlymin} [get_ports {sig}[i]]
    }}
'''

            #if re.search(r'\w+[\d+]|\w+',portnm):
            else:
                iodly_lines += f'''
    {iocmd} -add_delay -clock [get_clocks {clknm}] {clkfall} -max {dlymax} [get_ports {portnm}]
    {iocmd} -add_delay -clock [get_clocks {clknm}] {clkfall} -min {dlymin} [get_ports {portnm}]
'''


            if row['Vol']:
                iodly_lines += f'''
}}
'''

        self.save_text(iodly_lines,sdc_file)

    def get_dlymaxmin(self,clknm,alias,ndly):

        if re.search(r'IO_DLY_M',ndly):
            #self.get_dlymaxmin(clknm,alias,str(self._vardata[ndly]))
            fdly = str(self._vardata[ndly])
        else:
            fdly = ndly
        
        if re.match(r'(-)*0\.\d+#$',fdly):
            return fdly.replace('#','')
        
        elif re.match(r'(-)*\[expr \d+\w+\]#',fdly):
            return fdly.replace('#','')   
                   
        elif re.match(r'(-)*0\.0$|0$',fdly):
            return f'0.0'
        
        elif re.match(r'(-)*0\.\d+$',fdly):
            return f'[expr $SDCVAR(CYCLE,${{{alias}}},{clknm}) * {fdly}]'
                          
        elif re.match(r'(-)*\d+%$',fdly):
            xdly = int(fdly.split('%')[0])/100
            return f'[expr $SDCVAR(CYCLE,${{{alias}}},{clknm}) * {xdly}]'
     
        else:
            return fdly.replace('#','')
    
    def get_sig_num(self,portnm):
        sig = ''
        stn = ''
        edn = ''
        num = ''
        iopat = re.findall(r'(\w+)(\[\d+:\d+\])',portnm)
        if iopat:
            sig = iopat[0][0]
            stn,edn,num = self.cal_num(iopat[0][1])

        iopat = re.findall(r'(\w+)(\[\d+:\d+\])(\[\d+:\d+\])',portnm)
        if iopat:
            sig = iopat[0][0]
            stn1,edn1,num1 = self.cal_num(iopat[0][1])
            stn2,edn2,num2 = self.cal_num(iopat[0][2])
            num = str(int(num1) * int(num2) -1)
            # how to set stn/edn ??
            stn = '0'
            edn = num

        #input wire [3:0][5: 0] dat_up,  ====> dat_up[3:0][5:0]
        #input wire [3:0] datg_up[5: 0][2:0][7:0], ====> datg_up[5:0][2:0][7:0][3:0]

        iopat = re.findall(r'(\w+)(\[\d+:\d+\])(\[\d+:\d+\])(\[\d+:\d+\])',portnm)
        if iopat:
            sig = iopat[0][0]
            stn1,edn1,num1 = self.cal_num(iopat[0][1])
            stn2,edn2,num2 = self.cal_num(iopat[0][2])
            stn3,edn3,num3 = self.cal_num(iopat[0][3])
            num = str(int(num1) * int(num2) * int(num3) -1)
            # how to set stn/edn ??
            stn = '0'
            edn = num        

        iopat = re.findall(r'(\w+)(\[\d+:\d+\])(\[\d+:\d+\])(\[\d+:\d+\])(\[\d+:\d+\])',portnm)
        if iopat:
            sig = iopat[0][0]
            stn1,edn1,num1 = self.cal_num(iopat[0][1])
            stn2,edn2,num2 = self.cal_num(iopat[0][2])
            stn3,edn3,num3 = self.cal_num(iopat[0][3])
            stn4,edn4,num4 = self.cal_num(iopat[0][4])
            num = str(int(num1) * int(num2) * int(num3) * int(num4) -1)
            # how to set stn/edn ??
            stn = '0'
            edn = num
        
        return sig,stn,edn,num

    def cal_num(self,npat):
        #[n:m]
        stn = npat.split(':')[1].replace(']','')
        edn = npat.split(':')[0].replace('[','')
        num = str(int(edn) - int(stn) + 1)
        return stn,edn,num
        # sig = portnm.split('[')[0]
        # stn = portnm.split('[')[1].split(':')[1].replace(']','')
        # edn = portnm.split('[')[1].split(':')[0]
        # print(sig,stn,edn)
        # num = int(edn) - int(stn) + 1        









        