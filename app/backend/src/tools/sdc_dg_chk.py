#!/usr/bin/env python3


import os,sys
import yaml
import re
import json

from os.path import dirname, abspath, basename
import time

import shutil
import tkinter
import yaml

import argparse
from glob import glob

import openpyxl
from openpyxl import worksheet 
from pprint import pprint 
import pandas as pd
from openpyxl.utils import get_column_letter 

import tkinter as tk

from openpyxl.styles import Border, Side, PatternFill, Font, Alignment 
from openpyxl.worksheet.datavalidation import DataValidation
from collections import Counter

# Define some attributes and various methods for every block in hier tree
class BaseBlock():
    def __init__(self,name):
        self._name = name
        self._alias = ''
        self._hdlevel = 'blk'
        self._prime_pwr = ''
        self._constr_dir = ''
        self._insts = []
        self._mac_insts = []
        self._dig_insts = []
        self._cust_insts = {}
        self._proj = False

    def __repr__(self):
        return '<%s name=%s alias=%s>' % (self.__class__.__name__, self._name, self._alias)

    @property 
    def hdlevel(self):
        return self._hdlevel

    @hdlevel.setter
    def hdlevel(self, level):
        supported_lvs = ('top', 'sys', 'blk', 'soft', 'lib', 'crg', 'pll')
        if level not in supported_lvs:
            sdc_error(f'Unsupported HDLEVEL {level} of block {self._name}, should be one of {supported_lvs}')
            return
        else:
            if level == 'sys' or level == 'top':
                self._hdlevel = 'sys'
            else:
                self._hdlevel = level

    @property
    def lvl_flat(self):
        if self._hdlevel == 'sys':
            return 'IS_CHIP'
        else:
            return 'IS_FLAT'
        
    @property
    def alias(self):
        return self._alias

    @alias.setter
    def alias(self, alias):
        self._alias = alias

    @property
    def prime_pwr(self):
        return self._prime_pwr

    @prime_pwr.setter
    def prime_pwr(self, pwr):
        self._prime_pwr = pwr

    @property
    def insts(self):
        return self._insts

    @insts.setter
    def insts(self, insts):
        self._insts = insts

    @property
    def mac_insts(self):
        return self._mac_insts

    @mac_insts.setter
    def mac_insts(self, mac_insts):
        self._mac_insts = mac_insts

    @property
    def dig_insts(self):
        return self._dig_insts

    @dig_insts.setter
    def dig_insts(self, dig_insts):
        self._dig_insts = dig_insts

    @property
    def constr_dir(self):
        return self._constr_dir
    
    @constr_dir.setter
    def constr_dir(self, consdir):
        self._constr_dir = consdir

    @property
    def proj(self):
        return self._proj

    @proj.setter
    def proj(self, proj):
        self._proj = proj

    def get_curcust_by_name(self, inst_type,flg=''):
        _hier = []
        _ref = []
        #_lvl = []
        _alias = []
        _pwr = []
        _usersdc = []
        if inst_type == 'insts' and self.insts:
            for i in range(0,len(self.insts)):
                finst = self.insts[i].split(',')
                if len(finst) == 3:
                    _hier.append(finst[0].strip())
                    _ref.append(finst[1].strip())
                    _alias.append(None)
                    _pwr.append(finst[2].strip())
                    _usersdc.append(None)
                if len(finst) == 4:
                    _hier.append(finst[0].strip())
                    _ref.append(finst[1].strip())
                    if not flg:
                        _alias.append(finst[2].strip().replace('#',''))
                    else:
                        _alias.append(finst[2].strip())
                    _pwr.append(finst[3].strip())
                    _usersdc.append(None)
            self._cust_insts['insthier'] = _hier
            self._cust_insts['instref'] = _ref
            self._cust_insts['instalias'] = _alias
            self._cust_insts['instpwr'] = _pwr
            self._cust_insts['instuser'] = _usersdc
        
        if inst_type == 'mac_insts' and self.mac_insts:
            for i in range(0,len(self.mac_insts)):
                if isinstance(self.mac_insts[i],str):
                    fmac = self.mac_insts[i].split(',')  
                    if len(fmac) == 3:
                        _hier.append(fmac[0].strip())
                        _ref.append(fmac[1].strip())
                        _alias.append(None)
                        _pwr.append(fmac[2].strip())
                        _usersdc.append(None)
                    if len(fmac) == 4:
                        _hier.append(fmac[0].strip())
                        _ref.append(fmac[1].strip())
                        if not flg:
                            _alias.append(fmac[2].strip().replace('#',''))
                        else:
                            _alias.append(fmac[2].strip())
                        _pwr.append(fmac[3].strip())
                        _usersdc.append(None)
                if isinstance(self.mac_insts[i],dict):
                    fmac = ''.join(self.mac_insts[i].keys()).split(',')
                    _hier.append(fmac[0].strip())
                    _ref.append(fmac[1].strip()) # + '_USR' )
                    _alias.append(None)
                    _pwr.append(fmac[2].strip())
                    _usersdc.append(''.join(self.mac_insts[i].values()))  
            self._cust_insts['machier'] = _hier
            self._cust_insts['macref'] = _ref
            self._cust_insts['macalias'] = _alias
            self._cust_insts['macpwr'] = _pwr
            self._cust_insts['macuser'] = _usersdc

        if inst_type == 'dig_insts' and self.dig_insts:
            for i in range(0,len(self.dig_insts)):
                if isinstance(self.dig_insts[i],str):
                    fdig = self.dig_insts[i].split(',')  
                    if len(fdig) == 3:
                        _hier.append(fdig[0].strip())
                        _ref.append(fdig[1].strip())
                        _alias.append(None)
                        _pwr.append(fdig[2].strip())
                        _usersdc.append(None)
                    if len(fdig) == 4:
                        _hier.append(fdig[0].strip())
                        _ref.append(fdig[1].strip())
                        if not flg:
                            _alias.append(fdig[2].strip().replace('#',''))
                        else:
                            _alias.append(fdig[2].strip())
                        _pwr.append(fdig[3].strip())
                        _usersdc.append(None)
                if isinstance(self.dig_insts[i],dict):
                    fdig = ''.join(self.dig_insts[i].keys()).split(',')
                    _hier.append(fdig[0].strip())
                    _ref.append(fdig[1].strip()) # + '_USR')
                    _alias.append(None)
                    _pwr.append(fdig[2].strip())
                    _usersdc.append(''.join(self.dig_insts[i].values()))
            self._cust_insts['dighier'] = _hier
            self._cust_insts['digref'] = _ref
            self._cust_insts['digalias'] = _alias
            self._cust_insts['digpwr'] = _pwr
            self._cust_insts['diguser'] = _usersdc       

        return self._cust_insts

    def get_curhd_by_name(self):
        #return self.name.split() + self._cust_insts['instref']
        self.get_curcust_by_name('insts')
        if 'instref' in self._cust_insts:
            return self._cust_insts['instref']

    def get_curmac_by_name(self,flg=''):
        self.get_curcust_by_name('mac_insts',flg)
        if 'macref' in self._cust_insts:
            return self._cust_insts['macref']

    def get_curdig_by_name(self,flg=''):
        self.get_curcust_by_name('dig_insts',flg)
        if 'digref' in self._cust_insts:
            return self._cust_insts['digref']
    
    def get_curuser_by_name(self, inst_type):
        self.get_curcust_by_name('mac_insts')
        self.get_curcust_by_name('dig_insts')
        if inst_type == 'mac_insts' and 'macuser' in self._cust_insts:           
            return self._cust_insts['macuser']       
        elif inst_type == 'dig_insts'and 'diguser' in self._cust_insts:
            return self._cust_insts['diguser']
        else:
            return None

    # ref
    def get_curhd_by_name(self):
        # return self.name.split() + self._cust_insts['instref']
        self.get_curcust_by_name('insts')
        if 'instref' in self._cust_insts:
            return self._cust_insts['instref']

    def get_curmac_by_name(self, flg=''):
        self.get_curcust_by_name('mac_insts', flg)
        if 'macref' in self._cust_insts:
            return self._cust_insts['macref']

    def get_curdig_by_name(self, flg=''):
        self.get_curcust_by_name('dig_insts', flg)
        if 'digref' in self._cust_insts:
            return self._cust_insts['digref']

    def get_curuser_by_name(self, inst_type):
        self.get_curcust_by_name('mac_insts')
        self.get_curcust_by_name('dig_insts')
        if inst_type == 'mac_insts' and 'macuser' in self._cust_insts:
            return self._cust_insts['macuser']
        elif inst_type == 'dig_insts' and 'diguser' in self._cust_insts:
            return self._cust_insts['diguser']
        else:
            return None

    # alias val
    def get_curhdval_by_name(self):
        # return self.name.split() + self._cust_insts['instalias']
        self.get_curcust_by_name('insts')
        if 'instalias' in self._cust_insts:
            return self._cust_insts['instalias']

    def get_curmacval_by_name(self, flg=''):
        self.get_curcust_by_name('mac_insts', flg)
        if 'macalias' in self._cust_insts:
            return self._cust_insts['macalias']

    def get_curdigval_by_name(self, flg=''):
        self.get_curcust_by_name('dig_insts', flg)
        if 'digalias' in self._cust_insts:
            return self._cust_insts['digalias']

    # pwr
    def get_curhdpwr_by_name(self):
        # return self.name.split() + self._cust_insts['instpwr']
        self.get_curcust_by_name('insts')
        if 'instpwr' in self._cust_insts:
            return self._cust_insts['instpwr']

    def get_curmacpwr_by_name(self, flg=''):
        self.get_curcust_by_name('mac_insts', flg)
        if 'macpwr' in self._cust_insts:
            return self._cust_insts['macpwr']

    def get_curdigpwr_by_name(self, flg=''):
        self.get_curcust_by_name('dig_insts', flg)
        if 'digpwr' in self._cust_insts:
            return self._cust_insts['digpwr']

class HierPwrTree():
    def __init__(self,yaml_file):
        self.yaml_file = yaml_file
        self._blocks = {}
        self._primepwr = {}
        self._yaml_data = {}
        self._hierdata = {}
        self._pwrdata = {}
        #self._blktrees = {}
        self.build_hier_tree(yaml_file)
        

    def build_hier_tree(self, yaml_file):

        # get yaml_data
        yaml_data = {}
        if not os.path.exists(yaml_file):
            raise FileExistsError(f'{yaml_file} does not exists')
        with open(yaml_file, 'r') as fh:
            yaml_data = yaml.load(fh, yaml.FullLoader)

        if 'hier' not in yaml_data:
            print('Missing hier keyword in yaml file.')
            upf_fatal(f'Must include keyword <hier>')
        if 'pwr' not in yaml_data:
            print('Missing pwr keyword in yaml file.')
            upf_fatal(f'Must include keyword <pwr>')

        # get '_primepwr'
        for pwr_name in yaml_data['pwr'].keys():
            if yaml_data['pwr'][pwr_name]:
                self._primepwr[pwr_name] = yaml_data['pwr'][pwr_name]   

        for blk_name in yaml_data['hier'].keys():

            self._blocks[blk_name] = BaseBlock(blk_name)

            if 'alias' in yaml_data['hier'][blk_name] and yaml_data['hier'][blk_name]['alias']:
                self._blocks[blk_name].alias = yaml_data['hier'][blk_name]['alias']
            else:
                self._blocks[blk_name].alias = None

            if 'hdlevel' in yaml_data['hier'][blk_name] and yaml_data['hier'][blk_name]['hdlevel']:
                self._blocks[blk_name].hdlevel = yaml_data['hier'][blk_name]['hdlevel']
            else:
                self._blocks[blk_name].hdlevel = None            
            
            if 'prime_pwr' in yaml_data['hier'][blk_name] and yaml_data['hier'][blk_name]['prime_pwr'] in self._primepwr and yaml_data['hier'][blk_name]['prime_pwr']:
                self._blocks[blk_name].prime_pwr = yaml_data['hier'][blk_name]['prime_pwr'] + ' ' + self._primepwr[yaml_data['hier'][blk_name]['prime_pwr']]
            else:
                self._blocks[blk_name].prime_pwr = None 

            if 'constr_dir' in yaml_data['hier'][blk_name] and yaml_data['hier'][blk_name]['constr_dir']:
                self._blocks[blk_name].constr_dir = yaml_data['hier'][blk_name]['constr_dir']
            else:
                self._blocks[blk_name].constr_dir = None

            if 'insts' in yaml_data['hier'][blk_name] and yaml_data['hier'][blk_name]['insts']:
                self._blocks[blk_name].insts = yaml_data['hier'][blk_name]['insts']
            else:
                self._blocks[blk_name].insts = None

            if 'mac_insts' in yaml_data['hier'][blk_name] and yaml_data['hier'][blk_name]['mac_insts']:
                self._blocks[blk_name].mac_insts = yaml_data['hier'][blk_name]['mac_insts']
            else:
                self._blocks[blk_name].mac_insts = None

            if 'dig_insts' in yaml_data['hier'][blk_name] and yaml_data['hier'][blk_name]['dig_insts']:
                self._blocks[blk_name].dig_insts = yaml_data['hier'][blk_name]['dig_insts']
            else:
                self._blocks[blk_name].dig_insts = None

        self._hierdata = yaml_data['hier'] 
        self._pwrdata = yaml_data['pwr']
        self._yaml_data = yaml_data

    def get_block_by_name(self, name) -> BaseBlock:
        if name in self._blocks:
            return self._blocks[name]
        else:
            return None    

    def get_curblks(self,name):
        curblks = []
        
        allblks = list(self._hierdata.keys())
        if name in allblks:
            curblk = self.get_block_by_name(name)

            if curblk.get_curhd_by_name():
                curblks = [x for x in curblk.get_curhd_by_name() if x is not None]
            if curblk.get_curmac_by_name():
                curblks += [x for x in curblk.get_curmac_by_name() if x is not None]
            if curblk.get_curdig_by_name():
                curblks += [x for x in curblk.get_curdig_by_name() if x is not None]  
        else:
            sdc_warn(f'{name} is not expanded in hier_pwr yaml file.')
        
        return  curblks
    
    def get_hiertrees(self, name, blktrees={}, valstyle=None) -> dict:

        curblks = self.get_curblks(name)
        #blktrees = {}
        if curblks:
            new_curblks = [elem.replace('_USR', '') if re.search(r'_USR$',elem) else elem for elem in curblks]
            blktrees[name] = new_curblks
        else:
            if valstyle:
                blktrees[name] = None   

        for blk in curblks:
            if re.search(r'_USR$',blk):
                blk = blk.replace('_USR','')
                sdc_warn(f'{blk} is not expanded in hier_pwr yaml file.')
            elif len(curblks) > 0:
                self.get_hiertrees(blk,blktrees)    
        
        return blktrees

    def get_hierblks(self, name) -> list:

        blktrees = self.get_hiertrees(name)
        result = []
        for key, value in blktrees.items():
            if key not in result:
                result.append(key)
            if isinstance(value, list):
                for element in value:
                    if element not in result:
                        result.append(element)
        return result        

    def get_hierblks_infos(self,name) -> dict:

        blks = self.get_hierblks(name)
        blksinfo = {}

        allblks = list(self._hierdata.keys())
        for val in blks:
            if val in allblks:
                blk = self.get_block_by_name(val)
                alias = blk.alias
                lvl = blk.hdlevel
                condir = blk.constr_dir
                pwr = blk.prime_pwr
                pwrg = pwr.split(' ')[0].strip()
                blksinfo[val] = val + f' {alias}' + f' {lvl}' + f' {pwrg}' + f' {condir}'
            else:
                blksinfo[val] = val + ' (NOT EXPEND)'

        return blksinfo 

    def get_hierblks_infos_chk(self, name) -> dict:
        """
        get hier block info
        :param name:
        :return:
        """
        blks = self.get_hierblks(name)
        if not blks:
            blks = [f'{name}']
        blksinfo = {}

        allblks = list(self._hierdata.keys())
        # print('______BLKDSF: ', blks, allblks)
        for sbk in blks:
            if sbk in allblks:
                blk = self.get_block_by_name(sbk)
                alias = blk.alias
                lvl = blk.hdlevel
                condir = blk.constr_dir
                pwr = blk.prime_pwr
                # nlb = blk.intg_nlb
                nlb = 'net'
                inst = blk.insts
                mac = blk.mac_insts
                dig = blk.dig_insts
                # pwrg = pwr.split(' ')[0].strip()
                blksinfo[sbk] = [f'{alias}', f'{lvl}', f'{pwr}', f'{nlb}', f'{inst}', f'{mac}', f'{dig}', f'{condir}']
            else:
                blksinfo[sbk] = ['NOT_EXPANDED']
                sdc_error(f'{sbk} NOT expanded in hier yaml.')

        return blksinfo

    def get_full_hierblks_info(self):
        allblks = list(self._hierdata.keys())
        blksinfo = {}

        for sbk in allblks:
            blk = self.get_block_by_name(sbk)
            alias = blk.alias
            lvl = blk.hdlevel
            condir = blk.constr_dir
            pwr = blk.prime_pwr
            # nlb = blk.intg_nlb
            nlb = 'net'
            inst = blk.insts
            mac = blk.mac_insts
            dig = blk.dig_insts
            # pwrg = pwr.split(' ')[0].strip()
            blksinfo[sbk] = [f'{alias}', f'{lvl}', f'{pwr}', f'{nlb}', f'{inst}', f'{mac}', f'{dig}', f'{condir}']

        return blksinfo

    def get_cursub_alias(self, name, outtype='hd', htype='cur') -> list:
        '''
        get different harden insts, mac insts, dig insts under current design
        outtype is hd/lib/soft/crg/pll
        '''
        allhierblks = []
        insts_als = []
        ips_als = []
        digs_als = []
        plls_als = []
        crgs_als = []

        if htype == 'flt':
            allhierblks = self.get_hierblks(name)
            allhierblks = [x for x in allhierblks if x != name]
        if htype == 'cur':
            allhierblks = self.get_curblks(name)
        if htype == 'fcur':
            allhierblks.append(name)
            allhierblks.extend(self.get_curblks(name))
        # alltreeblks = self.get_hiertrees(name)
        # allhierblks.append(name)
        if allhierblks:
            for curblk in allhierblks:
                # if curblk in allblks:
                blk = self.get_block_by_name(curblk)
                lvl = blk.hdlevel
                if lvl in ['blk', 'sys', 'top']:
                    insts_als.append(blk.alias)
                elif lvl in ['crg']:
                    crgs_als.append(blk.alias)
                elif lvl in ['lib', 'soft']:
                    ips_als.append(blk.alias)
                elif lvl in ['pll']:
                    plls_als.append(blk.alias)
                else:
                    sdc_error(f'hdlevel of {curblk} is wrong.')

        if outtype == 'hd':
            return insts_als
        if outtype == 'ip':
            return ips_als
        # if outtype == 'lib':
        #     return macs
        # if outtype == 'soft':
        #     return digs  
        # if outtype == 'pll':
        #     return plls
        if outtype == 'crg':
            return crgs_als
        if outtype == 'crgip':
            return crgs_als + ips_als

    def get_curlvlblks(self, name, outtype='hd', htype='cur') -> list:
        """
        get different harden insts, mac insts, dig insts under current design
        outtype is hd/lib/soft/crg/pll
        """
        allhierblks = []
        insts = []
        ips = []
        digs = []
        plls = []
        crgs = []

        if htype == 'cur':
            allhierblks.append(name)
            allhierblks.extend(self.get_curblks(name))
        if htype == 'slf':
            allhierblks.extend(self.get_curblks(name))
        # alltreeblks = self.get_hiertrees(name)
        if allhierblks:
            for curblk in allhierblks:
                # if curblk in allblks:
                blk = self.get_block_by_name(curblk)
                lvl = blk.hdlevel
                if lvl in ['blk', 'sys', 'top']:
                    insts.append(curblk)
                elif lvl in ['crg']:
                    crgs.append(curblk)
                elif lvl in ['lib', 'soft']:
                    ips.append(curblk)
                elif lvl in ['pll']:
                    plls.append(curblk)
                else:
                    sdc_error(f'hdlevel of {curblk} is wrong.')

        if outtype == 'hd':
            return insts
        if outtype == 'ip':
            return ips
        # if outtype == 'lib':
        #     return macs
        # if outtype == 'soft':
        #     return digs  
        # if outtype == 'pll':
        #     return plls
        if outtype == 'crg':
            return crgs
        if outtype == 'crgip':
            return ips + crgs


    def get_hierlvlblks(self, name, outtype='hd') -> list:
        '''
        get different harden insts, mac insts, dig insts under current design
        outtype is hd/lib/soft/crg/pll
        '''
        allhierblks = []
        allhierblks = self.get_hierblks(name)
        #alltreeblks = self.get_hiertrees(name)
        allhierblks.append(name)

        #allblks = list(self._hierdata.keys())

        insts = []
        macs = []
        digs = []
        plls = []
        crgs = []
        if allhierblks:
            for curblk in allhierblks:
                #if curblk in allblks:
                blk = self.get_block_by_name(curblk)
            
                lvl = blk.hdlevel
                if lvl in ['blk', 'sys', 'top']:
                    insts.append(curblk)
                if lvl in ['lib']:
                    macs.append(curblk)  
                if lvl in ['pll']:
                    plls.append(curblk) 
                if lvl in ['soft']:
                    digs.append(curblk)  
                if lvl in ['crg']:
                    crgs.append(curblk)
                # else:
                #     #print(f'XYZ_{curblk}')
                #     for ky,vl in alltreeblks.items():
                #         if curblk in vl:
                #             parent = ky
                #     blkg = self.get_block_by_name(parent)
                #     if curblk in blkg.get_curhd_by_name():
                #         insts.append(curblk)
                #         sdc_warn(f'{curblk} is harden block need expand it in hier_pwr yaml file.')
                #     if curblk in blkg.get_curmac_by_name() or f'{curblk}_USR' in blkg.get_curmac_by_name():
                #         macs.append(curblk)
                #         sdc_warn(f'{curblk} is macro block, if not user constraint, need expand it in hier_pwr yaml file.')
                #     if curblk in blkg.get_curdig_by_name() or f'{curblk}_USR' in blkg.get_curdig_by_name():
                #         digs.append(curblk)
                #         sdc_warn(f'{curblk} is digital block, if not user constraint, need expand it in hier_pwr yaml file.')

        if outtype == 'hd':
            return insts
        if outtype == 'lib':
            return macs
        if outtype == 'soft':
            return digs  
        if outtype == 'pll':
            return plls
        if outtype == 'crg':
            return crgs  

    def get_hierdepth(self,dic,key):
        
        allblks = list(self._hierdata.keys())
    
        if key not in dic:
            return 1
        else:
            max_depth = 0
            for sub_key in dic[key]:
                if sub_key not in allblks:
                    sdc_warn(f'{sub_key} is not expanded in hier yaml file.')
                else:
                    current_depth = self.get_hierdepth(dic, sub_key) + 1
                    max_depth = max(max_depth, current_depth)
            return max_depth

        # # blktrees = self.get_hiertrees(name)
        # if key not in dic:
        #     return 0
        
        # depths = []
        # for child_key in dic[key]:
        #     if child_key not in allblks:
        #         print(f'{child_key} is not expanded in hier yaml file.')
        #     else:
        #         depths.append(self.get_hierdepth(dic,child_key))
        #         #max_depth = max(max_depth, depth)
        
        # return max(depths) + 1

        # max_depth = depth  # 记录最大深度
        
        # if isinstance(blktrees, dict):
        #     for child_node in blktrees.values():
        #         if child_node not in allblks:
        #             print(f'{child_node} is not expanded in hier yaml file.')
        #     else:
        #         child_depth = self.get_hierdepth(child_node, depth + 1)
        #         max_depth = max(max_depth, child_depth)
        
        # return max_depth

    def get_alias_by_name(self, name):
        return self._blocks[name].alias

    def parse_env_path(self, epath):
        interp = tkinter.Tcl()
        interp.tk.eval('proc user_path {args} {return $args}')
        rsdcdir = f'user_path {epath}'

        try:
            rsdcdir = interp.tk.eval(rsdcdir)
        except:
            sdc_fatal(f'{epath} can not be parsed. Must set env(variable) in the path.')

        return rsdcdir

    def get_sblk_rdir(self, sblk, als, kw='outputs/'):
        if not self._proj:
            # sblk_rdir = f'$SDCVAR(SDC_DIR,${{{als}}})../../{sblk}/sdc/{kw}'
            vardata = self.get_vars_from_vardef(sblk, self._vfdir, 'vardef.json')
            sdir = vardata['SDC_DIR']
            sblk_rdir = f'{sdir}../../../{sblk}/sdc/{kw}'
        else:
            sblk_rdir = f'{self.get_block_by_name(sblk).constr_dir}/sdc/{kw}'

        if re.search(r'\$env\(', sblk_rdir):
            sblk_rdir = self.post_process_dir(sblk_rdir)
            sblk_rdir = self.parse_env_path(sblk_rdir)
        else:
            sblk_rdir = self.post_process_dir(sblk_rdir)

        return sblk_rdir

    def post_process_dir(self, filepath):
        filepath = re.sub(r'\/+', '/', filepath)
        return filepath

    def get_vars_from_vardef(self, blk, vfdir, fname):
        jdata = {}
        tmvars = {}

        # cblk = self.get_block_by_name(blk)
        hdir = f'{vfdir}/../json'

        if not os.path.exists(hdir):
            sdc_warn(f'Missing directory {hdir} for {blk} json check in hierpwr.')
        else:
            json_file = f'{hdir}/{fname}'
            if os.path.exists(json_file):
                jdata = self.read_json(json_file)
            else:
                sdc_warn(f'Missing {fname} file of {blk}')

        mvars = {k: v for k, v in jdata.items() if 'TMVAR' in k}
        # tmhier = {k:v for k,v in jdata.items() if 'TMHIER' in k}
        if mvars:
            for k, v in mvars.items():
                tmvars[v['Variable']] = v['Value']
        else:
            sdc_warn(f'Missing TMVAR info. in {blk} vardef.json')

        return tmvars


    ########################################################################
    def check_hym(self, name='', fkw='chk_hym'):
        # tals = self.get_alias_by_name(name)
        if name:
            tbk = self.get_block_by_name(name)
            if tbk.hdlevel == 'top':
                tcblk = tbk.get_curhd_by_name()
                if tcblk:
                    for cblk in tcblk:
                        cbk = self.get_block_by_name(cblk)
                        # cals = self.get_alias_by_name(cblk)
                        if cbk.hdlevel == 'sys':
                            scrgals = self.get_cursub_alias(cblk, 'crg')
                            if not scrgals:
                                sdc_error(
                                    f'"hdlevel" with "sys" of {cblk} under top {name} must include crg in dig_insts.',
                                    f'{fkw}')
                        if cbk.hdlevel == 'blk':
                            scrgals = self.get_cursub_alias(cblk, 'crg')
                            if scrgals:
                                sdc_error(
                                    f'"hdlevel" with "blk" of {cblk}  under top {name} must NOT include crg in dig_insts.',
                                    f'{fkw}')

        # blk: [f'{alias}',f'{lvl}',f'{pwr}',f'{nlb}',f'{inst}',f'{mac}',f'{dig}',f'{condir}']
        curflg = True
        if name == '':
            hblksinfo = self.get_full_hierblks_info()
            curflg = False
        else:
            curflg = True
            hblksinfo = self.get_hierblks_infos_chk(name)
            crgs = self.get_hierlvlblks(name, 'crg')
            ips = self.get_hierlvlblks(name, 'ip')
            digs = self.get_hierlvlblks(name, 'soft')
            macs = self.get_hierlvlblks(name, 'lib')
            insts = self.get_hierlvlblks(name)
        # allblks = list(self._hierdata.keys())
        # print('hblksinfo: ',hblksinfo)
        #{'jpeg_top_wrap': ['JPEG', 'blk', 'VDDM_CLPS TT0P800V TT0P750V TT0P700V', 'net', 'None', 'None', "['u_jpeg_top/u_arv_top, arv_top, VDDM_CLPS']", '/home/master/stone/tmp/const/tpy/vscode/sdcgen/template/inputs/'], 
        # 'arv_top': ['AVE', 'soft', 'VDDM_CLPS TT0P800V TT0P750V TT0P700V', 'net', 'None', '[{\'u_cr8_top, cr8_top,VDDM_CLPS\': \'if {[file exists $SDCVAR(SDC_DIR,${AVE})/arv_top/sdc/inputs/mdblk/cr8_top.sdc]} {\\n    source -echo -verbose $SDCVAR(SDC_DIR,${AVE})/arv_top/sdc/inputs/mdblk/cr8_top.sdc\\n} else {\\n    puts "SDC_WARN: Missing cr8_top SDC file. Please check it."\\n}\\n\'}]', 'None', '/home/master/stone/tmp/const/tpy/vscode/sdcgen/template/inputs/'], 'cr8_top': ['CR8', 'lib', 'VDDM_CLPS TT0P800V TT0P750V TT0P700V', 'net', 'None', 'None', 'None', '$env(SYS_CAMERA_ICDIR)/de/impl/spg_top/']}

        # print('checkenvpath: ',self.parse_env_path('$env(MCUJPEG_ICDIR)/case/mcu_jpeg_top_wrap/'))
        # pwr check: supply vol must from big to small
        if self._pwrdata:
            for sp, sv in self._pwrdata.items():
                xsv = sorted(sv.split(' '))[::-1]
                if xsv != sv.split(' '):
                    sdc_warn(f'The voltage of "{sp}" in "pwr" NOT follow decreasing order.', f'{fkw}')
        else:
            sdc_error(f'Missing "pwr" definition in hier yaml file.', f'{fkw}')

        if hblksinfo:
            hdblks = list(hblksinfo.keys())
            # pwr check: supply name must cover all supplys from hier
            hpwr = [x[2] for x in hblksinfo.values()]
            #hpwr : ['VDD_CORE TT0P750V TT0P700V TT0P650V', 'VDD_LPI TT0P850V TT0P650V TT0P600V']
            if hdblks:
                xpwr = []
                for hbk in hdblks:
                    sbk = self.get_block_by_name(hbk)
                    if sbk.get_curhdpwr_by_name():
                        xpwr.extend(sbk.get_curhdpwr_by_name())
                    if sbk.get_curmacpwr_by_name():
                        xpwr.extend(sbk.get_curmacpwr_by_name())
                    if sbk.get_curdigpwr_by_name():
                        xpwr.extend(sbk.get_curdigpwr_by_name())
                # xpwr: ['VDD_CORE', 'VDD_LPI', 'VDD_CORE', 'VDDM_CLPS', 'VDD_MM_CSS']
                # if hpwr:    xpwr.extend(hpwr)
                # print('_cxpwr: ',hpwr,xpwr)
                if hpwr:
                    # self._pwrdata: {'VDD_CORE': 'TT0P750V TT0P700V TT0P650V', 'VDD_MM_CSS': 'TT0P650V TT0P600V TT0P550V'}
                    mpwr = [x for x in list(set(xpwr)) if x not in self._pwrdata.keys()]
                    if mpwr:
                        sdc_error(f'Supply {mpwr} missing definition in "pwr" of hier yaml file.', f'{fkw}')

            # als var must be unique
            als_var = [x[0] for x in hblksinfo.values()]
            als_var_u = list(set(als_var))
            if len(als_var) != len(als_var_u):
                counts = Counter(als_var)
                rvars = [ix for ix, cnt in counts.items() if cnt > 1]
                sdc_error(f'Alias variable {rvars} is not unique.', f'{fkw}')

            # als val must be unique
            if hdblks:
                als_val = []
                for hbk in hdblks:
                    sbk = self.get_block_by_name(hbk)
                    if sbk.get_curhdval_by_name():
                        als_val.extend(sbk.get_curhdval_by_name())
                    if sbk.get_curmacval_by_name():
                        als_val.extend(sbk.get_curmacval_by_name())
                    if sbk.get_curdigval_by_name():
                        als_val.extend(sbk.get_curdigval_by_name())
                als_val = [x for x in als_val if x != None and re.search(r'\$|\:|\{|\}', x)]
                als_val_u = list(set(als_val))
                if len(als_val) != len(als_val_u):
                    counts = Counter(als_val)
                    rvals = [ix for ix, cnt in counts.items() if cnt > 1]
                    sdc_error(f'Alias value {rvals} is not unique.', f'{fkw}')

            for blk, alst in hblksinfo.items():
                # can not cover in full hier yaml check
                if alst == 'NOT EXPANDED':
                    sdc_error(f'{blk} Not expanded in hier yaml.', f'{fkw}')

                # alias
                if not alst[0] or alst[0] == 'None':
                    sdc_error(f'{blk} Not found <ALIAS> value in hier yaml. Must start upper string.', f'{fkw}')
                if re.search(r'^\d+', alst[0]):
                    sdc_error(
                        f'Found {blk} <ALIAS> value begins with digital number in hier yaml. Must begin with upper string.', f'{fkw}')
                if re.search(r'[a-z]', alst[0]):
                    sdc_error(f'Found {blk} <ALIAS> value includes lower string in hier yaml. Must use upper string.', f'{fkw}')
                if re.search(r'\_|\-|\$|\%|\@|\#|\&|\[|\]|\(|\)|\*|\:', alst[0]):
                    sdc_error(
                        f'Found {blk} <ALIAS> value includes "_|-|$|%|@|#|&|[|]|(|)|*|:" in hier yaml. Must NOT use special string.', f'{fkw}')

                # hdlevel
                if not alst[1] or alst[1] == 'None':
                    sdc_error(
                        f'{blk} Not found <HDLEVEL> value in hier yaml. Must be "blk/sys/top/crg/lib/soft" value.',
                        f'{fkw}')
                lvslst = ('top', 'sys', 'blk', 'soft', 'lib', 'crg')
                if alst[1] not in lvslst:
                    sdc_error(f'Not Found {blk} <HDLEVEL> value includs "blk/sys/top/crg/lib/soft" in hier yaml.',
                                f'{fkw}')
                if curflg:
                    if crgs:
                        if blk in crgs:
                            if alst[1] not in ('crg'):
                                sdc_error(f'{blk} belongs to crgs, but hdlevel NOT found in "(crg)"', f'{fkw}')
                    if ips:
                        if blk in ips:
                            if alst[1] not in ('soft', 'lib'):
                                sdc_error(
                                    f'{blk} belongs to mac_inst or dig_inst, but hdlevel NOT found in "(soft, lib)"', f'{fkw}')

                    if digs:
                        if blk in digs:
                            if alst[1] not in ('soft'):
                                sdc_error(f'{blk} belongs to dig_inst, but hdlevel NOT found in "(soft)"', f'{fkw}')

                    if macs:
                        if blk in macs:
                            if alst[1] not in ('lib'):
                                sdc_error(f'{blk} belongs to mac_inst, but hdlevel NOT found in "(lib)"', f'{fkw}')

                    if insts:
                        if blk in insts:
                            if alst[1] not in ('top', 'sys', 'blk'):
                                sdc_error(f'{blk} belongs to harden insts, but hdlevel NOT found in "(top, blk, sys)"', f'{fkw}')

                # prime_pwr
                if not alst[2] or alst[2] == 'None':
                    sdc_error(f'{blk} Not found <PRIME_PWR> value in hier yaml. Must set in <pwr> and <prime_pwr>.', f'{fkw}')
                # need add mim supply check???

                # intg_nlb
                # if not alst[3] or alst[3] == 'None':
                #     sdc_error(f'{blk} Not found <INTG_NLB> value in hier yaml. Must be "net/lib/bbx" value.', f'{fkw}')
                # if alst[3] not in ('net', 'lib', 'bbx'):
                #     sdc_error(f'Not found {blk} <INTG_NLB> value includes "net/lib/bbx" in hier yaml.', f'{fkw}')

                # constr_dir
                # if alst[1] in ('blk', 'sys', 'top'):
                #     if not alst[7] or alst[7] == 'None':
                #         sdc_warn(f'Not found {blk} constr_dir value for project mode.', f'{fkw}')
                #     else:
                #         rcdir = self.parse_env_path(alst[7])
                #         if not os.path.exists(rcdir):
                #             sdc_error(f'Not found {blk} constr_dir real path. Must be the absolute existed path.', f'{fkw}')
                #         else:
                #             if not re.search(r'\/$', alst[7]):
                #                 sdc_error(f'Not found {blk} constr_dir value include "/" for the end string', f'{fkw}')

                #             if alst[7].endswith('sdc') or alst[7].endswith('sdc/'):
                #                 sdc_error(f'The end strings of {blk} constr_dir value must NOT be "sdc/" or "sdc"', f'{fkw}')

                #             vardata = self.get_vars_from_vardef(blk, self._vfdir, 'vardef.json')
                #             print('__var: ', blk, self._vfdir, vardata)
                #             if re.search(r'\$env(', vardata['SDC_DIR']):
                #                 sdir = self.parse_env_path(vardata['SDC_DIR'])
                #             else:
                #                 sdir = vardata['SDC_DIR']
                #             if re.search(r'\$env(', alst[7]):
                #                 gdir = self.parse_env_path(alst[7])
                #             else:
                #                 gdir = alst[7]
                #             if sdir != gdir:
                #                 sdc_warn(f'NOT consistency bet <SDC_DIR> in vardef and <constr_dir> in hier yaml. If proj mode, Need be consistent.')

                # if alst[1] == 'crg':
                #     if not alst[7] or alst[7] == 'None':
                #         sdc_error(f'{blk} is crg module. Must specify constr_dir value.', f'{fkw}')
                #     else:
                #         if not os.path.exists(alst[7]):
                #             sdc_error(f'Not found {blk} constr_dir real path. If proj mode, Must be the absolute existed path.', f'{fkw}')
                #         else:
                #             if not re.search(r'\/$', alst[7]):
                #                 sdc_error(f'Not found {blk} constr_dir value include "/" for the end string', f'{fkw}')
                #             if alst[7].endswith('sdc') or alst[7].endswith('sdc/'):
                #                 sdc_error(f'The end strings of {blk} constr_dir value must NOT be "sdc/" or "sdc"', f'{fkw}')




class SDC_DG(object):
    def __init__(self):
        self._sheets = {}
        self._hier_tree = {}
        self._vardef = {}
        self._wb = {}
        self._sdcdir = ''
        self._mdname = ''
        self._alias = '' #self._hier_tree._blocks[self._mdname].alias
        self._hdlvl = ''
        self._pwr = ''
        self._inputs = BaseInputs()
        # self._iodly = VIODly()

        #self.proj_mode = False
        self._vfile_data = None
        self._vfile_list = None
        self._data = None
        self._lvl = 'blk'
        self._flt = 'IS_FLAT'

        self._vardata = {}
        # self._sdc_dir = ''
        # self._com_dir= ''
        # self._dft_dir = ''
        # self._hd_process = ''
        # # self._hd_mod_name = self._mdname
        # self._cycle_list = '[list CYCLE500M]'


    @property
    def hier_tree(self):
        return self._hier_tree
    
    @hier_tree.setter
    def hier_tree(self, hier_tree):
        self._hier_tree = hier_tree
 
    def load_design_guide(self,dg_file,kwd=''):
        self._sdcdir = abspath(dirname(dirname(dg_file)))
        self._wb = openpyxl.load_workbook(dg_file)

        valdef_sheet = self._wb['VarDef']
        start = 0
        for i in range(1, valdef_sheet.max_row+1):
            if valdef_sheet.cell(i,1).value == 'Variable':
                start = i + 1
                break
        for i in range(start, valdef_sheet.max_row+1):
            key = valdef_sheet.cell(row=i, column=1).value
            val = valdef_sheet.cell(row=i, column=2).value
            self._vardef[key] = val

        self._sheets = {
            'VarDef'            : VarDefSheet(self, 'VarDef'),
            'ClkDef'            : ClkDefSheet(self, 'ClkDef'),
            'IODly'             : IODlySheet(self, 'IODly'),
            'Exp'               : ExpSheet(self, 'Exp'),
            # 'IOExp'             : IOExpSheet(self, 'IOExp'),
            # 'IntExp'            : IntExpSheet(self, 'IntExp'),
        }

        # read and convert excel to json data
        for sheetname,sheet in self._sheets.items():
            sheet.read_data()
        # if kwd == 'json':
        #     for sheetname,sheet in self._sheets.items():
        #         if sheetname == 'ClkDef':
        #             sheet.read_data('json')
        #         else:
        #             sheet.read_data()            
                # json_file = dirname(dirname(dg_file)) + '/json' + f'/{sheetname.lower()}.json'
                # # json_file = dirname(dirname(dg_file)) + os.path.join('json', f'{sheetname.lower()}.json')
                # sheet.dump_json(json_file)

            # json_file = dirname(dirname(dg_file)) + '/json' + f'/hier_pwr.json'
            # # json_file = dirname(dirname(dg_file)) + os.path.join('json', 'hier.json')
            # self._data = self._hier_tree._yaml_data
            # self.write_json(json_file)
            

    def read_vfile(self,vfile,kwd=''):
        self._vfile_list, self._vfile_data = self._inputs.read_vfile(vfile)
        self._mdname = self._vfile_data['module_name']
        # print(self._vfile_list)
        # print(self._vfile_data)

        # if kwd == 'json':
        #     self._data = self._vfile_data
        #     json_file = dirname(dirname(vfile)) + '/json' + f'/vfile.json'
        #     self.write_json(json_file)    

    def update_dg(self):
        for sht in self._sheets.values():
            sht.update_sheet()

    def check_dg(self):
        for sht in self._sheets.values():
            sht.check_sheet()

    def change_dg(self,dgfile):
        for sht in self._sheets.values():
            sht.change_sheet(dgfile)

    def read_json(self,file_path):
        sblk_data = {}
        if os.path.exists(file_path):
            with open(file_path,'r') as fw:
                content = fw.read()
                sblk_data = json.loads(content)

        #print('sblk_data:',sblk_data)
        return sblk_data


    def write_json(self,filepath):
        os.makedirs(dirname(filepath), exist_ok=True)
        jsonstr = json.dumps(self._data, indent=4)
        with open(filepath,'w') as fw:
            print(jsonstr, file=fw) 

    # def save_text(self,context,file):
    #     with open(file, 'w') as fw:
    #         fw.write(context)

    def save_text(self, context,file,kw='xyz'):
        if os.path.exists(file) and 'proc' in kw:
            with open(file, 'a') as fw:
                fw.write(context)
        else:
            with open(file, 'w') as fw:
                fw.write(context)

    def save_workbook(self,output):
        self._wb.save(output)

    def read_text(self, file):
        if not os.path.exists(file):
            raise FileExistsError(f'{file} does not exists')
            # sdc_error(f'{file} not exist. Please check it.')
            # exit(1)
        else:
            txt_list = []
            with open(file,'r') as fh:
                for line in fh.readlines():
                    txt_list.append(line)
        
            return txt_list


# sdcdg is XsdcDesignGuide object
class BaseSheet(object):
    def __init__(self, sdcdg, sheetname):
        self._sdcdg = sdcdg
        self._sheetname = sheetname
        self._data = []
        #self._vardef = {}
        self._pdnmdict = {}
    
    def get_sheet(self):
        return self._sdcdg._wb[self._sheetname]

    def read_data(self):
        raise NotImplementedError(self.__class__.__name__ + ' read_data not implemented yet')

    def write_json(self, filepath):
        os.makedirs(dirname(filepath), exist_ok=True)
        jsonstr = json.dumps(self._data, indent=4)
        with open(filepath,'w') as fw:
            print(jsonstr, file=fw)

    def find_sheet(self, sheet, skw):
        start_rowg = 1
        # TABCONST = ['TMVAR','TMHIER','TMCLK','TMIODLY','TMIOEXP','TMINOUT','TMINTEXP','TMSTPGATE']
        TABCONST = ['TMVAR','TMCLK','TMIODLY','TMIOEXP','TMINOUT','TMINTEXP','TMSTPGATE']
        #print(skw,sheet)
        for i in range(1,sheet.max_row+1):
            if skw in TABCONST and sheet.cell(i,1).value == skw:
                start_rowg = i + 1
                break  
        return  start_rowg 

    def get_vardef_value(self, sheet):
        vardef = {}
        start_rowg = self.find_sheet(sheet, 'TMVAR')
        # end_rowg = self.find_sheet(sheet, 'TMHIER')
        for i in range(start_rowg + 1, start_rowg + 15):
            key = sheet.cell(row=i, column=1).value
            val = sheet.cell(row=i, column=2).value
            vardef[key] = val

        vardef['SDC_DIR'] = self._sdcdg._sdcdir
        vardef['COM_DIR'] = self._sdcdg._sdcdir
        vardef['DFT_DIR'] = ''
        # vardef['HD_MOD_NAME'] = self._sdcdg._mdname
        vardef['HD_PROCESS'] = ''
        vardef['CYCLE_LIST'] = '[list CYCLE500M]'

        # print('vardef:', vardef)
        return vardef   

    def set_name_style(self, kw):
        #time_stamp = time.strftime("%Y%m%d%H%M%S", time.localtime())
        #CONST = f'Generic_Xsdc_{time_stamp}'
        CONST = f'Generic_XSDC'
        return kw + '_' + CONST

    # showErrorMessage=False,showDropDown=True
    def add_dropdown(self, sheet, options, start, end):       
        dv = DataValidation(type="list", formula1=options, showErrorMessage=False)
        sheet.add_data_validation(dv)
        if len(start) == 2 and len(end) == 2:
            for i in range(start[0], end[0] + 1):
                for j in range(start[1], end[1] + 1):
                    dv.add(sheet.cell(i,j))
        if len(start) == 1 and len(end) == 1:
            dv.add(sheet.cell(start[0],end[0]))


    def cell_style1(self, sheet, start, end):
        border=Border(left=Side(border_style='thin', color='000000'),
                      right=Side(border_style='thin', color='000000'),
                      top=Side(border_style='thin', color='000000'),
                      bottom=Side(border_style='thin', color='000000'))
        #bgfill = PatternFill(fill_type='solid', start_color='fff2cc', end_color='fff2cc') 
        #bgfill = PatternFill(fill_type = 'solid', start_color='197e00',end_color='197e00')
        bgfill = PatternFill(fill_type = 'solid', start_color='FF385724',end_color='FF333300')
        font = Font(name='等线', size=11, color='FFFFFF')
        for i in range(start[0], end[0] + 1):
            for j in range(start[1], end[1] + 1):
                sheet.cell(i,j).border=border 
                sheet.cell(i,j).fill=bgfill
                sheet.cell(i,j).font=font
                sheet.cell(i,j).alignment = Alignment(horizontal='left', vertical='center',wrapText=True) 
                sheet.cell(i,j).alignment = Alignment(horizontal='left', vertical='center',wrapText=True) 
                sheet.cell(i,j).alignment=Alignment(horizontal='left', vertical='center') 

    def cell_style2(self, sheet, start, end):
        border=Border(left=Side(border_style='thin', color='000000'),
                      right=Side(border_style='thin', color='000000'),
                      top=Side(border_style='thin', color='000000'),
                      bottom=Side(border_style='thin', color='000000'))
        #bgfill = PatternFill(fill_type='solid', start_color='fff2cc', end_color='fff2cc') 
        #bgfill = PatternFill(fill_type = 'solid', start_color='197e00',end_color='197e00')
        bgfill = PatternFill(fill_type = 'solid', start_color='FFFFFF',end_color='FFFFFF')
        #font = Font(name='等线', size=11, color='FFFFFF')
        for i in range(start[0], end[0] + 1):
            for j in range(start[1], end[1] + 1):
                sheet.cell(i,j).border=border 
                sheet.cell(i,j).fill=bgfill
                #sheet.cell(i,j).font=font
                sheet.cell(i,j).alignment = Alignment(horizontal='left', vertical='center',wrapText=True) 
                sheet.cell(i,j).alignment = Alignment(horizontal='left', vertical='center',wrapText=True) 
                sheet.cell(i,j).alignment=Alignment(horizontal='left', vertical='center',wrapText=True)

    def get_supply_infos(self):

        delkeys = ['module_name', 'ISO_CTRL', 'RET_SAVE', 'RET_RES', 'PSO_CTRL', 'PSO_ACK']
        supply_datag = self._sdcdg.vfile_data
        supply_data = {}

        for ky,vl in supply_datag.items():
            if ky not in delkeys:
                supply_data[ky] = vl

        supply_kw = []
        supply_val = []
        supply_vss = []
        supply_tmp = ''
        for key,val in supply_data.items():
            if '0v' in val or '0.0v' in val:
                supply_vss.append(key)
                sdc_info(f'Ground pin is {key}')
            elif 'PSO' in val:
                supply_kw.append(key)
                for i in range(1, int(val[-1]) + 1):
                    supply_kw.append(key + f'_PSW{i}')
                supply_tmp = supply_tmp + ' ' + supply_data[key].split('PSO')[0].strip()
            else:
                supply_kw.append(key)
                supply_tmp = supply_tmp + ' ' + supply_data[key].strip()

        #print(supply_tmp.strip().split(','))
        float_list = [float(x.strip('v')) for x in supply_tmp.strip().split()]
        unique_floats = set(float_list)
        sorted_floats = sorted(unique_floats, reverse=True)
        supply_val = [str(x) + 'v' for x in sorted_floats]
        supply_val.append('off')
        supply_val.append('0v')

        return supply_kw,supply_val,supply_vss,supply_data
    

    def get_ctl_sig(self, ctsig):
        ctrl = []
        for ct in ctsig:
            if re.search(r'\[\d+:\d+\]', ct):
                sig = ct.split('[')[0].strip()
                st = int(ct.split(':')[0].strip()[-1])
                ed = int(ct.split(':')[1].strip()[0])
                for i in range(ed,st+1):
                    ctrl.append(sig + '[' + str(i) + ']')
            else:
                ctrl.append(ct)
        return ctrl

    def get_table_loc(self,sheet) -> dict:

        if self._sheetname == 'VarDef':
            TABCONST = ['TMVAR']
        if self._sheetname == 'ClkDef':
            TABCONST = ['TMCLK']
        if self._sheetname == 'IODly':
            TABCONST = ['TMIODLY']
        # if self._sheetname == 'IOExp':
        #     TABCONST = ['TMIOEXP','TMINOUT']
        # if self._sheetname == 'IntExp':
        #     TABCONST = ['TMINTEXP','TMSTPGATE']
        if self._sheetname == 'Exp':
            TABCONST = ['TMIOEXP','TMINOUT','TMINTEXP','TMSTPGATE']

        # row_start max_col              
        row_start = ''        
        max_row = ''
        max_col = ''
        # row_start max_col max_row
        table_row_loc = {}
        for kw in TABCONST:
            strow = self.find_sheet(sheet, kw)
            row_start = str(strow)
            for i in range(1,sheet.max_column + 1):
                if sheet.cell(strow,i).value == 'Comment':
                    max_col = str(i)
                    #print(row_start[kw])
                    break              

            # print(kw,row_start)
            # if kw in ['TMHIER','TMCLK','TMIODLY','TMINOUT','TMSTPGATE']:
            if kw in ['TMVAR','TMCLK','TMIODLY','TMSTPGATE']:
                #table_row_loc[kw] = row_start[kw] + ' ' + str(int(row_start[kw].split()[0]) + 20)
                table_row_loc[kw] = row_start + ' ' + str(sheet.max_row + 2) + ' ' + max_col
            else:
                idx = TABCONST.index(kw) + 1
                max_row = self.find_sheet(sheet,TABCONST[idx]) - 1
                # print(kw,idx,max_row)
                #table_row_loc[kw] = row_start[kw] + ' ' + str(int(row_start[TABCONST[idx]].split()[0]) - 2)
                table_row_loc[kw] = row_start + ' ' + str(max_row) + ' ' + max_col

        #print(table_row_loc)
        return table_row_loc
  
    def get_table_contxt(self,sheet) -> dict:
        # row_start max_col max_row
        tab_loc = self.get_table_loc(sheet)
        # print(tab_loc)

        TABCONST = []
        if self._sheetname == 'VarDef':
            TABCONST = ['TMVAR']
        if self._sheetname == 'ClkDef':
            TABCONST = ['TMCLK']
        if self._sheetname == 'IODly':
            TABCONST = ['TMIODLY']
        # if self._sheetname == 'IOExp':
        #     TABCONST = ['TMIOEXP','TMINOUT']
        # if self._sheetname == 'IntExp':
        #     TABCONST = ['TMINTEXP','TMSTPGATE']
        if self._sheetname == 'Exp':
            TABCONST = ['TMIOEXP','TMINOUT','TMINTEXP','TMSTPGATE']

        table_contxt = {}
        #row_contxt = {}
        if TABCONST:
            for kw in TABCONST:
                start_row = int(tab_loc[kw].split(' ')[0])
                end_row = int(tab_loc[kw].split(' ')[1])
                end_col = int(tab_loc[kw].split(' ')[2])
                # if kw == 'TMSTPGATE':
                #     print('TMSTPGATE:',start_row,end_row,end_col)
                if kw == 'PMVAR':
                    for i in range(start_row, end_row + 1):
                        key = sheet.cell(i + 1, 1).value
                        val = str(sheet.cell(i + 1, 2).value)
                        if key:
                            table_contxt[key] = val.strip()
                        # print('PMVARdfd: ', table_contxt)
                        # if key and val:
                        #     table_contxt[key] = val
                else:
                    table_contxt.update(self.get_row_txt(sheet,kw,start_row,end_row,end_col))
                # if kw == 'TMSTPGATE':
                #     print('TMSTPGATE:',table_contxt)

        return table_contxt

    def get_row_txt(self,sheet,kw,start_row,end_row,end_col):
        row_contxt = {}
        table_contxt = {}
        for i in range(1,end_row-start_row):
            for j in range(1,end_col+1):
                key = sheet.cell(start_row,j).value
                val = sheet.cell(start_row+i,j).value
                val_col1 = sheet.cell(start_row+i,1).value
                if val_col1:
                    if re.search(r'^#',val_col1.strip()):
                        continue
                if key:     key = str(key).strip()
                if val:     val = str(val).strip()
                row_contxt[key] = val
                # if key and val:
                #     row_contxt[key] = val
            all_none = all(ele is None for ele in list(row_contxt.values()))
            if not all_none and row_contxt:
                table_contxt[f'{kw}_Row{start_row+i}'] = row_contxt
            row_contxt = {}
            # for key in table_contxt.keys():
            #     if 'TMCLK' in key:
            #         print(table_contxt)
        
        return table_contxt

        
    def save_text(self, context,file):
        with open(file, 'w') as fw:
            fw.write(context)

    def get_rows(self,pmdata,keyrow,kwd,ckwd):
        pmdict = {}
        pmlist = [(key, val) for key, val in pmdata.items() if re.search(r'{keyrow}\d+',key) and not re.search(r'^#',val[f'{ckwd}'].strip()) and val[f'{kwd}']]
        for k,v in pmlist:
            pmdict[k] = v
        pmkeys = [x for x in pmlist if re.search(r'{keyrow}\d+',x)]
        pmkeys.sort() 

        return pmdict, pmkeys     


class BaseInputs(object):
    def __init__(self):
        self.vfile_data = {}
        self.vfile_list = []
               
    def read_vfile(self, vfile) -> dict:

        lines = self.read_text(vfile)

        relclknum = 0
        for line in lines:
            line = line.replace('\n','').replace('\r','').replace('\t',' ').strip()
            if re.search(r'^\/\/', line) and '#RelClock:' not in line:
                continue

            if re.search(r'^module', line):
                self.vfile_data['module_name']= re.split(' +',line)[1].strip().replace('(','')
                self.vfile_list.append('module_name')
                continue

            if '#RelClock:' in line:
                relclknum += 1
                relclk = line.split('#RelClock:')[1].strip().replace('#','')
                self.vfile_list.append(f'RelClock{relclknum}')
                self.vfile_data[f'RelClock{relclknum}'] = relclk
                continue
                
            if re.search(r'^\);$',line):
                break

            dirc = ''
            portnum = ''
            kwd = ''
            if re.search(r'^input|^output|^inout',line):
                tline = line.split(' ')
                sline = [x for x in tline if x != '']
                dirc = sline[0]
                dircg = dirc
                
                if re.search(r'wire|logic|byte|bit|reg|tri1|tri0',line):               
                    if re.search(r'\[\d+:\d+\]',line):
                        lineg = ' '.join(sline[3:])
                        portnum = sline[2]
                        # print(sline)
                    else:
                        lineg = ' '.join(sline[2:])
                        portnum = '1'             
                else:
                    if re.search(r'\[\d+:\d+\]',line):
                        lineg = ' '.join(sline[2:])
                        portnum = sline[1]
                    else:
                        lineg = ' '.join(sline[1:])
                        portnum = '1'
                portnumg = portnum
                

                sigchar = lineg.replace(' ','')
                # print('inputoroutput:', sigchar)
                if re.search(r'\/\/#\w+#',sigchar):
                    #kwdg = ''.join(re.findall(r'\/\/(#\w+#)+', sigchar)).strip()
                    if '##' in sigchar:
                        kwd = sigchar.replace('##',' ').split('#')[1]
                    else:
                        kwd = sigchar.split('#')[1]
                    
                    if ',' in sigchar:
                        sigcharg = sigchar.split(',')
                        for ich in sigcharg:
                            if r'#\w+#' not in ich and '//' not in ich:
                                self.vfile_data[ich] = [dircg,portnumg,kwd]
                                self.vfile_list.append(ich)
                    else:
                        sdc_warn(f'{sigchar} not found , symbol ...')
                        ish = sigchar.split(r'//')[0].strip()
                        self.vfile_data[ish] = [dircg,portnumg,kwd]
                        self.vfile_list.append(ish)
                else:
                    kwd = 'None'
                    if ',' in sigchar:
                        sigcharg = sigchar.split(',')                        
                        for ich in sigcharg:
                            if '//' not in ich and ich != '':
                                self.vfile_data[ich] = [dircg,portnumg,kwd]
                                self.vfile_list.append(ich)
                    else:
                        sdc_warn(f'{sigchar} not found , symbol ...')
                        if '//' in sigchar:
                            ish = sigchar.split(r'//')[0].strip()
                        else:
                            ish = sigchar.strip()
                            self.vfile_data[ish] = [dircg,portnumg,kwd]
                            self.vfile_list.append(ish)

            else:
                # print('NO_inputoroutput:', line)
                if re.search(r'^\S+,$',line) and '//' not in line:
                    sline = line.split(',')
                    if re.search(r'\/\/#\w+#',line):
                        #kwd = ''.join(re.findall(r'\/\/#\w+#', line)).strip().split('#')[1]
                        if '##' in line:
                            kwd = line.replace('##',' ').split('#')[1]
                        else:
                            kwd = line.split('#')[1]
                        
                        for ich in sline:
                            if r'#\w+#' not in ich and '//' not in ich:
                                self.vfile_data[ich] = [dircg,portnumg,kwd]
                                self.vfile_list.append(ich)
                    else:
                        kwd = 'None'
                        for ich in sline:
                            if '//' not in ich and ich != '':
                                self.vfile_data[ich] = [dircg,portnumg,kwd]
                                self.vfile_list.append(ich)                       
                else:
                    #sline = line.split(' +')
                    tline = line.split(' ')
                    sline = [x for x in tline if x != '']
                    if re.search(r'wire|logic|byte|bit|reg|tri1|tri0',line):               
                        if re.search(r'\[\d+:\d+\]',line):
                            lineg = ' '.join(sline[2:])
                            portnumg = sline[1]
                        else:
                            lineg = ' '.join(sline[1:])
                            #portnum = '1'             
                    else:
                        if re.search(r'\[\d+:\d+\]',line):
                            lineg = ' '.join(sline[1:])
                            portnumg = sline[0]
                        else:
                            lineg = ' '.join(line[0:])
                            #portnum = '1'

                    sigchar = lineg.replace(' ','')
                    if re.search(r'\/\/#\w+#',sigchar):
                        #kwd = ''.join(re.findall(r'\/\/#\w+#', sigchar)).strip().split('#')[1]
                        #kwdg = ''.join(re.findall(r'\/\/(#\w+#)+', sigchar)).strip()
                        if '##' in sigchar:
                            kwd = sigchar.replace('##',' ').split('#')[1]
                        else:
                            kwd = sigchar.split('#')[1]
                        
                        if ',' in sigchar:
                            sigcharg = sigchar.split(',')
                            #kwd = ''.join(re.findall(r'\/\/#\w+#', sigchar)).strip().split('#')[1]
                            for ich in sigcharg:
                                if r'#\w+#' not in ich and '//' not in ich:
                                    self.vfile_data[ich] = [dircg,portnumg,kwd]
                                    self.vfile_list.append(ich)
                        else:
                            sdc_warn(f'{sigchar} not found , symbol ...')
                            ish = sigchar.split('//')[0].strip()
                            self.vfile_data[ish] = [dircg,portnumg,kwd]
                            self.vfile_list.append(ish)
                    else:
                        kwd = 'None'
                        if ',' in sigchar:
                            sigcharg = sigchar.split(',')                           
                            for ich in sigcharg:
                                if '//' not in ich and ich != '':
                                    self.vfile_data[ich] = [dircg,portnumg,kwd]
                                    self.vfile_list.append(ich)
                        else:
                            sdc_warn(f'{sigchar} not found , symbol ...')
                            if '//' in sigchar:
                                ish = sigchar.split('//')[0].strip()
                            else:
                                ish = sigchar.strip()
                            self.vfile_data[ish] = [dircg,portnumg,kwd]
                            self.vfile_list.append(ish)
                # print('vfile_list:',self.vfile_list)
        return self.vfile_list, self.vfile_data

        


    def read_yaml(self, yaml_file):

        yaml_data = {}
        if not os.path.exists(yaml_file):
            raise FileExistsError(f'{yaml_file} does not exists')
        with open(yaml_file, 'r') as fh:
            yaml_data = yaml.load(fh, yaml.FullLoader)

        return yaml_data
    


    def read_text(self, file):
        if not os.path.exists(file):
            raise FileExistsError(f'{file} does not exists')
            # sdc_error(f'{file} not exist. Please check it.')
            # exit(1)
        else:
            txt_list = []
            with open(file,'r') as fh:
                for line in fh.readlines():
                    if line.strip() == "":
                        continue
                    if line.strip().startswith("//") and '#RelClock:' not in line.strip():
                         continue   
                    line = re.sub(r"\[\s*(\d+)\s*:\s*(\d+)\s*\]", r"[\1:\2]", line)
                    txt_list.append(line.strip())
        
            return txt_list



class VarDefSheet(BaseSheet):
    def __init__(self,*args):
        super().__init__(*args)  
        self._valdata = {}  
        self._vardata = {}
        #self._clkdata = self.get_table_contxt(self._sdcdg._wb['ClkDef'])
        #self._clkdef = None

        self._vfdata = self._sdcdg._vfile_data
        self._hiertree = self._sdcdg._hier_tree
        self._sdcdir = self._sdcdg._sdcdir
        self._mdname = self._sdcdg._mdname

    ###########################################################
    def read_data(self):
        sheet = self.get_sheet()
        self._valdata = self.get_table_contxt(sheet)
        nvaldata = {}
        nvaldata["TMVAR_Row14"] = {
            "Variable": "SDC_DIR",
            "Value": f'{self._sdcdir}',
            "Comment": ''
        }
        nvaldata["TMVAR_Row15"] = {
            "Variable": "COM_DIR",
            "Value": f'{self._sdcdir}',
            "Comment": ''
        }
        nvaldata["TMVAR_Row16"] = {
            "Variable": "DFT_DIR",
            "Value": '',
            "Comment": ''
        }
        # nvaldata["TMVAR_Row17"] = {
        #     "Variable": "HD_MOD_NAME",
        #     "Value": f'{self._mdname}',
        #     "Comment": ''
        # }
        nvaldata["TMVAR_Row17"] = {
            "Variable": "HD_PROCESS",
            "Value": '',
            "Comment": ''
        }
        nvaldata["TMVAR_Row18"] = {
            "Variable": "CYCLE_LIST",
            "Value": '[list CYCLE500M]',
            "Comment": ''
        }
        self._valdata.update(nvaldata)
    
    def get_var_cell(self,sht):
        start = 0
        end = 0
        vardef_data = {}
        for i in range(1, sht.max_row+1):
            if sht.cell(i,1).value == 'Variable':
                start = i + 1
            # if sht.cell(i,1).value == 'TMHIER':
            #     end = i - 1
        for i in range(start, sht.max_row+1):
            key = sht.cell(row=i, column=1).value
            val = sht.cell(row=i, column=2).value
            if key is not None:
                vardef_data[key] = val

        tmp_vardata = {}
        for val in list(self._valdata.values()):
          tmp_vardata[val['Variable']] = val['Value']

        for ky in list(vardef_data.keys()):
            for k,v in tmp_vardata.items():
                if k == ky:
                    sdc_info(f'{k} is already defined in vardef sheet.')
                else:
                    vardef_data[k] = v

        return vardef_data
    
    def check_sheet(self):
        sheet = self.get_sheet()
        vardata = self.get_var_cell(sheet)
        # print('vardata:', vardata)

        # module name 
        mdname = vardata['HD_MOD_NAME']
        if mdname != self._sdcdg._vfile_data['module_name']:
            sdc_error(f'Module name from vardef is different from the empty vfile.','chk_sht')

        # # hier variable name
        # cblk = self._hiertree.get_block_by_name(mdname)
        # lvl = cblk.hdlevel
        # hierlst = [x for x in vardata.keys() if 'HIER_' in x]
        # if hierlst:
        #     for hier in hierlst:
        #         if lvl in ('blk','sys','top'):
        #             if not f'HIER_{lvl.upper()}_' in hier and hier != 'HIER_EXPD_STYLE':
        #                 sdc_error(f'Naming of variable "{hier}" does not start with "HIER_{lvl.upper()}_".')
        #             if not vardata[hier].endswith('/') and hier != 'HIER_EXPD_STYLE':
        #                 sdc_error(f'Variable "{hier}" value does not end with "/".')

        # check SDC_DIR/COM_DIR/DFT_DIR with the end of '/'
        # for hir in ('SDC_DIR','COM_DIR','DFT_DIR'):
        #     if not vardata[hir].endswith('/'):
        #         sdc_error(f'Variable "{hir}" value does not end with "/".')



    def change_sheet(self):
        pass

    def dump_json(self,json_file):
        self._data = self._valdata
        self.write_json(json_file)

    def read_json(self,file_path):
        sblk_clk_list = {}
        if os.path.exists(file_path):
            with open(file_path,'r') as fw:
                content = fw.read()
                sblk_clk_list = json.loads(content)

        return sblk_clk_list



    def change_sheet(self):
        pass

    def dump_json(self,json_file):
        self._data = self._valdata
        self.write_json(json_file)


class ClkDefSheet(BaseSheet):
    def __init__(self,*args):
        super().__init__(*args)  
        self._hiertree = self._sdcdg._hier_tree
        
        self._sdcdir = self._sdcdg._sdcdir
        self._mdname = self._sdcdg._mdname
        self._alias = self._hiertree._blocks[self._mdname].alias #self._sdcdg._alias

        self._clkdata = {}  
        self._clknmdata = {}
        self._clknmlst = []

        self._clkvardata = {}
        #self._clkvarlinegs = ''

        self._rowcrgdata = {}
        self._iptcrgdata = {}
        self._iptcrglst  = {}
        self._crgals = {}

        self._rowipdata = {}
        self._iptipdata = {}
        self._iptiplst  = {}
        self._ipals = {}

        self._crgflg = 0
        self._ipflg = 0

        self._crgalsiptval = {}
        self._ipalsiptval = {}

        self._tclkdata = {}
        self._tclklst = {}

        # for crgip mstclk/srcpin/clkgrp
        #self._clkinfolst = []
        self._cycle_clkdeflst = []
        self._cycle_crgiplst = []
        #self._intgportlst = []

        self._crgipclknmals = {}

        self._hdportclks = {}
        self._hdportclksinfo = {}

        self._curhd_portclks = {}
        self._curhd_portclksinfo = {}

        # self._clkdef = {}
        # self._crgipclkdef = {}
        # self._hdportclkdef = {}
        self._curclkdef = {}

        self._lvl = 'blk'
        self._flt = 'IS_FLAT'   


    def check_sheet(self):
        # mstclk position
        rclkdata = {}
        rcnetdata = {}
        xrcnetdata = {}
        for kw, vl in self._clkdata.items():
            if re.search(r'TMCLKDEF_Row', kw):
                rclkdata[kw] = vl
            if re.search(r'TMCLKCNET_Row', kw):
                xrcnetdata[kw] = vl
            if xrcnetdata:
                for k, v in xrcnetdata.items():
                    tmpcnet = {}
                    for n, m in v.items():
                        if not n is None:
                            tmpcnet[n] = m
                    rcnetdata[k] = tmpcnet
        if rclkdata:
            self.get_clkdata_by_clkname(rclkdata)
            # print('chkkkclknmlst: ',self._clknmlst,self._clknmdata)

            # not cover mstclk from crg/ip/hd
            crtclk = [x for x in self._clknmlst if self.is_crtclk(x)]
            genclk = [x for x in self._clknmlst if self.is_genclk(x)]
            xgclklst = []
            for clk in genclk:
                idx = self._clknmlst.index(clk)
                nclklst = self._clknmlst[0:idx]
                mclk = self._clknmdata[clk][4]
                # print('idnclklstx: ',idx,mclk,nclklst)
                if mclk not in nclklst:
                    sdc_error(f'Mstclk {mclk} of {clk} NOT from defined clk before.', 'chk_sht')
                    xgclklst.append(clk)

            # dvfs number bet mstclk and genclk
            crtdvfs = [x for x in self._clknmlst if self.is_dvfsclk(x) and self.is_crtclk(x)]
            gendvfs = [x for x in self._clknmlst if self.is_dvfsclk(x) and self.is_genclk(x) and x not in xgclklst]
            if gendvfs:
                for gclk in gendvfs:
                    dnum = len(self._clknmdata[gclk][3].split('|'))
                    mclk = self._clknmdata[gclk][4]
                    pin = self._clknmdata[gclk][5]
                    if '|' in pin:
                        pnum = len(pin.split('|'))
                        if dnum != pnum:
                            sdc_error(
                                f'Dvfs number of genClk "{gclk}" is NOT same  bet divedge {dnum} with pinport {pnum}.',
                                'chk_sht')
                    if self.is_crtclk(mclk):
                        if '|' not in self._clknmdata[mclk][1]:
                            sdc_error(
                                f'Dvfs number of MstClk "{mclk}" with 1 is NOT same as genclk "{gclk}" with {dnum}.',
                                'chk_sht')
                        else:
                            fnum = len(self._clknmdata[mclk][1].split('|'))
                            if dnum != fnum:
                                sdc_error(
                                    f'Dvfs number of MstClk "{mclk}" with {fnum} is NOT same as genclk "{gclk}" with {dnum}.',
                                    'chk_sht')

            if crtdvfs:
                for rclk in crtdvfs:
                    if '|' not in self._clknmdata[rclk][1]:
                        sdc_error(f'Crtclk {mclk} is NOT dvfs clk.')
                    else:
                        freq = self._clknmdata[rclk][1].split('|')
                        if self._clknmdata[rclk][2]:
                            if '|' in self._clknmdata[rclk][2]:
                                wav = self._clknmdata[rclk][2].split('|')
                                if len(freq) != len(wav):
                                    sdc_error(
                                        f'Crtclk "{rclk}" freq number {len(freq)} is NOT same as wav number {wav}.',
                                        'chk_sht')

            xgclks = [x for x in self._clknmlst if
                      not self.is_inps_crtclk(clk) and self.is_pll_crtclk(x) and not self.is_genclk(x)]
            if xgclks:
                xclkpin = [self._clknmdata[x][5] for x in xgclks]
                if rcnetdata:
                    rclkpin = []
                    for k, v in rcnetdata.items():
                        for n, t in v.items():
                            if n == self._mdname:
                                rclkpin.append(t)
                    rclkpin = [x for x in rclkpin if x]
                    if rclkpin:
                        rxdif = [x for x in xclkpin if x not in rclkpin]
                        if rxdif:
                            for t in rxdif:
                                ck = [x for x in self._clknmlst if self._clknmdata[x][5] == t]
                                # print('cksdssx: ',ck,rxdif,rclkpin,xclkpin)
                                if self.is_pll_crtclk(ck[0]):
                                    sdc_warn(
                                        f'{ck[0]} definition point {t} can not find connection with crg or ip or other harden blk.')
                                else:
                                    sdc_error(
                                        f'{ck[0]} definition point {t} can not find connection with crg or ip or other harden blk.')

            # ClkGrp	Freq	WaveForm	DivEdge	MstClk	PortPin	ClkIntg	Vol	Comment
            if crtclk:
                for sclk in self._clknmlst:
                    # clkgrp
                    clkgrp = self._clknmdata[sclk][0]
                    if not clkgrp:
                        sdc_error(f'{sclk} does NOT set ClkGrp.', 'chk_sht')
                    else:
                        if not re.search(r'CGP\d+', clkgrp):
                            sdc_error(f'{sclk} ClkGrp does NOT match "CGP\\d+" format.', 'chk_sht')
                    # freq
                    freq = self._clknmdata[sclk][1]
                    if not freq:
                        if self.is_crtclk(sclk):
                            sdc_error(f'{sclk} does NOT set Freq')
                    else:
                        if not re.search(r'[0-9]+M[0-9]?', freq):
                            sdc_error(f'{sclk} Freq does NOT match "[0-9]+M[0-9]?" format.', 'chk_sht')
                    # wave
                    wave = self._clknmdata[sclk][2]
                    if self.is_crtclk(sclk) and wave:
                        if not re.search(r'\{[0-9.]+\s+[0-9.]+\}', wave):
                            sdc_error(f'{sclk} waveform does NOT match "[0-9]+\\s+[0-9]+" format.', 'chk_sht')
                    # div /mst
                    # gaclks = [x for x in self._clknmlst if not self.is_crtclk(x)]
                    div = self._clknmdata[gclk][3]
                    mclk = self._clknmdata[gclk][4]
                    if not self.is_crtclk(sclk):
                        if not div:
                            sdc_error(f'{sclk} does NOT set DivEdge.', 'chk_sht')
                        else:
                            if not re.search(r'\d+|\d+ inv|\d+\/\d+|comb|\{[0-9]+\s+[0-9]+\s+[0-9]+\}', div):
                                sdc_error(f'{sclk} divedge does NOT match reference format.', 'chk_sht')
                        # if mclk:
                        #     if not re.search(r'NAME_|[a-z]+[0-9]+')
                    # cmt
                    cmt = self._clknmdata[sclk][7]
                    if cmt:
                        if not re.search(
                                r'PLL_CRT|PLL_GEN|TOSYS_OUT|TOTOP_OUT|TOPAD_OUT|PHYGRP_[A-Z]+_\d+|LOGGRP_[A-Z]+_\d+|CDC',
                                cmt):
                            sdc_error(f'{sclk} comment does NOT match reference value.', 'chk_sht')

    def is_crtclk(self,clk):
        divedge = self._clknmdata[clk][3]
        mstclk = self._clknmdata[clk][4]
        if not divedge and not mstclk:
            return True
        else:
            return False
        
    def is_inps_crtclk(self,clknmdata,clk):
        portpin = clknmdata[clk][5]
        if portpin and  portpin in self._sdcdg._vfile_data:
            portdir = self._sdcdg._vfile_data[portpin][0]
            if portdir == 'input' and self.is_crtclk(clk):
                return True
            else:
                return False        

    def is_oups_genclk(self,clknmdata,clk):
        portpin = clknmdata[clk][5]
        if portpin and portpin in self._sdcdg._vfile_data:
            portdir = self._sdcdg._vfile_data[portpin][0]
            if portdir == 'output' and self.is_genclk(clk):
                return True
            else:
                return False
           
    def has_wave(self,clk):
        wave = self._clknmdata[clk][2]
        if wave:
            return True
        else:
            return False

    # include all of genclk which pll genclk, out genclk, internal genclk
    def is_genclk(self,clk):
        divedge = self._clknmdata[clk][3]
        mstclk = self._clknmdata[clk][4]
        if divedge and mstclk:
            return True
        else:
            return False

    def is_pllclk(self,clk):
        cmt = self._clknmdata[clk][8]
        if re.match(r'.+(_pll_crt|_pll_gen)$', clk):
            return True
        elif cmt:
            if 'PLL' in cmt:
                return True
        else:
            return False
                    
    def is_pll_crtclk(self,clk):
        cmt = self._clknmdata[clk][8]
        if re.match(r'.+(_pll_crt)$', clk) and self.is_crtclk(clk):
            return True
        elif cmt:
            if 'PLL_CRT' in cmt and self.is_crtclk(clk):
                return True
        else:
            return False
        
    def is_pll_genclk(self,clk):
        cmt = self._clknmdata[clk][8]
        if re.match(r'.+(_pll_gen)$', clk) and self.is_genclk(clk):
            return True
        elif cmt:
            if 'PLL_GEN' in cmt and self.is_genclk(clk):
             return True
        else:
            return False      

    def is_virclk(self,clk):
        portpin = self._clknmdata[clk][5]
        if re.match(r'.+(_vir_crt|_virtual_crt)$', clk):
            return True
        elif not portpin:
            return True
        else:
            return False
        
    def is_dvfsclk(self,clk):
        divedge = self._clknmdata[clk][3]
        freq = self._clknmdata[clk][1]

        if divedge:
            if '|' in divedge:
                return True
        elif freq:
            if '|' in freq:
                return True
        else:
            return False
                
    def is_outclk(self,clk):
        cmt = self._clknmdata[clk][8]

        if re.match(r'.+(_topad_out|_totop_out|_tosys_out)$', clk):
            return True
        elif cmt:
            if re.search(r'TOTOP|TOPAD|TOSYS',cmt):
                return True
        else:
            return False


    def chg_divedge_format(self,div):
        if re.match(r'^\d+$',div):
            rdivedge = f'-divide_by {div}'
        elif re.match(r'^\d+\s+inv$',div):
            dv = div.split(' ')[0]
            rdivedge = f'-divide_by {dv} -invert'
        elif re.match(r'^\{\d+(\s+\d+)*\}',div):
            rdivedge = f'-edges {div}'
        elif re.match(r'^comb',div):
            rdivedge = f'-divide_by 1 -combinational'
        elif re.match(r'^\d+/\d+',div):
            mul = div.split('/')[1]
            rdivedge = f'-multiply_by {mul}'

        return rdivedge

    def get_srcclk(self,tclkdata,genclk,srclst=[]):
        srclst.append(genclk)
        #divedge = self._clknmdata[genclk][3]
        #mstclk = self._clknmdata[genclk][4]
        if genclk in tclkdata:
            divedge = tclkdata[genclk][3]
            mstclk = tclkdata[genclk][4]
            if divedge and mstclk:
                self.get_srcclk(tclkdata,mstclk,srclst)
        # else:
        #     sdc_info(f'Not Found {genclk} in current hdhier clk list.')

        return srclst[::-1]

    def cal_genclk_div(self,div):
        ndiv = ''
        if re.match(r'\d+$',div):
            ndiv = f'{div}d'
        if re.match(r'^1/\d+$',div):
            ndiv = div.split('/')[1] + 'm'
        if re.match(r'^\{(\d+)(\s+\d+)*\}$',div):
            nums = re.sub(r'\D','',div)
            ndiv = str(int((int(nums[2]) - int(nums[0]))/2)) + 'd'
        if re.match(r'^\d+\s+inv$',div):
            ndiv = div.split(' ')[0] + 'd'
        if div == 'comb':
            ndiv = '1d'

        return ndiv

    # besides clkdef, and also cover all of genclk from crg/ip outclk
    def cal_genclk_div_freq(self,genclk):
        # crgip_clknm = self.get_clkinfo_from_crgip('0','GEN')
        # crgip_clkperd = self.get_clkinfo_from_crgip('4','GEN')
        
        # ciclknm = [vl for x in crgip_clknm for ky,vl in x.items()]
        # ciclkperd = [vl for x in crgip_clkperd for ky,vl in x.items()]
        # for key in crgip_clknm.keys():
        tclkdata = {}
        tclklst = []
        tclkdata,tclklst = self.concat_clkdef_crgiphd_gen_outclk()
        #print('cal_genclk_div_freq:tclkdata',tclkdata)

        clknmlst = self.get_srcclk(tclkdata,genclk,srclst=[])
        #print('cal_genclk_div_freq:clknmlst',genclk,clknmlst,tclkdata)
        if clknmlst:
            #freq = self._clknmdata[clknmlst[0]][1].strip()
            if tclkdata[clknmlst[-2]][1]:
                freq = tclkdata[clknmlst[-2]][1].strip()
            else:
                freq = tclkdata[clknmlst[0]][1].strip()
            #print('cal_genclk_div_freq:freq:clknmlst',freq,clknmlst[-2])
            freq = freq.replace('CYCLE','')
            cfreq = ''
            if '|' in freq:
                for frq in freq.split('|'):
                    if re.match(r'^\d+M\d+$',frq):
                        cfreq += frq.replace('M','.') + ' '
                    if re.match(r'^\d+M$',frq):
                        cfreq += frq.replace('M','') + ' '
            else:
                if re.match(r'^\d+M\d+$',freq):
                    cfreq = freq.replace('M','.')
                if re.match(r'^\d+M$',freq):
                    cfreq = freq.replace('M','')   
        else:
            sdc_error(f'Can not find master clock of {genclk}')   

        gfreq = []
        gdiv = []
        mdiv = []
        
        for clknm in clknmlst[1:]:
            #print(clknm)
            #ndiv = self._clknmdata[clknm][3] 
            ndiv = tclkdata[clknm][3]
            if '|' in ndiv:
                pdiv = []
                for div in ndiv.split('|'):
                    pdiv.append(self.cal_genclk_div(div.strip()))
                mdiv.append(pdiv)
            else:
                mdiv.append(self.cal_genclk_div(ndiv.strip()))
            #print('clkfreq',clknm,mdiv)
        
        if isinstance(mdiv[0],str):
            ffreq = ''
            fdiv = '1'
            for dv in mdiv:
                flg = ''
                if 'd' in dv:
                    fdiv = str(int(fdiv) * int(dv.replace('d',''))) 
                    flg = 'div'
                if 'm' in dv:
                    fdiv = str(int(fdiv) * int(dv.replace('m','')))  
                    flg = 'mul' 
            if flg == 'div':
                ffreq = str(round(float(cfreq) / int(fdiv),2))
            if flg == 'mul':
                fdiv = '1/' + fdiv
                ffreq = str(round(float(cfreq) * int(fdiv),2))
            #print(fdiv,ffreq)
            #print('oooooop')

            gdiv = list(fdiv)
            if '.0' in ffreq:
                ffreq = ffreq.replace('.0','M')
            elif '.00' in ffreq:
                ffreq = ffreq.replace('.00','M')
            else:
                ffreq = ffreq.replace('.','M')
            gfreq = f'CYCLE{ffreq}'.split()
            
        if isinstance(mdiv[0],list):
            ffreq = ''            
            fdivg = ''
            f1 = []
            f2 = []
            if len(mdiv) == 1:
                f1 = mdiv[0]
                f2 = [[x] for x in f1]
            else:               
                for i in range(0,len(mdiv[0])):
                    f1 = []
                    for j in range(0,len(mdiv)):
                        f1.append(mdiv[j][i])
                    f2.append(f1)
            #print(f1,f2)
            #print('etewqghhg')

            for sdv,feq in zip(f2,cfreq.split(' ')):
                fdiv = '1'
                for odv in sdv:
                    flg = ''                    
                    if 'd' in odv:
                        fdiv = str(int(fdiv) * int(odv.replace('d','')))
                        flg = 'div'
                    if 'm' in odv:
                        fdiv = str(int(fdiv) * int(odv.replace('m',''))) 
                        flg = 'mul' 
                #print('cal_genclk_div_freqfdiv:',cfreq,f2,fdiv)   
                #print('cal_genclk_div_freqfeq:',genclk,feq)              
                if flg == 'div':
                    fdivg += fdiv + ' '                   
                    ffreq += str(round(float(feq) / int(fdiv),2)) + ' ' 
                if flg == 'mul':
                    fdivg += '1/' + fdiv + ' '
                    ffreq += str(round(float(feq) * int(fdiv),2)) + ' '
                #print(fdivg,ffreq)

            gdiv = [x for x in fdivg.strip().split(' ')]  
            for tfq in ffreq.strip().split(' '):
                if '.0' in tfq:
                    tfq = tfq.replace('.0','M')
                elif '.00' in tfq:
                    tfq = tfq.replace('.00','M')
                else:
                    tfq = tfq.replace('.','M')
                gfreq.append(f'CYCLE{tfq}')   

        #print(gdiv,gfreq)
        return gfreq, gdiv

    # ClkNm	ClkGrp	Freq	WaveForm	DivEdge	MstClk	PortPin	ClkIntg	Vol	Comment
    def get_clkdata_by_clkname(self,clkdata):
        #clkkeys = list(clkdata.keys())
        #if not self._clknmlst:
        clknmlst = []
        if clkdata:
            #clknmlst = [v['ClkNm'] for k,v in clkdata.items()]
            ncrg = {x:x.split('Row')[-1] for x,y in clkdata.items()}
            xcrg = sorted(ncrg.items(), key = lambda x:int(x[1]))           
            clst = [k for k,_ in xcrg]
            #print('get_clkdata_by_clkname++++++++++++++++++++++++:',ncrg,xcrg,clst)
            clknmlst = [clkdata[k]['ClkNm'] for k in clst]

            for key,clkrow in clkdata.items():
                kwlst = ['ClkGrp','Freq','WaveForm','DivEdge','MstClk',	'PortPin','ClkIntg','Vol','Comment']
                vallst = []

                for kw in kwlst:
                    #vallst.append(clkrow[kw])
                    if kw == 'MstClk':
                        mclk = clkrow['MstClk']

                        # hblk = self._hiertree.get_block_by_name(self._mdname)
                        # hdblks = hblk.get_curhd_by_name() 
                        # hdals = [] 
                        # for blk in  hdblks:
                        #     bk = self._hiertree.get_block_by_name(blk)
                        #     hdals(bk.alias)
                        if mclk:
                            if not mclk in clknmlst:
                                if ' ' in mclk.strip():
                                    vallst.append(mclk.strip())
                                else:
                                    sp = mclk.split('_')
                                    if 'NAME_' in mclk:
                                        val = sp[1]
                                    else:
                                        val = sp[0]
                                    var = self.get_als_var(self._mdname,val)
                                    #print('get_clkdata_by_clkname:mclk',val,var)
                                    vallst.append(f'{var} {mclk.strip()}')
                            else:
                                vallst.append(mclk)
                        else:
                            vallst.append(mclk)
                    else:
                        vallst.append(clkrow[kw])
        
                self._clknmdata[clkrow['ClkNm']] = vallst     
                #self._clknmlst.append(clkrow['ClkNm'])
            self._clknmlst = clknmlst
           

    def read_data(self,kwd=''):
        sheet = self.get_sheet()
        self._clkdata = self.get_table_contxt(sheet)

class IODlySheet(BaseSheet):
    def __init__(self,*args):
        super().__init__(*args)  
        self._iodlydata = {}  
        self._hiertree = self._sdcdg._hier_tree
        #self._valdata = self._sdcdg._sheets['VarDef']._valdata
        self._vardata = self.get_vardef_value(self._sdcdg._wb['VarDef'])
        
        #self._sdcdir = self._sdcdg._sdcdir
        self._mdname = self._sdcdg._mdname

        #self._clkdef = None
        self._vfdata = self._sdcdg._vfile_data
        self._hiertree = self._sdcdg._hier_tree   


    ######################################################
    def read_data(self):
        sheet = self.get_sheet()
        self._iodlydata = self.get_table_contxt(sheet)
        #print(self._iodlydata)
        
    def check_sheet(self):
        clkdef = self._sdcdg._sheets['ClkDef']
        clkdef.get_clkdata_by_clkname(clkdef._clkdata)
        if self._iodlydata:
            # port consistency
            #vfile_ports = [f'{k}{v[1]}' for k,v in self._sdcdg._vfile_data.items() if not re.search(r'FP|CASE|ANA|DFT|IDEAL|TCLK',v[2])]           
            vfile_ports = []
            for k,v in self._sdcdg._vfile_data.items():
                # if not re.search(r'FP|CASE|ANA|DFT|IDEAL|TCLK',v[2]):
                if v[1] == '1':
                    vfile_ports.append(k)
                else:
                    vfile_ports.append(f'{k}{v[1]}')
            
            # print('vfile_ports: ',vfile_ports)
            # print('self._iodlydata: ',self._iodlydata)
            iodly_ports = [v['PortNm'] for k,v in self._iodlydata.items()]
            if iodly_ports and vfile_ports:
                for iopt in iodly_ports:
                    if iopt not in vfile_ports:
                        sdc_error(f'IODly {iopt} NOT in empty vfile but in iodly table.','chk_sht')

            crgipclks = {}
            if clkdef._iptcrgdata:  crgipclks.update(clkdef._iptcrgdata)
            if clkdef._iptipdata:  crgipclks.update(clkdef._iptipdata)
            ciclks = []
            if crgipclks:
                for k,v in crgipclks.items():
                    for n,t in v.items():
                        if 'OUT' in n:
                            ciclks.append(t[0])
            aclks = []
            if clkdef._clknmlst:    aclks.extend(clkdef._clknmlst)
            # print('aclks: ',aclks)
            if ciclks:  aclks.extend(ciclks)
            iodly_clks = [v['ClkNm'] for k,v in self._iodlydata.items()]
            if iodly_clks:
                for ioclk in list(set(iodly_clks)):
                    # maybe hdout clk ???
                    if ioclk not in aclks:
                        sdc_error(f'IODly Related clock "{ioclk}" NOT found from clkdef/crg/ip.','chk_sht')


            for kw,vl in self._iodlydata.items():
                num = kw.split('Row')[1]

                # PortNm	Direction	ClkNm	ClkFall	DlyMax	DlyMin	Vol	Comment
                ptnm = vl['PortNm']
                if ptnm:
                    if not re.search(r'\w+\[\d+:\d+\]+|\w+',ptnm):
                        sdc_error(f'IODly line{num} PortNm "{ptnm}" does NOT follow reference format.','chk_sht')
                dirc = vl['Direction']
                if dirc:
                    if not re.search(r'input|output',dirc):
                        sdc_error(f'IODly line{num} Direction "{dirc}" does NOT follow "input|output" format.','chk_sht')
                dmax = vl['DlyMax']
                if dmax:
                    if not re.search(r'IO_DLY_MAX|[-]?\d+%|[-]?\d+[.0-9]*#|[-]?0[.0-9]*|[-]?\[expr.*\]',dmax):
                        sdc_error(f'IODly line{num} DlyMax "{dmax}" does NOT follow reference format.','chk_sht')
                dmin = vl['DlyMin']
                if dmin:
                    if not re.search(r'IO_DLY_MIN|[-]?\d+[.0-9]*%|[-]?\d+[.0-9]*#|[-]?0[.0-9]*|[-]?\[expr.*\]',dmin):
                        sdc_error(f'IODly line{num} DlyMin "{dmin}" does NOT follow reference format.','chk_sht')
        else:
            sdc_info(f'{self._mdname} has NO IO delay constraints.','chk_sht')
    
    
    def change_sheet(self):
        pass

    def dump_json(self,json_file):
        self._data = self._iodlydata
        self.write_json(json_file)


class ExpSheet(BaseSheet):
    def __init__(self ,*args):
        super().__init__(*args)
        self._expdata = {}
        self._hier_tree = self._sdcdg._hier_tree
        # self._clkdef = None
        self._vardata = self.get_vardef_value(self._sdcdg._wb['VarDef'])

        self._lvl = 'blk'
        self._flt = 'IS_FLAT'

        self._mdname = ''
        self._intexpdata = {}
        self._ioexpdata = {}


    ######################################################
    def read_data(self):
        sheet = self.get_sheet()
        self._expdata = self.get_table_contxt(sheet)

    def check_sheet(self):
        for kw, vl in self._expdata.items():
            # num = kw.split('Row')[1]
            if re.search(r'TMINTEXP_Row|TMSTPGATE_Row', kw):
                # FP	MCP	CaseVal	CasePin	From	Through	To	Comment
                # StopClk	StopPin	DisClkGating	Comment
                self._intexpdata[kw] = vl
            if re.search(r'TMIOEXP_Row|TMINOUT_Row', kw):
                # PortNm	Direction	Ideal	CaseVal	FP	MCP	From	Through	To	Comment
                # PortIn	PortOut	DlyIn	DlyOut	RealDly	ClkNm	Vol	Comment
                self._ioexpdata[kw] = vl

        # IOEXP
        clkdef = self._sdcdg._sheets['ClkDef']
        crgipclks = {}
        if clkdef._iptcrgdata:  crgipclks.update(clkdef._iptcrgdata)
        if clkdef._iptipdata:  crgipclks.update(clkdef._iptipdata)
        ciclks = []
        if crgipclks:
            for k,v in crgipclks.items():
                for n,t in v.items():
                    if 'OUT' in n:
                        ciclks.append(t[0])
        aclks = []
        if clkdef._clknmlst:    aclks.extend(clkdef._clknmlst)
        if ciclks:  aclks.extend(ciclks)

        #vfile_ports = [f'{k}{v[1]}' for k,v in self._sdcdg._vfile_data.items() if re.search(r'FP|CASE|MCP',v[2])]
        vfile_ports = []
        for k,v in self._sdcdg._vfile_data.items():
            if re.search(r'FP|CASE|MCP|IDEAL',v[2]):
                if v[1] == '1':
                    vfile_ports.append(k)
                else:
                    vfile_ports.append(f'{k}{v[1]}')
        ioexp_ports = [v['PortNm'] for k,v in self._ioexpdata.items() if re.search(r'TMIOEXP_Row',k)]
        if ioexp_ports and vfile_ports:
            for iopt in ioexp_ports:
                if iopt not in vfile_ports:
                    sdc_error(f'IOEXP "{iopt}" NOT in empty vfile but in ioexp table. Maybe not set "FP|CASE|MCP|IDEAL".','chk_sht')

        if self._ioexpdata:
            for kw,vl in self._ioexpdata.items():
                num = kw.split('Row')[1]
                if re.search(r'TMIOEXP_Row',kw):
                    # FP	MCP	CaseVal	CasePin	From	Through	To	Comment  
                    dirc = vl['Direction']
                    if dirc:
                        if not re.search(r'input|output',dirc):
                            sdc_error(f'IOExp line{num} Direction "{dirc}" does NOT follow "input|output" format.','chk_sht')
                    fp = vl['FP']
                    if fp:
                        if not re.search(r'all|setup|hold',fp):
                            sdc_error(f'IOExp line{num} FP "{fp}" does NOT follow "all|setup|hold" format.','chk_sht')
                    mcp = vl['MCP']
                    if mcp:
                        if not re.search(r'start \d+ \d+|end \d+ \d+|start NA \d+|end NA \d+|start \d+ NA|end \d+ NA',mcp):
                            sdc_error(f'IOExp line{num} MCP "{mcp}" does NOT follow reference format.','chk_sht')
                    casval = vl['CaseVal']
                    if casval:
                        if not re.search(r'0|1',casval):
                            sdc_error(f'IOExp line{num} CaseVal "{casval}" does NOT match "0|1" value.','chk_sht')

                    fr = vl['From']
                    if fr:
                        if not re.search(r'pin \[list.*\]|clk \[list.*\]',fr):
                            sdc_error(f'IOExp line{num} From "{fr}" does NOT follow reference format.','chk_sht')
                    thr = vl['Through']
                    if thr:
                        if not re.search(r'pin \[list.*\]',thr):
                            sdc_error(f'IOExp line{num} Through "{thr}" does NOT follow reference format.','chk_sht')
                    to = vl['To']
                    if to:
                        if not re.search(r'pin \[list.*\]|clk \[list.*\]',to):
                            sdc_error(f'IOExp line{num} To "{to}" does NOT follow reference format.','chk_sht')
                    
                    if fp and mcp:
                        sdc_error(f'IOExp line{num} FP and MCP cannot be set at the same time.','chk_sht')
                    if fp and casval:
                        sdc_error(f'IOExp line{num} FP and CaseVal cannot be set at the same time.','chk_sht')
                    if mcp and casval:
                        sdc_error(f'IOExp line{num} MCP and CaseVal cannot be set at the same time.','chk_sht')

                if re.search(r'TMINOUT_Row',kw):
                    #PortIn	PortOut	DlyIn	DlyOut	RealDly	ClkNm	Vol	Comment
                    ptin = vl['PortIn']
                    if ptin:
                        if ptin not in vfile_ports:
                            sdc_error(f'TMINOUT line{num} PortIn "{ptin}" NOT in empty vfile but in inout table','chk_sht')
                    ptout = vl['PortOut']
                    if ptout:
                        if ptout not in vfile_ports:
                            sdc_error(f'TMINOUT line{num} PortOut "{ptout}" NOT in empty vfile but in inout table','chk_sht')
                    dlin = vl['DlyIn']
                    if dlin:
                        if not re.search(r'IO_DLY_MAX|IO_DLY_MIN|[-]?\d+[.0-9]*%|[-]?\d+[.0-9]*#|[-]?0\.[0-9]*|[-]?\[expr.*\]#',dlin):
                            sdc_error(f'TMINOUT line{num} DlyIn "{dlin}" does NOT follow reference format.','chk_sht')
                    dlout = vl['DlyOut']
                    if dlout:
                        if not re.search(r'IO_DLY_MAX|IO_DLY_MIN|[-]?\d+[.0-9]*%|[-]?\d+[.0-9]*#|[-]?0\.[0-9]*|[-]?\[expr.*\]#',dlout):
                            sdc_error(f'TMINOUT line{num} DlyOut "{dlout}" does NOT follow reference format.','chk_sht')
                    rdly = vl['RealDly']
                    if rdly:
                        if not re.search(r'[-]?\d+[.0-9#]*|[-]?\[expr.*\][#]?',rdly):
                            sdc_error(f'TMINOUT line{num} RealDly "{rdly}" does NOT follow reference format.','chk_sht')
                    else:
                        sdc_error(f'TMINOUT line{num} RealDly "{rdly}" NOT found.','chk_sht')
                    clknm = vl['ClkNm']
                    # maybe hdout clk ???
                    if clknm not in aclks:
                        sdc_warn(f'TMINOUT line{num} ClkNm "{clknm}" NOT found from clkdef/crg/ip.','chk_sht')

        else:
            sdc_info(f'{self._mdname} has NO io exception.','chk_sht')
    
        # INTEXP
        if self._intexpdata:
            for kw,vl in self._intexpdata.items():
                num = kw.split('Row')[1]
                if re.search(r'TMINTEXP_Row',kw):
                    # FP	MCP	CaseVal	CasePin	From	Through	To	Comment  
                    fp = vl['FP']
                    if fp:
                        if not re.search(r'all|setup|hold',fp):
                            sdc_error(f'IntExp line{num} FP "{fp}" does NOT follow "all|setup|hold" format.','chk_sht')
                    mcp = vl['MCP']
                    if mcp:
                        if not re.search(r'start \d+ \d+|end \d+ \d+|start NA \d+|end NA \d+|start \d+ NA|end \d+ NA',mcp):
                            sdc_error(f'IntExp line{num} MCP "{mcp}" does NOT follow reference format.','chk_sht')

                    casval = vl['CaseVal']
                    if casval:
                        if not re.search(r'0|1',casval):
                            sdc_error(f'IntExp line{num} CaseVal "{casval}" does NOT match "0|1" value.','chk_sht')
                    caspin = vl['CasePin']
                    if caspin:
                        if not re.search(r'pin \[list.*\]',caspin):
                            sdc_error(f'IntExp line{num} CasePin "{caspin}" does NOT follow reference format.','chk_sht')
                    if casval and not caspin:
                        sdc_error(f'IntExp line{num} CaseVal "{casval}" does NOT find CasePin value.','chk_sht')
                    if not casval and caspin:
                        sdc_error(f'IntExp line{num} CasePin "{caspin}" does NOT find CaseVal value.','chk_sht')

                    fr = vl['From']
                    if fr:
                        if not re.search(r'pin \[list.*\]|clk \[list.*\]',fr):
                            sdc_error(f'IntExp line{num} From "{fr}" does NOT follow reference format.','chk_sht')
                    thr = vl['Through']
                    if thr:
                        if not re.search(r'pin \[list.*\]',thr):
                            sdc_error(f'IntExp line{num} "{thr}" Through does NOT follow reference format.','chk_sht')
                    to = vl['To']
                    if to:
                        if not re.search(r'pin \[list.*\]|clk \[list.*\]',to):
                            sdc_error(f'IntExp line{num} To "{to}" does NOT follow reference format.','chk_sht')

                    if fp and mcp:
                        sdc_error(f'IOExp line{num} FP and MCP cannot be set at the same time.','chk_sht')
                    if fp and casval:
                        sdc_error(f'IOExp line{num} FP and CaseVal cannot be set at the same time.','chk_sht')
                    if mcp and casval:
                        sdc_error(f'IOExp line{num} MCP and CaseVal cannot be set at the same time.','chk_sht')

                if re.search(r'TMSTPGATE_Row',kw):
                    # StopClk	StopPin	DisClkGating	Comment
                    stpclk = vl['StopClk']
                    if stpclk:
                        if not re.search(r'clk \[list.*\]',stpclk):
                            sdc_error(f'TMSTPGATE line{num} StopClk "{stpclk}" does NOT follow reference format.','chk_sht')
                    stpin = vl['StopPin']
                    if stpin:
                        if not re.search(r'pin \[list.*\]|clk \[list.*\]',stpin):
                            sdc_error(f'TMSTPGATE line{num} StopPin "{stpin}" does NOT follow reference format.','chk_sht')
                    dclkgt = vl['DisClkGating']
                    if dclkgt:
                        if not re.search(r'pin \[list.*\]|inst \[list.*\]',dclkgt):
                            sdc_error(f'TMSTPGATE line{num} DisClkGating "{dclkgt}" does NOT follow reference format.','chk_sht')

        else:
            sdc_info(f'{self._mdname} has NO internal exception.','chk_sht')
    

    def change_sheet(self):
        pass

        
    def read_json(self,file_path):
        sblk_clk_list = {}
        if os.path.exists(file_path):
            with open(file_path,'r') as fw:
                content = fw.read()
                sblk_clk_list = json.loads(content)

        #print('sblk_clk_list:',sblk_clk_list)
        return sblk_clk_list

    def dump_json(self ,json_file):
        self._data = self._expdata
        self.write_json(json_file)



def modify_line_in_file(file_path, search_pattern, replacement):
    # 打开文件并逐行读取内容
    with open(file_path, 'r') as file:
        lines = file.readlines()

    # 遍历每一行并进行匹配和替换
    modified_lines = []
    for line in lines:
        if search_pattern in line:
            modified_line = line.replace(search_pattern, replacement)
            modified_lines.append(modified_line)
        else:
            modified_lines.append(line)

    # 将修改后的内容写回文件
    with open(file_path, 'w') as file:
        file.writelines(modified_lines)

def rm_exist_log(logdir):
    # LOG_DIR = os.getenv('TASK_LOGS_DIR')
    # logdir = os.path.join(LOG_DIR,taskid)
    # logdir = r'E:\stone\work\smalltool\pycharm\const0203_sysflatok_full\logs\adfeicc67ere'
    rlogfiles = os.listdir(logdir)
    for logfile in rlogfiles:
        # if logfile.endswith('.log') or logfile.endswith('.rpt'):
        if logfile in ['chk_hym.rpt','chk_sht.rpt','full_chk.rpt']:
            os.remove(os.path.join(logdir,logfile))


def printlog(context, file='sdc_gen.log',logdir=''):
    if not logdir:
        LOG_DIR = os.getenv('TASK_LOGS_DIR')
        taskid = os.getenv('CURRENT_TASK_ID', 'default_task')
        if LOG_DIR:
            logdir = os.path.join(LOG_DIR, taskid)
            os.makedirs(logdir, exist_ok=True)
            rlogdir = os.path.join(logdir, file)
        else:
            # 如果环境变量不存在，使用当前目录
            rlogdir = file
    else:
        rlogdir = logdir

    if os.path.exists(rlogdir):
        with open(rlogdir, 'a') as fw:
            fw.write(context)
    else:
        with open(rlogdir, 'w') as fw:
            fw.write(context)

def movelogrpt(msgnm, flog, fdir):
    if os.path.exists(f'{msgnm}'):
        if os.path.exists(flog):
            # with open('sdc_gen.log','r') as fh:
            #     for line in fh.readlines():
            #         #txt_list.append(line.strip())
            #         txt_list += f'{line.strip()} \n'
            # with open(logfile, 'a') as fw:
            #     fw.write(txt_list)
            os.system(f'rm -f {flog}')
            os.system(f'mv {msgnm} {fdir}')
        else:
            os.system(f'mv {msgnm} {fdir}')
    else:
        sdc_warn(f'Can not find {msgnm}.')


#########################################################################################################################
# sdc message
full_log_message_list =  []
full_rpt_message_list =  []

def sdc_log(level, msg, out=sys.stdout):
    print(f'{level.upper()}: {msg}', flush=True, file=out)

def sdc_info(msg, kw='log'):
    # if msg not in full_log_message_list:
    if kw == 'log':
        full_log_message_list.append(msg)
        printlog(f'SDC_INFO: {msg} \n', 'full_msg.log')
    else:
        full_rpt_message_list.append(msg)
        printlog(f'SDC_INFO: {msg} \n', 'full_chk.rpt')
    # sdc_log('SDC_INFO', msg)
    if kw == 'log':
        printlog(f'SDC_INFO: {msg} \n')
    if kw == 'chk_hym':
        printlog(f'SDC_INFO: {msg} \n', 'chk_hym.rpt')
    if kw == 'chk_sht':
        printlog(f'SDC_INFO: {msg} \n', 'chk_sht.rpt')
    if kw == 'chk_dti':
        printlog(f'SDC_INFO: {msg} \n', 'chk_dti.rpt')
    if kw == 'chk_sdc':
        printlog(f'SDC_INFO: {msg} \n', 'chk_sdc.rpt')

def sdc_warn(msg, kw='log'):
    # if msg not in full_log_message_list:
    if kw == 'log':
        full_log_message_list.append(msg)
        printlog(f'SDC_WARN: {msg} \n', 'full_msg.log')
    else:
        full_rpt_message_list.append(msg)
        printlog(f'SDC_WARN: {msg} \n', 'full_chk.rpt')
    print(f'\033[0:31mSDC_WARN\033[0m: {msg}', flush=True)
    # print(f'SDC_WARN: {msg}', flush=True)
    if kw == 'log':
        printlog(f'SDC_WARN: {msg} \n')
    if kw == 'chk_hym':
        printlog(f'SDC_WARN: {msg} \n', 'chk_hym.rpt')
    if kw == 'chk_sht':
        printlog(f'SDC_WARN: {msg} \n', 'chk_sht.rpt')
    if kw == 'chk_dti':
        printlog(f'SDC_WARN: {msg} \n', 'chk_dti.rpt')
    if kw == 'chk_sdc':
        printlog(f'SDC_WARN: {msg} \n', 'chk_sdc.rpt')
         

def sdc_error(msg, kw='log'):
    # if msg not in full_log_message_list:
    if kw == 'log':
        full_log_message_list.append(msg)
        printlog(f'SDC_ERROR: {msg} \n', 'full_msg.log')
    else:
        full_rpt_message_list.append(msg)
        printlog(f'SDC_ERROR: {msg} \n', 'full_chk.rpt')
    print(f'\033[0:31mSDC_ERROR\033[0m: {msg}', flush=True)
    # print(f'SDC_ERROR: {msg}', flush=True)
    if kw == 'log':
        printlog(f'SDC_ERROR: {msg} \n')
    if kw == 'chk_hym':
        printlog(f'SDC_ERROR: {msg} \n', 'chk_hym.rpt')
    if kw == 'chk_sht':
        printlog(f'SDC_ERROR: {msg} \n', 'chk_sht.rpt')
    if kw == 'chk_dti':
        printlog(f'SDC_ERROR: {msg} \n', 'chk_dti.rpt')
    if kw == 'chk_sdc':
        printlog(f'SDC_ERROR: {msg} \n', 'chk_sdc.rpt')

def sdc_fatal(msg, kw='log'):
    # if msg not in full_log_message_list:
    if kw == 'log':
        full_log_message_list.append(msg)
        printlog(f'SDC_FATAL: {msg} \n', 'full_msg.log')
    else:
        full_rpt_message_list.append(msg)
        printlog(f'SDC_FATAL: {msg} \n', 'full_chk.rpt')
    print(f'\033[0:31mSDC_FATAL\033[0m: {msg}', flush=True)
    # print(f'SDC_FATAL: {msg}', flush=True)
    if kw == 'log':
        printlog(f'SDC_FATAL: {msg} \n')
    if kw == 'chk_hym':
        printlog(f'SDC_FATAL: {msg} \n', 'chk_hym.rpt')
    if kw == 'chk_sht':
        printlog(f'SDC_FATAL: {msg} \n', 'chk_sht.rpt')
    if kw == 'chk_dti':
        printlog(f'SDC_FATAL: {msg} \n', 'chk_dti.rpt')
    if kw == 'chk_sdc':
        printlog(f'SDC_FATAL: {msg} \n', 'chk_sdc.rpt')
           
    # sys.exit(1)

def sdc_args():
    print(f'##++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++')
    print(f'## SDC Generation Optional Arguments Presentation:')
    print(f'## -hlp:        All of flow option usage reference.')
    print(f'## -usg:        Flow step command usage reference.')
    print(f'## -tmp:        Write out template SDCs for reference')
    print(f'## -gen_dir:    SDC generationg directory')
    print(f'## -hier_yaml:  Design hierarchy info. from user input file')
    print(f'## -setup:      Build setup directories from blocks defined in hier yaml')
    print(f'## -blocks:     Indicate only current design name for SDC generation')
    print(f'## -flat:       Generate flatten SDC based on current design. If not flat, only generate current design level only SDC')
    print(f'## -dg:         Generate design guide file to be present SDC request format')
    print(f'## -idg:        Incrementally update design guide file based on existed dg file and updated input files')
    print(f'## -sdc:        Write out SDC files')
    print(f'## -check_hym: Check hier yaml data before generating SDC')
    print(f'## -check_only: Check input data before generating SDC')
    print(f'## -check_sdc:  Check SDC consistency after generating SDC')
    print(f'## -proj:       Open project mode. Maybe need set some related project environment variables')
    print(f'##++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++')


def sdc_usage():
    print(f'##++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++')
    print(f'## Before flow run, user must provide input data including hier yaml and empty vfile, dg file and crg/ip SDCs to be integrated ')
    print(f'## For hier yaml and empty vfile, must follow format of hier_pwr.yaml and tmempty.v in template/ftemp')
    print(f'## For crg/ip SDCs, must follow header format of crg.sdc and userip.sdc in template/ftemp')
    print(f'##++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++')
    print(f'##++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++')
    print(f'## Step I: Build the whole SDC generation directory structure.')
    print(f'## Cmd Line: xconst sdgen -gen_dir <sdcdir> -hier_yaml <hier_file> -setup -blocks <blk_name> [-tmp]')
    print(f'##++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++')
    print(f'##++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++')
    print(f'## Step II: Generate initial design guide file according to current input data.')
    print(f'## Cmd Line: xconst sdgen -gen_dir <sdcdir> -hier_yaml <hier_file> -dg -blocks <blk_name>')
    print(f'##++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++')
    print(f'##++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++')
    print(f'## Step III: Check and debug input data after users provide all of input files.')
    print(f'## Cmd  Line: xconst sdgen -gen_dir <sdcdir> -hier_yaml <hier_file> -chk_only -blocks <blk_name>')
    print(f'##++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++')
    print(f'##++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++')
    print(f'## Step IV: Generate SDC files if all of input files are OK.')
    print(f'## Cmd  Line: xconst sdgen -gen_dir <sdcdir> -hier_yaml <hier_file> -sdc -blocks <blk_name> [-proj] [-flat]')
    print(f'##++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++')
    print(f'##++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++')
    print(f'## Step V: Check and debug SDC output files.')
    print(f'## Cmd  Line: xconst sdgen -gen_dir <sdcdir> -hier_yaml <hier_file> -chk_sdc -blocks <blk_name>')
    print(f'##++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++')




###########################################################################################################################
# upf meesage
upf_message_list =  []

def upf_log(level, msg, out=sys.stdout):
    print(f'{level.upper()}: {msg}', flush=True, file=out)

def upf_info(msg,kw='log'):
    if msg not in upf_message_list:
        upf_message_list.append(msg)
        upf_log('UPF_INFO', msg)
        if kw == 'log':
            printlog(f'UPF_INFO: {msg} \n')
        if kw == 'chk_hym':
            printlog(f'UPF_INFO: {msg} \n','chk_hym.rpt')
        if kw == 'chk_sht':
            printlog(f'UPF_INFO: {msg} \n','chk_sht.rpt')
        

def upf_warn(msg,kw='log'):
    if msg not in upf_message_list:
        upf_message_list.append(msg)
        print(f'\033[0:31mUPF_WARN\033[0m: {msg}', flush=True)
        if kw == 'log':
            printlog(f'UPF_WARN: {msg} \n')
        if kw == 'chk_hym':
            printlog(f'UPF_WARN: {msg} \n','chk_hym.rpt')
        if kw == 'chk_sht':
            printlog(f'UPF_WARN: {msg} \n','chk_sht.rpt')
         

def upf_error(msg,kw='log'):
    if msg not in upf_message_list:
        upf_message_list.append(msg)
        print(f'\033[0:31mUPF_ERROR\033[0m: {msg}', flush=True)
        if kw == 'log':
            printlog(f'UPF_ERROR: {msg} \n')
        if kw == 'chk_hym':
            printlog(f'UPF_ERROR: {msg} \n','chk_hym.rpt')
        if kw == 'chk_sht':
            printlog(f'UPF_ERROR: {msg} \n','chk_sht.rpt')

def upf_fatal(msg,kw='log'):
    if msg not in upf_message_list:
        upf_message_list.append(msg)
        print(f'\033[0:31mUPF_FATAL\033[0m: {msg}', flush=True)  
        if kw == 'log':
            printlog(f'UPF_FATAL: {msg} \n')
        if kw == 'chk_hym':
            printlog(f'UPF_FATAL: {msg} \n','chk_hym.rpt')
        if kw == 'chk_sht':
            printlog(f'UPF_FATAL: {msg} \n','chk_sht.rpt')
           
    sys.exit(1)    

def upf_args():
    print(f'##++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++')
    print(f'## UPF Generation Optional Arguments Presentation:')
    print(f'## -hlp:        All of flow option usage reference.')
    print(f'## -usg:        Flow step command usage reference.')
    print(f'## -tmp:        Write out template UPFs for reference')
    print(f'## -gen_dir:    UPF generationg directory')
    print(f'## -hier_yaml:  Design hierarchy info. from user input file')
    print(f'## -setup:      Build setup directories from blocks defined in hier yaml')
    print(f'## -blocks:     Indicate only current design name for UPF generation')
    #print(f'## -flat:       Generate flatten UPF based on current design. If not flat, only generate current design level only UPF')
    print(f'## -dg:         Generate design guide file to be present UPF request format')
    #print(f'## -idg:        Incrementally update design guide file based on existed dg file and updated input files')
    print(f'## -upf:        Write out UPF files')
    print(f'## -check_hym: Check hier yaml data before generating UPF')
    print(f'## -check_only: Check input data before generating UPF')
    print(f'## -check_upf:  Check UPF consistency after generating UPF')
    print(f'## -proj:       Open project mode. Maybe need set some related project environment variables')
    print(f'##++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++')


def upf_usage():
    print(f'##++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++')
    print(f'## Before flow run, user must provide input data including hier yaml and empty vfile, dg file and crg/ip UPFs to be integrated ')
    print(f'## For hier yaml and empty vfile, must follow format of hier_pwr.yaml and pmempty.v in template/ftemp')
    print(f'## For crg/ip UPFs, must follow header format of crg.upf and userip.upf in template/ftemp')
    print(f'##++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++')
    print(f'##++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++')
    print(f'## Step I: Build the whole UPF generation directory structure.')
    print(f'## Cmd Line: xconst sdgen -gen_dir <upfdir> -hier_yaml <hier_file> -setup -blocks <blk_name> [-tmp]')
    print(f'##++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++')
    print(f'##++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++')
    print(f'## Step II: Generate initial design guide file according to current input data.')
    print(f'## Cmd Line: xconst sdgen -gen_dir <upfdir> -hier_yaml <hier_file> -dg -blocks <blk_name>')
    print(f'##++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++')
    print(f'##++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++')
    print(f'## Step III: Check and debug input data after users provide all of input files.')
    print(f'## Cmd  Line: xconst sdgen -gen_dir <upfdir> -hier_yaml <hier_file> -chk_only -blocks <blk_name>')
    print(f'##++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++')
    print(f'##++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++')
    print(f'## Step IV: Generate UPF files if all of input files are OK.')
    print(f'## Cmd  Line: xconst sdgen -gen_dir <upfdir> -hier_yaml <hier_file> -upf -blocks <blk_name> [-proj] [-flat]')
    print(f'##++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++')
    print(f'##++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++')
    print(f'## Step V: Check and debug UPF output files.')
    print(f'## Cmd  Line: xconst sdgen -gen_dir <upfdir> -hier_yaml <hier_file> -chk_upf -blocks <blk_name>')
    print(f'##++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++')


# ORG_COMDIR = os.getenv('ECS_TEMPLATES_DIR')
TEMP_DIR = os.getenv('TEMP_UPLOAD_DIR')
LOG_DIR = os.getenv('TASK_LOGS_DIR')
# ORG_COMDIR = r'E:\stone\work\smalltool\pycharm\const0203_sysflatok_full\templates'
# TEMP_DIR = r'E:\stone\work\smalltool\pycharm\const0203_sysflatok_full\temp'
# LOG_DIR = r'E:\stone\work\smalltool\pycharm\const0203_sysflatok_full\logs'
# adfeicc67ere


def sdc_dg_chk(*arglist):
    if len(arglist) == 0: arglist = ['-h']
    parser = argparse.ArgumentParser(prog='sdc_dg_chk', description='design constraint excel file generation script')
    parser.add_argument('-taskid', help='Task ID name', default='abc123xyesergtyjht', required='False')
    parser.add_argument('-chk', help='Write or update design guide files', action='store_true')

    args = parser.parse_args(args=arglist)
    taskid = args.taskid

    if args.chk:
        hier_yaml = os.path.join(TEMP_DIR, taskid, 'hier.yaml')
        vfile = os.path.join(TEMP_DIR, taskid, 'vlog.v')
        logdir = os.path.join(LOG_DIR,taskid)

        # tfile = os.path.join(TEMP_DIR, 'tune.sdc')
        # tfile = os.path.join(ORG_COMDIR, 'tune.sdc')

        dg_file = os.path.join(TEMP_DIR, taskid,'dcont.xlsx')
        dg_file = re.sub(r'/+', '/', dg_file)
        # lock_file = f'{TEMP_DIR}/.~lock.dcont.xlsx#'
        # lock_file = re.sub(r'/+', '/', lock_file)

        # check hier yaml existence
        if not os.path.exists(hier_yaml):
            sdc_error(f'hier yaml file not found {hier_yaml}')
            exit(1)

        # check vfile existence
        if not os.path.exists(vfile):
            sdc_error(f'Empty vfile not found {vfile}')
            exit(1)

        # check dcont existence
        if not os.path.exists(dg_file):
            sdc_error(f'Design constraint file not found {dg_file}')
            exit(1)

        # check logdir existence
        if not os.path.exists(logdir):
            sdc_error(f'logdir not found {logdir}')
            exit(1)

        rm_exist_log(logdir)
        # if os.path.exists(lock_file):
        #     sdc_fatal('dcont.xlsx is in edit mode. Please close it')

        hier_tree = HierPwrTree(hier_yaml)
        sdcdg = SDC_DG()
        sdcdg.hier_tree = hier_tree
        sdcdg.read_vfile(vfile)

        if os.path.exists(dg_file):
            sdcdg.load_design_guide(dg_file)
        else:
            sdc_error(f'Can not find {dg_file}')
            exit(1)

        mdname = sdcdg._mdname
        sdcdg.hier_tree.check_hym(mdname)
        sdcdg.check_dg()
        # sdcdg.save_workbook(dg_file)

        print(f'Design guide file {dg_file} is verified.')



if __name__ == '__main__':
    if len(sys.argv) < 2 or (len(sys.argv) > 2 and 'sdc_dg_chk' not in sys.argv[1]):
        sdc_error('Missing some parameters for SDC generation')
        locals()['sdc_dg_chk']('-h')
        exit(1)
    app_name = sys.argv[1]
    if app_name in locals():
        locals()[app_name](*sys.argv[2:])
    else:
        raise NameError(f'The application of DataBase generation {app_name} not found')





