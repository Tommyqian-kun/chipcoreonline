
#!/usr/bin/env python3


import os,sys
import yaml
import re
import json

from os.path import dirname, abspath, basename
import time

import argparse
from glob import glob

import shutil

import yaml
#from itertools import chain

import openpyxl
from openpyxl import worksheet 
from pprint import pprint 
import pandas as pd
from openpyxl.utils import get_column_letter 

import tkinter as tk

from openpyxl.styles import Border, Side, PatternFill, Font, Alignment 
from openpyxl.worksheet.datavalidation import DataValidation




# Define some attributes and various methods for every block in hier tree
class BaseBlock():
    def __init__(self,name):
        self._name = name
        self._alias = ''
        self._hdlevel = 'blk'
        self._prime_pwr = ''
        self._constr_dir = ''
        self._insts = []
        self._mac_insts = []
        self._dig_insts = []
        self._cust_insts = {}
        self._proj = False

    def __repr__(self):
        return '<%s name=%s alias=%s>' % (self.__class__.__name__, self._name, self._alias)

    @property 
    def hdlevel(self):
        return self._hdlevel

    @hdlevel.setter
    def hdlevel(self, level):
        supported_lvs = ('top', 'sys', 'blk', 'soft', 'lib', 'crg', 'pll')
        if level not in supported_lvs:
            upf_error(f'Unsupported HDLEVEL {level} of block {self._name}, should be one of {supported_lvs}')
            return
        else:
            if level == 'sys' or level == 'top':
                self._hdlevel = 'sys'
            else:
                self._hdlevel = level

    @property
    def lvl_flat(self):
        if self._hdlevel == 'sys':
            return 'IS_CHIP'
        else:
            return 'IS_FLAT'
        
    @property
    def alias(self):
        return self._alias

    @alias.setter
    def alias(self, alias):
        self._alias = alias

    @property
    def prime_pwr(self):
        return self._prime_pwr

    @prime_pwr.setter
    def prime_pwr(self, pwr):
        self._prime_pwr = pwr

    @property
    def insts(self):
        return self._insts

    @insts.setter
    def insts(self, insts):
        self._insts = insts

    @property
    def mac_insts(self):
        return self._mac_insts

    @mac_insts.setter
    def mac_insts(self, mac_insts):
        self._mac_insts = mac_insts

    @property
    def dig_insts(self):
        return self._dig_insts

    @dig_insts.setter
    def dig_insts(self, dig_insts):
        self._dig_insts = dig_insts

    @property
    def constr_dir(self):
        return self._constr_dir
    
    @constr_dir.setter
    def constr_dir(self, consdir):
        self._constr_dir = consdir

    @property
    def proj(self):
        return self._proj

    @proj.setter
    def proj(self, proj):
        self._proj = proj

    def get_curcust_by_name(self, inst_type,flg=''):
        _hier = []
        _ref = []
        #_lvl = []
        _alias = []
        _pwr = []
        _usersdc = []
        if inst_type == 'insts' and self.insts:
            for i in range(0,len(self.insts)):
                finst = self.insts[i].split(',')
                if len(finst) == 3:
                    _hier.append(finst[0].strip())
                    _ref.append(finst[1].strip())
                    _alias.append(None)
                    _pwr.append(finst[2].strip())
                    _usersdc.append(None)
                if len(finst) == 4:
                    _hier.append(finst[0].strip())
                    _ref.append(finst[1].strip())
                    if not flg:
                        _alias.append(finst[2].strip().replace('#',''))
                    else:
                        _alias.append(finst[2].strip())
                    _pwr.append(finst[3].strip())
                    _usersdc.append(None)
            self._cust_insts['insthier'] = _hier
            self._cust_insts['instref'] = _ref
            self._cust_insts['instalias'] = _alias
            self._cust_insts['instpwr'] = _pwr
            self._cust_insts['instuser'] = _usersdc
        
        if inst_type == 'mac_insts' and self.mac_insts:
            for i in range(0,len(self.mac_insts)):
                if isinstance(self.mac_insts[i],str):
                    fmac = self.mac_insts[i].split(',')  
                    if len(fmac) == 3:
                        _hier.append(fmac[0].strip())
                        _ref.append(fmac[1].strip())
                        _alias.append(None)
                        _pwr.append(fmac[2].strip())
                        _usersdc.append(None)
                    if len(fmac) == 4:
                        _hier.append(fmac[0].strip())
                        _ref.append(fmac[1].strip())
                        if not flg:
                            _alias.append(fmac[2].strip().replace('#',''))
                        else:
                            _alias.append(fmac[2].strip())
                        _pwr.append(fmac[3].strip())
                        _usersdc.append(None)
                if isinstance(self.mac_insts[i],dict):
                    fmac = ''.join(self.mac_insts[i].keys()).split(',')
                    _hier.append(fmac[0].strip())
                    _ref.append(fmac[1].strip()) # + '_USR' )
                    _alias.append(None)
                    _pwr.append(fmac[2].strip())
                    _usersdc.append(''.join(self.mac_insts[i].values()))  
            self._cust_insts['machier'] = _hier
            self._cust_insts['macref'] = _ref
            self._cust_insts['macalias'] = _alias
            self._cust_insts['macpwr'] = _pwr
            self._cust_insts['macuser'] = _usersdc

        if inst_type == 'dig_insts' and self.dig_insts:
            for i in range(0,len(self.dig_insts)):
                if isinstance(self.dig_insts[i],str):
                    fdig = self.dig_insts[i].split(',')  
                    if len(fdig) == 3:
                        _hier.append(fdig[0].strip())
                        _ref.append(fdig[1].strip())
                        _alias.append(None)
                        _pwr.append(fdig[2].strip())
                        _usersdc.append(None)
                    if len(fdig) == 4:
                        _hier.append(fdig[0].strip())
                        _ref.append(fdig[1].strip())
                        if not flg:
                            _alias.append(fdig[2].strip().replace('#',''))
                        else:
                            _alias.append(fdig[2].strip())
                        _pwr.append(fdig[3].strip())
                        _usersdc.append(None)
                if isinstance(self.dig_insts[i],dict):
                    fdig = ''.join(self.dig_insts[i].keys()).split(',')
                    _hier.append(fdig[0].strip())
                    _ref.append(fdig[1].strip()) # + '_USR')
                    _alias.append(None)
                    _pwr.append(fdig[2].strip())
                    _usersdc.append(''.join(self.dig_insts[i].values()))
            self._cust_insts['dighier'] = _hier
            self._cust_insts['digref'] = _ref
            self._cust_insts['digalias'] = _alias
            self._cust_insts['digpwr'] = _pwr
            self._cust_insts['diguser'] = _usersdc       

        return self._cust_insts

    def get_curhd_by_name(self):
        #return self.name.split() + self._cust_insts['instref']
        self.get_curcust_by_name('insts')
        if 'instref' in self._cust_insts:
            return self._cust_insts['instref']

    def get_curmac_by_name(self,flg=''):
        self.get_curcust_by_name('mac_insts',flg)
        if 'macref' in self._cust_insts:
            return self._cust_insts['macref']

    def get_curdig_by_name(self,flg=''):
        self.get_curcust_by_name('dig_insts',flg)
        if 'digref' in self._cust_insts:
            return self._cust_insts['digref']
    
    def get_curuser_by_name(self, inst_type):
        self.get_curcust_by_name('mac_insts')
        self.get_curcust_by_name('dig_insts')
        if inst_type == 'mac_insts' and 'macuser' in self._cust_insts:           
            return self._cust_insts['macuser']       
        elif inst_type == 'dig_insts'and 'diguser' in self._cust_insts:
            return self._cust_insts['diguser']
        else:
            return None



class HierPwrTree():
    def __init__(self,yaml_file):
        self.yaml_file = yaml_file
        self._blocks = {}
        self._primepwr = {}
        self._yaml_data = {}
        self._hierdata = {}
        self._pwrdata = {}
        #self._blktrees = {}
        self.build_hier_tree(yaml_file)
        

    def build_hier_tree(self, yaml_file):

        # get yaml_data
        yaml_data = {}
        if not os.path.exists(yaml_file):
            raise FileExistsError(f'{yaml_file} does not exists')
        with open(yaml_file, 'r') as fh:
            yaml_data = yaml.load(fh, yaml.FullLoader)

        if 'hier' not in yaml_data:
            print('Missing hier keyword in yaml file.')
            upf_fatal(f'Must include keyword <hier>')
        if 'pwr' not in yaml_data:
            print('Missing pwr keyword in yaml file.')
            upf_fatal(f'Must include keyword <pwr>')

        # get '_primepwr'
        for pwr_name in yaml_data['pwr'].keys():
            if yaml_data['pwr'][pwr_name]:
                self._primepwr[pwr_name] = yaml_data['pwr'][pwr_name]   

        for blk_name in yaml_data['hier'].keys():

            self._blocks[blk_name] = BaseBlock(blk_name)

            if 'alias' in yaml_data['hier'][blk_name] and yaml_data['hier'][blk_name]['alias']:
                self._blocks[blk_name].alias = yaml_data['hier'][blk_name]['alias']
            else:
                self._blocks[blk_name].alias = None

            if 'hdlevel' in yaml_data['hier'][blk_name] and yaml_data['hier'][blk_name]['hdlevel']:
                self._blocks[blk_name].hdlevel = yaml_data['hier'][blk_name]['hdlevel']
            else:
                self._blocks[blk_name].hdlevel = None            
            
            if 'prime_pwr' in yaml_data['hier'][blk_name] and yaml_data['hier'][blk_name]['prime_pwr'] in self._primepwr and yaml_data['hier'][blk_name]['prime_pwr']:
                self._blocks[blk_name].prime_pwr = yaml_data['hier'][blk_name]['prime_pwr'] + ' ' + self._primepwr[yaml_data['hier'][blk_name]['prime_pwr']]
            else:
                self._blocks[blk_name].prime_pwr = None 

            if 'constr_dir' in yaml_data['hier'][blk_name] and yaml_data['hier'][blk_name]['constr_dir']:
                self._blocks[blk_name].constr_dir = yaml_data['hier'][blk_name]['constr_dir']
            else:
                self._blocks[blk_name].constr_dir = None

            if 'insts' in yaml_data['hier'][blk_name] and yaml_data['hier'][blk_name]['insts']:
                self._blocks[blk_name].insts = yaml_data['hier'][blk_name]['insts']
            else:
                self._blocks[blk_name].insts = None

            if 'mac_insts' in yaml_data['hier'][blk_name] and yaml_data['hier'][blk_name]['mac_insts']:
                self._blocks[blk_name].mac_insts = yaml_data['hier'][blk_name]['mac_insts']
            else:
                self._blocks[blk_name].mac_insts = None

            if 'dig_insts' in yaml_data['hier'][blk_name] and yaml_data['hier'][blk_name]['dig_insts']:
                self._blocks[blk_name].dig_insts = yaml_data['hier'][blk_name]['dig_insts']
            else:
                self._blocks[blk_name].dig_insts = None

        self._hierdata = yaml_data['hier'] 
        self._pwrdata = yaml_data['pwr']
        self._yaml_data = yaml_data

    def get_block_by_name(self, name) -> BaseBlock:
        if name in self._blocks:
            return self._blocks[name]
        else:
            return None    

    def get_curblks(self,name):
        curblks = []
        
        allblks = list(self._hierdata.keys())
        if name in allblks:
            curblk = self.get_block_by_name(name)

            if curblk.get_curhd_by_name():
                curblks = [x for x in curblk.get_curhd_by_name() if x is not None]
            if curblk.get_curmac_by_name():
                curblks += [x for x in curblk.get_curmac_by_name() if x is not None]
            if curblk.get_curdig_by_name():
                curblks += [x for x in curblk.get_curdig_by_name() if x is not None]  
        else:
            upf_warn(f'{name} is not expanded in hier_pwr yaml file.')
        
        return  curblks
    
    def get_hiertrees(self, name, blktrees={}, valstyle=None) -> dict:

        curblks = self.get_curblks(name)
        #blktrees = {}
        if curblks:
            new_curblks = [elem.replace('_USR', '') if re.search(r'_USR$',elem) else elem for elem in curblks]
            blktrees[name] = new_curblks
        else:
            if valstyle:
                blktrees[name] = None   

        for blk in curblks:
            if re.search(r'_USR$',blk):
                blk = blk.replace('_USR','')
                upf_warn(f'{blk} is not expanded in hier_pwr yaml file.')
            elif len(curblks) > 0:
                self.get_hiertrees(blk,blktrees)    
        
        return blktrees

    def get_hierblks(self, name) -> list:

        blktrees = self.get_hiertrees(name)
        result = []
        for key, value in blktrees.items():
            if key not in result:
                result.append(key)
            if isinstance(value, list):
                for element in value:
                    if element not in result:
                        result.append(element)
        return result        

    def get_hierblks_infos(self,name) -> dict:

        blks = self.get_hierblks(name)
        blksinfo = {}

        allblks = list(self._hierdata.keys())
        for val in blks:
            if val in allblks:
                blk = self.get_block_by_name(val)
                alias = blk.alias
                lvl = blk.hdlevel
                condir = blk.constr_dir
                pwr = blk.prime_pwr
                pwrg = pwr.split(' ')[0].strip()
                blksinfo[val] = val + f' {alias}' + f' {lvl}' + f' {pwrg}' + f' {condir}'
            else:
                blksinfo[val] = val + ' (NOT EXPEND)'

        return blksinfo 


    def get_hierlvlblks(self, name, outtype='hd') -> list:
        '''
        get different harden insts, mac insts, dig insts under current design
        outtype is hd/lib/soft/crg/pll
        '''
        allhierblks = []
        allhierblks = self.get_hierblks(name)
        #alltreeblks = self.get_hiertrees(name)
        allhierblks.append(name)

        #allblks = list(self._hierdata.keys())

        insts = []
        macs = []
        digs = []
        plls = []
        crgs = []
        if allhierblks:
            for curblk in allhierblks:
                #if curblk in allblks:
                blk = self.get_block_by_name(curblk)
            
                lvl = blk.hdlevel
                if lvl in ['blk', 'sys', 'top']:
                    insts.append(curblk)
                if lvl in ['lib']:
                    macs.append(curblk)  
                if lvl in ['pll']:
                    plls.append(curblk) 
                if lvl in ['soft']:
                    digs.append(curblk)  
                if lvl in ['crg']:
                    crgs.append(curblk)
                # else:
                #     #print(f'XYZ_{curblk}')
                #     for ky,vl in alltreeblks.items():
                #         if curblk in vl:
                #             parent = ky
                #     blkg = self.get_block_by_name(parent)
                #     if curblk in blkg.get_curhd_by_name():
                #         insts.append(curblk)
                #         upf_warn(f'{curblk} is harden block need expand it in hier_pwr yaml file.')
                #     if curblk in blkg.get_curmac_by_name() or f'{curblk}_USR' in blkg.get_curmac_by_name():
                #         macs.append(curblk)
                #         upf_warn(f'{curblk} is macro block, if not user constraint, need expand it in hier_pwr yaml file.')
                #     if curblk in blkg.get_curdig_by_name() or f'{curblk}_USR' in blkg.get_curdig_by_name():
                #         digs.append(curblk)
                #         upf_warn(f'{curblk} is digital block, if not user constraint, need expand it in hier_pwr yaml file.')

        if outtype == 'hd':
            return insts
        if outtype == 'lib':
            return macs
        if outtype == 'soft':
            return digs  
        if outtype == 'pll':
            return plls
        if outtype == 'crg':
            return crgs  

    def get_hierdepth(self,dic,key):
        
        allblks = list(self._hierdata.keys())
    
        if key not in dic:
            return 1
        else:
            max_depth = 0
            for sub_key in dic[key]:
                if sub_key not in allblks:
                    upf_warn(f'{sub_key} is not expanded in hier yaml file.')
                else:
                    current_depth = self.get_hierdepth(dic, sub_key) + 1
                    max_depth = max(max_depth, current_depth)
            return max_depth

        # # blktrees = self.get_hiertrees(name)
        # if key not in dic:
        #     return 0
        
        # depths = []
        # for child_key in dic[key]:
        #     if child_key not in allblks:
        #         print(f'{child_key} is not expanded in hier yaml file.')
        #     else:
        #         depths.append(self.get_hierdepth(dic,child_key))
        #         #max_depth = max(max_depth, depth)
        
        # return max(depths) + 1

        # max_depth = depth  # 记录最大深度
        
        # if isinstance(blktrees, dict):
        #     for child_node in blktrees.values():
        #         if child_node not in allblks:
        #             print(f'{child_node} is not expanded in hier yaml file.')
        #     else:
        #         child_depth = self.get_hierdepth(child_node, depth + 1)
        #         max_depth = max(max_depth, child_depth)
        
        # return max_depth

    def get_alias_by_name(self, name):
        return self._blocks[name].alias


class UPF_DG(object):
    def __init__(self):
        self._sheets = {}
        self._hier_tree = {}
        self._vardefdata = {}
        self._vardefpcell = {}
        self._wb = {}
        self._inputs = BaseInputs()
        # self._vardef = VarDefSheet()

        self.proj_mode = False
        self._vfile_data = None
        self._pmfile_data = None
        self._objfile_data = None
        self._objfile_list = None
        self._data = None
        self._upfdir = ''
        self._mdname = ''
        self._alias = ''
        
        # self._mdname = ''
        # self._blkalias = ''
        # self._blklvl = ''

    @property
    def hier_tree(self):
        return self._hier_tree
    
    @hier_tree.setter
    def hier_tree(self, hier_tree):
        self._hier_tree = hier_tree
        # if self._mdname and self._hier_tree._blocks[self._mdname]:
        #     self._blkalias = self._hier_tree._blocks[self._mdname].alias
        #     self._blklvl = self._hier_tree._blocks[self._mdname].hdlevel
 
    def load_design_guide(self,dg_file,kwd=''):
        # self._upfdir = abspath(dirname(dirname(dg_file)))
        self._wb = openpyxl.load_workbook(dg_file)

        valdef_sheet = self._wb['VarDef']
        # start = 0
        # for i in range(1, valdef_sheet.max_row+1):
        #     if valdef_sheet.cell(i,1).value == 'Variable':
        #         start = i + 1
        #         break
        # for i in range(start, valdef_sheet.max_row+1):
        #     key = valdef_sheet.cell(row=i, column=1).value
        #     val = valdef_sheet.cell(row=i, column=2).value
        #     self._vardef[key] = val


        self._sheets = {
            'VarDef'        : VarDefSheet(self, 'VarDef'),
            'PDomain'       : PDomainSheet(self, 'PDomain'),
            'PStrategy'     : PStrategySheet(self, 'PStrategy'),     
            'PMode'         : PModeSheet(self, 'PMode'),
        }

        self._vardefdata = self._sheets['VarDef'].get_vardef_value(valdef_sheet)
        self._vardefpcell = self._sheets['VarDef'].get_table_contxt(valdef_sheet, 'VarDef',['PMCELL'])

        # # read and convert excel to json data
        # for sheetname,sheet in self._sheets.items():
        #     sheet.read_data()
        # if kwd == 'json':
        #     for sheetname,sheet in self._sheets.items():
        #         json_file = dirname(dirname(dg_file)) + '/json' + f'/{sheetname.lower()}.json'
        #         sheet.read_data()
        #         sheet.dump_json(json_file)
            

    def read_vfile(self, vfile,kwd=''):
        self._upfdir = abspath(dirname(dirname(vfile)))
        self._vfile_data = self._inputs.read_vfile(vfile)
        self._mdname = self._vfile_data['module_name']
    #    print(self._vfile_data)
        if kwd == 'json':
            self._data = self._vfile_data
            json_file = dirname(dirname(vfile)) + '/json' + f'/pvlog.json'
            self.write_json(json_file)

    def read_pmfile(self, pmfile,kwd=''):
        self._pmfile_data = self._inputs.read_pmfile(pmfile)
        #print(self._pmfile_data)
        if kwd == 'json':
            self._data = self._pmfile_data
            json_file = dirname(dirname(pmfile)) + '/json' + f'/pcell.json'
            self.write_json(json_file)
        
    def read_objfile(self, objfile,kwd=''):
        self._objfile_data, self._objfile_list = self._inputs.read_objfile(objfile)
        #print(self._objfile_data)
        #print(self._objfile_list)
        if kwd == 'json':
            self._data = self._objfile_data
            json_file = dirname(dirname(objfile)) + '/json' + f'/pobj.json'
            self.write_json(json_file)
      

    def read_data(self):
        for sht in self._sheets.values():
            sht.read_data()       

    def update_dg(self):
        for sht in self._sheets.values():
            sht.update_sheet()

    def check_dg(self):
        for sht in self._sheets.values():
            sht.check_sheet()

    def change_dg(self,dgfile):
        for sht in self._sheets.values():
            sht.change_sheet(dgfile)

    def read_json(self,file_path):
        sblk_data = {}
        if os.path.exists(file_path):
            with open(file_path,'r') as fw:
                content = fw.read()
                sblk_data = json.loads(content)

        #print('sblk_data:',sblk_data)
        return sblk_data


    def write_json(self,filepath):
        os.makedirs(dirname(filepath), exist_ok=True)
        jsonstr = json.dumps(self._data, indent=4)
        with open(filepath,'w') as fw:
            print(jsonstr, file=fw) 

    # def save_text(self,context,file):
    #     with open(file, 'w') as fw:
    #         fw.write(context)

    def save_text(self, context,file,kw='xyz'):
        if os.path.exists(file) and 'proc' in kw:
            with open(file, 'a') as fw:
                fw.write(context)
        else:
            with open(file, 'w') as fw:
                fw.write(context)

    def save_workbook(self,output):
        self._wb.save(output)

    def read_text(self, file):
        if not os.path.exists(file):
            raise FileExistsError(f'{file} does not exists')
            # sdc_error(f'{file} not exist. Please check it.')
            # exit(1)
        else:
            txt_list = []
            with open(file,'r') as fh:
                for line in fh.readlines():
                    txt_list.append(line)
        
            return txt_list


# upfdg is XupfDesignGuide object
class BaseSheet(object):
    def __init__(self, upfdg, sheetname):
        self._upfdg = upfdg
        self._sheetname = sheetname
        self._data = []
        self._vardef = {}
        self._pdnmdict = {}
    
    def get_sheet(self):
        return self._upfdg._wb[self._sheetname]

    def read_data(self):
        raise NotImplementedError(self.__class__.__name__ + ' raad_data not implemented yet')

    def write_json(self, filepath):
        os.makedirs(dirname(filepath), exist_ok=True)
        jsonstr = json.dumps(self._data, indent=4)
        with open(filepath,'w') as fw:
            print(jsonstr, file=fw)

    def find_sheet(self, sheet, skw):
        start_rowg = 1
        # TABCONST = ['PMVAR','PMCELL','PMHIER','PMDOMAIN','PMNETWORK','PMBOUNDARY','PMISO','PMLS','PMRET','PMPSW','PMRPT','PMMODE']
        TABCONST = ['PMVAR','PMCELL','PMDOMAIN','PMNETWORK','PMBOUNDARY','PMISO','PMLS','PMRET','PMPSW','PMMODE']
        for i in range(1,sheet.max_row+1):
            if skw in TABCONST and sheet.cell(i,1).value == skw:
                start_rowg = i + 1
                break  
        return  start_rowg 

    def get_vardef_value(self, sheet):

        start_rowg = self.find_sheet(sheet, 'PMVAR')
        # end_rowg = self.find_sheet(sheet, 'PMHIER')
        end_rowg = self.find_sheet(sheet, 'PMCELL')
        for i in range(start_rowg + 1, end_rowg-1):
            key = sheet.cell(row=i, column=1).value
            val = sheet.cell(row=i, column=2).value
            self._vardef[key] = val

        self._vardef['UPF_DIR'] = self._upfdg._upfdir
        self._vardef['COM_DIR'] = self._upfdg._upfdir
        self._vardef['UPF_VERSION'] = '2.1'
        self._vardef['BOUNDARY_MODE'] = 'lower'
        # self._vardef['HD_MOD_NAME'] = self._upfdg._mdname
        # self._vardef['HD_PROCESS'] = ''
        self._vardef['SS_MODE'] = 'full'
        self._vardef['SCOPE_TYPE'] = 'parent; # self'

        return self._vardef   

    def set_name_style(self, kw):
        #time_stamp = time.strftime("%Y%m%d%H%M%S", time.localtime())
        #CONST = f'Generic_Xupf_{time_stamp}'
        CONST = f'Generic_XUPF'
        return kw + '_' + CONST

    # def change_space(self, dgfile):

    #     #sheet = self.get_sheet()
    #     shtname = self._sheetname
    #     sheet = self._upfdg._wb[shtname]

    #     # Find variable start row num, below "Variable" header 
    #     start =0       
    #     df = pd.read_excel(dgfile, sheet_name=shtname, engine='openpyxl') 
    #     df.loc[len(df)] = list(df.columns)
    #     for col in df.columns:
    #         index = list(df.columns).index(col)
    #         letter = get_column_letter(index + 1)
    #         collen = df[col].apply(lambda x: len(self.max_str(str(x)).encode())).max()
    #         # sheet.column dimensions[letter).width = collen*o.9 
    #         sheet.column_dimensions[letter].width = collen * 1.05

    def max_str(self, li):
        max = 0
        max_str = ''
        try:
            for i in li.split('\n'):
                if len(i) > max:
                    max =len(i)
                    max_str=i 
            return max_str 
        except:
            return li
        


    # showErrorMessage=False,showDropDown=True
    def add_dropdown_short(self, sheet, options, start, end):
        dv = DataValidation(type="list", formula1=options, showErrorMessage=False)
        # dv = DataValidation(type="list", formula1=f"Lists!$A$1:$A${len(options)}", showErrorMessage=False)
        sheet.add_data_validation(dv)
        if len(start) == 2 and len(end) == 2:
            for i in range(start[0], end[0] + 1):
                for j in range(start[1], end[1] + 1):
                    dv.add(sheet.cell(i,j))
        if len(start) == 1 and len(end) == 1:
            dv.add(sheet.cell(start[0],end[0]))

    def add_dropdown(self, sheet, options_str, start, end):
        """
        为指定单元格范围添加下拉列表验证
        根据选项长度自动选择直接列表或引用列表方式
        
        参数:
            sheet: 目标工作表
            options_str: 下拉选项字符串，格式为"选项1,选项2,选项3"（带双引号）
            start: 起始单元格坐标 (行, 列) 或 [行]
            end: 结束单元格坐标 (行, 列) 或 [列]
        """
        # 移除首尾的双引号，然后分割为列表
        options = options_str.strip('"').split(',')
        
        # 计算选项总长度（包括逗号分隔符）
        total_length = sum(len(option) for option in options) + (len(options) - 1)
        
        # Excel直接列表的字符限制约为255，留些余量设为250
        if total_length <= 250 and len(options) > 0:
            # 使用直接列表方式（保持原有的字符串格式）
            dv = DataValidation(
                type="list",
                formula1=options_str,  # 直接使用传入的带引号字符串
                showErrorMessage=False
            )
        else:
            # 使用引用列表方式
            wb = sheet.parent
            
            # 获取或创建存储列表的工作表
            if "Lists" not in wb.sheetnames:
                list_sheet = wb.create_sheet("Lists")
                list_sheet.sheet_state = "hidden"  # 隐藏列表工作表
            else:
                list_sheet = wb["Lists"]
            
            # 找到第一个空行来存储新的选项列表
            next_row = 1
            while list_sheet.cell(row=next_row, column=1).value is not None:
                next_row += 1
            
            # 写入选项数据
            for idx, option in enumerate(options, next_row):
                list_sheet.cell(row=idx, column=1, value=option)
            
            # 创建引用公式
            formula = f"Lists!$A${next_row}:$A${next_row + len(options) - 1}"
            dv = DataValidation(
                type="list",
                formula1=formula,
                showErrorMessage=False
            )
        
        # 添加数据验证到工作表
        sheet.add_data_validation(dv)
        
        # 应用数据验证到指定单元格范围
        # 处理坐标格式，确保start和end都是(行, 列)格式
        if len(start) == 1:
            # 如果start只有一个元素，视为行号，列号使用end的值
            start = (start[0], end[0])
            end = start  # 单个单元格
        
        if len(start) == 2 and len(end) == 2:
            # 处理单元格范围
            for i in range(start[0], end[0] + 1):
                for j in range(start[1], end[1] + 1):
                    dv.add(sheet.cell(i, j))
        elif len(start) == 1 and len(end) == 1:
            # 处理单个单元格
            dv.add(sheet.cell(start[0], end[0]))
            
        return dv
        

    def cell_style1(self, sheet, start, end):
        border=Border(left=Side(border_style='thin', color='000000'),
                    right=Side(border_style='thin', color='000000'),
                    top=Side(border_style='thin', color='000000'),
                    bottom=Side(border_style='thin', color='000000'))
        #bgfill = PatternFill(fill_type='solid', start_color='fff2cc', end_color='fff2cc') 
        #bgfill = PatternFill(fill_type = 'solid', start_color='197e00',end_color='197e00')
        bgfill = PatternFill(fill_type = 'solid', start_color='FF385724',end_color='FF333300')
        font = Font(name='等线', size=11, color='FFFFFF')
        for i in range(start[0], end[0] + 1):
            for j in range(start[1], end[1] + 1):
                sheet.cell(i,j).border=border 
                sheet.cell(i,j).fill=bgfill
                sheet.cell(i,j).font=font
                sheet.cell(i,j).alignment = Alignment(horizontal='left', vertical='center',wrapText=True) 
                sheet.cell(i,j).alignment = Alignment(horizontal='left', vertical='center',wrapText=True) 
                sheet.cell(i,j).alignment=Alignment(horizontal='left', vertical='center') 

    def cell_style2(self, sheet, start, end):
        border=Border(left=Side(border_style='thin', color='000000'),
                    right=Side(border_style='thin', color='000000'),
                    top=Side(border_style='thin', color='000000'),
                    bottom=Side(border_style='thin', color='000000'))
        #bgfill = PatternFill(fill_type='solid', start_color='fff2cc', end_color='fff2cc') 
        #bgfill = PatternFill(fill_type = 'solid', start_color='197e00',end_color='197e00')
        bgfill = PatternFill(fill_type = 'solid', start_color='FFFFFF',end_color='FFFFFF')
        #font = Font(name='等线', size=11, color='FFFFFF')
        for i in range(start[0], end[0] + 1):
            for j in range(start[1], end[1] + 1):
                sheet.cell(i,j).border=border 
                sheet.cell(i,j).fill=bgfill
                #sheet.cell(i,j).font=font
                sheet.cell(i,j).alignment = Alignment(horizontal='left', vertical='center',wrapText=True) 
                sheet.cell(i,j).alignment = Alignment(horizontal='left', vertical='center',wrapText=True) 
                sheet.cell(i,j).alignment=Alignment(horizontal='left', vertical='center',wrapText=True)

    def get_supply_infos(self):

        delkeys = ['module_name', 'ISO_CTRL', 'RET_SAVE', 'RET_RES', 'PSO_CTRL', 'PSO_ACK']
        supply_datag = self._upfdg._vfile_data
        supply_data = {}
        if not supply_datag:
            upf_warn('No supply info in vfile')

        for ky,vl in supply_datag.items():
            if ky not in delkeys:
                supply_data[ky] = vl

        # print('supply_data: ', supply_data)
        supply_kw = []
        supply_val = []
        supply_vss = []
        supply_tmp = ''
        for key,val in supply_data.items():
            if '0v' in val or '0.0v' in val:
                if key not in supply_vss:
                    supply_vss.append(key) 
                    upf_info(f'Ground pin is {key}')
            elif 'PSO' in val:
                supply_kw.append(key)
                for i in range(1, int(val[-1]) + 1):
                    supply_kw.append(key + f'_PSW{i}')
                supply_tmp = supply_tmp + ' ' + supply_data[key].split('PSO')[0].strip()
            else:
                supply_kw.append(key)
                supply_tmp = supply_tmp + ' ' + supply_data[key].strip()

        #print(supply_tmp.strip().split(','))
        float_list = [float(x.strip('v')) for x in supply_tmp.strip().split()]
        unique_floats = set(float_list)
        sorted_floats = sorted(unique_floats, reverse=True)
        supply_val = [str(x) + 'v' for x in sorted_floats]
        supply_val.append('off')
        supply_val.append('0v')

        return supply_kw,supply_val,supply_vss,supply_data
    

    def get_ctl_sig(self, ctsig):
        ctrl = []
        for ct in ctsig:
            if re.search(r'\[\d+:\d+\]', ct):
                sig = ct.split('[')[0].strip()
                st = int(ct.split(':')[0].strip()[-1])
                ed = int(ct.split(':')[1].strip()[0])
                for i in range(ed,st+1):
                    ctrl.append(sig + '[' + str(i) + ']')
            else:
                ctrl.append(ct)
        return ctrl

    def get_table_loc(self,sheet, shnm='') -> dict:

        #TABCONSTT = ['PMVAR','PMCELL','PMHIER','PMDOMAIN','PMNETWORK','PMBOUNDARY','PMISO','PMLS','PMPSW','PMRET','PMRPT','PMMODE']

        if not shnm:
            sheetname = self._sheetname
        else:
            sheetname = shnm

        if sheetname == 'VarDef':
            TABCONST = ['PMVAR','PMCELL']
        if sheetname == 'PDomain':
            TABCONST = ['PMDOMAIN','PMNETWORK','PMBOUNDARY']
        if sheetname == 'PStrategy':
            TABCONST = ['PMISO','PMLS','PMPSW','PMRET']
        if sheetname == 'PMode':
            TABCONST = ['PMMODE']

        # row_start max_col              
        row_start = ''
        max_row = ''
        max_col = ''
        # row_start max_col max_row
        table_row_loc = {}
        for kw in TABCONST:
            strow = self.find_sheet(sheet, kw)
            row_start = str(strow)
            for i in range(1,sheet.max_column + 1):
                if sheet.cell(strow,i).value == 'Comment':
                    max_col = str(i)
                    break

            if kw in ['PMCELL','PMBOUNDARY','PMRET','PMMODE']:
                #table_row_loc[kw] = row_start[kw] + ' ' + str(int(row_start[kw].split()[0]) + 20)
                table_row_loc[kw] = row_start + ' ' + str(sheet.max_row + 2) + ' ' + max_col
            else:
                idx = TABCONST.index(kw) + 1
                # print('dfsg: ', TABCONST,idx)
                max_row = self.find_sheet(sheet,TABCONST[idx]) - 1
                #table_row_loc[kw] = row_start[kw] + ' ' + str(int(row_start[TABCONST[idx]].split()[0]) - 2)
                table_row_loc[kw] = row_start + ' ' + str(max_row) + ' ' + max_col

        return table_row_loc

    def get_table_contxt(self,sheet, shnm='', tabnm=[]) -> dict:
        # row_start max_col max_row
        tab_loc = self.get_table_loc(sheet,shnm)
        print('sheet: tab_loc: ', tab_loc)

        if not shnm:
            sheetname = self._sheetname
        else:
            sheetname = shnm

        #TABCONST = ['PMVAR','PMCELL','PMHIER','PMDOMAIN','PMNETWORK','PMBOUNDARY','PMISO','PMLS','PMRET','PMPSW','PMRPT','PMMODE']
        if sheetname == 'VarDef':
            # TABCONST = ['PMVAR','PMCELL','PMHIER']
            if not tabnm:
                TABCONST = ['PMVAR','PMCELL']
            else:
                TABCONST = tabnm
        if sheetname == 'PDomain':
            if not tabnm:
                TABCONST = ['PMDOMAIN','PMNETWORK','PMBOUNDARY']
            else:
                TABCONST = tabnm
        if sheetname == 'PStrategy':
            # TABCONST = ['PMISO','PMLS','PMRET','PMPSW','PMRPT']
            if not tabnm:
                TABCONST = ['PMISO','PMLS','PMRET','PMPSW']
            else:
                TABCONST = tabnm
        if sheetname == 'PMode':
            if not tabnm:
                TABCONST = ['PMMODE']
            else:
                TABCONST = tabnm

        table_contxt = {}
        #row_contxt = {}
        if TABCONST:
            for kw in TABCONST:
                start_row = int(tab_loc[kw].split(' ')[0])
                end_row = int(tab_loc[kw].split(' ')[1])
                end_col = int(tab_loc[kw].split(' ')[2])
                if kw == 'XPMVAR':
                    for i in range(start_row,end_row+1):
                        key = sheet.cell(i+1,1).value
                        val = str(sheet.cell(i+1,2).value)
                        if key:
                            table_contxt[key] = val.strip()
                        # print('PMVARdfd: ', table_contxt)
                        # if key and val:
                        #     table_contxt[key] = val
                else:
                    table_contxt.update(self.get_row_txt(sheet,kw,start_row,end_row,end_col))

        return table_contxt

    def get_row_txt(self, sheet, kw, start_row, end_row, end_col):
        row_contxt = {}
        table_contxt = {}
        for i in range(1, end_row - start_row):
            for j in range(1, end_col + 1):
                key = sheet.cell(start_row, j).value
                val = sheet.cell(start_row + i, j).value
                val_col1 = sheet.cell(start_row + i, 1).value
                if val_col1:
                    if re.search(r'^#', val_col1.strip()):
                        continue
                if key:     key = str(key).strip()
                if val:     val = str(val).strip()
                row_contxt[key] = val
                # if key and val:
                #     row_contxt[key] = val

            all_none = all(ele is None for ele in list(row_contxt.values()))
            if not all_none and row_contxt:
                table_contxt[f'{kw}_Row{start_row + i}'] = row_contxt
            row_contxt = {}
            # for key in table_contxt.keys():
            #     if 'TMCLK' in key:
            #         print(table_contxt)

        return table_contxt


    def get_impl_obj(self, sheet, start_rowg, kwd):
        keywds = list(self._upfdg._objfile_data.keys())
        TABCONST = ['PMVAR','PMCELL','PMDOMAIN','PMNETWORK','PMBOUNDARY','PMISO','PMLS','PMPSW','PMRET','PMRPT','PMMODE']        
        
        supply_kw, supply_vol,supply_vss, supply_data = self.get_supply_infos()
        supply_kw.extend(supply_vss)
        notvss = [x for x in supply_vss if not x in ['VSS']]

        if kwd == 'PMNETWORK':
            # SupplyPortNet	NPwellNet	InstList	MapSupplyList	Comment
            row_tmp = start_rowg
            row_tot = start_rowg + 10
            if len(supply_kw) > 8:
                sheet.insert_rows(start_rowg + 3, 8)
                row_tot += 8
                #self.cell_style2(sheet,[start_rowg + 3,1], [start_rowg + 12,5])
            if len(keywds) > 4:
                sheet.insert_rows(start_rowg + 11, 6)
                row_tot += 6
            self.cell_style2(sheet,[start_rowg + 1,1], [row_tot + 1,5])

            virpwr = [x for x in supply_kw if re.search(r'_PSW\d+',x)]
            relpwr = [x for x in supply_kw if not re.search(r'_PSW\d+',x)]
            cmt = ['PAL','PAL OUT','OUT','SNE1','SNE2','SNE3','SNE4','SNE5']
            if len(notvss) > 0:
                ncmt = notvss.extend(cmt)
            else:
                ncmt = cmt
            #print(cmt)
            self.add_dropdown(sheet, '"' + ','.join(supply_kw) + '"', [start_rowg + 1, 1], [row_tot + 1 , 1])
            self.add_dropdown(sheet, '"' + ','.join(relpwr) + '"', [start_rowg + 1, 2], [row_tot + 1 , 2])
            self.add_dropdown(sheet, '"' + ','.join(ncmt) + '"', [start_rowg + 1, 5], [row_tot + 1, 5])
            
            for i in range(1, len(supply_kw) + 1):               
                sheet.cell(start_rowg + i,1).value = supply_kw[i-1]
                if re.search(r'_PSW\d+',supply_kw[i-1]):
                    nwell = supply_kw[i-1].split('_PSW')[0].strip()
                    sheet.cell(start_rowg + i,2).value = nwell
                else:
                    sheet.cell(start_rowg + i,2).value = supply_kw[i-1]
                row_tmp += 1
            #print(row_tmp)

            for ky in keywds:
                #row_tmp += 1
                if '_conspy_insts' in ky:
                    row_tmp += 1
                    nky = ky.split('_conspy_insts')[0].strip()
                    sheet.cell(row_tmp,1).value = self._upfdg._objfile_data[nky + '_outer_spy']
                    sheet.cell(row_tmp,3).value = nky + '_conspy_insts' #self._upfdg._objfile_data[nky + '_conspy_insts']
                    #self.add_dropdown(sheet,'"' + ky + '"',)
                    sheet.cell(row_tmp,4).value = self._upfdg._objfile_data[nky + '_inner_spy']
                if '_conspy_hinsts' in ky:
                    row_tmp += 1
                    nky = ky.split('_conspy_hinsts')[0].strip()
                    sheet.cell(row_tmp,1).value = self._upfdg._objfile_data[nky + '_outer_spy']
                    #sheet.cell(row_tmp,3).value = self._upfdg._objfile_data[nky + '_conspy_hinsts']
                    sheet.cell(row_tmp,3).value = nky + '_conspy_hinsts'
                    sheet.cell(row_tmp,4).value = self._upfdg._objfile_data[nky + '_inner_spy']
            
            row_tmp = 0

        if kwd == 'PMBOUNDARY':
            bd_ele = []
            bd_exd = []
            flg_ele = 0
            flg_exd = 0
            for ky in keywds:
                if re.search(r'_spa_inport|_spa_outport|_spa_inhpin|_spa_outhpin', ky):
                    bd_ele.append(ky)
                    flg_ele = 1
                if re.search(r'_exdspa_inport|_exdspa_outport|_exdspa_inhpin|_exdspa_outhpin', ky):
                    bd_exd.append(ky)
                    flg_exd = 1
                if re.search(r'_spa_in$|_spa_out$|_spa_input|_spa_output|_spa_inpin|_spa_outpin', ky):
                    bd_ele.append(ky)
                    flg_ele = 1
                if re.search(r'_exdspa_in$|_exdspa_out$|_exdspa_input|_exdspa_output|_exdspa_inpin|_exdspa_outpin', ky):
                    bd_exd.append(ky)
                    flg_exd = 1                    
            if not flg_ele:     bd_ele = None
            if not flg_exd:     bd_exd = None

            return bd_ele,bd_exd

        if kwd == 'PMISO':
            iso_ele = []
            iso_exd = []
            iso_no = []
            flg_ele = 0
            flg_exd = 0
            flg_no = 0
            for ky in keywds:
                if re.search(r'_iso_inport|_iso_outport|_iso_inhpin|_iso_outhpin|_ctliso_inport|_ctliso_inhpin|_fdthiso_inport|_fdthiso_outport|_fdthiso_inhpin|_fdthiso_outhpin', ky):
                    iso_ele.append(ky)
                    flg_ele = 1
                if re.search(r'_exdiso_inport|_exdiso_outport|_exdiso_inhpin|_exdiso_outhpin', ky):
                    iso_exd.append(ky)
                    flg_exd = 1
                if re.search(r'_noiso_inport|_noiso_outport|_noiso_inhpin|_noiso_outhpin', ky):
                    iso_no.append(ky)
                    flg_no = 1

                if re.search(r'_iso_in$|_iso_out$|_ctliso_in$|_fdthiso_in$|_fdthiso_out$', ky):
                    iso_ele.append(ky)
                    flg_ele = 1
                if re.search(r'_exdiso_in$|_exdiso_out$', ky):
                    iso_exd.append(ky)
                    flg_exd = 1
                if re.search(r'_noiso_in$|_noiso_out$', ky):
                    iso_no.append(ky)
                    flg_no = 1

                if re.search(r'_iso_input|_iso_output|_ctliso_input|_fdthiso_input|_fdthiso_output', ky):
                    iso_ele.append(ky)
                    flg_ele = 1
                if re.search(r'_iso_inpin|_iso_outpin|_ctliso_inpin|_fdthiso_inpin|_fdthiso_outpin', ky):
                    iso_ele.append(ky)
                    flg_ele = 1                    
                if re.search(r'_exdiso_input|_exdiso_output|_exdiso_outpin', ky):
                    iso_exd.append(ky)
                    flg_exd = 1
                if re.search(r'_noiso_input|_noiso_output|_noiso_outpin', ky):
                    iso_no.append(ky)
                    flg_no = 1

            if not flg_ele:     iso_ele = None
            if not flg_exd:     iso_exd = None
            if not flg_no:      iso_no = None

            return iso_ele,iso_exd,iso_no
                

        if kwd == 'PMLS':
            ls_ele = []
            ls_exd = []
            ls_no = []
            flg_ele = 0
            flg_exd = 0
            flg_no = 0                
            for ky in keywds:
                if re.search(r'_ls_inport|_ls_outport|_ls_inhpin|_ls_outhpin|_fdthls_inport|_fdthls_outport|_fdthls_inhpin|_fdthls_outhpin', ky):
                    ls_ele.append(ky)
                    flg_ele = 1
                if re.search(r'_exdls_inport|_exdls_outport|_exdls_inhpin|_exdls_outhpin', ky):
                    ls_exd.append(ky)
                    flg_exd = 1
                if re.search(r'_nols_inport|_nols_outport|_nols_inhpin|_nols_outhpin', ky):
                    ls_no.append(ky) 
                    flg_no = 1

                if re.search(r'_ls_in$|_ls_out$|_fdthls_in$|_fdthls_out$|_ls_input|_ls_output|_fdthls_input|_fdthls_output|_ls_inpin|_ls_outpin|_fdthls_inpin|_fdthls_outpin', ky):
                    ls_ele.append(ky)
                    flg_ele = 1
                if re.search(r'_exdls_in$|_exdls_out$|_exdls_input|_exdls_output|_exdls_inpin|_exdls_outpin', ky):
                    ls_exd.append(ky)
                    flg_exd = 1
                if re.search(r'_nols_in$|_nols_out$|_nols_input|_nols_output|_nols_inpin|_nols_outpin', ky):
                    ls_no.append(ky) 
                    flg_no = 1                     
            if not flg_ele:     iso_ele = None
            if not flg_exd:     iso_exd = None
            if not flg_no:      iso_no = None

            return ls_ele,ls_exd,ls_no
        
        if kwd == 'PMPSW':
            psw_ele = []
            psw_exd = []
            flg_ele = 0
            flg_exd = 0
            for ky in keywds:
                if re.search(r'_ctlpsw_inport|_ctlpsw_inhpin', ky):
                    psw_ele.append(ky)
                    flg_ele = 1
                if re.search(r'_ackpsw_inport|_ackpsw_inhpin', ky):
                    psw_exd.append(ky)
                    flg_exd = 1

                if re.search(r'_ctlpsw_in$|_ctlpsw_input|_ctlpsw_inpin', ky):
                    psw_ele.append(ky)
                    flg_ele = 1
                if re.search(r'_ackpsw_in$|_ackpsw_input|_ackpsw_inpin', ky):
                    psw_exd.append(ky)
                    flg_exd = 1
            if not flg_ele:     psw_ele = None
            if not flg_exd:     psw_exd = None    

            return psw_ele,psw_exd

        if kwd == 'PMRET':
            ret_save = []
            ret_res = []
            ret_ele = []
            ret_exd = []
            ret_no = []
            flg_save = 0
            flg_res = 0
            flg_ele = 0
            flg_exd = 0
            flg_no = 0                
            for ky in keywds:
                if re.search(r'_saveret_inport|_saveret_inhpin', ky):
                    ret_save.append(ky)
                    flg_save = 1
                if re.search(r'_resret_inport|_resret_inhpin', ky):
                    ret_res.append(ky)
                    flg_res = 1
                if re.search(r'_ret_insts|_ret_hinsts', ky):
                    ret_ele.append(ky)
                    flg_ele = 1
                if re.search(r'_exdret_insts|_exdret_hinsts', ky):
                    ret_exd.append(ky)
                    flg_exd = 1
                if re.search(r'_noret_insts|_noret_hinsts', ky):
                    ret_no.append(ky)
                    flg_no = 1

                if re.search(r'_saveret_in$|_saveret_input|_saveret_inpin', ky):
                    ret_save.append(ky)
                    flg_save = 1
                if re.search(r'_resret_in$|_resret_input|_resret_inpin', ky):
                    ret_res.append(ky)
                    flg_res = 1


            if not flg_save:    ret_save = None
            if not flg_res:     ret_res = None 
            if not flg_ele:     ret_ele = None
            if not flg_exd:     ret_exd = None
            if not flg_no:      ret_no = None

            return ret_save,ret_res,ret_ele,ret_exd,ret_no

        if kwd == 'PMRPT':
            rpt_ele = []
            rpt_exd = []
            flg_ele = 0
            flg_exd = 0
            for ky in keywds:
                if re.search(r'_rpt_inport|_rpt_outport|_rpt_inhpin|_rpt_outhpin', ky):
                    rpt_ele.append(ky)
                    flg_ele = 1
                if re.search(r'_exdrpt_inport|_exdrpt_outport|_exdrpt_inhpin|_exdrpt_outhpin', ky):
                    rpt_exd.append(ky)
                    flg_exd = 1

            for ky in keywds:
                if re.search(r'_rpt_in$|_rpt_out$|_rpt_input|_rpt_output|_rpt_inpin|_rpt_outpin', ky):
                    rpt_ele.append(ky)
                    flg_ele = 1
                if re.search(r'_exdrpt_in$|_exdrpt_out$|_exdrpt_input|_exdrpt_output|_exdrpt_inpin|_exdrpt_outpin', ky):
                    rpt_exd.append(ky)
                    flg_exd = 1

            if not flg_ele:     rpt_ele = None
            if not flg_exd:     rpt_exd = None     

            return rpt_ele,rpt_exd
        
    def save_text(self, context,file):
        with open(file, 'w') as fw:
            fw.write(context)

    def get_rows(self,pmdata,keyrow,kwd,ckwd):
        pmdict = {}
        pmkeys = []
        pmlist = []
        tab = keyrow.split('_')[0]
        # print('pmdata: ', pmdata)

        # pmlist = [(key, val) for key, val in pmdata.items() if keyrow in key and not re.search(r'^#',val[f'{ckwd}'].strip()) and val[f'{kwd}']]
        for key, val in pmdata.items():
            if keyrow in key and not re.search(r'^#', val[f'{ckwd}'].strip()):
                if '|' in kwd:
                    kwlst = kwd.split('|')
                    kflg = False
                    for kl in kwlst:
                        if val[f'{kl}']: kflg = True
                    if kflg:
                        pmlist.append((key, val))
                else:
                    if val[f'{kwd}']:
                        pmlist.append((key, val))

        if pmlist:
            for k,v in pmlist:
                pmdict[k] = v
                pmkeys.append(k)
            pmkeys.sort(key=lambda x: int(x.split("Row")[1]))
        else:
            print(f'The table {tab} is empty.')

        # print(f'real tabale {tab}', pmkeys,pmdict)

        return pmdict, pmkeys     

    def get_pdname(self,blkalias,sheet):
        i = 0
        pddic = {}

        # print('get_pdname--sheet: ', sheet)
        contxt = self.get_table_contxt(sheet,'PDomain',['PMDOMAIN'])
        pdkeys = list(contxt.keys())
        # print('pdfsmcxc: ', contxt)
        pdict, prows = self.get_rows(contxt,'PMDOMAIN_Row','PDName','PDName')
        # print('pdict: prows', pdict, prows)
        if prows:
            for prow in prows:
                rowdict = pdict[prow]
                pdname = rowdict['PDName']
                i += 1

                prmdic = [(key,val) for key,val in rowdict.items() if val == 'PRM']
                if len(prmdic) > 1:
                    upf_error(f'{prow} has two or more primary power. Please check it.')
                    prm = 'NOPRM'
                else:
                    prm = prmdic[0][0]

                nprm = ''.join(prm.split('_'))
                self._pdnmdict[f'{pdname}'] = f'PD{i}_${{{blkalias}}}_{nprm}_{pdname}'
        else:
            upf_error(f'Can not find power domain definition. Please check it.')

        return self._pdnmdict


class BaseInputs(object):
    def __init__(self):
        self.vfile_data = {}
        self.pmfile_data = {}
        self.objfile_data = {}
               
    def read_vfile(self, vfile) -> dict:

        lines = self.read_text(vfile)
        for line in lines:
            line = line.replace('\n','').replace('\r','').replace('\t',' ').strip()
            if re.search(r'^\/\/', line):
                continue
            if re.search(r'^module', line):
                self.vfile_data['module_name']= re.split(' +',line)[1].strip().replace('(','')
            if re.search(r'\/\/#(\d+[vV]#|\d+.\d+[vV]#|\d+.\d+[vV]:)', line):
                #print(line)
                port_pwrg = re.split('//#',line)[0].strip().replace(',', '')
                port_pwrg = [x for x in re.split(' ', port_pwrg) if x]
                # port_pwr = re.split(' ', port_pwrg)[1].strip()
                port_pwr = port_pwrg[1]
                port_volx = re.split('//#',line)[1].strip() # maybe include 'PSO'
                if re.search(r'#PSO \d+#', port_volx):                    
                    #port_volg = ' PSO' + re.split(r'#PSO ',port_volx)[1].replace('#','')
                    port_volg = ' ' + re.split(r'#',port_volx)[1].strip().replace(' ','')
                    port_volx = re.split(r'#PSO ', port_volx)[0].strip()
                    #print(port_volg)
                else:
                    port_volg = ''

                if re.search(r':', port_volx):
                    port_vol = ' '.join(re.split(':',port_volx)).strip().lower().replace('#', '') + port_volg
                else:
                    port_vol = port_volx.lower().replace('#', '') + port_volg
                self.vfile_data[port_pwr] = port_vol
       
        self.vfile_data['ISO_CTRL'] = self.get_ctrl_ports(lines, 'ISO_CTRL')
        self.vfile_data['RET_SAVE'] = self.get_ctrl_ports(lines, 'RET_SAVE')
        self.vfile_data['RET_RES'] = self.get_ctrl_ports(lines, 'RET_RES')
        self.vfile_data['PSO_CTRL'] = self.get_ctrl_ports(lines, 'PSO_CTRL')
        self.vfile_data['PSO_ACK'] = self.get_ctrl_ports(lines, 'PSO_ACK')

        return self.vfile_data



    def read_pmfile(self, pmfile) -> dict:

        pm_data = self.read_yaml(pmfile)

        # ISOROW1, ELSROW1,
        isocells = self.get_pmcell_info(pm_data, 'ISOPowerMCell')
        lscells = self.get_pmcell_info(pm_data, 'LSPowerMCell')
        elscells = self.get_pmcell_info(pm_data, 'ELSPowerMCell')
        retcells = self.get_pmcell_info(pm_data, 'RETPowerMCell')
        pswcells = self.get_pmcell_info(pm_data, 'PSWPowerMCell')

        isocells.update(lscells)
        isocells.update(elscells)
        isocells.update(retcells)
        isocells.update(pswcells)
        self.pmfile_data = isocells

        return self.pmfile_data



    def read_objfile(self, objfile) -> dict:
        
        obj_lines = self.read_text(objfile)

        objfile_data = {}
        var_list = []
        var_nline = ''

        for line in obj_lines:
            if re.search(r'^#', line):
                var_nline += f'\n{line}'
                #var_list += line
            elif re.search(r'^set\s+', line):
                var_list.append(re.split(' +', line)[1].strip())
                if re.search(r'[|]',line):
                    nline = line.replace('[','\\[')
                    nline = nline.replace(']','\\]')
                else:
                    nline = line
                var_nline += f'\n{nline}'
            elif re.search(r'\w+[|\d+]',line):
                nline = line.replace('[','\\[')
                nline = nline.replace(']','\\]')
                var_nline += f'\n{nline}'              
            else:
                var_nline += f'\n{line}'
        #var_list = [re.split(' +', line)[1].strip() for line in obj_lines if re.search(r'^set\s+', line)]
        #print(f'objfile var list {var_list}')

        # tmpfile = dirname(objfile) + '/objtmp.tcl'
        tmpfile = os.path.join(os.path.dirname(objfile),'objtmp.tcl')
        # 确保目录存在
        os.makedirs(os.path.dirname(tmpfile), exist_ok=True)
        with open(tmpfile, 'w') as fw:
            fw.write(var_nline)        
        #print(var_nline)

        tcl_intp = tk.Tcl()
        # tcl_intp.eval(f'source {tmpfile}')
        #tcl_intp.eval(f'source {objfile}')
        #tcl_intp.eval(' '.join(var_nline))

        # 使用TCL的文件路径语法来source文件，确保跨平台兼容
        # 将路径转换为TCL格式
        tcl_path = tmpfile.replace(os.sep, '/')
        tcl_intp.eval(f'source [file normalize {{{tcl_path}}}]')
        
        for var in var_list:
            nvar = tcl_intp.getvar(var).strip()
            if re.search(r'\[|\]',nvar):
                nvar = nvar.replace('\\[','[')
                nvar = nvar.replace('\\]',']')
            self.objfile_data[var] = nvar

        if os.path.exists(tmpfile):
            os.system(f'rm -f {tmpfile}')

        return self.objfile_data,var_list
        


    def read_yaml(self, yaml_file):

        yaml_data = {}
        if not os.path.exists(yaml_file):
            raise FileExistsError(f'{yaml_file} does not exists')
        with open(yaml_file, 'r') as fh:
            yaml_data = yaml.load(fh, yaml.FullLoader)

        return yaml_data
        
    def get_pmcell_info(self, pmdata, kwd) -> dict:

        pm_num = len(pmdata[kwd].keys())
        kws = kwd.replace('PowerMCell', '')

        pm_data = {}
        if pm_num == 0:
            upf_warn(f'{kws} power cell not found.')
        elif pm_num == 1:
            kws = kws + 'Row1'
            pm_data[kws] = pmdata[kwd]['TypeIndex1']
        else:
            for num in range(1,pm_num + 1):
                kwn = kws + f'Row{num}'
                pm_data[kwn] = pmdata[kwd][f'TypeIndex{num}']

        return pm_data
    


    def read_text(self, file):
        if not os.path.exists(file):
            raise FileExistsError(f'{file} does not exists')
            # upf_error(f'{file} not exist. Please check it.')
            # exit(1)
        else:
            txt_list = []
            with open(file,'r') as fh:
                for line in fh.readlines():
                    #pattern = r"\[\s*(\d+)\s*:\s*(\d+)\]"
                    if line.strip() == "":
                        continue
                    # if line.strip().startswith("//"):
                    #     continue                     
                    line = re.sub(r"\[\s*(\d+)\s*:\s*(\d+)\s*\]", r"[\1:\2]", line)
                    txt_list.append(line.strip())
        
            return txt_list

    def get_ctrl_ports(self, lnlist, kwd) -> list:

        full_list = []
        stwp = ''
        stw = []

        nkw = r'//#' + f'{kwd}' + '#'
        #print(nkw)
        for line in lnlist:
            if re.search(r'^\/\/', line):
                continue
            if re.search(f'{nkw}', line):
                # kw_loc = re.split(' +',line).index(r'//#ISO_CTRL#')
                # iso_ctrl.append(re.split(' +',line)[kw_loc - 1].strip().replace(',', ' ').strip())
                #stw = re.split('//#ISO_CTRL',line)[-2].strip()
                stwp = re.split(f'{nkw}',line)[-2].strip()
                if re.search(r'\/\/$', stwp):
                    stwp.replace('//','')
                elif re.search(r'\/\/\w*\s*', stwp):
                    stwp = re.split('//', stwp)[-2].strip()

                #patn = r'\[(\d+:\d+)|(\d+\s+:\d+)|(\d+:\s+\d+)|(\s+\d+:\d+)\]'    
                if re.search(r'\[\d+:\d+\]', stwp):
                    portnum = ''.join(re.findall(r'\[\d+:\d+\]', stwp)).strip()
                    stwp = re.split(r'\[\d+:\d+\]', stwp)[-1].strip()
                    #stw = re.split('wire|logic|byte|bit|reg', stw)[-1].strip()
                    #stw = re.split('input|output', stw)[-1].strip()                 
                    stw = re.split(',', stwp)[:-1] 
                    stw = [st + portnum for st in stw]
                elif re.search(r'wire|logic|byte|bit|reg', stwp):
                    stwp = re.split('wire|logic|byte|bit|reg', stwp)[-1].strip()
                    stw = re.split(',', stwp)[:-1]
                elif re.search(r'input|output', stwp):
                    stwp = re.split('input|output', stwp)[-1].strip()
                    stw = re.split(',', stwp)[:-1].strip()

                #nested_list.append(stw)
                full_list.extend(stw)

        #return list(chain(*nested_list))
        return full_list


    # def find_start_cell_location(self,sheet,var):
    #     start = 0
    #     for i in range(1, sheet.max_row+1):
    #         if sheet.cell(i,1).value == var:
    #             start = i + 1
    #             break
    #     for i in range(start, sheet.max_row+1):
    #         key = sheet.cell(row=i, column=1).value
    #         val = sheet.cell(row=i, column=2).value
    #         self._vardef[key] = val


class VarDefSheet(BaseSheet):
    def __init__(self,*args):
        super().__init__(*args)  
        self._vardata = {}     

    def update_sheet(self):
        '''
        # only during -dg option
        # addition of module name value from vfile
        # addition of user_defined variables
        # addition of pmcell table from pmfile, vfile
        # addition of block hier tree expanded table from hier yaml
        '''
        sheet = self.get_sheet()

        hiertree = self._upfdg._hier_tree

        # find PMVAR table
        start_rowg = self.find_sheet(sheet, 'PMVAR')
        
        mdname = self._upfdg._vfile_data['module_name']
        if mdname:
            sheet.cell(start_rowg + 1, 2).value = mdname
        
        vardef = self.get_vardef_value(sheet)

        varlist = ['T28','T12','T7','T5', 'T3']
        self.add_dropdown(sheet, '"' + ','.join(varlist) + '"', [start_rowg + 2], [2])
        varlist = ['RTL','SYN','DFT_SYN','SIM','PLA','CTS','PnR','SIGNOFF']
        self.add_dropdown(sheet, '"' + ','.join(varlist) + '"', [start_rowg + 3], [2])
        varlist = ['1','0']
        self.add_dropdown(sheet, '"' + ','.join(varlist) + '"', [start_rowg + 4], [2])
        # varlist = ['full','fast'] # SS_MODE
        # self.add_dropdown(sheet, '"' + ','.join(varlist) + '"', [start_rowg + 5], [2])
        # varlist = ['parent','self'] # SCOPE_TYPE
        # self.add_dropdown(sheet, '"' + ','.join(varlist) + '"', [start_rowg + 6], [2])
        # varlist = ['lower','higher']
        # self.add_dropdown(sheet, '"' + ','.join(varlist) + '"', [start_rowg + 5], [2])
        varlist = ['DC','FC','PT','VCS','ICC2','GNS','INN','TPS']
        self.add_dropdown(sheet, '"' + ','.join(varlist) + '"', [start_rowg + 5], [2])
        # varlist = ['2.0','2.1','3.0']
        # self.add_dropdown(sheet, '"' + ','.join(varlist) + '"', [start_rowg + 11], [2])
        # varlist = ['full','local']
        # self.add_dropdown(sheet, '"' + ','.join(varlist) + '"', [start_rowg + 12], [2])

        block = hiertree.get_block_by_name(mdname)
        for i in range(start_rowg, sheet.max_row + 1):
            if sheet.cell(i, 1).value == 'UPF_DIR':
                sheet.cell(i, 2).value = block.constr_dir
        
        for i in range(1,start_rowg+11):
            for j in range(1,4):
                sheet.cell(start_rowg+i,j).alignment = Alignment(horizontal='left',vertical='center',wrapText=True)


        # find PMCELL table
        start_rowg = self.find_sheet(sheet, 'PMCELL')
        # PMType	PMCtrlSig	PMCell	PMSupplyPin	PMCtrlPin	PDFunction	PathType	NameFormat Comment
        # ISO         +          +        +           +            +           *            + 
        # LS          -          +        +           -            +           *            +      
        # ELS         +          +        +           +            +           *            +
        # RET         +          +        +           +            +           *            
        # PSO         +          +        +           +            +           *            

        # iso
        isoctl = ' '.join(self._upfdg._vfile_data['ISO_CTRL'])
        #print(self._upfdg._pmfile_data.keys())
        isokw = re.findall(r'ISORow\d+', ' '.join(self._upfdg._pmfile_data.keys()))
        for i in range(1, len(isokw) + 1):
            sheet.cell(start_rowg + i, 1).value = self._upfdg._pmfile_data[f'ISORow{i}']['CellType']
            #print(self._upfdg._pmfile_data[f'ISORow{i}']['CellType'])
            sheet.cell(start_rowg + i, 2).value = isoctl
            sheet.cell(start_rowg + i, 3).value = self._upfdg._pmfile_data[f'ISORow{i}']['CellPatn']
            supplypin = self._upfdg._pmfile_data[f'ISORow{i}']['PrimaryPower'] + ' ' + self._upfdg._pmfile_data[f'ISORow{i}']['BackupPower'] + ' ' + self._upfdg._pmfile_data[f'ISORow{i}']['NwellPower'] + ' ' + self._upfdg._pmfile_data[f'ISORow{i}']['PwellPower'] + ' ' + self._upfdg._pmfile_data[f'ISORow{i}']['GroundPin']
            sheet.cell(start_rowg + i, 4).value = supplypin
            sheet.cell(start_rowg + i, 5).value = self._upfdg._pmfile_data[f'ISORow{i}']['CtrlPin']
            sheet.cell(start_rowg + i, 6).value = self._upfdg._pmfile_data[f'ISORow{i}']['PDFunction']
            sheet.cell(start_rowg + i, 7).value = self._upfdg._pmfile_data[f'ISORow{i}']['PathType']
            sheet.cell(start_rowg + i, 8).value = self.set_name_style('ISO')  # 'ISO_Generic_Xconst_09292003'
        # ls
        start_rowg = start_rowg  + len(isokw)
        #isoctl = ' '.join(self._upfdg._vfile_data['ISO_CTRL'])
        #print(self._upfdg._pmfile_data.keys())
        lskw = re.findall(r'\s+LSRow\d+', ' '.join(self._upfdg._pmfile_data.keys()))
        #print(lskw)
        for i in range(1, len(lskw) + 1):
            sheet.cell(start_rowg + i, 1).value = self._upfdg._pmfile_data[f'LSRow{i}']['CellType']
            #sheet.cell(start_rowg + i, 2).value = isoctl
            sheet.cell(start_rowg + i, 3).value = self._upfdg._pmfile_data[f'LSRow{i}']['CellPatn']
            supplypin = self._upfdg._pmfile_data[f'LSRow{i}']['PrimaryPower'] + ' ' + self._upfdg._pmfile_data[f'LSRow{i}']['BackupPower'] + ' ' + self._upfdg._pmfile_data[f'LSRow{i}']['NwellPower'] + ' ' + self._upfdg._pmfile_data[f'LSRow{i}']['PwellPower'] + ' ' + self._upfdg._pmfile_data[f'LSRow{i}']['GroundPin']
            sheet.cell(start_rowg + i, 4).value = supplypin
            #sheet.cell(start_rowg + i, 5).value = self._upfdg._pmfile_data[f'LSRow{i}']['CtrlPin']
            sheet.cell(start_rowg + i, 6).value = self._upfdg._pmfile_data[f'LSRow{i}']['PDFunction']
            sheet.cell(start_rowg + i, 7).value = self._upfdg._pmfile_data[f'LSRow{i}']['PathType']
            sheet.cell(start_rowg + i, 8).value = self.set_name_style('LS')  # 'LS_Generic_Xconst_09292003'      

        # els
        start_rowg = start_rowg + len(lskw)
        elsctl = ' '.join(self._upfdg._vfile_data['ISO_CTRL'])
        elskw = re.findall(r'ELSRow\d+', ' '.join(self._upfdg._pmfile_data.keys()))
        for i in range(1, len(elskw) + 1):
            sheet.cell(start_rowg + i, 1).value = self._upfdg._pmfile_data[f'ELSRow{i}']['CellType']
            sheet.cell(start_rowg + i, 2).value = elsctl
            sheet.cell(start_rowg + i, 3).value = self._upfdg._pmfile_data[f'ELSRow{i}']['CellPatn']
            supplypin = self._upfdg._pmfile_data[f'ELSRow{i}']['PrimaryPower'] + ' ' + self._upfdg._pmfile_data[f'ELSRow{i}']['BackupPower'] + ' ' + self._upfdg._pmfile_data[f'ELSRow{i}']['NwellPower'] + ' ' + self._upfdg._pmfile_data[f'ELSRow{i}']['PwellPower'] + ' ' + self._upfdg._pmfile_data[f'ELSRow{i}']['GroundPin']
            sheet.cell(start_rowg + i, 4).value = supplypin
            sheet.cell(start_rowg + i, 5).value = self._upfdg._pmfile_data[f'ELSRow{i}']['CtrlPin']
            sheet.cell(start_rowg + i, 6).value = self._upfdg._pmfile_data[f'ELSRow{i}']['PDFunction']
            sheet.cell(start_rowg + i, 7).value = self._upfdg._pmfile_data[f'ELSRow{i}']['PathType']
            sheet.cell(start_rowg + i, 8).value = self.set_name_style('ELS')  # 'ELS_Generic_Xconst_09292003'

        # ret
        start_rowg = start_rowg + len(elskw)
        retsave = ' '.join(self._upfdg._vfile_data['RET_SAVE'])
        retres = ' '.join(self._upfdg._vfile_data['RET_RES'])
        retctl = retsave + ' | ' + retsave
        retkw = re.findall(r'RETRow\d+', ' '.join(self._upfdg._pmfile_data.keys()))
        for i in range(1, len(retkw) + 1):
            sheet.cell(start_rowg + i, 1).value = self._upfdg._pmfile_data[f'RETRow{i}']['CellType']
            sheet.cell(start_rowg + i, 2).value = retctl
            sheet.cell(start_rowg + i, 3).value = self._upfdg._pmfile_data[f'RETRow{i}']['CellPatn']
            supplypin = self._upfdg._pmfile_data[f'RETRow{i}']['PrimaryPower'] + ' ' + self._upfdg._pmfile_data[f'RETRow{i}']['BackupPower'] + ' ' + self._upfdg._pmfile_data[f'RETRow{i}']['NwellPower'] + ' ' + self._upfdg._pmfile_data[f'RETRow{i}']['PwellPower'] + ' ' + self._upfdg._pmfile_data[f'RETRow{i}']['GroundPin']
            sheet.cell(start_rowg + i, 4).value = supplypin
            sheet.cell(start_rowg + i, 5).value = self._upfdg._pmfile_data[f'RETRow{i}']['SavePin'] + ' | ' + self._upfdg._pmfile_data[f'RETRow{i}']['ResPin']
            sheet.cell(start_rowg + i, 6).value = self._upfdg._pmfile_data[f'RETRow{i}']['PDFunction']
            #sheet.cell(start_rowg + i, 7).value = self._upfdg._pmfile_data[f'RETRow{i}']['PathType']
            #sheet.cell(start_rowg + i, 8).value = self.set_name_style('RET')  # 'RET_Generic_Xconst_09292003'


        # pso
        start_rowg = start_rowg + len(retkw)
        psoctl = ' '.join(self._upfdg._vfile_data['PSO_CTRL'])
        psoack = ' '.join(self._upfdg._vfile_data['PSO_ACK'])
        psoctl = psoctl + ' | ' + psoack
        psokw = re.findall(r'PSWRow\d+', ' '.join(self._upfdg._pmfile_data.keys()))
        #print(psokw)
        for i in range(1, len(psokw) + 1):
            sheet.cell(start_rowg + i, 1).value = self._upfdg._pmfile_data[f'PSWRow{i}']['CellType']
            sheet.cell(start_rowg + i, 2).value = psoctl
            sheet.cell(start_rowg + i, 3).value = self._upfdg._pmfile_data[f'PSWRow{i}']['CellPatn']
            supplypin = self._upfdg._pmfile_data[f'PSWRow{i}']['InputPower'] + ' ' + self._upfdg._pmfile_data[f'PSWRow{i}']['OutputPower']  + ' ' + self._upfdg._pmfile_data[f'PSWRow{i}']['GroundPin']
            sheet.cell(start_rowg + i, 4).value = supplypin
            sheet.cell(start_rowg + i, 5).value = self._upfdg._pmfile_data[f'PSWRow{i}']['CtrlPin'] + ' | ' + self._upfdg._pmfile_data[f'PSWRow{i}']['AckPin']
            sheet.cell(start_rowg + i, 6).value = self._upfdg._pmfile_data[f'PSWRow{i}']['PDFunction']
            #sheet.cell(start_rowg + i, 7).value = self._upfdg._pmfile_data[f'PSWRow{i}']['PathType']
            #sheet.cell(start_rowg + i, 8).value = 'PSO_Generic_Xconst_09292003'

        # repeater

    def read_data(self):
        sheet = self.get_sheet()
        self._vardata = self.get_table_contxt(sheet,'VarDef')
        # nvaldata = {}
        # nvaldata["TMVAR_Row16"] = {
        #     "Variable": "UPF_DIR",
        #     "Value": f'{self._upfdg._upfdir}',
        #     "Comment": ''
        # }
        # nvaldata["TMVAR_Row17"] = {
        #     "Variable": "COM_DIR",
        #     "Value": f'{self._upfdg._upfdir}',
        #     "Comment": ''
        # }
        # nvaldata["TMVAR_Row18"] = {
        #     "Variable": "UPF_VERSION",
        #     "Value": '2.1',
        #     "Comment": ''
        # }
        # nvaldata["TMVAR_Row19"] = {
        #     "Variable": "HD_MOD_NAME",
        #     "Value": f'{self._upfdg._mdname}',
        #     "Comment": ''
        # }
        # nvaldata["TMVAR_Row20"] = {
        #     "Variable": "HD_PROCESS",
        #     "Value": '',
        #     "Comment": ''
        # }
        # nvaldata["TMVAR_Row21"] = {
        #     "Variable": "SCOPE_TYPE",
        #     "Value": 'parent; # self',
        #     "Comment": ''
        # }
        # self._vardata.update(nvaldata)
        # print('_vardata: ', self._vardata)

    def change_sheet(self):
        pass

    def dump_json(self,json_file):
        self._data = self._vardata
        self.write_json(json_file)


class PDomainSheet(BaseSheet):
    def __init__(self,*args):
        super().__init__(*args)
        self._pddata = {}
        #self._pdnmdic = {}
        
    def update_sheet(self):
        '''
        # only during -dg option
        # addition of supply infos in PMDOMAIN table
        # addition of dropdown infos from pmempty and pmobj file
        '''

        sheet = self.get_sheet()

        supply_kw, supply_vol, supply_vss, supply_data = self.get_supply_infos()

        # find PMDOMAIN table 
        start_rowg = self.find_sheet(sheet, 'PMDOMAIN') 

        #sheet.insert_cols(2, amount = len(supply_kw) + 1)
        #print(start_rowg)
        #rgb_color = PatternFill(start_color=cell_color, end_color=cell_color).start_color.rgb
        #print("单元格A1的填充颜色RGB值:", rgb_color)
        # print(sheet.cell(start_rowg,1).fill.start_color.index)
        # print(sheet.cell(start_rowg,1).fill.end_color.index)

        for i in range(1, len(supply_kw) + 1):
            sheet.cell(start_rowg, i + 2).value = supply_kw[i-1]
        sheet.cell(start_rowg, len(supply_kw) + 3).value = 'Comment'
        self.cell_style1(sheet, [start_rowg, 3], [start_rowg, len(supply_kw) + 3])
        #print(sheet.cell(start_rowg,1).fill.start_color.index)
        #print(sheet.cell(start_rowg,1).fill.end_color.index)
        self.cell_style2(sheet, [start_rowg + 1, 3], [start_rowg + 11, len(supply_kw) + 3])

        # formulal=f'"{iodly_str},40%,50%, 60%, 70%, 80%"'
        # formula2=r'[-+]?[0-9]*\.?[0-9]+'
        # f'"{",".join(choices)}"'
        self.add_dropdown(sheet, '"PRM"', [start_rowg + 1, 3], [start_rowg + 11, len(supply_kw) + 2])
        self.add_dropdown(sheet, '"-update"', [start_rowg + 1, len(supply_kw) + 3], [start_rowg + 11, len(supply_kw) + 3])

        # find PMNETWORK table
        # SupplyPortNet	NPwellNet	InstList	MapSupplyList	Comment
        # f'"{",".join(supply_kw)}"'
        start_rowg = self.find_sheet(sheet, 'PMNETWORK')
        # self.add_dropdown(sheet, '"' + ','.join(supply_kw) + '"', [start_rowg + 1, 1], [start_rowg + 10, 1])
        # obj_data = list(self._upfdg._objfile_data.keys())
        # self.add_dropdown(sheet, '"' + ','.join(obj_data) + '"', [start_rowg + 1, 3], [start_rowg + 10, 3])
        self.get_impl_obj(sheet, start_rowg, 'PMNETWORK')

        # find PMBOUNDARY table
        # ApplyPorts	Elements	ExcludeList	DriverSupply	ReceiverSupply	Attribute	Comment
        start_rowg = self.find_sheet(sheet, 'PMBOUNDARY')
        self.add_dropdown(sheet, '"inputs,outputs,both"', [start_rowg + 1, 1], [start_rowg + 11, 1])
        #print(self.get_impl_obj(sheet, start_rowg, 'PMBOUNDARY'))
        kw_ele,kw_exd = self.get_impl_obj(sheet, start_rowg, 'PMBOUNDARY')
        if kw_ele:
            self.add_dropdown(sheet, '"' + ','.join(kw_ele) + '"', [start_rowg + 1, 2], [start_rowg + 11, 2])
        if kw_exd:  
            self.add_dropdown(sheet, '"' + ','.join(kw_exd) + '"', [start_rowg + 1, 3], [start_rowg + 11, 3])

        self.add_dropdown(sheet, '"' + ','.join(supply_kw) + '"', [start_rowg + 1, 4], [start_rowg + 11, 5])
        kw_bd = ['iso_sink', 'iso_source', 'snps_derived', 'related_supply_default_primary', 'resolved_iso_strategy']
        self.add_dropdown(sheet, '"' + ','.join(kw_bd) + '"', [start_rowg + 1, 6], [start_rowg + 11, 6])
        kw_bd = ['feedthrough', 'unconnected', 'is_analog']
        self.add_dropdown(sheet, '"' + ','.join(kw_bd) + '"', [start_rowg + 1, 7], [start_rowg + 11, 7])        
  

    def read_data(self):
        sheet = self.get_sheet()
        self._pddata = self.get_table_contxt(sheet,'PDomain')
        # print('_pddata: ', self._pddata)

    def change_sheet(self):
        pass

    def dump_json(self,json_file):
        self._data = self._pddata
        self.write_json(json_file)



class PStrategySheet(BaseSheet):
    def __init__(self,*args):
        super().__init__(*args)
        self._psdata = {}
        #self._pdname = {}
        self._pmels = {}
        

    def update_sheet(self):
        sheet = self.get_sheet()
        kw_iso = []

        supply_kw, supply_vol, supply_vss, supply_data = self.get_supply_infos()
        print(supply_kw)

        # PDName	Location	SrcSupply	SinkSupply	DiffSupply	SupplyIn	EnCtrlSens	ClampVal	ApplyPorts	Elements	ExcludeList	NoISO	Comment
        # find PMISO table 
        start_rowg = self.find_sheet(sheet, 'PMISO') 
        print('PStrategySheet start_rowg: ', start_rowg)
        kw_iso = ['self', 'parent','fanout']
        self.add_dropdown(sheet, '"' + ','.join(kw_iso) + '"', [start_rowg + 1, 2], [start_rowg + 11, 2])
        for i in (3,4,6):
            self.add_dropdown(sheet, '"' + ','.join(supply_kw) + '"', [start_rowg + 1, i], [start_rowg + 11, i])
        kw_iso = ['true', 'false']
        self.add_dropdown(sheet, '"' + ','.join(kw_iso) + '"', [start_rowg + 1, 5], [start_rowg + 11, 5])
        iso_ctl = self._upfdg._vfile_data['ISO_CTRL']
        isoctl = self.get_ctl_sig(iso_ctl)
        # print('ISOER ' + ' '.join(isoctl))
        kw_iso = []
        for i in isoctl:
            kw_iso.append(i + ' high')
            kw_iso.append(i + ' low')
        # print('kw_iso: ', kw_iso)
        self.add_dropdown(sheet, '"' + ','.join(kw_iso) + '"', [start_rowg + 1, 7], [start_rowg + 11, 7])
        kw_iso = ['0', '1','latch']
        self.add_dropdown(sheet, '"' + ','.join(kw_iso) + '"', [start_rowg + 1, 8], [start_rowg + 11, 8])
        kw_iso = ['inputs', 'outputs','both']
        self.add_dropdown(sheet, '"' + ','.join(kw_iso) + '"', [start_rowg + 1, 9], [start_rowg + 11, 9]) 

        iso_ele,iso_exd,iso_no = self.get_impl_obj(sheet, start_rowg, 'PMISO') 
        # print(iso_exd)
        if iso_ele:
            self.add_dropdown(sheet, '"' + ','.join(iso_ele) + '"', [start_rowg + 1, 10], [start_rowg + 11, 10])
        if iso_exd:
            self.add_dropdown(sheet, '"' + ','.join(iso_exd) + '"', [start_rowg + 1, 11], [start_rowg + 11, 11])
        if iso_no:
            self.add_dropdown(sheet, '"' + ','.join(iso_no) + '"', [start_rowg + 1, 12], [start_rowg + 11, 12])

        kw_iso = ['ELS_A', 'ELS_B','ELS_C','ELS_D', 'ELS_E', 'force', '-update']
        self.add_dropdown(sheet, '"' + ','.join(kw_iso) + '"', [start_rowg + 1, 13], [start_rowg + 11, 13])

        # PDName	Location	SrcSupply	SinkSupply	SupplyIn SupplyOut Rule	ApplyPorts	Elements	ExcludeList	NoLS	Comment
        # find PMLS table 
        start_rowg = self.find_sheet(sheet, 'PMLS') 
        kw_ls = ['self', 'parent','other','fanout','automatic']
        self.add_dropdown(sheet, '"' + ','.join(kw_ls) + '"', [start_rowg + 1, 2], [start_rowg + 11, 2])
        self.add_dropdown(sheet, '"' + ','.join(supply_kw) + '"', [start_rowg + 1, 3], [start_rowg + 11, 3])
        self.add_dropdown(sheet, '"' + ','.join(supply_kw) + '"', [start_rowg + 1, 4], [start_rowg + 11, 4])  
        self.add_dropdown(sheet, '"' + ','.join(supply_kw) + '"', [start_rowg + 1, 5], [start_rowg + 11, 5])
        self.add_dropdown(sheet, '"' + ','.join(supply_kw) + '"', [start_rowg + 1, 6], [start_rowg + 11, 6])            
        kw_ls = ['low_to_high', 'high_to_low','both']
        self.add_dropdown(sheet, '"' + ','.join(kw_ls) + '"', [start_rowg + 1, 7], [start_rowg + 11, 7])
        kw_ls = ['inputs', 'outputs','both']
        self.add_dropdown(sheet, '"' + ','.join(kw_ls) + '"', [start_rowg + 1, 8], [start_rowg + 11, 8])

        ls_ele,ls_exd,ls_no = self.get_impl_obj(sheet, start_rowg, 'PMLS') 
        if ls_ele:
            self.add_dropdown(sheet, '"' + ','.join(ls_ele) + '"', [start_rowg + 1, 9], [start_rowg + 11, 9])
        if ls_exd:
            self.add_dropdown(sheet, '"' + ','.join(ls_exd) + '"', [start_rowg + 1, 10], [start_rowg + 11, 10])
        if ls_no:
            self.add_dropdown(sheet, '"' + ','.join(ls_no) + '"', [start_rowg + 1, 11], [start_rowg + 11, 11])

        kw_ls = ['ELS_A', 'ELS_B','ELS_C','ELS_D', 'ELS_E', 'force', '-update']
        self.add_dropdown(sheet, '"' + ','.join(kw_ls) + '"', [start_rowg + 1, 12], [start_rowg + 11, 12])

        # PDName	SupplyIn	SupplyOut	EnCtrl	AckResp	OnState	OffState	CtrlAckSupply Comment
        # find PMPSW table 
        start_rowg = self.find_sheet(sheet, 'PMPSW')
        in_psw = []
        out_psw = []
        inpsw = []
        outpsw = []        
        psokw = re.findall(r'PSWRow\d+', ' '.join(self._upfdg._pmfile_data.keys()))
        for i in range(1, len(psokw) + 1):
            inpsw.append(self._upfdg._pmfile_data[f'PSWRow{i}']['InputPower'])
            outpsw.append(self._upfdg._pmfile_data[f'PSWRow{i}']['OutputPower'])
        for i in inpsw:
            for j in supply_kw:
                if re.search(r'_PSW', j):
                    in_psw.append(i + ' ' + j.split('_PSW')[0].strip())
        for i in outpsw:
            for j in supply_kw:
                if re.search(r'_PSW', j):
                    out_psw.append(i + ' ' + j)

        self.add_dropdown(sheet, '"' + ','.join(in_psw) + '"', [start_rowg + 1, 2], [start_rowg + 11, 2])
        self.add_dropdown(sheet, '"' + ','.join(out_psw) + '"', [start_rowg + 1, 3], [start_rowg + 11, 3])
        pso_ctl = []
        pso_ack = []

        psoctl = []
        psoack = []
        psw_ctl,psw_ack = self.get_impl_obj(sheet, start_rowg, 'PMPSW')
        #psoctl = self._upfdg._vfile_data['PSO_CTRL']
        psoctl1 = self.get_ctl_sig(self._upfdg._vfile_data['PSO_CTRL'])
        #psoack = self._upfdg._vfile_data['PSO_ACK']
        psoack1 = self.get_ctl_sig(self._upfdg._vfile_data['PSO_ACK'])
        if psw_ctl:
            psoctl.extend(psoctl1)
            psoctl.extend(psw_ctl)
        else:
            psoctl = psoctl1
        if psw_ack:
            psoack.extend(psoack1)
            psoack.extend(psw_ack)
        else:
            psoack = psoack1

        ctlpin = []
        ackpin = []       
        for i in range(1, len(psokw) + 1):
            ctlpin.append(self._upfdg._pmfile_data[f'PSWRow{i}']['CtrlPin'])
            ackpin.append(self._upfdg._pmfile_data[f'PSWRow{i}']['AckPin'])
        for i in ctlpin:
            for j in psoctl:
                pso_ctl.append(i + ' ' + j)
        for i in ackpin:
            for j in psoack:
                pso_ack.append(i + ' ' + j)       
        self.add_dropdown(sheet, '"' + ','.join(pso_ctl) + '"', [start_rowg + 1, 4], [start_rowg + 11, 4])
        self.add_dropdown(sheet, '"' + ','.join(pso_ack) + '"', [start_rowg + 1, 5], [start_rowg + 11, 5])

        on_st = []
        for i in inpsw:
            for j in psoctl:
                on_st.append('ONST' + ' ' + i + ' ' + j)
        self.add_dropdown(sheet, '"' + ','.join(on_st) + '"', [start_rowg + 1, 6], [start_rowg + 11, 6])
            
        off_st = []
        # ?? ! ~
        for j in psoctl:
            off_st.append('OFFST' + ' ' + j)
        self.add_dropdown(sheet, '"' + ','.join(off_st) + '"', [start_rowg + 1, 7], [start_rowg + 11, 7])  
        # self.add_dropdown(sheet, '"' + ','.join(supply_kw) + '"', [start_rowg + 1, 8], [start_rowg + 11, 8])
        self.add_dropdown(sheet, '"-update"', [start_rowg + 1, 8], [start_rowg + 11, 8])    

        # PDName	SupplyIn	SaveCtrl	RestCtrl	Elements	ExcludeList	NoRET	RetRegs	Comment
        # find PMRET table 
        start_rowg = self.find_sheet(sheet, 'PMRET')
        #retsave = ' '.join(self._upfdg._vfile_data['RET_SAVE'])
        retsave1 = self.get_ctl_sig(self._upfdg._vfile_data['RET_SAVE'])
        #retres = ' '.join(self._upfdg._vfile_data['RET_RES'])
        retres1 = self.get_ctl_sig(self._upfdg._vfile_data['RET_RES'])
        # retkw = re.findall(r'RETRow\d+', ' '.join(self._upfdg._pmfile_data.keys()))
        # self._upfdg._pmfile_data[f'RETRow{i}']['SavePin'] + ' | ' + self._upfdg._pmfile_data[f'RETRow{i}']['ResPin']
        # for i in range(1, len(retkw) + 1)
        self.add_dropdown(sheet, '"' + ','.join(supply_kw) + '"', [start_rowg + 1, 2], [start_rowg + 11, 2])

        retsave = []
        retres = []
        save_ret,res_ret,ele_ret,exd_ret,no_ret = self.get_impl_obj(sheet, start_rowg, 'PMRET')

        if  save_ret:
            retsave.extend(retsave1)
            retsave.extend(save_ret)
        else:
            retsave = retsave1
        if  res_ret:
            retres.extend(retres1)
            retres.extend(res_ret)
        else:
            retres = retres1

        ret_save = []
        for i in retsave:
            ret_save.append(i + ' high')
            ret_save.append(i + ' low')
            ret_save.append(i + ' positive')
            ret_save.append(i + ' negtive')
        self.add_dropdown(sheet, '"' + ','.join(ret_save) + '"', [start_rowg + 1, 3], [start_rowg + 11, 3])
        ret_res = []
        for i in retres:
            ret_res.append(i + ' high')
            ret_res.append(i + ' low')
            ret_res.append(i + ' positive')
            ret_res.append(i + ' negtive')
        self.add_dropdown(sheet, '"' + ','.join(ret_res) + '"', [start_rowg + 1, 4], [start_rowg + 11, 4])

        if ele_ret:
            self.add_dropdown(sheet, '"' + ','.join(ele_ret) + '"', [start_rowg + 1, 5], [start_rowg + 11, 5])
        if exd_ret:
            self.add_dropdown(sheet, '"' + ','.join(exd_ret) + '"', [start_rowg + 1, 6], [start_rowg + 11, 6])
        if no_ret:
            self.add_dropdown(sheet, '"' + ','.join(no_ret) + '"', [start_rowg + 1, 7], [start_rowg + 11, 7])
        self.add_dropdown(sheet, '"-update"', [start_rowg + 1, 9], [start_rowg + 11, 9])

        # # PDName	SupplyIn	ApplyPorts	Elements	ExcludeList	Comment
        # # find PMRPT table 
        # start_rowg = self.find_sheet(sheet, 'PMRPT')
        # self.add_dropdown(sheet, '"' + ','.join(supply_kw) + '"', [start_rowg + 1, 2], [start_rowg + 11, 2])
        # kw_rpt = ['inputs', 'outputs','both']
        # self.add_dropdown(sheet, '"' + ','.join(kw_rpt) + '"', [start_rowg + 1, 3], [start_rowg + 11, 3])

        # ele_rpt,exd_rpt = self.get_impl_obj(sheet, start_rowg, 'PMRPT')
        # if ele_rpt:
        #     self.add_dropdown(sheet, '"' + ','.join(ele_rpt) + '"', [start_rowg + 1, 4], [start_rowg + 11, 4])
        # if exd_rpt:
        #     self.add_dropdown(sheet, '"' + ','.join(exd_rpt) + '"', [start_rowg + 1, 5], [start_rowg + 11, 5])


    def read_data(self):
        sheet = self.get_sheet()
        self._psdata = self.get_table_contxt(sheet,'PStrategy')
        # print('_psdata: ', self._psdata)

    def dump_json(self,json_file):
        self._data = self._psdata
        self.write_json(json_file)


    def change_sheet(self):
        pass


class PModeSheet(BaseSheet):
    def __init__(self,*args):
        super().__init__(*args)
        self._pmdata = {}
        


    def update_sheet(self):
        
        sheet = self.get_sheet()

        supply_kw, supply_vol, supply_vss, supply_data = self.get_supply_infos()

        # find PMMODE table 
        start_rowg = self.find_sheet(sheet, 'PMMODE') 
        for i in range(1, len(supply_kw) + 1):
            sheet.cell(start_rowg, i + 1).value = supply_kw[i-1]
        sheet.cell(start_rowg, len(supply_kw) + 2).value = 'Comment'
        self.cell_style1(sheet, [start_rowg, 2], [start_rowg, len(supply_kw) + 2])

        self.cell_style2(sheet, [start_rowg + 1, 2], [start_rowg + 11, len(supply_kw) + 2])

        self.add_dropdown(sheet, '"' + ','.join(supply_vol) + '"', [start_rowg + 1, 2], [start_rowg + 11, len(supply_kw) + 2])
        self.add_dropdown(sheet, '"-update"', [start_rowg + 1, len(supply_kw) + 2], [start_rowg + 11, len(supply_kw) + 2])

    def read_data(self):
        sheet = self.get_sheet()
        self._pmdata = self.get_table_contxt(sheet,'PMode')
        # print('_pmdata: ', self._pmdata)

    def dump_json(self,json_file):
        self._data = self._pmdata
        self.write_json(json_file)

    def change_sheet(self):
        pass


#####################################################################################################

def modify_line_in_file(file_path, search_pattern, replacement):
    # 打开文件并逐行读取内容
    with open(file_path, 'r') as file:
        lines = file.readlines()

    # 遍历每一行并进行匹配和替换
    modified_lines = []
    for line in lines:
        if search_pattern in line:
            modified_line = line.replace(search_pattern, replacement)
            modified_lines.append(modified_line)
        else:
            modified_lines.append(line)

    # 将修改后的内容写回文件
    with open(file_path, 'w') as file:
        file.writelines(modified_lines)


def rm_exist_log(logdir):
    # LOG_DIR = os.getenv('TASK_LOGS_DIR')
    # logdir = os.path.join(LOG_DIR,taskid)
    # logdir = r'E:\stone\work\smalltool\pycharm\const0203_sysflatok_full\logs\adfeicc67ere'
    rlogfiles = os.listdir(logdir)
    for logfile in rlogfiles:
        # if logfile.endswith('.log') or logfile.endswith('.rpt'):
        if logfile in ['upf_dg.log','upf_gen.log','upf_gen.rpt','full_msg.log']:
            os.remove(os.path.join(logdir,logfile))


def printlog(context, file='upf_dg.log',logdir=''):
    # if not logdir:
    #     # LOG_DIR = os.getenv('TASK_LOGS_DIR')
    #     # logdir = os.path.join(LOG_DIR,taskid)
    #     logdir = r'E:\stone\work\smalltool\pycharm\const0203_sysflatok_full\logs\adfeicc67ere'
    #     rlogdir = os.path.join(logdir,file)
    # else:
    #     rlogdir = logdir

    if not logdir:
        LOG_DIR = os.getenv('TASK_LOGS_DIR')
        taskid = os.getenv('CURRENT_TASK_ID', 'default_task')
        if LOG_DIR:
            logdir = os.path.join(LOG_DIR, taskid)
            os.makedirs(logdir, exist_ok=True)
            rlogdir = os.path.join(logdir, file)
        else:
            # 如果环境变量不存在，使用当前目录
            rlogdir = file
    else:
        rlogdir = logdir

    if os.path.exists(rlogdir):
        with open(rlogdir, 'a') as fw:
            fw.write(context)
    else:
        with open(rlogdir, 'w') as fw:
            fw.write(context)

def movelogrpt(msgnm, flog, fdir):
    if os.path.exists(f'{msgnm}'):
        if os.path.exists(flog):
            # with open('upf_gen.log','r') as fh:
            #     for line in fh.readlines():
            #         #txt_list.append(line.strip())
            #         txt_list += f'{line.strip()} \n'
            # with open(logfile, 'a') as fw:
            #     fw.write(txt_list)
            os.system(f'rm -f {flog}')
            os.system(f'mv {msgnm} {fdir}')
        else:
            os.system(f'mv {msgnm} {fdir}')
    else:
        upf_warn(f'Can not find {msgnm}.')


#########################################################################################################################
# upf message
full_log_message_list =  []
full_rpt_message_list =  []

def upf_log(level, msg, out=sys.stdout):
    print(f'{level.upper()}: {msg}', flush=True, file=out)

def upf_info(msg, kw='log'):
    # if msg not in full_log_message_list:
    if kw == 'log':
        full_log_message_list.append(msg)
        printlog(f'UPF_INFO: {msg} \n', 'full_msg.log')
    else:
        full_rpt_message_list.append(msg)
        printlog(f'UPF_INFO: {msg} \n', 'full_chk.rpt')
    # upf_log('UPF_INFO', msg)
    if kw == 'log':
        printlog(f'UPF_INFO: {msg} \n')
    if kw == 'chk_hym':
        printlog(f'UPF_INFO: {msg} \n', 'chk_hym.rpt')
    if kw == 'chk_sht':
        printlog(f'UPF_INFO: {msg} \n', 'chk_sht.rpt')
    if kw == 'chk_dti':
        printlog(f'UPF_INFO: {msg} \n', 'chk_dti.rpt')
    if kw == 'chk_upf':
        printlog(f'UPF_INFO: {msg} \n', 'chk_upf.rpt')

def upf_warn(msg, kw='log'):
    # if msg not in full_log_message_list:
    if kw == 'log':
        full_log_message_list.append(msg)
        printlog(f'UPF_WARN: {msg} \n', 'full_msg.log')
    else:
        full_rpt_message_list.append(msg)
        printlog(f'UPF_WARN: {msg} \n', 'full_chk.rpt')
    print(f'\033[0:31mUPF_WARN\033[0m: {msg}', flush=True)
    # print(f'UPF_WARN: {msg}', flush=True)
    if kw == 'log':
        printlog(f'UPF_WARN: {msg} \n')
    if kw == 'chk_hym':
        printlog(f'UPF_WARN: {msg} \n', 'chk_hym.rpt')
    if kw == 'chk_sht':
        printlog(f'UPF_WARN: {msg} \n', 'chk_sht.rpt')
    if kw == 'chk_dti':
        printlog(f'UPF_WARN: {msg} \n', 'chk_dti.rpt')
    if kw == 'chk_upf':
        printlog(f'UPF_WARN: {msg} \n', 'chk_upf.rpt')
         

def upf_error(msg, kw='log'):
    # if msg not in full_log_message_list:
    if kw == 'log':
        full_log_message_list.append(msg)
        printlog(f'UPF_ERROR: {msg} \n', 'full_msg.log')
    else:
        full_rpt_message_list.append(msg)
        printlog(f'UPF_ERROR: {msg} \n', 'full_chk.rpt')
    print(f'\033[0:31mUPF_ERROR\033[0m: {msg}', flush=True)
    # print(f'UPF_ERROR: {msg}', flush=True)
    if kw == 'log':
        printlog(f'UPF_ERROR: {msg} \n')
    if kw == 'chk_hym':
        printlog(f'UPF_ERROR: {msg} \n', 'chk_hym.rpt')
    if kw == 'chk_sht':
        printlog(f'UPF_ERROR: {msg} \n', 'chk_sht.rpt')
    if kw == 'chk_dti':
        printlog(f'UPF_ERROR: {msg} \n', 'chk_dti.rpt')
    if kw == 'chk_upf':
        printlog(f'UPF_ERROR: {msg} \n', 'chk_upf.rpt')

def upf_fatal(msg, kw='log'):
    # if msg not in full_log_message_list:
    if kw == 'log':
        full_log_message_list.append(msg)
        printlog(f'UPF_FATAL: {msg} \n', 'full_msg.log')
    else:
        full_rpt_message_list.append(msg)
        printlog(f'UPF_FATAL: {msg} \n', 'full_chk.rpt')
    print(f'\033[0:31mUPF_FATAL\033[0m: {msg}', flush=True)
    # print(f'UPF_FATAL: {msg}', flush=True)
    if kw == 'log':
        printlog(f'UPF_FATAL: {msg} \n')
    if kw == 'chk_hym':
        printlog(f'UPF_FATAL: {msg} \n', 'chk_hym.rpt')
    if kw == 'chk_sht':
        printlog(f'UPF_FATAL: {msg} \n', 'chk_sht.rpt')
    if kw == 'chk_dti':
        printlog(f'UPF_FATAL: {msg} \n', 'chk_dti.rpt')
    if kw == 'chk_upf':
        printlog(f'UPF_FATAL: {msg} \n', 'chk_upf.rpt')
           
    # sys.exit(1)


def upf_args():
    print(f'##++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++')
    print(f'## UPF Generation Optional Arguments Presentation:')
    print(f'## -hlp:        All of flow option usage reference.')
    print(f'## -usg:        Flow step command usage reference.')
    print(f'## -tmp:        Write out template UPFs for reference')
    print(f'## -gen_dir:    UPF generationg directory')
    print(f'## -hier_yaml:  Design hierarchy info. from user input file')
    print(f'## -setup:      Build setup directories from blocks defined in hier yaml')
    print(f'## -blocks:     Indicate only current design name for UPF generation')
    #print(f'## -flat:       Generate flatten UPF based on current design. If not flat, only generate current design level only UPF')
    print(f'## -dg:         Generate design guide file to be present UPF request format')
    #print(f'## -idg:        Incrementally update design guide file based on existed dg file and updated input files')
    print(f'## -upf:        Write out UPF files')
    print(f'## -check_hym: Check hier yaml data before generating UPF')
    print(f'## -check_only: Check input data before generating UPF')
    print(f'## -check_upf:  Check UPF consistency after generating UPF')
    print(f'## -proj:       Open project mode. Maybe need set some related project environment variables')
    print(f'##++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++')


def upf_usage():
    print(f'##++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++')
    print(f'## Before flow run, user must provide input data including hier yaml and empty vfile, dg file and crg/ip UPFs to be integrated ')
    print(f'## For hier yaml and empty vfile, must follow format of hier_pwr.yaml and pmempty.v in template/ftemp')
    print(f'## For crg/ip UPFs, must follow header format of crg.upf and userip.upf in template/ftemp')
    print(f'##++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++')
    print(f'##++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++')
    print(f'## Step I: Build the whole UPF generation directory structure.')
    print(f'## Cmd Line: xconst sdgen -gen_dir <upfdir> -hier_yaml <hier_file> -setup -blocks <blk_name> [-tmp]')
    print(f'##++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++')
    print(f'##++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++')
    print(f'## Step II: Generate initial design guide file according to current input data.')
    print(f'## Cmd Line: xconst sdgen -gen_dir <upfdir> -hier_yaml <hier_file> -dg -blocks <blk_name>')
    print(f'##++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++')
    print(f'##++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++')
    print(f'## Step III: Check and debug input data after users provide all of input files.')
    print(f'## Cmd  Line: xconst sdgen -gen_dir <upfdir> -hier_yaml <hier_file> -chk_only -blocks <blk_name>')
    print(f'##++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++')
    print(f'##++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++')
    print(f'## Step IV: Generate UPF files if all of input files are OK.')
    print(f'## Cmd  Line: xconst sdgen -gen_dir <upfdir> -hier_yaml <hier_file> -upf -blocks <blk_name> [-proj] [-flat]')
    print(f'##++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++')
    print(f'##++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++')
    print(f'## Step V: Check and debug UPF output files.')
    print(f'## Cmd  Line: xconst sdgen -gen_dir <upfdir> -hier_yaml <hier_file> -chk_upf -blocks <blk_name>')
    print(f'##++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++')


ORG_COMDIR = os.getenv('ECS_TEMPLATES_DIR')
TEMP_DIR = os.getenv('TEMP_UPLOAD_DIR')
LOG_DIR = os.getenv('TASK_LOGS_DIR')
# ORG_COMDIR = r'E:\stone\work\smalltool\pycharm\const0203_sysflatok_full\templates'
# TEMP_DIR = r'E:\stone\work\smalltool\pycharm\const0203_sysflatok_full\temp'
# LOG_DIR = r'E:\stone\work\smalltool\pycharm\const0203_sysflatok_full\logs'
# ufpadfeicc67ere

def upf_dg_gen(*arglist):
    if len(arglist) == 0: arglist = ['-h']
    parser = argparse.ArgumentParser(prog='upf_dg_gen', description='design constraint excel file generation script')
    parser.add_argument('-taskid', help='Task ID name', default='ufpadfeicc67ere', required='False')
    parser.add_argument('-dg', help='Write or update design guide files', action='store_true')
    parser.add_argument('-usr', help='User permission for upf generation')

    args = parser.parse_args(args=arglist)
    taskid = args.taskid

    if args.dg:
        hier_yaml = os.path.join(TEMP_DIR, taskid, 'hier.yaml')
        pvfile = os.path.join(TEMP_DIR, taskid, 'pvlog.v')
        logdir = os.path.join(LOG_DIR,taskid)
        pmfile = os.path.join(TEMP_DIR, taskid, 'pcell.yaml')
        objfile = os.path.join(TEMP_DIR, taskid, 'pobj.tcl')

        # tfile = os.path.join(TEMP_DIR, 'tune.sdc')
        # tfile = os.path.join(ORG_COMDIR, 'tune.sdc')

        dg_temp = os.path.join(ORG_COMDIR, 'upfgen','pcont_org.xlsx')
        dg_file = os.path.join(TEMP_DIR, taskid,'pcont.xlsx')
        dg_file = re.sub(r'/+', '/', dg_file)
        lock_file = f'{TEMP_DIR}/.~lock.pcont.xlsx#'
        lock_file = re.sub(r'/+', '/', lock_file)

        # check pcont excel file existence
        if not os.path.exists(dg_temp):
            upf_error(f'pcont excel file not found {dg_temp}')
            exit(1)

        # check hier yaml existence
        if not os.path.exists(hier_yaml):
            upf_error(f'hier yaml file not found {hier_yaml}')
            exit(1)

        # check pvfile existence
        if not os.path.exists(pvfile):
            upf_error(f'Empty vfile not found {pvfile}')
            exit(1)

        # check pobj existence
        if not os.path.exists(objfile):
            upf_error(f'object file not found {objfile}')
            exit(1)

        # check pmfile existence
        if not os.path.exists(pmfile):
            upf_error(f'pmcell file not found {pmfile}')
            exit(1)

        if os.path.exists(dg_temp):
            # os.system(f'cp -f {dg_temp} {dg_file}')

            # 使用 shutil 复制文件，自动处理不同操作系统的差异
            shutil.copy2(dg_temp, dg_file)  # 保留文件元数据
            upf_info(f'SDC design guide not found and copy from template design guide file at first time.')
        else:
            upf_error(f'SDC original design guide NOT been found in template directory.')
            exit(1)

        # check logdir existence
        if not os.path.exists(logdir):
            upf_error(f'logdir not found {logdir}')
            exit(1)

        if os.path.exists(lock_file):
            upf_fatal('pcont.xlsx is in edit mode. Please close it')

        rm_exist_log(logdir)
        hier_tree = HierPwrTree(hier_yaml)
        upfdg = UPF_DG()
        upfdg.hier_tree = hier_tree

        upfdg.read_vfile(pvfile)
        upfdg.read_pmfile(pmfile)
        upfdg.read_objfile(objfile)

        if os.path.exists(dg_file):
            upfdg.load_design_guide(dg_file)

        upfdg.update_dg()
        upfdg.save_workbook(dg_file)

        upf_info(f'Design guide file {dg_file} is updated.')

if __name__ == '__main__':
    if len(sys.argv) < 2 or (len(sys.argv) > 2 and 'upf_dg_gen' not in sys.argv[1]):
        upf_error('Missing some parameters for SDC generation')
        locals()['upf_dg_gen']('-h')
        exit(1)
    app_name = sys.argv[1]
    if app_name in locals():
        locals()[app_name](*sys.argv[2:])
    else:
        raise NameError(f'The application of DataBase generation {app_name} not found')

