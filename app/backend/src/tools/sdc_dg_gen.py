#!/usr/bin/env python3


import os,sys
import yaml
import re
import json

from os.path import dirname, abspath, basename
import time

import argparse
from glob import glob

import shutil

import yaml
#from itertools import chain

import openpyxl
from openpyxl import worksheet
from pprint import pprint
import pandas as pd
from openpyxl.utils import get_column_letter

import tkinter as tk

from openpyxl.styles import Border, Side, PatternFill, Font, Alignment 
from openpyxl.worksheet.datavalidation import DataValidation

# Define some attributes and various methods for every block in hier tree
class BaseBlock():
    def __init__(self,name):
        self._name = name
        self._alias = ''
        self._hdlevel = 'blk'
        self._prime_pwr = ''
        self._constr_dir = ''
        self._insts = []
        self._mac_insts = []
        self._dig_insts = []
        self._cust_insts = {}
        self._proj = False

    def __repr__(self):
        return '<%s name=%s alias=%s>' % (self.__class__.__name__, self._name, self._alias)

    @property 
    def hdlevel(self):
        return self._hdlevel

    @hdlevel.setter
    def hdlevel(self, level):
        supported_lvs = ('top', 'sys', 'blk', 'soft', 'lib', 'crg', 'pll')
        if level not in supported_lvs:
            sdc_error(f'Unsupported HDLEVEL {level} of block {self._name}, should be one of {supported_lvs}')
            return
        else:
            if level == 'sys' or level == 'top':
                self._hdlevel = 'sys'
            else:
                self._hdlevel = level

    @property
    def lvl_flat(self):
        if self._hdlevel == 'sys':
            return 'IS_CHIP'
        else:
            return 'IS_FLAT'
        
    @property
    def alias(self):
        return self._alias

    @alias.setter
    def alias(self, alias):
        self._alias = alias

    @property
    def prime_pwr(self):
        return self._prime_pwr

    @prime_pwr.setter
    def prime_pwr(self, pwr):
        self._prime_pwr = pwr

    @property
    def insts(self):
        return self._insts

    @insts.setter
    def insts(self, insts):
        self._insts = insts

    @property
    def mac_insts(self):
        return self._mac_insts

    @mac_insts.setter
    def mac_insts(self, mac_insts):
        self._mac_insts = mac_insts

    @property
    def dig_insts(self):
        return self._dig_insts

    @dig_insts.setter
    def dig_insts(self, dig_insts):
        self._dig_insts = dig_insts

    @property
    def constr_dir(self):
        return self._constr_dir
    
    @constr_dir.setter
    def constr_dir(self, consdir):
        self._constr_dir = consdir

    @property
    def proj(self):
        return self._proj

    @proj.setter
    def proj(self, proj):
        self._proj = proj

    def get_curcust_by_name(self, inst_type,flg=''):
        _hier = []
        _ref = []
        #_lvl = []
        _alias = []
        _pwr = []
        _usersdc = []
        if inst_type == 'insts' and self.insts:
            for i in range(0,len(self.insts)):
                finst = self.insts[i].split(',')
                if len(finst) == 3:
                    _hier.append(finst[0].strip())
                    _ref.append(finst[1].strip())
                    _alias.append(None)
                    _pwr.append(finst[2].strip())
                    _usersdc.append(None)
                if len(finst) == 4:
                    _hier.append(finst[0].strip())
                    _ref.append(finst[1].strip())
                    if not flg:
                        _alias.append(finst[2].strip().replace('#',''))
                    else:
                        _alias.append(finst[2].strip())
                    _pwr.append(finst[3].strip())
                    _usersdc.append(None)
            self._cust_insts['insthier'] = _hier
            self._cust_insts['instref'] = _ref
            self._cust_insts['instalias'] = _alias
            self._cust_insts['instpwr'] = _pwr
            self._cust_insts['instuser'] = _usersdc
        
        if inst_type == 'mac_insts' and self.mac_insts:
            for i in range(0,len(self.mac_insts)):
                if isinstance(self.mac_insts[i],str):
                    fmac = self.mac_insts[i].split(',')  
                    if len(fmac) == 3:
                        _hier.append(fmac[0].strip())
                        _ref.append(fmac[1].strip())
                        _alias.append(None)
                        _pwr.append(fmac[2].strip())
                        _usersdc.append(None)
                    if len(fmac) == 4:
                        _hier.append(fmac[0].strip())
                        _ref.append(fmac[1].strip())
                        if not flg:
                            _alias.append(fmac[2].strip().replace('#',''))
                        else:
                            _alias.append(fmac[2].strip())
                        _pwr.append(fmac[3].strip())
                        _usersdc.append(None)
                if isinstance(self.mac_insts[i],dict):
                    fmac = ''.join(self.mac_insts[i].keys()).split(',')
                    _hier.append(fmac[0].strip())
                    _ref.append(fmac[1].strip()) # + '_USR' )
                    _alias.append(None)
                    _pwr.append(fmac[2].strip())
                    _usersdc.append(''.join(self.mac_insts[i].values()))  
            self._cust_insts['machier'] = _hier
            self._cust_insts['macref'] = _ref
            self._cust_insts['macalias'] = _alias
            self._cust_insts['macpwr'] = _pwr
            self._cust_insts['macuser'] = _usersdc

        if inst_type == 'dig_insts' and self.dig_insts:
            for i in range(0,len(self.dig_insts)):
                if isinstance(self.dig_insts[i],str):
                    fdig = self.dig_insts[i].split(',')  
                    if len(fdig) == 3:
                        _hier.append(fdig[0].strip())
                        _ref.append(fdig[1].strip())
                        _alias.append(None)
                        _pwr.append(fdig[2].strip())
                        _usersdc.append(None)
                    if len(fdig) == 4:
                        _hier.append(fdig[0].strip())
                        _ref.append(fdig[1].strip())
                        if not flg:
                            _alias.append(fdig[2].strip().replace('#',''))
                        else:
                            _alias.append(fdig[2].strip())
                        _pwr.append(fdig[3].strip())
                        _usersdc.append(None)
                if isinstance(self.dig_insts[i],dict):
                    fdig = ''.join(self.dig_insts[i].keys()).split(',')
                    _hier.append(fdig[0].strip())
                    _ref.append(fdig[1].strip()) # + '_USR')
                    _alias.append(None)
                    _pwr.append(fdig[2].strip())
                    _usersdc.append(''.join(self.dig_insts[i].values()))
            self._cust_insts['dighier'] = _hier
            self._cust_insts['digref'] = _ref
            self._cust_insts['digalias'] = _alias
            self._cust_insts['digpwr'] = _pwr
            self._cust_insts['diguser'] = _usersdc       

        return self._cust_insts

    def get_curhd_by_name(self):
        #return self.name.split() + self._cust_insts['instref']
        self.get_curcust_by_name('insts')
        if 'instref' in self._cust_insts:
            return self._cust_insts['instref']

    def get_curmac_by_name(self,flg=''):
        self.get_curcust_by_name('mac_insts',flg)
        if 'macref' in self._cust_insts:
            return self._cust_insts['macref']

    def get_curdig_by_name(self,flg=''):
        self.get_curcust_by_name('dig_insts',flg)
        if 'digref' in self._cust_insts:
            return self._cust_insts['digref']
    
    def get_curuser_by_name(self, inst_type):
        self.get_curcust_by_name('mac_insts')
        self.get_curcust_by_name('dig_insts')
        if inst_type == 'mac_insts' and 'macuser' in self._cust_insts:           
            return self._cust_insts['macuser']       
        elif inst_type == 'dig_insts'and 'diguser' in self._cust_insts:
            return self._cust_insts['diguser']
        else:
            return None



class HierPwrTree():
    def __init__(self,yaml_file):
        self.yaml_file = yaml_file
        self._blocks = {}
        self._primepwr = {}
        self._yaml_data = {}
        self._hierdata = {}
        self._pwrdata = {}
        #self._blktrees = {}
        self.build_hier_tree(yaml_file)
        

    def build_hier_tree(self, yaml_file):

        # get yaml_data
        yaml_data = {}
        if not os.path.exists(yaml_file):
            raise FileExistsError(f'{yaml_file} does not exists')
        with open(yaml_file, 'r') as fh:
            yaml_data = yaml.load(fh, yaml.FullLoader)

        if 'hier' not in yaml_data:
            print('Missing hier keyword in yaml file.')
            upf_fatal(f'Must include keyword <hier>')
        if 'pwr' not in yaml_data:
            print('Missing pwr keyword in yaml file.')
            upf_fatal(f'Must include keyword <pwr>')

        # get '_primepwr'
        for pwr_name in yaml_data['pwr'].keys():
            if yaml_data['pwr'][pwr_name]:
                self._primepwr[pwr_name] = yaml_data['pwr'][pwr_name]   

        for blk_name in yaml_data['hier'].keys():

            self._blocks[blk_name] = BaseBlock(blk_name)

            if 'alias' in yaml_data['hier'][blk_name] and yaml_data['hier'][blk_name]['alias']:
                self._blocks[blk_name].alias = yaml_data['hier'][blk_name]['alias']
            else:
                self._blocks[blk_name].alias = None

            if 'hdlevel' in yaml_data['hier'][blk_name] and yaml_data['hier'][blk_name]['hdlevel']:
                self._blocks[blk_name].hdlevel = yaml_data['hier'][blk_name]['hdlevel']
            else:
                self._blocks[blk_name].hdlevel = None            
            
            if 'prime_pwr' in yaml_data['hier'][blk_name] and yaml_data['hier'][blk_name]['prime_pwr'] in self._primepwr and yaml_data['hier'][blk_name]['prime_pwr']:
                self._blocks[blk_name].prime_pwr = yaml_data['hier'][blk_name]['prime_pwr'] + ' ' + self._primepwr[yaml_data['hier'][blk_name]['prime_pwr']]
            else:
                self._blocks[blk_name].prime_pwr = None 

            if 'constr_dir' in yaml_data['hier'][blk_name] and yaml_data['hier'][blk_name]['constr_dir']:
                self._blocks[blk_name].constr_dir = yaml_data['hier'][blk_name]['constr_dir']
            else:
                self._blocks[blk_name].constr_dir = None

            if 'insts' in yaml_data['hier'][blk_name] and yaml_data['hier'][blk_name]['insts']:
                self._blocks[blk_name].insts = yaml_data['hier'][blk_name]['insts']
            else:
                self._blocks[blk_name].insts = None

            if 'mac_insts' in yaml_data['hier'][blk_name] and yaml_data['hier'][blk_name]['mac_insts']:
                self._blocks[blk_name].mac_insts = yaml_data['hier'][blk_name]['mac_insts']
            else:
                self._blocks[blk_name].mac_insts = None

            if 'dig_insts' in yaml_data['hier'][blk_name] and yaml_data['hier'][blk_name]['dig_insts']:
                self._blocks[blk_name].dig_insts = yaml_data['hier'][blk_name]['dig_insts']
            else:
                self._blocks[blk_name].dig_insts = None

        self._hierdata = yaml_data['hier'] 
        self._pwrdata = yaml_data['pwr']
        self._yaml_data = yaml_data

    def get_block_by_name(self, name) -> BaseBlock:
        if name in self._blocks:
            return self._blocks[name]
        else:
            return None    

    def get_curblks(self,name):
        curblks = []
        
        allblks = list(self._hierdata.keys())
        if name in allblks:
            curblk = self.get_block_by_name(name)

            if curblk.get_curhd_by_name():
                curblks = [x for x in curblk.get_curhd_by_name() if x is not None]
            if curblk.get_curmac_by_name():
                curblks += [x for x in curblk.get_curmac_by_name() if x is not None]
            if curblk.get_curdig_by_name():
                curblks += [x for x in curblk.get_curdig_by_name() if x is not None]  
        else:
            sdc_warn(f'{name} is not expanded in hier_pwr yaml file.')
        
        return  curblks
    
    def get_hiertrees(self, name, blktrees={}, valstyle=None) -> dict:

        curblks = self.get_curblks(name)
        #blktrees = {}
        if curblks:
            new_curblks = [elem.replace('_USR', '') if re.search(r'_USR$',elem) else elem for elem in curblks]
            blktrees[name] = new_curblks
        else:
            if valstyle:
                blktrees[name] = None   

        for blk in curblks:
            if re.search(r'_USR$',blk):
                blk = blk.replace('_USR','')
                sdc_warn(f'{blk} is not expanded in hier_pwr yaml file.')
            elif len(curblks) > 0:
                self.get_hiertrees(blk,blktrees)    
        
        return blktrees

    def get_hierblks(self, name) -> list:

        blktrees = self.get_hiertrees(name)
        result = []
        for key, value in blktrees.items():
            if key not in result:
                result.append(key)
            if isinstance(value, list):
                for element in value:
                    if element not in result:
                        result.append(element)
        return result        

    def get_hierblks_infos(self,name) -> dict:

        blks = self.get_hierblks(name)
        blksinfo = {}

        allblks = list(self._hierdata.keys())
        for val in blks:
            if val in allblks:
                blk = self.get_block_by_name(val)
                alias = blk.alias
                lvl = blk.hdlevel
                condir = blk.constr_dir
                pwr = blk.prime_pwr
                pwrg = pwr.split(' ')[0].strip()
                blksinfo[val] = val + f' {alias}' + f' {lvl}' + f' {pwrg}' + f' {condir}'
            else:
                blksinfo[val] = val + ' (NOT EXPEND)'

        return blksinfo 


    def get_hierlvlblks(self, name, outtype='hd') -> list:
        '''
        get different harden insts, mac insts, dig insts under current design
        outtype is hd/lib/soft/crg/pll
        '''
        allhierblks = []
        allhierblks = self.get_hierblks(name)
        #alltreeblks = self.get_hiertrees(name)
        allhierblks.append(name)

        #allblks = list(self._hierdata.keys())

        insts = []
        macs = []
        digs = []
        plls = []
        crgs = []
        if allhierblks:
            for curblk in allhierblks:
                #if curblk in allblks:
                blk = self.get_block_by_name(curblk)
            
                lvl = blk.hdlevel
                if lvl in ['blk', 'sys', 'top']:
                    insts.append(curblk)
                if lvl in ['lib']:
                    macs.append(curblk)  
                if lvl in ['pll']:
                    plls.append(curblk) 
                if lvl in ['soft']:
                    digs.append(curblk)  
                if lvl in ['crg']:
                    crgs.append(curblk)
                # else:
                #     #print(f'XYZ_{curblk}')
                #     for ky,vl in alltreeblks.items():
                #         if curblk in vl:
                #             parent = ky
                #     blkg = self.get_block_by_name(parent)
                #     if curblk in blkg.get_curhd_by_name():
                #         insts.append(curblk)
                #         sdc_warn(f'{curblk} is harden block need expand it in hier_pwr yaml file.')
                #     if curblk in blkg.get_curmac_by_name() or f'{curblk}_USR' in blkg.get_curmac_by_name():
                #         macs.append(curblk)
                #         sdc_warn(f'{curblk} is macro block, if not user constraint, need expand it in hier_pwr yaml file.')
                #     if curblk in blkg.get_curdig_by_name() or f'{curblk}_USR' in blkg.get_curdig_by_name():
                #         digs.append(curblk)
                #         sdc_warn(f'{curblk} is digital block, if not user constraint, need expand it in hier_pwr yaml file.')

        if outtype == 'hd':
            return insts
        if outtype == 'lib':
            return macs
        if outtype == 'soft':
            return digs  
        if outtype == 'pll':
            return plls
        if outtype == 'crg':
            return crgs  

    def get_hierdepth(self,dic,key):
        
        allblks = list(self._hierdata.keys())
    
        if key not in dic:
            return 1
        else:
            max_depth = 0
            for sub_key in dic[key]:
                if sub_key not in allblks:
                    sdc_warn(f'{sub_key} is not expanded in hier yaml file.')
                else:
                    current_depth = self.get_hierdepth(dic, sub_key) + 1
                    max_depth = max(max_depth, current_depth)
            return max_depth

        # # blktrees = self.get_hiertrees(name)
        # if key not in dic:
        #     return 0
        
        # depths = []
        # for child_key in dic[key]:
        #     if child_key not in allblks:
        #         print(f'{child_key} is not expanded in hier yaml file.')
        #     else:
        #         depths.append(self.get_hierdepth(dic,child_key))
        #         #max_depth = max(max_depth, depth)
        
        # return max(depths) + 1

        # max_depth = depth  # 记录最大深度
        
        # if isinstance(blktrees, dict):
        #     for child_node in blktrees.values():
        #         if child_node not in allblks:
        #             print(f'{child_node} is not expanded in hier yaml file.')
        #     else:
        #         child_depth = self.get_hierdepth(child_node, depth + 1)
        #         max_depth = max(max_depth, child_depth)
        
        # return max_depth

    def get_alias_by_name(self, name):
        return self._blocks[name].alias



class SDC_DG(object):
    def __init__(self):
        self._sheets = {}
        self._hier_tree = {}
        self._vardef = {}
        self._wb = {}
        self._sdcdir = ''
        self._mdname = ''
        self._alias = '' #self._hier_tree._blocks[self._mdname].alias
        self._hdlvl = ''
        self._pwr = ''
        self._inputs = BaseInputs()
        # self._iodly = VIODly()

        #self.proj_mode = False
        self._vfile_data = None
        self._vfile_list = None
        self._data = None
        self._lvl = 'blk'
        self._flt = 'IS_FLAT'

        self._vardata = {}
        # self._sdc_dir = ''
        # self._com_dir= ''
        # self._dft_dir = ''
        # self._hd_process = ''
        # # self._hd_mod_name = self._mdname
        # self._cycle_list = '[list CYCLE500M]'


    @property
    def hier_tree(self):
        return self._hier_tree
    
    @hier_tree.setter
    def hier_tree(self, hier_tree):
        self._hier_tree = hier_tree
 
    def load_design_guide(self,dg_file,kwd=''):
        self._sdcdir = abspath(dirname(dirname(dg_file)))
        self._wb = openpyxl.load_workbook(dg_file)

        valdef_sheet = self._wb['VarDef']
        start = 0
        for i in range(1, valdef_sheet.max_row+1):
            if valdef_sheet.cell(i,1).value == 'Variable':
                start = i + 1
                break
        for i in range(start, valdef_sheet.max_row+1):
            key = valdef_sheet.cell(row=i, column=1).value
            val = valdef_sheet.cell(row=i, column=2).value
            self._vardef[key] = val

        self._sheets = {
            'VarDef'            : VarDefSheet(self, 'VarDef'),
            'ClkDef'            : ClkDefSheet(self, 'ClkDef'),
            'IODly'             : IODlySheet(self, 'IODly'),
            'Exp'               : ExpSheet(self, 'Exp'),
            # 'IOExp'             : IOExpSheet(self, 'IOExp'),
            # 'IntExp'            : IntExpSheet(self, 'IntExp'),
        }

        # read and convert excel to json data
        if kwd == 'json':
            for sheetname,sheet in self._sheets.items():
                if sheetname == 'ClkDef':
                    sheet.read_data('json')
                else:
                    sheet.read_data()            
                # json_file = dirname(dirname(dg_file)) + '/json' + f'/{sheetname.lower()}.json'
                # # json_file = dirname(dirname(dg_file)) + os.path.join('json', f'{sheetname.lower()}.json')
                # sheet.dump_json(json_file)

            # json_file = dirname(dirname(dg_file)) + '/json' + f'/hier_pwr.json'
            # # json_file = dirname(dirname(dg_file)) + os.path.join('json', 'hier.json')
            # self._data = self._hier_tree._yaml_data
            # self.write_json(json_file)
            

    def read_vfile(self,vfile,kwd=''):
        self._vfile_list, self._vfile_data = self._inputs.read_vfile(vfile)
        self._mdname = self._vfile_data['module_name']
        # print(self._vfile_list)
        # print(self._vfile_data)

        # if kwd == 'json':
        #     self._data = self._vfile_data
        #     json_file = dirname(dirname(vfile)) + '/json' + f'/vfile.json'
        #     self.write_json(json_file)    

    def update_dg(self):
        for sht in self._sheets.values():
            sht.update_sheet()

    def check_dg(self):
        for sht in self._sheets.values():
            sht.check_sheet()

    def change_dg(self,dgfile):
        for sht in self._sheets.values():
            sht.change_sheet(dgfile)

    def read_json(self,file_path):
        sblk_data = {}
        if os.path.exists(file_path):
            with open(file_path,'r') as fw:
                content = fw.read()
                sblk_data = json.loads(content)

        #print('sblk_data:',sblk_data)
        return sblk_data


    def write_json(self,filepath):
        os.makedirs(dirname(filepath), exist_ok=True)
        jsonstr = json.dumps(self._data, indent=4)
        with open(filepath,'w') as fw:
            print(jsonstr, file=fw) 

    # def save_text(self,context,file):
    #     with open(file, 'w') as fw:
    #         fw.write(context)

    def save_text(self, context,file,kw='xyz'):
        if os.path.exists(file) and 'proc' in kw:
            with open(file, 'a') as fw:
                fw.write(context)
        else:
            with open(file, 'w') as fw:
                fw.write(context)

    def save_workbook(self,output):
        self._wb.save(output)

    def read_text(self, file):
        if not os.path.exists(file):
            raise FileExistsError(f'{file} does not exists')
            # sdc_error(f'{file} not exist. Please check it.')
            # exit(1)
        else:
            txt_list = []
            with open(file,'r') as fh:
                for line in fh.readlines():
                    txt_list.append(line)
        
            return txt_list


# sdcdg is XsdcDesignGuide object
class BaseSheet(object):
    def __init__(self, sdcdg, sheetname):
        self._sdcdg = sdcdg
        self._sheetname = sheetname
        self._data = []
        #self._vardef = {}
        self._pdnmdict = {}
    
    def get_sheet(self):
        return self._sdcdg._wb[self._sheetname]

    def read_data(self):
        raise NotImplementedError(self.__class__.__name__ + ' read_data not implemented yet')

    def write_json(self, filepath):
        os.makedirs(dirname(filepath), exist_ok=True)
        jsonstr = json.dumps(self._data, indent=4)
        with open(filepath,'w') as fw:
            print(jsonstr, file=fw)

    def find_sheet(self, sheet, skw):
        start_rowg = 1
        # TABCONST = ['TMVAR','TMHIER','TMCLK','TMIODLY','TMIOEXP','TMINOUT','TMINTEXP','TMSTPGATE']
        TABCONST = ['TMVAR','TMCLK','TMIODLY','TMIOEXP','TMINOUT','TMINTEXP','TMSTPGATE']
        #print(skw,sheet)
        for i in range(1,sheet.max_row+1):
            if skw in TABCONST and sheet.cell(i,1).value == skw:
                start_rowg = i + 1
                break  
        return  start_rowg 

    def get_vardef_value(self, sheet):
        vardef = {}
        start_rowg = self.find_sheet(sheet, 'TMVAR')
        # end_rowg = self.find_sheet(sheet, 'TMHIER')
        for i in range(start_rowg + 1, start_rowg + 15):
            key = sheet.cell(row=i, column=1).value
            val = sheet.cell(row=i, column=2).value
            vardef[key] = val

        vardef['SDC_DIR'] = self._sdcdg._sdcdir
        vardef['COM_DIR'] = self._sdcdg._sdcdir
        vardef['DFT_DIR'] = ''
        # vardef['HD_MOD_NAME'] = self._sdcdg._mdname
        vardef['HD_PROCESS'] = ''
        vardef['CYCLE_LIST'] = '[list CYCLE500M]'

        # print('vardef:', vardef)
        return vardef   

    def set_name_style(self, kw):
        #time_stamp = time.strftime("%Y%m%d%H%M%S", time.localtime())
        #CONST = f'Generic_Xsdc_{time_stamp}'
        CONST = f'Generic_XSDC'
        return kw + '_' + CONST

    # showErrorMessage=False,showDropDown=True
    def add_dropdown_short(self, sheet, options, start, end):       
        dv = DataValidation(type="list", formula1=options, showErrorMessage=False)
        sheet.add_data_validation(dv)
        if len(start) == 2 and len(end) == 2:
            for i in range(start[0], end[0] + 1):
                for j in range(start[1], end[1] + 1):
                    dv.add(sheet.cell(i,j))
        if len(start) == 1 and len(end) == 1:
            dv.add(sheet.cell(start[0],end[0]))


    def add_dropdown(self, sheet, options_str, start, end):
        """
        为指定单元格范围添加下拉列表验证
        根据选项长度自动选择直接列表或引用列表方式
        
        参数:
            sheet: 目标工作表
            options_str: 下拉选项字符串，格式为"选项1,选项2,选项3"（带双引号）
            start: 起始单元格坐标 (行, 列) 或 [行]
            end: 结束单元格坐标 (行, 列) 或 [列]
        """
        # 移除首尾的双引号，然后分割为列表
        options = options_str.strip('"').split(',')
        
        # 计算选项总长度（包括逗号分隔符）
        total_length = sum(len(option) for option in options) + (len(options) - 1)
        
        # Excel直接列表的字符限制约为255，留些余量设为250
        if total_length <= 250 and len(options) > 0:
            # 使用直接列表方式（保持原有的字符串格式）
            dv = DataValidation(
                type="list",
                formula1=options_str,  # 直接使用传入的带引号字符串
                showErrorMessage=False
            )
        else:
            # 使用引用列表方式
            wb = sheet.parent
            
            # 获取或创建存储列表的工作表
            if "Lists" not in wb.sheetnames:
                list_sheet = wb.create_sheet("Lists")
                list_sheet.sheet_state = "hidden"  # 隐藏列表工作表
            else:
                list_sheet = wb["Lists"]
            
            # 找到第一个空行来存储新的选项列表
            next_row = 1
            while list_sheet.cell(row=next_row, column=1).value is not None:
                next_row += 1
            
            # 写入选项数据
            for idx, option in enumerate(options, next_row):
                list_sheet.cell(row=idx, column=1, value=option)
            
            # 创建引用公式
            formula = f"Lists!$A${next_row}:$A${next_row + len(options) - 1}"
            dv = DataValidation(
                type="list",
                formula1=formula,
                showErrorMessage=False
            )
        
        # 添加数据验证到工作表
        sheet.add_data_validation(dv)
        
        # 应用数据验证到指定单元格范围
        # 处理坐标格式，确保start和end都是(行, 列)格式
        if len(start) == 1:
            # 如果start只有一个元素，视为行号，列号使用end的值
            start = (start[0], end[0])
            end = start  # 单个单元格
        
        if len(start) == 2 and len(end) == 2:
            # 处理单元格范围
            for i in range(start[0], end[0] + 1):
                for j in range(start[1], end[1] + 1):
                    dv.add(sheet.cell(i, j))
        elif len(start) == 1 and len(end) == 1:
            # 处理单个单元格
            dv.add(sheet.cell(start[0], end[0]))
            
        return dv
        

    def cell_style1(self, sheet, start, end):
        border=Border(left=Side(border_style='thin', color='000000'),
                      right=Side(border_style='thin', color='000000'),
                      top=Side(border_style='thin', color='000000'),
                      bottom=Side(border_style='thin', color='000000'))
        #bgfill = PatternFill(fill_type='solid', start_color='fff2cc', end_color='fff2cc') 
        #bgfill = PatternFill(fill_type = 'solid', start_color='197e00',end_color='197e00')
        bgfill = PatternFill(fill_type = 'solid', start_color='FF385724',end_color='FF333300')
        font = Font(name='等线', size=11, color='FFFFFF')
        for i in range(start[0], end[0] + 1):
            for j in range(start[1], end[1] + 1):
                sheet.cell(i,j).border=border 
                sheet.cell(i,j).fill=bgfill
                sheet.cell(i,j).font=font
                sheet.cell(i,j).alignment = Alignment(horizontal='left', vertical='center',wrapText=True) 
                sheet.cell(i,j).alignment = Alignment(horizontal='left', vertical='center',wrapText=True) 
                sheet.cell(i,j).alignment=Alignment(horizontal='left', vertical='center') 

    def cell_style2(self, sheet, start, end):
        border=Border(left=Side(border_style='thin', color='000000'),
                      right=Side(border_style='thin', color='000000'),
                      top=Side(border_style='thin', color='000000'),
                      bottom=Side(border_style='thin', color='000000'))
        #bgfill = PatternFill(fill_type='solid', start_color='fff2cc', end_color='fff2cc') 
        #bgfill = PatternFill(fill_type = 'solid', start_color='197e00',end_color='197e00')
        bgfill = PatternFill(fill_type = 'solid', start_color='FFFFFF',end_color='FFFFFF')
        #font = Font(name='等线', size=11, color='FFFFFF')
        for i in range(start[0], end[0] + 1):
            for j in range(start[1], end[1] + 1):
                sheet.cell(i,j).border=border 
                sheet.cell(i,j).fill=bgfill
                #sheet.cell(i,j).font=font
                sheet.cell(i,j).alignment = Alignment(horizontal='left', vertical='center',wrapText=True) 
                sheet.cell(i,j).alignment = Alignment(horizontal='left', vertical='center',wrapText=True) 
                sheet.cell(i,j).alignment=Alignment(horizontal='left', vertical='center',wrapText=True)

    def get_supply_infos(self):

        delkeys = ['module_name', 'ISO_CTRL', 'RET_SAVE', 'RET_RES', 'PSO_CTRL', 'PSO_ACK']
        supply_datag = self._sdcdg.vfile_data
        supply_data = {}

        for ky,vl in supply_datag.items():
            if ky not in delkeys:
                supply_data[ky] = vl

        supply_kw = []
        supply_val = []
        supply_vss = []
        supply_tmp = ''
        for key,val in supply_data.items():
            if '0v' in val or '0.0v' in val:
                supply_vss.append(key)
                sdc_info(f'Ground pin is {key}')
            elif 'PSO' in val:
                supply_kw.append(key)
                for i in range(1, int(val[-1]) + 1):
                    supply_kw.append(key + f'_PSW{i}')
                supply_tmp = supply_tmp + ' ' + supply_data[key].split('PSO')[0].strip()
            else:
                supply_kw.append(key)
                supply_tmp = supply_tmp + ' ' + supply_data[key].strip()

        #print(supply_tmp.strip().split(','))
        float_list = [float(x.strip('v')) for x in supply_tmp.strip().split()]
        unique_floats = set(float_list)
        sorted_floats = sorted(unique_floats, reverse=True)
        supply_val = [str(x) + 'v' for x in sorted_floats]
        supply_val.append('off')
        supply_val.append('0v')

        return supply_kw,supply_val,supply_vss,supply_data
    

    def get_ctl_sig(self, ctsig):
        ctrl = []
        for ct in ctsig:
            if re.search(r'\[\d+:\d+\]', ct):
                sig = ct.split('[')[0].strip()
                st = int(ct.split(':')[0].strip()[-1])
                ed = int(ct.split(':')[1].strip()[0])
                for i in range(ed,st+1):
                    ctrl.append(sig + '[' + str(i) + ']')
            else:
                ctrl.append(ct)
        return ctrl

    def get_table_loc(self,sheet) -> dict:

        if self._sheetname == 'VarDef':
            TABCONST = ['TMVAR']
        if self._sheetname == 'ClkDef':
            TABCONST = ['TMCLK']
        if self._sheetname == 'IODly':
            TABCONST = ['TMIODLY']
        # if self._sheetname == 'IOExp':
        #     TABCONST = ['TMIOEXP','TMINOUT']
        # if self._sheetname == 'IntExp':
        #     TABCONST = ['TMINTEXP','TMSTPGATE']
        if self._sheetname == 'Exp':
            TABCONST = ['TMIOEXP','TMINOUT','TMINTEXP','TMSTPGATE']

        # row_start max_col              
        row_start = ''        
        max_row = ''
        max_col = ''
        # row_start max_col max_row
        table_row_loc = {}
        for kw in TABCONST:
            strow = self.find_sheet(sheet, kw)
            row_start = str(strow)
            for i in range(1,sheet.max_column + 1):
                if sheet.cell(strow,i).value == 'Comment':
                    max_col = str(i)
                    #print(row_start[kw])
                    break              

            # print(kw,row_start)
            # if kw in ['TMHIER','TMCLK','TMIODLY','TMINOUT','TMSTPGATE']:
            if kw in ['TMVAR','TMCLK','TMIODLY','TMSTPGATE']:
                #table_row_loc[kw] = row_start[kw] + ' ' + str(int(row_start[kw].split()[0]) + 20)
                table_row_loc[kw] = row_start + ' ' + str(sheet.max_row + 2) + ' ' + max_col
            else:
                idx = TABCONST.index(kw) + 1
                max_row = self.find_sheet(sheet,TABCONST[idx]) - 1
                # print(kw,idx,max_row)
                #table_row_loc[kw] = row_start[kw] + ' ' + str(int(row_start[TABCONST[idx]].split()[0]) - 2)
                table_row_loc[kw] = row_start + ' ' + str(max_row) + ' ' + max_col

        #print(table_row_loc)
        return table_row_loc
  
    def get_table_contxt(self,sheet) -> dict:
        # row_start max_col max_row
        tab_loc = self.get_table_loc(sheet)
        # print(tab_loc)

        TABCONST = []
        if self._sheetname == 'VarDef':
            TABCONST = ['TMVAR']
        if self._sheetname == 'ClkDef':
            TABCONST = ['TMCLK']
        if self._sheetname == 'IODly':
            TABCONST = ['TMIODLY']
        # if self._sheetname == 'IOExp':
        #     TABCONST = ['TMIOEXP','TMINOUT']
        # if self._sheetname == 'IntExp':
        #     TABCONST = ['TMINTEXP','TMSTPGATE']
        if self._sheetname == 'Exp':
            TABCONST = ['TMIOEXP','TMINOUT','TMINTEXP','TMSTPGATE']

        table_contxt = {}
        #row_contxt = {}
        if TABCONST:
            for kw in TABCONST:
                start_row = int(tab_loc[kw].split(' ')[0])
                end_row = int(tab_loc[kw].split(' ')[1])
                end_col = int(tab_loc[kw].split(' ')[2])
                # if kw == 'TMSTPGATE':
                #     print('TMSTPGATE:',start_row,end_row,end_col)
                if kw == 'PMVAR':
                    for i in range(start_row, end_row + 1):
                        key = sheet.cell(i + 1, 1).value
                        val = str(sheet.cell(i + 1, 2).value)
                        if key:
                            table_contxt[key] = val.strip()
                        # print('PMVARdfd: ', table_contxt)
                        # if key and val:
                        #     table_contxt[key] = val
                else:
                    table_contxt.update(self.get_row_txt(sheet,kw,start_row,end_row,end_col))
                # if kw == 'TMSTPGATE':
                #     print('TMSTPGATE:',table_contxt)

        return table_contxt

    def get_row_txt(self,sheet,kw,start_row,end_row,end_col):
        row_contxt = {}
        table_contxt = {}
        for i in range(1,end_row-start_row):
            for j in range(1,end_col+1):
                key = sheet.cell(start_row,j).value
                val = sheet.cell(start_row+i,j).value
                val_col1 = sheet.cell(start_row+i,1).value
                if val_col1:
                    if re.search(r'^#',val_col1.strip()):
                        continue
                if key:     key = str(key).strip()
                if val:     val = str(val).strip()
                row_contxt[key] = val
                # if key and val:
                #     row_contxt[key] = val
            all_none = all(ele is None for ele in list(row_contxt.values()))
            if not all_none and row_contxt:
                table_contxt[f'{kw}_Row{start_row+i}'] = row_contxt
            row_contxt = {}
            # for key in table_contxt.keys():
            #     if 'TMCLK' in key:
            #         print(table_contxt)
        
        return table_contxt

        
    def save_text(self, context,file):
        with open(file, 'w') as fw:
            fw.write(context)

    def get_rows(self,pmdata,keyrow,kwd,ckwd):
        pmdict = {}
        pmlist = [(key, val) for key, val in pmdata.items() if re.search(r'{keyrow}\d+',key) and not re.search(r'^#',val[f'{ckwd}'].strip()) and val[f'{kwd}']]
        for k,v in pmlist:
            pmdict[k] = v
        pmkeys = [x for x in pmlist if re.search(r'{keyrow}\d+',x)]
        pmkeys.sort() 

        return pmdict, pmkeys     


class BaseInputs(object):
    def __init__(self):
        self.vfile_data = {}
        self.vfile_list = []
               
    def read_vfile(self, vfile) -> dict:

        lines = self.read_text(vfile)

        relclknum = 0
        for line in lines:
            line = line.replace('\n','').replace('\r','').replace('\t',' ').strip()
            if re.search(r'^\/\/', line) and '#RelClock:' not in line:
                continue

            if re.search(r'^module', line):
                self.vfile_data['module_name']= re.split(' +',line)[1].strip().replace('(','')
                self.vfile_list.append('module_name')
                continue

            if '#RelClock:' in line:
                relclknum += 1
                relclk = line.split('#RelClock:')[1].strip().replace('#','')
                self.vfile_list.append(f'RelClock{relclknum}')
                self.vfile_data[f'RelClock{relclknum}'] = relclk
                continue
                
            if re.search(r'^\);$',line):
                break

            dirc = ''
            portnum = ''
            kwd = ''
            if re.search(r'^input|^output|^inout',line):
                tline = line.split(' ')
                sline = [x for x in tline if x != '']
                dirc = sline[0]
                dircg = dirc
                
                if re.search(r'wire|logic|byte|bit|reg|tri1|tri0',line):               
                    if re.search(r'\[\d+:\d+\]',line):
                        lineg = ' '.join(sline[3:])
                        portnum = sline[2]
                        # print(sline)
                    else:
                        lineg = ' '.join(sline[2:])
                        portnum = '1'             
                else:
                    if re.search(r'\[\d+:\d+\]',line):
                        lineg = ' '.join(sline[2:])
                        portnum = sline[1]
                    else:
                        lineg = ' '.join(sline[1:])
                        portnum = '1'
                portnumg = portnum
                

                sigchar = lineg.replace(' ','')
                # print('inputoroutput:', sigchar)
                if re.search(r'\/\/#\w+#',sigchar):
                    #kwdg = ''.join(re.findall(r'\/\/(#\w+#)+', sigchar)).strip()
                    if '##' in sigchar:
                        kwd = sigchar.replace('##',' ').split('#')[1]
                    else:
                        kwd = sigchar.split('#')[1]
                    
                    if ',' in sigchar:
                        sigcharg = sigchar.split(',')
                        for ich in sigcharg:
                            if r'#\w+#' not in ich and '//' not in ich:
                                self.vfile_data[ich] = [dircg,portnumg,kwd]
                                self.vfile_list.append(ich)
                    else:
                        sdc_warn(f'{sigchar} not found , symbol ...')
                        ish = sigchar.split(r'//')[0].strip()
                        self.vfile_data[ish] = [dircg,portnumg,kwd]
                        self.vfile_list.append(ish)
                else:
                    kwd = 'None'
                    if ',' in sigchar:
                        sigcharg = sigchar.split(',')                        
                        for ich in sigcharg:
                            if '//' not in ich and ich != '':
                                self.vfile_data[ich] = [dircg,portnumg,kwd]
                                self.vfile_list.append(ich)
                    else:
                        sdc_warn(f'{sigchar} not found , symbol ...')
                        if '//' in sigchar:
                            ish = sigchar.split(r'//')[0].strip()
                        else:
                            ish = sigchar.strip()
                            self.vfile_data[ish] = [dircg,portnumg,kwd]
                            self.vfile_list.append(ish)

            else:
                # print('NO_inputoroutput:', line)
                if re.search(r'^\S+,$',line) and '//' not in line:
                    sline = line.split(',')
                    if re.search(r'\/\/#\w+#',line):
                        #kwd = ''.join(re.findall(r'\/\/#\w+#', line)).strip().split('#')[1]
                        if '##' in line:
                            kwd = line.replace('##',' ').split('#')[1]
                        else:
                            kwd = line.split('#')[1]
                        
                        for ich in sline:
                            if r'#\w+#' not in ich and '//' not in ich:
                                self.vfile_data[ich] = [dircg,portnumg,kwd]
                                self.vfile_list.append(ich)
                    else:
                        kwd = 'None'
                        for ich in sline:
                            if '//' not in ich and ich != '':
                                self.vfile_data[ich] = [dircg,portnumg,kwd]
                                self.vfile_list.append(ich)                       
                else:
                    #sline = line.split(' +')
                    tline = line.split(' ')
                    sline = [x for x in tline if x != '']
                    if re.search(r'wire|logic|byte|bit|reg|tri1|tri0',line):               
                        if re.search(r'\[\d+:\d+\]',line):
                            lineg = ' '.join(sline[2:])
                            portnumg = sline[1]
                        else:
                            lineg = ' '.join(sline[1:])
                            #portnum = '1'             
                    else:
                        if re.search(r'\[\d+:\d+\]',line):
                            lineg = ' '.join(sline[1:])
                            portnumg = sline[0]
                        else:
                            lineg = ' '.join(line[0:])
                            #portnum = '1'

                    sigchar = lineg.replace(' ','')
                    if re.search(r'\/\/#\w+#',sigchar):
                        #kwd = ''.join(re.findall(r'\/\/#\w+#', sigchar)).strip().split('#')[1]
                        #kwdg = ''.join(re.findall(r'\/\/(#\w+#)+', sigchar)).strip()
                        if '##' in sigchar:
                            kwd = sigchar.replace('##',' ').split('#')[1]
                        else:
                            kwd = sigchar.split('#')[1]
                        
                        if ',' in sigchar:
                            sigcharg = sigchar.split(',')
                            #kwd = ''.join(re.findall(r'\/\/#\w+#', sigchar)).strip().split('#')[1]
                            for ich in sigcharg:
                                if r'#\w+#' not in ich and '//' not in ich:
                                    self.vfile_data[ich] = [dircg,portnumg,kwd]
                                    self.vfile_list.append(ich)
                        else:
                            sdc_warn(f'{sigchar} not found , symbol ...')
                            ish = sigchar.split('//')[0].strip()
                            self.vfile_data[ish] = [dircg,portnumg,kwd]
                            self.vfile_list.append(ish)
                    else:
                        kwd = 'None'
                        if ',' in sigchar:
                            sigcharg = sigchar.split(',')                           
                            for ich in sigcharg:
                                if '//' not in ich and ich != '':
                                    self.vfile_data[ich] = [dircg,portnumg,kwd]
                                    self.vfile_list.append(ich)
                        else:
                            sdc_warn(f'{sigchar} not found , symbol ...')
                            if '//' in sigchar:
                                ish = sigchar.split('//')[0].strip()
                            else:
                                ish = sigchar.strip()
                            self.vfile_data[ish] = [dircg,portnumg,kwd]
                            self.vfile_list.append(ish)
                # print('vfile_list:',self.vfile_list)
        return self.vfile_list, self.vfile_data

        


    def read_yaml(self, yaml_file):

        yaml_data = {}
        if not os.path.exists(yaml_file):
            raise FileExistsError(f'{yaml_file} does not exists')
        with open(yaml_file, 'r') as fh:
            yaml_data = yaml.load(fh, yaml.FullLoader)

        return yaml_data
    


    def read_text(self, file):
        if not os.path.exists(file):
            raise FileExistsError(f'{file} does not exists')
            # sdc_error(f'{file} not exist. Please check it.')
            # exit(1)
        else:
            txt_list = []
            with open(file,'r') as fh:
                for line in fh.readlines():
                    if line.strip() == "":
                        continue
                    if line.strip().startswith("//") and '#RelClock:' not in line.strip():
                         continue   
                    line = re.sub(r"\[\s*(\d+)\s*:\s*(\d+)\s*\]", r"[\1:\2]", line)
                    txt_list.append(line.strip())
        
            return txt_list



class VarDefSheet(BaseSheet):
    def __init__(self,*args):
        super().__init__(*args)  
        self._valdata = {}  
        self._vardata = {}
        #self._clkdata = self.get_table_contxt(self._sdcdg._wb['ClkDef'])
        #self._clkdef = None

        self._vfdata = self._sdcdg._vfile_data
        self._hiertree = self._sdcdg._hier_tree
        self._sdcdir = self._sdcdg._sdcdir
        self._mdname = self._sdcdg._mdname

    def update_sheet(self):
        '''
        # only during -dg option
        # addition of module name value from vfile
        # addition of user_defined variables
        # addition of block hier tree expanded table from hier yaml
        '''
        sheet = self.get_sheet()

        hiertree = self._sdcdg._hier_tree

        # find TMVAR table
        start_rowg = self.find_sheet(sheet, 'TMVAR')
        
        mdname = self._sdcdg._vfile_data['module_name']
        # if mdname:
        #     sheet.cell(start_rowg + 1, 2).value = mdname

        # vardef = self.get_vardef_value(sheet)
        sheet.cell(start_rowg + 1, 2).value = self._mdname
     
        # varlist = ['T28','T16','T7','T4']
        # self.add_dropdown(sheet, '"' + ','.join(varlist) + '"', [start_rowg + 2], [2])
        varlist = ['RTL','SYN','DFT_SYN','SIM','PLA','CTS','PnR','SIGNOFF']
        self.add_dropdown(sheet, '"' + ','.join(varlist) + '"', [start_rowg + 2], [2])
        varlist = ['DC WLM','DC SPG','GNS PLE','GNS ISP']
        self.add_dropdown(sheet, '"' + ','.join(varlist) + '"', [start_rowg + 3], [2])
        # varlist = ['full','local']
        # self.add_dropdown(sheet, '"' + ','.join(varlist) + '"', [start_rowg + 9], [2])
        varlist = ['70%','60%','50%','40%','30%','0','-10%','-20%']
        self.add_dropdown(sheet, '"' + ','.join(varlist) + '"', [start_rowg + 4], [2])
        varlist = ['20%','10%','0','-10%','-20%']
        self.add_dropdown(sheet, '"' + ','.join(varlist) + '"', [start_rowg + 5], [2])

        # block = hiertree.get_block_by_name(mdname)
        # for i in range(start_rowg, sheet.max_row + 1):
        #     if sheet.cell(i, 1).value == 'SDC_DIR':
        #         sheet.cell(i, 2).value = block.constr_dir
        
        for i in range(1,start_rowg+11):
            for j in range(1,4):
                sheet.cell(start_rowg+i,j).alignment = Alignment(horizontal='left',vertical='center',wrapText=True)

    def read_data(self):
        sheet = self.get_sheet()
        self._valdata = self.get_table_contxt(sheet)
        nvaldata = {}
        nvaldata["TMVAR_Row14"] = {
            "Variable": "SDC_DIR",
            "Value": f'{self._sdcdir}',
            "Comment": ''
        }
        nvaldata["TMVAR_Row15"] = {
            "Variable": "COM_DIR",
            "Value": f'{self._sdcdir}',
            "Comment": ''
        }
        nvaldata["TMVAR_Row16"] = {
            "Variable": "DFT_DIR",
            "Value": '',
            "Comment": ''
        }
        # nvaldata["TMVAR_Row17"] = {
        #     "Variable": "HD_MOD_NAME",
        #     "Value": f'{self._mdname}',
        #     "Comment": ''
        # }
        nvaldata["TMVAR_Row17"] = {
            "Variable": "HD_PROCESS",
            "Value": '',
            "Comment": ''
        }
        nvaldata["TMVAR_Row18"] = {
            "Variable": "CYCLE_LIST",
            "Value": '[list CYCLE500M]',
            "Comment": ''
        }
        self._valdata.update(nvaldata)

    def check_sheet(self):
        pass

    def change_sheet(self):
        pass

    def dump_json(self,json_file):
        self._data = self._valdata
        self.write_json(json_file)


class ClkDefSheet(BaseSheet):
    def __init__(self,*args):
        super().__init__(*args)  
        self._hiertree = self._sdcdg._hier_tree
        
        self._sdcdir = self._sdcdg._sdcdir
        self._mdname = self._sdcdg._mdname
        self._alias = self._hiertree._blocks[self._mdname].alias #self._sdcdg._alias

        self._clkdata = {}  
        self._clknmdata = {}
        self._clknmlst = []

        self._clkvardata = {}
        #self._clkvarlinegs = ''

        self._rowcrgdata = {}
        self._iptcrgdata = {}
        self._iptcrglst  = {}
        self._crgals = {}

        self._rowipdata = {}
        self._iptipdata = {}
        self._iptiplst  = {}
        self._ipals = {}

        self._crgflg = 0
        self._ipflg = 0

        self._crgalsiptval = {}
        self._ipalsiptval = {}

        self._tclkdata = {}
        self._tclklst = {}

        # for crgip mstclk/srcpin/clkgrp
        #self._clkinfolst = []
        self._cycle_clkdeflst = []
        self._cycle_crgiplst = []
        #self._intgportlst = []

        self._crgipclknmals = {}

        self._hdportclks = {}
        self._hdportclksinfo = {}

        self._curhd_portclks = {}
        self._curhd_portclksinfo = {}

        # self._clkdef = {}
        # self._crgipclkdef = {}
        # self._hdportclkdef = {}
        self._curclkdef = {}

        self._lvl = 'blk'
        self._flt = 'IS_FLAT'   

    # def set_curclk_attr(self):

    def update_sheet(self):
        '''
        # only during -dg option
        # addition of crg files
        # addition of block hier tree expanded table from hier yaml
        '''
        sheet = self.get_sheet()
        # self.read_crgip_data()

        # hiertree = self._sdcdg._hier_tree
        # alias = self._sdcdg._hier_tree._blocks[self._mdname].alias
        #indir = self._sdcdir + '/inputs'

        #if not self._tclklst:
        # tclklst,tclkdata = self.concat_curhd_crgiphd_connect()
        

        # find TMCLK table
        start_rowg = self.find_sheet(sheet, 'TMCLK')

        varlist = ['clk_mcu_crt','clk_mcu_gen','clk_mcu_pll_crt','cllk_mcu_pll_gen','clk_mcu_virtual_crt','clk_mcu_totop_out','clk_mcu_topad_out','clk_mcu_tosys_out','clk_mcu_fdth_topad_out']
        self.add_dropdown(sheet, '"' + ','.join(varlist) + '"', [start_rowg + 1,1], [start_rowg + 10,1])

        # clkgrp
        #nvarlst = self.get_clkinfo_from_crgip(indir,self._mdname,'2')
        varlist = ['CGP1','CGP2','CGP3','CGP4','CGP5']
        #varlist += nvarlst
        self.add_dropdown(sheet, '"' + ','.join(varlist) + '"', [start_rowg + 1,2], [start_rowg + 10,2])

        varlist = ['200M','400M|200M','1666M|800M|76M8']
        self.add_dropdown(sheet, '"' + ','.join(varlist) + '"', [start_rowg + 1,3], [start_rowg + 10,3])

        # waveform
        varlist = ['{0 2.5}','{1.0 4.2}','{0 4.0}|{0 5.0}']
        self.add_dropdown(sheet, '"' + ','.join(varlist) + '"', [start_rowg + 1,4], [start_rowg + 10,4])

        # divedge
        varlist = ['1','comb','1//2','1|2|{1 3 5}|comb','1//2|2 inv|{2 4 6}']
        self.add_dropdown(sheet, '"' + ','.join(varlist) + '"', [start_rowg + 1,5], [start_rowg + 10,5])

        # mstclk from crg_out/ip_out in header
        # varlistg = self.get_clkinfo_from_crgip('0','GEN')
        # nvar = [x.values() for x in varlistg]
        # xvar = [list(vl) for vl in nvar]
        # varlist = [i for g in xvar for i in g]
        #varlist = self._clkinfolst
        #print(varlist)

        # msclk
        # varlist = []
        # clknm = self.get_clkinfo_from_crgip('0','GEN')
        # ipclk = []
        # for x in clknm:
        #     for k,v in x.items():
        #         cials = k.split(' ')[0].split('_')[1]
        #         nalsck = [f'{cials} {x}' for x in v]
        #         ipclk.extend(nalsck)
        # hdmstclk,hdintgclk,mdmstclk,mdintgclk = self.get_hdclk_dropdown()
        # if ipclk:
        #     varlist.extend(ipclk)
        # if hdmstclk:
        #     varlist.extend(hdmstclk)
        # if varlist:
        #     self.add_dropdown(sheet, '"' + ','.join(varlist) + '"', [start_rowg + 1,6], [start_rowg + 10,6])

        # portpin from vfile clk port
        varlist = []
        vdata = self._sdcdg._vfile_data
        vlist = self._sdcdg._vfile_list
        for kwd in vlist:
            if 'module_name' not in kwd and 'RelClock' not in kwd:
                if re.search(r'TCLK',vdata[kwd][2]):
                    varlist.append(kwd)
        self.add_dropdown(sheet, '"' + ','.join(varlist) + '"', [start_rowg + 1,7], [start_rowg + 10,7]) 

        # clkintg from crg/ip header
        # varlist = self.get_intgport_from_crgip()
        # #varlist = [f'{alias}_{x}' for x in varlistg]
        # #varlist = self._intgportlst
        # if hdintgclk:
        #     varlist.extend(hdintgclk)
        # if mdintgclk:
        #     varlist.extend(mdintgclk)
        # if varlist:
        #     self.add_dropdown(sheet, '"' + ','.join(varlist) + '"', [start_rowg + 1,8], [start_rowg + 10,8]) 
        # else:
        #     sdc_info(f'Can not find clock connection from or to crg and ip.')

        # vol from hier yaml
        varlist = []
        for blknm in self._hiertree.get_curblks(self._mdname):
            blk = self._hiertree.get_block_by_name(blknm)
            pwrg = blk.prime_pwr.split(' ')[0].strip()
            if f'{pwrg}' not in varlist:
                varlist.append(pwrg)      
        self.add_dropdown(sheet, '"' + ','.join(varlist) + '"', [start_rowg + 1,9], [start_rowg + 10,9]) 

        # comment 
        varlist = ['PLL_CRT','PLL_GEN','TOTOP_OUT','TOPAD_OUT','TOSYS_OUT','PHYGRP_A_1','PHYGRP_A_2','LOGGRP_B_1','LOGGRP_B_2']
        self.add_dropdown(sheet, '"' + ','.join(varlist) + '"', [start_rowg + 1,10], [start_rowg + 10,10]) 



class IODlySheet(BaseSheet):
    def __init__(self,*args):
        super().__init__(*args)  
        self._iodlydata = {}  
        self._hiertree = self._sdcdg._hier_tree
        #self._valdata = self._sdcdg._sheets['VarDef']._valdata
        self._vardata = self.get_vardef_value(self._sdcdg._wb['VarDef'])
        
        #self._sdcdir = self._sdcdg._sdcdir
        self._mdname = self._sdcdg._mdname

        #self._clkdef = None
        self._vfdata = self._sdcdg._vfile_data
        self._hiertree = self._sdcdg._hier_tree   

    def update_sheet(self):
        '''
        # only during -dg option
        # addition of module name value from vfile
        # addition of block hier tree expanded table from hier yaml
        '''
        sheet = self.get_sheet()

        #hiertree = self._sdcdg._hier_tree

        # find TMIODLY table
        start_rowg = self.find_sheet(sheet, 'TMIODLY')

        vdata = self._sdcdg._vfile_data
        vlist = self._sdcdg._vfile_list
        
        n = 0
        for kwd in vlist:
            if 'module_name' not in kwd and 'RelClock' not in kwd:
                if re.search(r'IDEAL|CASEXP|FPEXP|ANA|TCLK',vdata[kwd][2]):
                    n += 1
        sheet.insert_rows(start_rowg + 9, len(vlist) - n)
        self.cell_style2(sheet, [start_rowg + 9, 1], [start_rowg + 9 + len(vlist) - n, 8])

        varlist = ['input','output']
        self.add_dropdown(sheet, '"' + ','.join(varlist) + '"', [start_rowg + 1,2], [start_rowg + 9 + len(vlist) - n + 1,2])
        
        #indir = self._sdcdir + '/inputs'
        clkdef = self._sdcdg._sheets['ClkDef']
        #varlist = clkdef.get_clkinfo_from_crgip('IO','0')
        # clknm = clkdef.get_clkinfo_from_crgip('0','IO')
        alsck = []
        varlist = []
        varlistg = []
        # for x in clknm:
        #     for k,v in x.items():
        #         cials = k.split(' ')[0].split('_')[1]
        #         nalsck = [f'{cials} {p}' for p in v]
        #         alsck.extend(nalsck)
        # varlist.extend(alsck)
        for kwd in vlist:
            if 'RelClock' in kwd:
                varlistg.append(vdata[kwd])
        if varlistg:
            varlistg = list(set(varlistg))
            varlist.extend(varlistg)
        self.add_dropdown(sheet, '"' + ','.join(varlist) + '"', [start_rowg + 1,3], [start_rowg + 9 + len(vlist) - n + 1,3])
             
        varlist = ['Y']
        self.add_dropdown(sheet, '"' + ','.join(varlist) + '"', [start_rowg + 1,4], [start_rowg + 9 + len(vlist) - n + 1,4])
        
        maxlst = []
        minlst = []
        for kw in self._vardata.keys():
            if re.match(r'IO_DLY_MAX?',str(kw)):
                maxlst.append(kw)
            if re.match(r'IO_DLY_MIN?',str(kw)):
                minlst.append(kw)
        varlist = maxlst + ['70%','60%','50%','40%','30%','0','-10%']
        self.add_dropdown(sheet, '"' + ','.join(varlist) + '"', [start_rowg + 1,5], [start_rowg + 9 + len(vlist) - n + 1,5])
        varlist = minlst + ['20%','10%','0','-10%','-20%']
        self.add_dropdown(sheet, '"' + ','.join(varlist) + '"', [start_rowg + 1,6], [start_rowg + 9 + len(vlist) - n + 1,6])

        # vol
        varlist = []
        for blknm in self._hiertree.get_curblks(self._mdname):
            blk = self._hiertree.get_block_by_name(blknm)
            pwrg = blk.prime_pwr.split(' ')[0].strip()
            if f'{pwrg}' not in varlist:
                varlist.append(pwrg)
        varlistg = [f'{x}=TT0P750V' for x in varlist]
        self.add_dropdown(sheet, '"' + ','.join(varlistg) + '"', [start_rowg + 1,7], [start_rowg + 9 + len(vlist) - n + 1,7])

        i = 0
        grp_flag  = 0
        ioclknm = ''
        for kwd in vlist:
            if i < len(vlist) + 1:
                
                if 'module_name' not in kwd and 'RelClock' not in kwd:
                    if not re.search(r'IDEAL|CASEXP|FPEXP|ANA|TCLK|DFT',vdata[kwd][2]):
                        i += 1
                        if vdata[kwd][1] != '1':
                            sheet.cell(start_rowg + i,1).value = kwd + vdata[kwd][1]
                        else:
                            sheet.cell(start_rowg + i,1).value = kwd
                        sheet.cell(start_rowg + i,2).value = vdata[kwd][0]
                        sheet.cell(start_rowg + i,5).value = 'IO_DLY_MAXA'
                        sheet.cell(start_rowg + i,6).value = 'IO_DLY_MINA'

                        if grp_flag:
                            sheet.cell(start_rowg + i,3).value = ioclknm

                if 'RelClock' in kwd:
                    if not grp_flag:
                        grp_flag  = 1
                        ioclknm = vdata[kwd]
                    else:
                        grp_flag  = 0
                        ioclknm = ''

    ######################################################
    def read_data(self):
        sheet = self.get_sheet()
        self._iodlydata = self.get_table_contxt(sheet)
        #print(self._iodlydata)
        
    def check_sheet(self):
        pass
    
    def change_sheet(self):
        pass

    def dump_json(self,json_file):
        self._data = self._iodlydata
        self.write_json(json_file)


class ExpSheet(BaseSheet):
    def __init__(self ,*args):
        super().__init__(*args)
        self._expdata = {}
        self._hier_tree = self._sdcdg._hier_tree
        # self._clkdef = None
        self._vardata = self.get_vardef_value(self._sdcdg._wb['VarDef'])

        self._lvl = 'blk'
        self._flt = 'IS_FLAT'

        self._mdname = ''

    # 处理包含特殊字符的选项列表
    def escape_options(self,varlist):
        # 给每个选项添加双引号以转义特殊字符
        return ','.join([f'"{opt}"' for opt in varlist])

    def update_sheet(self):
        '''
        # only during -dg option
        # addition of module name value from vfile
        # addition of block hier tree expanded table from hier yaml
        '''
        sheet = self.get_sheet()

        hiertree = self._hier_tree
        vdata = self._sdcdg._vfile_data
        vlist = self._sdcdg._vfile_list

        # find TMIOEXP table
        start_rowg = self.find_sheet(sheet, 'TMIOEXP')

        nvdata = {}
        nvlist = []
        for kwd in vlist:
            if 'module_name' not in kwd and 'RelClock' not in kwd:
                if 'TCLK' not in vdata[kwd][2] and re.search(r'IDEAL|CASE|FP|MCP' ,vdata[kwd][2]):
                    # 'IDEAL' in vdata[kwd][2] or 'FPEXP' in vdata[kwd][2] or 'MCPEXP' in vdata[kwd][2] or 'CASEXP' in vdata[kwd][2]:
                    nvlist.append(kwd)
                    nvdata[kwd] = vdata[kwd]

        sheet.insert_rows(start_rowg + 9, len(nvlist) + 1)
        self.cell_style2(sheet, [start_rowg + 9, 1], [start_rowg + 9 + len(nvlist) + 1, 10])

        varlist = ['input','output']
        self.add_dropdown(sheet, '"' + ','.join(varlist) + '"', [start_rowg + 1 ,2], [start_rowg + len(nvlist) + 11 ,2])

        varlist = ['Y']
        self.add_dropdown(sheet, '"' + ','.join(varlist) + '"', [start_rowg + 1 ,3], [start_rowg + len(nvlist) + 11 ,3])
        varlist = ['0' ,'1']
        self.add_dropdown(sheet, '"' + ','.join(varlist) + '"', [start_rowg + 1 ,4], [start_rowg + len(nvlist) + 11 ,4])
        varlist = ['setup' ,'hold' ,'all']
        self.add_dropdown(sheet, '"' + ','.join(varlist) + '"', [start_rowg + 1 ,5], [start_rowg + len(nvlist) + 11 ,5])
        varlist = ['start 2 1' ,'end 2 1' ,'start NA 1' ,'end 2 NA']
        self.add_dropdown(sheet, '"' + ','.join(varlist) + '"', [start_rowg + 1 ,6], [start_rowg + len(nvlist) + 11 ,6])

        varlist = ['pin [list ]' ,'clk [list ]']
        self.add_dropdown(sheet, '"' + ','.join(varlist) + '"', [start_rowg + 1 ,7], [start_rowg + len(nvlist) + 11 ,7])
        varlist = ['pin [list ]']
        self.add_dropdown(sheet, '"' + ','.join(varlist) + '"', [start_rowg + 1 ,8], [start_rowg + len(nvlist) + 11 ,8])
        varlist = ['pin [list ]' ,'clk [list ]']
        self.add_dropdown(sheet, '"' + ','.join(varlist) + '"', [start_rowg + 1 ,9], [start_rowg + len(nvlist) + 11 ,9])


        i = 0
        for kwd in nvlist:
            if i < len(nvlist) + 1:
                # expkw = 'module_name' not in kwd and 'RelClock' not in kwd and 'IDEAL' not in vdata[kwd][2] and 'TCLK' not in vdata[kwd][2] and 'ANA' not in vdata[kwd][2]

                if 'CASE'  in nvdata[kwd][2] and 'FP'  in nvdata[kwd][2]:
                    sdc_warn(f'port {kwd} includes CASE and FP.')
                elif 'MCP'  in nvdata[kwd][2] and 'FP'  in nvdata[kwd][2]:
                    sdc_error(f'port {kwd} includes MCP and FP.')
                elif 'CASE'  in nvdata[kwd][2] and 'MCP'  in nvdata[kwd][2]:
                    sdc_warn(f'port {kwd} includes CASE and MCP.')

                if re.search(r'IDEAL|CASE|FP|MCP' ,nvdata[kwd][2]):
                    i += 1
                    if vdata[kwd][1] != '1':
                        sheet.cell(start_rowg + i ,1).value = kwd + nvdata[kwd][1]
                    else:
                        sheet.cell(start_rowg + i ,1).value = kwd
                    sheet.cell(start_rowg + i ,2).value = nvdata[kwd][0]
                    if 'IDEAL' in nvdata[kwd][2]:
                        sheet.cell(start_rowg + i ,3).value = 'Y'


        # find TMINOUT table
        start_rowg = self.find_sheet(sheet, 'TMINOUT')

        varlist = ['70%' ,'60%' ,'50%' ,'40%' ,'30%' ,'0' ,'-10%']
        self.add_dropdown(sheet, '"' + ','.join(varlist) + '"', [start_rowg + 1 ,3], [start_rowg + 10 ,4])
        varlist = ['0.2' ,'0.1' ,'0.05']
        self.add_dropdown(sheet, '"' + ','.join(varlist) + '"', [start_rowg + 1 ,5], [start_rowg + 10 ,5])

        clkdef = self._sdcdg._sheets['ClkDef']
        # varlist = clkdef.get_clkinfo_from_crgip('IO','0')
        # varlist = clkdef.get_clkinfo_from_crgip('IO','0')
        # clknm = clkdef.get_clkinfo_from_crgip('0' ,'IO')
        alsck = []
        varlist = []
        varlistg = []
        # for x in clknm:
        #     for k ,v in x.items():
        #         cials = k.split(' ')[0].split('_')[1]
        #         nalsck = [f'{cials} {x}' for x in v]
        #         alsck.extend(nalsck)
        # varlist.extend(alsck)
        for kwd in vlist:
            if 'RelClock' in kwd:
                varlistg.append(vdata[kwd])
        if varlistg:
            varlistg = list(set(varlistg))
            varlist.extend(varlistg)
        self.add_dropdown(sheet, '"' + ','.join(varlist) + '"', [start_rowg + 1 ,6], [start_rowg + 10 ,6])

        # find TMINTEXP table
        start_rowg = self.find_sheet(sheet, 'TMINTEXP')

        varlist = ['0','1']
        self.add_dropdown(sheet, '"' + ','.join(varlist) + '"', [start_rowg + 1,3], [start_rowg + 10,3])
        varlist = ['setup','hold','all']
        self.add_dropdown(sheet, '"' + ','.join(varlist) + '"', [start_rowg + 1,1], [start_rowg + 10,1])
        varlist = ['start 2 1','end 2 1','start NA 1','end 2 NA']
        self.add_dropdown(sheet, '"' + ','.join(varlist) + '"', [start_rowg + 1,2], [start_rowg + 10,2])

        varlist = ['pin [list ]','clk [list ]']
        self.add_dropdown(sheet, '"' + ','.join(varlist) + '"', [start_rowg + 1,5], [start_rowg + 10,5])
        varlist = ['pin [list ]']
        self.add_dropdown(sheet, '"' + ','.join(varlist) + '"', [start_rowg + 1,6], [start_rowg + 10,6])
        varlist = ['pin [list ]','clk [list ]']
        self.add_dropdown(sheet, '"' + ','.join(varlist) + '"', [start_rowg + 1,7], [start_rowg + 10,7])

        # find TMSTPGATE table
        start_rowg = self.find_sheet(sheet, 'TMSTPGATE')
        varlist = ['clk [list ]']
        self.add_dropdown(sheet, '"' + ','.join(varlist) + '"', [start_rowg + 1,1], [start_rowg + 10,1])
        varlist = ['pin [list ]']
        self.add_dropdown(sheet, '"' + ','.join(varlist) + '"', [start_rowg + 1,2], [start_rowg + 10,2])
        varlist = ['inst [list ]','pin [list ]']
        self.add_dropdown(sheet, '"' + ','.join(varlist) + '"', [start_rowg + 1,3], [start_rowg + 10,3])


    ######################################################
    def read_data(self):
        sheet = self.get_sheet()
        self._expdata = self.get_table_contxt(sheet)

    def check_sheet(self):
        pass

    def change_sheet(self):
        pass

    def dump_json(self ,json_file):
        self._data = self._expdata
        self.write_json(json_file)



def modify_line_in_file(file_path, search_pattern, replacement):
    # 打开文件并逐行读取内容
    with open(file_path, 'r') as file:
        lines = file.readlines()

    # 遍历每一行并进行匹配和替换
    modified_lines = []
    for line in lines:
        if search_pattern in line:
            modified_line = line.replace(search_pattern, replacement)
            modified_lines.append(modified_line)
        else:
            modified_lines.append(line)

    # 将修改后的内容写回文件
    with open(file_path, 'w') as file:
        file.writelines(modified_lines)

def rm_exist_log(logdir):
    # LOG_DIR = os.getenv('TASK_LOGS_DIR')
    # logdir = os.path.join(LOG_DIR,taskid)
    # logdir = r'E:\stone\work\smalltool\pycharm\const0203_sysflatok_full\logs\adfeicc67ere'
    rlogfiles = os.listdir(logdir)
    for logfile in rlogfiles:
        # if logfile.endswith('.log') or logfile.endswith('.rpt'):
        if logfile in ['sdc_dg.log','sdc_gen.log','sdc_gen.rpt','full_msg.log']:
            os.remove(os.path.join(logdir,logfile))


def printlog(context, file='sdc_dg.log',logdir=''):
    if not logdir:
        LOG_DIR = os.getenv('TASK_LOGS_DIR')
        taskid = os.getenv('CURRENT_TASK_ID', 'default_task')
        if LOG_DIR:
            logdir = os.path.join(LOG_DIR, taskid)
            os.makedirs(logdir, exist_ok=True)
            rlogdir = os.path.join(logdir, file)
        else:
            # 如果环境变量不存在，使用当前目录
            rlogdir = file
    else:
        rlogdir = logdir

    if os.path.exists(rlogdir):
        with open(rlogdir, 'a') as fw:
            fw.write(context)
    else:
        with open(rlogdir, 'w') as fw:
            fw.write(context)

def movelogrpt(msgnm, flog, fdir):
    if os.path.exists(f'{msgnm}'):
        if os.path.exists(flog):
            # with open('sdc_gen.log','r') as fh:
            #     for line in fh.readlines():
            #         #txt_list.append(line.strip())
            #         txt_list += f'{line.strip()} \n'
            # with open(logfile, 'a') as fw:
            #     fw.write(txt_list)
            os.system(f'rm -f {flog}')
            os.system(f'mv {msgnm} {fdir}')
        else:
            os.system(f'mv {msgnm} {fdir}')
    else:
        sdc_warn(f'Can not find {msgnm}.')


#########################################################################################################################
# sdc message
full_log_message_list =  []
full_rpt_message_list =  []

def sdc_log(level, msg, out=sys.stdout):
    print(f'{level.upper()}: {msg}', flush=True, file=out)

def sdc_info(msg, kw='log'):
    # if msg not in full_log_message_list:
    if kw == 'log':
        full_log_message_list.append(msg)
        printlog(f'SDC_INFO: {msg} \n', 'full_msg.log')
    else:
        full_rpt_message_list.append(msg)
        printlog(f'SDC_INFO: {msg} \n', 'full_chk.rpt')
    # sdc_log('SDC_INFO', msg)
    if kw == 'log':
        printlog(f'SDC_INFO: {msg} \n')
    if kw == 'chk_hym':
        printlog(f'SDC_INFO: {msg} \n', 'chk_hym.rpt')
    if kw == 'chk_sht':
        printlog(f'SDC_INFO: {msg} \n', 'chk_sht.rpt')
    if kw == 'chk_dti':
        printlog(f'SDC_INFO: {msg} \n', 'chk_dti.rpt')
    if kw == 'chk_sdc':
        printlog(f'SDC_INFO: {msg} \n', 'chk_sdc.rpt')

def sdc_warn(msg, kw='log'):
    # if msg not in full_log_message_list:
    if kw == 'log':
        full_log_message_list.append(msg)
        printlog(f'SDC_WARN: {msg} \n', 'full_msg.log')
    else:
        full_rpt_message_list.append(msg)
        printlog(f'SDC_WARN: {msg} \n', 'full_chk.rpt')
    print(f'\033[0:31mSDC_WARN\033[0m: {msg}', flush=True)
    # print(f'SDC_WARN: {msg}', flush=True)
    if kw == 'log':
        printlog(f'SDC_WARN: {msg} \n')
    if kw == 'chk_hym':
        printlog(f'SDC_WARN: {msg} \n', 'chk_hym.rpt')
    if kw == 'chk_sht':
        printlog(f'SDC_WARN: {msg} \n', 'chk_sht.rpt')
    if kw == 'chk_dti':
        printlog(f'SDC_WARN: {msg} \n', 'chk_dti.rpt')
    if kw == 'chk_sdc':
        printlog(f'SDC_WARN: {msg} \n', 'chk_sdc.rpt')
         

def sdc_error(msg, kw='log'):
    # if msg not in full_log_message_list:
    if kw == 'log':
        full_log_message_list.append(msg)
        printlog(f'SDC_ERROR: {msg} \n', 'full_msg.log')
    else:
        full_rpt_message_list.append(msg)
        printlog(f'SDC_ERROR: {msg} \n', 'full_chk.rpt')
    print(f'\033[0:31mSDC_ERROR\033[0m: {msg}', flush=True)
    # print(f'SDC_ERROR: {msg}', flush=True)
    if kw == 'log':
        printlog(f'SDC_ERROR: {msg} \n')
    if kw == 'chk_hym':
        printlog(f'SDC_ERROR: {msg} \n', 'chk_hym.rpt')
    if kw == 'chk_sht':
        printlog(f'SDC_ERROR: {msg} \n', 'chk_sht.rpt')
    if kw == 'chk_dti':
        printlog(f'SDC_ERROR: {msg} \n', 'chk_dti.rpt')
    if kw == 'chk_sdc':
        printlog(f'SDC_ERROR: {msg} \n', 'chk_sdc.rpt')

def sdc_fatal(msg, kw='log'):
    # if msg not in full_log_message_list:
    if kw == 'log':
        full_log_message_list.append(msg)
        printlog(f'SDC_FATAL: {msg} \n', 'full_msg.log')
    else:
        full_rpt_message_list.append(msg)
        printlog(f'SDC_FATAL: {msg} \n', 'full_chk.rpt')
    print(f'\033[0:31mSDC_FATAL\033[0m: {msg}', flush=True)
    # print(f'SDC_FATAL: {msg}', flush=True)
    if kw == 'log':
        printlog(f'SDC_FATAL: {msg} \n')
    if kw == 'chk_hym':
        printlog(f'SDC_FATAL: {msg} \n', 'chk_hym.rpt')
    if kw == 'chk_sht':
        printlog(f'SDC_FATAL: {msg} \n', 'chk_sht.rpt')
    if kw == 'chk_dti':
        printlog(f'SDC_FATAL: {msg} \n', 'chk_dti.rpt')
    if kw == 'chk_sdc':
        printlog(f'SDC_FATAL: {msg} \n', 'chk_sdc.rpt')
           
    # sys.exit(1)

def sdc_args():
    print(f'##++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++')
    print(f'## SDC Generation Optional Arguments Presentation:')
    print(f'## -hlp:        All of flow option usage reference.')
    print(f'## -usg:        Flow step command usage reference.')
    print(f'## -tmp:        Write out template SDCs for reference')
    print(f'## -gen_dir:    SDC generationg directory')
    print(f'## -hier_yaml:  Design hierarchy info. from user input file')
    print(f'## -setup:      Build setup directories from blocks defined in hier yaml')
    print(f'## -blocks:     Indicate only current design name for SDC generation')
    print(f'## -flat:       Generate flatten SDC based on current design. If not flat, only generate current design level only SDC')
    print(f'## -dg:         Generate design guide file to be present SDC request format')
    print(f'## -idg:        Incrementally update design guide file based on existed dg file and updated input files')
    print(f'## -sdc:        Write out SDC files')
    print(f'## -check_hym: Check hier yaml data before generating SDC')
    print(f'## -check_only: Check input data before generating SDC')
    print(f'## -check_sdc:  Check SDC consistency after generating SDC')
    print(f'## -proj:       Open project mode. Maybe need set some related project environment variables')
    print(f'##++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++')


def sdc_usage():
    print(f'##++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++')
    print(f'## Before flow run, user must provide input data including hier yaml and empty vfile, dg file and crg/ip SDCs to be integrated ')
    print(f'## For hier yaml and empty vfile, must follow format of hier_pwr.yaml and tmempty.v in template/ftemp')
    print(f'## For crg/ip SDCs, must follow header format of crg.sdc and userip.sdc in template/ftemp')
    print(f'##++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++')
    print(f'##++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++')
    print(f'## Step I: Build the whole SDC generation directory structure.')
    print(f'## Cmd Line: xconst sdgen -gen_dir <sdcdir> -hier_yaml <hier_file> -setup -blocks <blk_name> [-tmp]')
    print(f'##++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++')
    print(f'##++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++')
    print(f'## Step II: Generate initial design guide file according to current input data.')
    print(f'## Cmd Line: xconst sdgen -gen_dir <sdcdir> -hier_yaml <hier_file> -dg -blocks <blk_name>')
    print(f'##++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++')
    print(f'##++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++')
    print(f'## Step III: Check and debug input data after users provide all of input files.')
    print(f'## Cmd  Line: xconst sdgen -gen_dir <sdcdir> -hier_yaml <hier_file> -chk_only -blocks <blk_name>')
    print(f'##++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++')
    print(f'##++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++')
    print(f'## Step IV: Generate SDC files if all of input files are OK.')
    print(f'## Cmd  Line: xconst sdgen -gen_dir <sdcdir> -hier_yaml <hier_file> -sdc -blocks <blk_name> [-proj] [-flat]')
    print(f'##++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++')
    print(f'##++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++')
    print(f'## Step V: Check and debug SDC output files.')
    print(f'## Cmd  Line: xconst sdgen -gen_dir <sdcdir> -hier_yaml <hier_file> -chk_sdc -blocks <blk_name>')
    print(f'##++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++')




###########################################################################################################################
# upf meesage
upf_message_list =  []

def upf_log(level, msg, out=sys.stdout):
    print(f'{level.upper()}: {msg}', flush=True, file=out)

def upf_info(msg,kw='log'):
    if msg not in upf_message_list:
        upf_message_list.append(msg)
        upf_log('UPF_INFO', msg)
        if kw == 'log':
            printlog(f'UPF_INFO: {msg} \n')
        if kw == 'chk_hym':
            printlog(f'UPF_INFO: {msg} \n','chk_hym.rpt')
        if kw == 'chk_sht':
            printlog(f'UPF_INFO: {msg} \n','chk_sht.rpt')
        

def upf_warn(msg,kw='log'):
    if msg not in upf_message_list:
        upf_message_list.append(msg)
        print(f'\033[0:31mUPF_WARN\033[0m: {msg}', flush=True)
        if kw == 'log':
            printlog(f'UPF_WARN: {msg} \n')
        if kw == 'chk_hym':
            printlog(f'UPF_WARN: {msg} \n','chk_hym.rpt')
        if kw == 'chk_sht':
            printlog(f'UPF_WARN: {msg} \n','chk_sht.rpt')
         

def upf_error(msg,kw='log'):
    if msg not in upf_message_list:
        upf_message_list.append(msg)
        print(f'\033[0:31mUPF_ERROR\033[0m: {msg}', flush=True)
        if kw == 'log':
            printlog(f'UPF_ERROR: {msg} \n')
        if kw == 'chk_hym':
            printlog(f'UPF_ERROR: {msg} \n','chk_hym.rpt')
        if kw == 'chk_sht':
            printlog(f'UPF_ERROR: {msg} \n','chk_sht.rpt')

def upf_fatal(msg,kw='log'):
    if msg not in upf_message_list:
        upf_message_list.append(msg)
        print(f'\033[0:31mUPF_FATAL\033[0m: {msg}', flush=True)  
        if kw == 'log':
            printlog(f'UPF_FATAL: {msg} \n')
        if kw == 'chk_hym':
            printlog(f'UPF_FATAL: {msg} \n','chk_hym.rpt')
        if kw == 'chk_sht':
            printlog(f'UPF_FATAL: {msg} \n','chk_sht.rpt')
           
    sys.exit(1)    

def upf_args():
    print(f'##++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++')
    print(f'## UPF Generation Optional Arguments Presentation:')
    print(f'## -hlp:        All of flow option usage reference.')
    print(f'## -usg:        Flow step command usage reference.')
    print(f'## -tmp:        Write out template UPFs for reference')
    print(f'## -gen_dir:    UPF generationg directory')
    print(f'## -hier_yaml:  Design hierarchy info. from user input file')
    print(f'## -setup:      Build setup directories from blocks defined in hier yaml')
    print(f'## -blocks:     Indicate only current design name for UPF generation')
    #print(f'## -flat:       Generate flatten UPF based on current design. If not flat, only generate current design level only UPF')
    print(f'## -dg:         Generate design guide file to be present UPF request format')
    #print(f'## -idg:        Incrementally update design guide file based on existed dg file and updated input files')
    print(f'## -upf:        Write out UPF files')
    print(f'## -check_hym: Check hier yaml data before generating UPF')
    print(f'## -check_only: Check input data before generating UPF')
    print(f'## -check_upf:  Check UPF consistency after generating UPF')
    print(f'## -proj:       Open project mode. Maybe need set some related project environment variables')
    print(f'##++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++')


def upf_usage():
    print(f'##++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++')
    print(f'## Before flow run, user must provide input data including hier yaml and empty vfile, dg file and crg/ip UPFs to be integrated ')
    print(f'## For hier yaml and empty vfile, must follow format of hier_pwr.yaml and pmempty.v in template/ftemp')
    print(f'## For crg/ip UPFs, must follow header format of crg.upf and userip.upf in template/ftemp')
    print(f'##++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++')
    print(f'##++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++')
    print(f'## Step I: Build the whole UPF generation directory structure.')
    print(f'## Cmd Line: xconst sdgen -gen_dir <upfdir> -hier_yaml <hier_file> -setup -blocks <blk_name> [-tmp]')
    print(f'##++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++')
    print(f'##++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++')
    print(f'## Step II: Generate initial design guide file according to current input data.')
    print(f'## Cmd Line: xconst sdgen -gen_dir <upfdir> -hier_yaml <hier_file> -dg -blocks <blk_name>')
    print(f'##++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++')
    print(f'##++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++')
    print(f'## Step III: Check and debug input data after users provide all of input files.')
    print(f'## Cmd  Line: xconst sdgen -gen_dir <upfdir> -hier_yaml <hier_file> -chk_only -blocks <blk_name>')
    print(f'##++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++')
    print(f'##++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++')
    print(f'## Step IV: Generate UPF files if all of input files are OK.')
    print(f'## Cmd  Line: xconst sdgen -gen_dir <upfdir> -hier_yaml <hier_file> -upf -blocks <blk_name> [-proj] [-flat]')
    print(f'##++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++')
    print(f'##++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++')
    print(f'## Step V: Check and debug UPF output files.')
    print(f'## Cmd  Line: xconst sdgen -gen_dir <upfdir> -hier_yaml <hier_file> -chk_upf -blocks <blk_name>')
    print(f'##++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++')


ORG_COMDIR = os.getenv('ECS_TEMPLATES_DIR')
TEMP_DIR = os.getenv('TEMP_UPLOAD_DIR')
LOG_DIR = os.getenv('TASK_LOGS_DIR')
# ORG_COMDIR = r'E:\stone\work\smalltool\pycharm\const0203_sysflatok_full\templates'
# TEMP_DIR = r'E:\stone\work\smalltool\pycharm\const0203_sysflatok_full\temp'
# LOG_DIR = r'E:\stone\work\smalltool\pycharm\const0203_sysflatok_full\logs'
# adfeicc67ere

def sdc_dg_gen(*arglist):
    if len(arglist) == 0: arglist = ['-h']
    parser = argparse.ArgumentParser(prog='sdc_dg_gen', description='design constraint excel file generation script')
    parser.add_argument('-taskid', help='Task ID name', default='abc123xyesergtyjht', required='False')
    parser.add_argument('-dg', help='Write or update design guide files', action='store_true')

    args = parser.parse_args(args=arglist)
    taskid = args.taskid

    if args.dg:
        hier_yaml = os.path.join(TEMP_DIR, taskid, 'hier.yaml')
        vfile = os.path.join(TEMP_DIR, taskid, 'vlog.v')
        logdir = os.path.join(LOG_DIR,taskid)

        # tfile = os.path.join(TEMP_DIR, 'tune.sdc')
        # tfile = os.path.join(ORG_COMDIR, 'tune.sdc')

        dg_temp = os.path.join(ORG_COMDIR, 'sdcgen','dcont_org.xlsx')
        dg_file = os.path.join(TEMP_DIR, taskid,'dcont.xlsx')
        dg_file = re.sub(r'/+', '/', dg_file)
        lock_file = f'{TEMP_DIR}/.~lock.dcont.xlsx#'
        lock_file = re.sub(r'/+', '/', lock_file)

        # check dcont excel file existence
        if not os.path.exists(dg_temp):
            sdc_error(f'dcont excel file not found {dg_temp}')
            exit(1)

        # check hier yaml existence
        if not os.path.exists(hier_yaml):
            sdc_error(f'hier yaml file not found {hier_yaml}')
            exit(1)

        # check vfile existence
        if not os.path.exists(vfile):
            sdc_error(f'Empty vfile not found {vfile}')
            exit(1)

        if os.path.exists(dg_temp):
            # os.system(f'cp -f {dg_temp} {dg_file}')

            # 使用 shutil 复制文件，自动处理不同操作系统的差异
            shutil.copy2(dg_temp, dg_file)  # 保留文件元数据
            sdc_info(f'SDC design guide not found and copy from template design guide file at first time.')
        else:
            sdc_error(f'SDC original design guide NOT been found in template directory.')
            exit(1)

        # check logdir existence
        if not os.path.exists(logdir):
            sdc_error(f'logdir not found {logdir}')
            exit(1)

        if os.path.exists(lock_file):
            sdc_fatal('dcont.xlsx is in edit mode. Please close it')

        rm_exist_log(logdir)
        hier_tree = HierPwrTree(hier_yaml)
        sdcdg = SDC_DG()
        sdcdg.hier_tree = hier_tree
        sdcdg.read_vfile(vfile)
        if os.path.exists(dg_file):
            sdcdg.load_design_guide(dg_file)
        sdcdg.update_dg()
        sdcdg.save_workbook(dg_file)

        sdc_info(f'Design guide file {dg_file} is updated.')

if __name__ == '__main__':
    if len(sys.argv) < 2 or (len(sys.argv) > 2 and 'sdc_dg_gen' not in sys.argv[1]):
        sdc_error('Missing some parameters for SDC generation')
        locals()['sdc_dg_gen']('-h')
        exit(1)
    app_name = sys.argv[1]
    if app_name in locals():
        locals()[app_name](*sys.argv[2:])
    else:
        raise NameError(f'The application of DataBase generation {app_name} not found')

