

#!/usr/bin/env python3


import os,sys
import yaml
import re
import json

from os.path import dirname, abspath, basename
import time

import argparse
from glob import glob

import shutil
import tkinter
import yaml

#from itertools import chain

import openpyxl
from openpyxl import worksheet 
from pprint import pprint 
import pandas as pd
from openpyxl.utils import get_column_letter 

import tkinter as tk

from openpyxl.styles import Border, Side, PatternFill, Font, Alignment 
from openpyxl.worksheet.datavalidation import DataValidation

from collections import Counter



# Define some attributes and various methods for every block in hier tree
class BaseBlock():
    def __init__(self,name):
        self._name = name
        self._alias = ''
        self._hdlevel = 'blk'
        self._prime_pwr = ''
        self._constr_dir = ''
        self._insts = []
        self._mac_insts = []
        self._dig_insts = []
        self._cust_insts = {}
        self._proj = False

    def __repr__(self):
        return '<%s name=%s alias=%s>' % (self.__class__.__name__, self._name, self._alias)

    @property 
    def hdlevel(self):
        return self._hdlevel

    @hdlevel.setter
    def hdlevel(self, level):
        supported_lvs = ('top', 'sys', 'blk', 'soft', 'lib', 'crg', 'pll')
        if level not in supported_lvs:
            upf_error(f'Unsupported HDLEVEL {level} of block {self._name}, should be one of {supported_lvs}')
            return
        else:
            if level == 'sys' or level == 'top':
                self._hdlevel = 'sys'
            else:
                self._hdlevel = level

    @property
    def lvl_flat(self):
        if self._hdlevel == 'sys':
            return 'IS_CHIP'
        else:
            return 'IS_FLAT'
        
    @property
    def alias(self):
        return self._alias

    @alias.setter
    def alias(self, alias):
        self._alias = alias

    @property
    def prime_pwr(self):
        return self._prime_pwr

    @prime_pwr.setter
    def prime_pwr(self, pwr):
        self._prime_pwr = pwr

    @property
    def insts(self):
        return self._insts

    @insts.setter
    def insts(self, insts):
        self._insts = insts

    @property
    def mac_insts(self):
        return self._mac_insts

    @mac_insts.setter
    def mac_insts(self, mac_insts):
        self._mac_insts = mac_insts

    @property
    def dig_insts(self):
        return self._dig_insts

    @dig_insts.setter
    def dig_insts(self, dig_insts):
        self._dig_insts = dig_insts

    @property
    def constr_dir(self):
        return self._constr_dir
    
    @constr_dir.setter
    def constr_dir(self, consdir):
        self._constr_dir = consdir

    @property
    def proj(self):
        return self._proj

    @proj.setter
    def proj(self, proj):
        self._proj = proj

    def get_curcust_by_name(self, inst_type,flg=''):
        _hier = []
        _ref = []
        #_lvl = []
        _alias = []
        _pwr = []
        _usersdc = []
        if inst_type == 'insts' and self.insts:
            for i in range(0,len(self.insts)):
                finst = self.insts[i].split(',')
                if len(finst) == 3:
                    _hier.append(finst[0].strip())
                    _ref.append(finst[1].strip())
                    _alias.append(None)
                    _pwr.append(finst[2].strip())
                    _usersdc.append(None)
                if len(finst) == 4:
                    _hier.append(finst[0].strip())
                    _ref.append(finst[1].strip())
                    if not flg:
                        _alias.append(finst[2].strip().replace('#',''))
                    else:
                        _alias.append(finst[2].strip())
                    _pwr.append(finst[3].strip())
                    _usersdc.append(None)
            self._cust_insts['insthier'] = _hier
            self._cust_insts['instref'] = _ref
            self._cust_insts['instalias'] = _alias
            self._cust_insts['instpwr'] = _pwr
            self._cust_insts['instuser'] = _usersdc
        
        if inst_type == 'mac_insts' and self.mac_insts:
            for i in range(0,len(self.mac_insts)):
                if isinstance(self.mac_insts[i],str):
                    fmac = self.mac_insts[i].split(',')  
                    if len(fmac) == 3:
                        _hier.append(fmac[0].strip())
                        _ref.append(fmac[1].strip())
                        _alias.append(None)
                        _pwr.append(fmac[2].strip())
                        _usersdc.append(None)
                    if len(fmac) == 4:
                        _hier.append(fmac[0].strip())
                        _ref.append(fmac[1].strip())
                        if not flg:
                            _alias.append(fmac[2].strip().replace('#',''))
                        else:
                            _alias.append(fmac[2].strip())
                        _pwr.append(fmac[3].strip())
                        _usersdc.append(None)
                if isinstance(self.mac_insts[i],dict):
                    fmac = ''.join(self.mac_insts[i].keys()).split(',')
                    _hier.append(fmac[0].strip())
                    _ref.append(fmac[1].strip()) # + '_USR' )
                    _alias.append(None)
                    _pwr.append(fmac[2].strip())
                    _usersdc.append(''.join(self.mac_insts[i].values()))  
            self._cust_insts['machier'] = _hier
            self._cust_insts['macref'] = _ref
            self._cust_insts['macalias'] = _alias
            self._cust_insts['macpwr'] = _pwr
            self._cust_insts['macuser'] = _usersdc

        if inst_type == 'dig_insts' and self.dig_insts:
            for i in range(0,len(self.dig_insts)):
                if isinstance(self.dig_insts[i],str):
                    fdig = self.dig_insts[i].split(',')  
                    if len(fdig) == 3:
                        _hier.append(fdig[0].strip())
                        _ref.append(fdig[1].strip())
                        _alias.append(None)
                        _pwr.append(fdig[2].strip())
                        _usersdc.append(None)
                    if len(fdig) == 4:
                        _hier.append(fdig[0].strip())
                        _ref.append(fdig[1].strip())
                        if not flg:
                            _alias.append(fdig[2].strip().replace('#',''))
                        else:
                            _alias.append(fdig[2].strip())
                        _pwr.append(fdig[3].strip())
                        _usersdc.append(None)
                if isinstance(self.dig_insts[i],dict):
                    fdig = ''.join(self.dig_insts[i].keys()).split(',')
                    _hier.append(fdig[0].strip())
                    _ref.append(fdig[1].strip()) # + '_USR')
                    _alias.append(None)
                    _pwr.append(fdig[2].strip())
                    _usersdc.append(''.join(self.dig_insts[i].values()))
            self._cust_insts['dighier'] = _hier
            self._cust_insts['digref'] = _ref
            self._cust_insts['digalias'] = _alias
            self._cust_insts['digpwr'] = _pwr
            self._cust_insts['diguser'] = _usersdc       

        return self._cust_insts

    def get_curhd_by_name(self):
        #return self.name.split() + self._cust_insts['instref']
        self.get_curcust_by_name('insts')
        if 'instref' in self._cust_insts:
            return self._cust_insts['instref']

    def get_curmac_by_name(self,flg=''):
        self.get_curcust_by_name('mac_insts',flg)
        if 'macref' in self._cust_insts:
            return self._cust_insts['macref']

    def get_curdig_by_name(self,flg=''):
        self.get_curcust_by_name('dig_insts',flg)
        if 'digref' in self._cust_insts:
            return self._cust_insts['digref']
    
    def get_curuser_by_name(self, inst_type):
        self.get_curcust_by_name('mac_insts')
        self.get_curcust_by_name('dig_insts')
        if inst_type == 'mac_insts' and 'macuser' in self._cust_insts:           
            return self._cust_insts['macuser']       
        elif inst_type == 'dig_insts'and 'diguser' in self._cust_insts:
            return self._cust_insts['diguser']
        else:
            return None

    # alias val
    def get_curhdval_by_name(self):
        # return self.name.split() + self._cust_insts['instalias']
        self.get_curcust_by_name('insts')
        if 'instalias' in self._cust_insts:
            return self._cust_insts['instalias']

    def get_curmacval_by_name(self, flg=''):
        self.get_curcust_by_name('mac_insts', flg)
        if 'macalias' in self._cust_insts:
            return self._cust_insts['macalias']

    def get_curdigval_by_name(self, flg=''):
        self.get_curcust_by_name('dig_insts', flg)
        if 'digalias' in self._cust_insts:
            return self._cust_insts['digalias']

    # pwr
    def get_curhdpwr_by_name(self):
        # return self.name.split() + self._cust_insts['instpwr']
        self.get_curcust_by_name('insts')
        if 'instpwr' in self._cust_insts:
            return self._cust_insts['instpwr']

    def get_curmacpwr_by_name(self, flg=''):
        self.get_curcust_by_name('mac_insts', flg)
        if 'macpwr' in self._cust_insts:
            return self._cust_insts['macpwr']

    def get_curdigpwr_by_name(self, flg=''):
        self.get_curcust_by_name('dig_insts', flg)
        if 'digpwr' in self._cust_insts:
            return self._cust_insts['digpwr']

class HierPwrTree():
    def __init__(self,yaml_file):
        self.yaml_file = yaml_file
        self._blocks = {}
        self._primepwr = {}
        self._yaml_data = {}
        self._hierdata = {}
        self._pwrdata = {}
        #self._blktrees = {}
        self.build_hier_tree(yaml_file)
        

    def build_hier_tree(self, yaml_file):

        # get yaml_data
        yaml_data = {}
        if not os.path.exists(yaml_file):
            raise FileExistsError(f'{yaml_file} does not exists')
        with open(yaml_file, 'r') as fh:
            yaml_data = yaml.load(fh, yaml.FullLoader)

        if 'hier' not in yaml_data:
            print('Missing hier keyword in yaml file.')
            upf_fatal(f'Must include keyword <hier>')
        if 'pwr' not in yaml_data:
            print('Missing pwr keyword in yaml file.')
            upf_fatal(f'Must include keyword <pwr>')

        # get '_primepwr'
        for pwr_name in yaml_data['pwr'].keys():
            if yaml_data['pwr'][pwr_name]:
                self._primepwr[pwr_name] = yaml_data['pwr'][pwr_name]   

        for blk_name in yaml_data['hier'].keys():

            self._blocks[blk_name] = BaseBlock(blk_name)

            if 'alias' in yaml_data['hier'][blk_name] and yaml_data['hier'][blk_name]['alias']:
                self._blocks[blk_name].alias = yaml_data['hier'][blk_name]['alias']
            else:
                self._blocks[blk_name].alias = None

            if 'hdlevel' in yaml_data['hier'][blk_name] and yaml_data['hier'][blk_name]['hdlevel']:
                self._blocks[blk_name].hdlevel = yaml_data['hier'][blk_name]['hdlevel']
            else:
                self._blocks[blk_name].hdlevel = None            
            
            if 'prime_pwr' in yaml_data['hier'][blk_name] and yaml_data['hier'][blk_name]['prime_pwr'] in self._primepwr and yaml_data['hier'][blk_name]['prime_pwr']:
                self._blocks[blk_name].prime_pwr = yaml_data['hier'][blk_name]['prime_pwr'] + ' ' + self._primepwr[yaml_data['hier'][blk_name]['prime_pwr']]
            else:
                self._blocks[blk_name].prime_pwr = None 

            if 'constr_dir' in yaml_data['hier'][blk_name] and yaml_data['hier'][blk_name]['constr_dir']:
                self._blocks[blk_name].constr_dir = yaml_data['hier'][blk_name]['constr_dir']
            else:
                self._blocks[blk_name].constr_dir = None

            if 'insts' in yaml_data['hier'][blk_name] and yaml_data['hier'][blk_name]['insts']:
                self._blocks[blk_name].insts = yaml_data['hier'][blk_name]['insts']
            else:
                self._blocks[blk_name].insts = None

            if 'mac_insts' in yaml_data['hier'][blk_name] and yaml_data['hier'][blk_name]['mac_insts']:
                self._blocks[blk_name].mac_insts = yaml_data['hier'][blk_name]['mac_insts']
            else:
                self._blocks[blk_name].mac_insts = None

            if 'dig_insts' in yaml_data['hier'][blk_name] and yaml_data['hier'][blk_name]['dig_insts']:
                self._blocks[blk_name].dig_insts = yaml_data['hier'][blk_name]['dig_insts']
            else:
                self._blocks[blk_name].dig_insts = None

        self._hierdata = yaml_data['hier'] 
        self._pwrdata = yaml_data['pwr']
        self._yaml_data = yaml_data

    def get_block_by_name(self, name) -> BaseBlock:
        if name in self._blocks:
            return self._blocks[name]
        else:
            return None    

    def get_curblks(self,name):
        curblks = []
        
        allblks = list(self._hierdata.keys())
        if name in allblks:
            curblk = self.get_block_by_name(name)

            if curblk.get_curhd_by_name():
                curblks = [x for x in curblk.get_curhd_by_name() if x is not None]
            if curblk.get_curmac_by_name():
                curblks += [x for x in curblk.get_curmac_by_name() if x is not None]
            if curblk.get_curdig_by_name():
                curblks += [x for x in curblk.get_curdig_by_name() if x is not None]  
        else:
            upf_warn(f'{name} is not expanded in hier_pwr yaml file.')
        
        return  curblks
    
    def get_hiertrees(self, name, blktrees={}, valstyle=None) -> dict:

        curblks = self.get_curblks(name)
        #blktrees = {}
        if curblks:
            new_curblks = [elem.replace('_USR', '') if re.search(r'_USR$',elem) else elem for elem in curblks]
            blktrees[name] = new_curblks
        else:
            if valstyle:
                blktrees[name] = None   

        for blk in curblks:
            if re.search(r'_USR$',blk):
                blk = blk.replace('_USR','')
                upf_warn(f'{blk} is not expanded in hier_pwr yaml file.')
            elif len(curblks) > 0:
                self.get_hiertrees(blk,blktrees)    
        
        return blktrees

    def get_hierblks(self, name) -> list:

        blktrees = self.get_hiertrees(name)
        result = []
        for key, value in blktrees.items():
            if key not in result:
                result.append(key)
            if isinstance(value, list):
                for element in value:
                    if element not in result:
                        result.append(element)
        return result        

    def get_hierblks_infos(self,name) -> dict:

        blks = self.get_hierblks(name)
        blksinfo = {}

        allblks = list(self._hierdata.keys())
        for val in blks:
            if val in allblks:
                blk = self.get_block_by_name(val)
                alias = blk.alias
                lvl = blk.hdlevel
                condir = blk.constr_dir
                pwr = blk.prime_pwr
                pwrg = pwr.split(' ')[0].strip()
                blksinfo[val] = val + f' {alias}' + f' {lvl}' + f' {pwrg}' + f' {condir}'
            else:
                blksinfo[val] = val + ' (NOT EXPEND)'

        return blksinfo 

    def get_hierblks_infos_chk(self, name) -> dict:
        """
        get hier block info
        :param name:
        :return:
        """
        blks = self.get_hierblks(name)
        if not blks:
            blks = [f'{name}']
        blksinfo = {}

        allblks = list(self._hierdata.keys())
        # print('______BLKDSF: ', blks, allblks)
        for sbk in blks:
            if sbk in allblks:
                blk = self.get_block_by_name(sbk)
                alias = blk.alias
                lvl = blk.hdlevel
                condir = blk.constr_dir
                pwr = blk.prime_pwr
                # nlb = blk.intg_nlb
                nlb = 'net'
                inst = blk.insts
                mac = blk.mac_insts
                dig = blk.dig_insts
                # pwrg = pwr.split(' ')[0].strip()
                blksinfo[sbk] = [f'{alias}', f'{lvl}', f'{pwr}', f'{nlb}', f'{inst}', f'{mac}', f'{dig}', f'{condir}']
            else:
                blksinfo[sbk] = ['NOT_EXPANDED']
                upf_error(f'{sbk} NOT expanded in hier yaml.')

        return blksinfo

    def get_full_hierblks_info(self):
        allblks = list(self._hierdata.keys())
        blksinfo = {}

        for sbk in allblks:
            blk = self.get_block_by_name(sbk)
            alias = blk.alias
            lvl = blk.hdlevel
            condir = blk.constr_dir
            pwr = blk.prime_pwr
            # nlb = blk.intg_nlb
            nlb = 'net'
            inst = blk.insts
            mac = blk.mac_insts
            dig = blk.dig_insts
            # pwrg = pwr.split(' ')[0].strip()
            blksinfo[sbk] = [f'{alias}', f'{lvl}', f'{pwr}', f'{nlb}', f'{inst}', f'{mac}', f'{dig}', f'{condir}']

        return blksinfo

    def get_cursub_alias(self, name, outtype='hd', htype='cur') -> list:
        '''
        get different harden insts, mac insts, dig insts under current design
        outtype is hd/lib/soft/crg/pll
        '''
        allhierblks = []
        insts_als = []
        ips_als = []
        digs_als = []
        plls_als = []
        crgs_als = []

        if htype == 'flt':
            allhierblks = self.get_hierblks(name)
            allhierblks = [x for x in allhierblks if x != name]
        if htype == 'cur':
            allhierblks = self.get_curblks(name)
        if htype == 'fcur':
            allhierblks.append(name)
            allhierblks.extend(self.get_curblks(name))
        # alltreeblks = self.get_hiertrees(name)
        # allhierblks.append(name)
        if allhierblks:
            for curblk in allhierblks:
                # if curblk in allblks:
                blk = self.get_block_by_name(curblk)
                lvl = blk.hdlevel
                if lvl in ['blk', 'sys', 'top']:
                    insts_als.append(blk.alias)
                elif lvl in ['crg']:
                    crgs_als.append(blk.alias)
                elif lvl in ['lib', 'soft']:
                    ips_als.append(blk.alias)
                elif lvl in ['pll']:
                    plls_als.append(blk.alias)
                else:
                    upf_error(f'hdlevel of {curblk} is wrong.')

        if outtype == 'hd':
            return insts_als
        if outtype == 'ip':
            return ips_als
        # if outtype == 'lib':
        #     return macs
        # if outtype == 'soft':
        #     return digs  
        # if outtype == 'pll':
        #     return plls
        if outtype == 'crg':
            return crgs_als
        if outtype == 'crgip':
            return crgs_als + ips_als

    def get_curlvlblks(self, name, outtype='hd', htype='cur') -> list:
        """
        get different harden insts, mac insts, dig insts under current design
        outtype is hd/lib/soft/crg/pll
        """
        allhierblks = []
        insts = []
        ips = []
        digs = []
        plls = []
        crgs = []

        if htype == 'cur':
            allhierblks.append(name)
            allhierblks.extend(self.get_curblks(name))
        if htype == 'slf':
            allhierblks.extend(self.get_curblks(name))
        # alltreeblks = self.get_hiertrees(name)
        if allhierblks:
            for curblk in allhierblks:
                # if curblk in allblks:
                blk = self.get_block_by_name(curblk)
                lvl = blk.hdlevel
                if lvl in ['blk', 'sys', 'top']:
                    insts.append(curblk)
                elif lvl in ['crg']:
                    crgs.append(curblk)
                elif lvl in ['lib', 'soft']:
                    ips.append(curblk)
                elif lvl in ['pll']:
                    plls.append(curblk)
                else:
                    upf_error(f'hdlevel of {curblk} is wrong.')

        if outtype == 'hd':
            return insts
        if outtype == 'ip':
            return ips
        # if outtype == 'lib':
        #     return macs
        # if outtype == 'soft':
        #     return digs  
        # if outtype == 'pll':
        #     return plls
        if outtype == 'crg':
            return crgs
        if outtype == 'crgip':
            return ips + crgs


    def get_hierlvlblks(self, name, outtype='hd') -> list:
        '''
        get different harden insts, mac insts, dig insts under current design
        outtype is hd/lib/soft/crg/pll
        '''
        allhierblks = []
        allhierblks = self.get_hierblks(name)
        #alltreeblks = self.get_hiertrees(name)
        allhierblks.append(name)

        #allblks = list(self._hierdata.keys())

        insts = []
        macs = []
        digs = []
        plls = []
        crgs = []
        if allhierblks:
            for curblk in allhierblks:
                #if curblk in allblks:
                blk = self.get_block_by_name(curblk)
            
                lvl = blk.hdlevel
                if lvl in ['blk', 'sys', 'top']:
                    insts.append(curblk)
                if lvl in ['lib']:
                    macs.append(curblk)  
                if lvl in ['pll']:
                    plls.append(curblk) 
                if lvl in ['soft']:
                    digs.append(curblk)  
                if lvl in ['crg']:
                    crgs.append(curblk)
                # else:
                #     #print(f'XYZ_{curblk}')
                #     for ky,vl in alltreeblks.items():
                #         if curblk in vl:
                #             parent = ky
                #     blkg = self.get_block_by_name(parent)
                #     if curblk in blkg.get_curhd_by_name():
                #         insts.append(curblk)
                #         upf_warn(f'{curblk} is harden block need expand it in hier_pwr yaml file.')
                #     if curblk in blkg.get_curmac_by_name() or f'{curblk}_USR' in blkg.get_curmac_by_name():
                #         macs.append(curblk)
                #         upf_warn(f'{curblk} is macro block, if not user constraint, need expand it in hier_pwr yaml file.')
                #     if curblk in blkg.get_curdig_by_name() or f'{curblk}_USR' in blkg.get_curdig_by_name():
                #         digs.append(curblk)
                #         upf_warn(f'{curblk} is digital block, if not user constraint, need expand it in hier_pwr yaml file.')

        if outtype == 'hd':
            return insts
        if outtype == 'lib':
            return macs
        if outtype == 'soft':
            return digs  
        if outtype == 'pll':
            return plls
        if outtype == 'crg':
            return crgs  

    def get_hierdepth(self,dic,key):
        
        allblks = list(self._hierdata.keys())
    
        if key not in dic:
            return 1
        else:
            max_depth = 0
            for sub_key in dic[key]:
                if sub_key not in allblks:
                    upf_warn(f'{sub_key} is not expanded in hier yaml file.')
                else:
                    current_depth = self.get_hierdepth(dic, sub_key) + 1
                    max_depth = max(max_depth, current_depth)
            return max_depth

        # # blktrees = self.get_hiertrees(name)
        # if key not in dic:
        #     return 0
        
        # depths = []
        # for child_key in dic[key]:
        #     if child_key not in allblks:
        #         print(f'{child_key} is not expanded in hier yaml file.')
        #     else:
        #         depths.append(self.get_hierdepth(dic,child_key))
        #         #max_depth = max(max_depth, depth)
        
        # return max(depths) + 1

        # max_depth = depth  # 记录最大深度
        
        # if isinstance(blktrees, dict):
        #     for child_node in blktrees.values():
        #         if child_node not in allblks:
        #             print(f'{child_node} is not expanded in hier yaml file.')
        #     else:
        #         child_depth = self.get_hierdepth(child_node, depth + 1)
        #         max_depth = max(max_depth, child_depth)
        
        # return max_depth

    def get_alias_by_name(self, name):
        return self._blocks[name].alias


    def parse_env_path(self, epath):
        interp = tkinter.Tcl()
        interp.tk.eval('proc user_path {args} {return $args}')
        rsdcdir = f'user_path {epath}'

        try:
            rsdcdir = interp.tk.eval(rsdcdir)
        except:
            upf_fatal(f'{epath} can not be parsed. Must set env(variable) in the path.')

        return rsdcdir

    def get_sblk_rdir(self, sblk, als, kw='outputs/'):
        if not self._proj:
            # sblk_rdir = f'$SDCVAR(SDC_DIR,${{{als}}})../../{sblk}/sdc/{kw}'
            vardata = self.get_vars_from_vardef(sblk, self._vfdir, 'vardef.json')
            sdir = vardata['SDC_DIR']
            sblk_rdir = f'{sdir}../../../{sblk}/sdc/{kw}'
        else:
            sblk_rdir = f'{self.get_block_by_name(sblk).constr_dir}/sdc/{kw}'

        if re.search(r'\$env\(', sblk_rdir):
            sblk_rdir = self.post_process_dir(sblk_rdir)
            sblk_rdir = self.parse_env_path(sblk_rdir)
        else:
            sblk_rdir = self.post_process_dir(sblk_rdir)

        return sblk_rdir

    def post_process_dir(self, filepath):
        filepath = re.sub(r'\/+', '/', filepath)
        return filepath

    def get_vars_from_vardef(self, blk, vfdir, fname):
        jdata = {}
        tmvars = {}

        # cblk = self.get_block_by_name(blk)
        hdir = f'{vfdir}/../json'

        if not os.path.exists(hdir):
            upf_warn(f'Missing directory {hdir} for {blk} json check in hierpwr.')
        else:
            json_file = f'{hdir}/{fname}'
            if os.path.exists(json_file):
                jdata = self.read_json(json_file)
            else:
                upf_warn(f'Missing {fname} file of {blk}')

        mvars = {k: v for k, v in jdata.items() if 'TMVAR' in k}
        # tmhier = {k:v for k,v in jdata.items() if 'TMHIER' in k}
        if mvars:
            for k, v in mvars.items():
                tmvars[v['Variable']] = v['Value']
        else:
            upf_warn(f'Missing TMVAR info. in {blk} vardef.json')

        return tmvars

   ########################################################################
    def check_hym(self, name='', fkw='chk_hym'):
        # tals = self.get_alias_by_name(name)
        if name:
            tbk = self.get_block_by_name(name)
            if tbk.hdlevel == 'top':
                tcblk = tbk.get_curhd_by_name()
                if tcblk:
                    for cblk in tcblk:
                        cbk = self.get_block_by_name(cblk)
                        # cals = self.get_alias_by_name(cblk)
                        if cbk.hdlevel == 'sys':
                            scrgals = self.get_cursub_alias(cblk, 'crg')
                            if not scrgals:
                                upf_error(
                                    f'"hdlevel" with "sys" of {cblk} under top {name} must include crg in dig_insts.',
                                    f'{fkw}')
                        if cbk.hdlevel == 'blk':
                            scrgals = self.get_cursub_alias(cblk, 'crg')
                            if scrgals:
                                upf_error(
                                    f'"hdlevel" with "blk" of {cblk}  under top {name} must NOT include crg in dig_insts.',
                                    f'{fkw}')

        # blk: [f'{alias}',f'{lvl}',f'{pwr}',f'{nlb}',f'{inst}',f'{mac}',f'{dig}',f'{condir}']
        curflg = True
        if name == '':
            hblksinfo = self.get_full_hierblks_info()
            curflg = False
        else:
            curflg = True
            hblksinfo = self.get_hierblks_infos_chk(name)
            crgs = self.get_hierlvlblks(name, 'crg')
            ips = self.get_hierlvlblks(name, 'ip')
            digs = self.get_hierlvlblks(name, 'soft')
            macs = self.get_hierlvlblks(name, 'lib')
            insts = self.get_hierlvlblks(name)
        # allblks = list(self._hierdata.keys())
        # print('hblksinfo: ',hblksinfo)
        #{'jpeg_top_wrap': ['JPEG', 'blk', 'VDDM_CLPS TT0P800V TT0P750V TT0P700V', 'net', 'None', 'None', "['u_jpeg_top/u_arv_top, arv_top, VDDM_CLPS']", '/home/master/stone/tmp/const/tpy/vscode/sdcgen/template/inputs/'], 
        # 'arv_top': ['AVE', 'soft', 'VDDM_CLPS TT0P800V TT0P750V TT0P700V', 'net', 'None', '[{\'u_cr8_top, cr8_top,VDDM_CLPS\': \'if {[file exists $SDCVAR(SDC_DIR,${AVE})/arv_top/sdc/inputs/mdblk/cr8_top.sdc]} {\\n    source -echo -verbose $SDCVAR(SDC_DIR,${AVE})/arv_top/sdc/inputs/mdblk/cr8_top.sdc\\n} else {\\n    puts "SDC_WARN: Missing cr8_top SDC file. Please check it."\\n}\\n\'}]', 'None', '/home/master/stone/tmp/const/tpy/vscode/sdcgen/template/inputs/'], 'cr8_top': ['CR8', 'lib', 'VDDM_CLPS TT0P800V TT0P750V TT0P700V', 'net', 'None', 'None', 'None', '$env(SYS_CAMERA_ICDIR)/de/impl/spg_top/']}

        # print('checkenvpath: ',self.parse_env_path('$env(MCUJPEG_ICDIR)/case/mcu_jpeg_top_wrap/'))
        # pwr check: supply vol must from big to small
        if self._pwrdata:
            for sp, sv in self._pwrdata.items():
                xsv = sorted(sv.split(' '))[::-1]
                if xsv != sv.split(' '):
                    upf_warn(f'The voltage of "{sp}" in "pwr" NOT follow decreasing order.', f'{fkw}')
        else:
            upf_error(f'Missing "pwr" definition in hier yaml file.', f'{fkw}')

        if hblksinfo:
            hdblks = list(hblksinfo.keys())
            # pwr check: supply name must cover all supplys from hier
            hpwr = [x[2] for x in hblksinfo.values()]
            #hpwr : ['VDD_CORE TT0P750V TT0P700V TT0P650V', 'VDD_LPI TT0P850V TT0P650V TT0P600V']
            if hdblks:
                xpwr = []
                for hbk in hdblks:
                    sbk = self.get_block_by_name(hbk)
                    if sbk.get_curhdpwr_by_name():
                        xpwr.extend(sbk.get_curhdpwr_by_name())
                    if sbk.get_curmacpwr_by_name():
                        xpwr.extend(sbk.get_curmacpwr_by_name())
                    if sbk.get_curdigpwr_by_name():
                        xpwr.extend(sbk.get_curdigpwr_by_name())
                # xpwr: ['VDD_CORE', 'VDD_LPI', 'VDD_CORE', 'VDDM_CLPS', 'VDD_MM_CSS']
                # if hpwr:    xpwr.extend(hpwr)
                # print('_cxpwr: ',hpwr,xpwr)
                if hpwr:
                    # self._pwrdata: {'VDD_CORE': 'TT0P750V TT0P700V TT0P650V', 'VDD_MM_CSS': 'TT0P650V TT0P600V TT0P550V'}
                    mpwr = [x for x in list(set(xpwr)) if x not in self._pwrdata.keys()]
                    if mpwr:
                        upf_error(f'Supply {mpwr} missing definition in "pwr" of hier yaml file.', f'{fkw}')

            # als var must be unique
            als_var = [x[0] for x in hblksinfo.values()]
            als_var_u = list(set(als_var))
            if len(als_var) != len(als_var_u):
                counts = Counter(als_var)
                rvars = [ix for ix, cnt in counts.items() if cnt > 1]
                upf_error(f'Alias variable {rvars} is not unique.', f'{fkw}')

            # als val must be unique
            if hdblks:
                als_val = []
                for hbk in hdblks:
                    sbk = self.get_block_by_name(hbk)
                    if sbk.get_curhdval_by_name():
                        als_val.extend(sbk.get_curhdval_by_name())
                    if sbk.get_curmacval_by_name():
                        als_val.extend(sbk.get_curmacval_by_name())
                    if sbk.get_curdigval_by_name():
                        als_val.extend(sbk.get_curdigval_by_name())
                als_val = [x for x in als_val if x != None and re.search(r'\$|\:|\{|\}', x)]
                als_val_u = list(set(als_val))
                if len(als_val) != len(als_val_u):
                    counts = Counter(als_val)
                    rvals = [ix for ix, cnt in counts.items() if cnt > 1]
                    upf_error(f'Alias value {rvals} is not unique.', f'{fkw}')

            for blk, alst in hblksinfo.items():
                # can not cover in full hier yaml check
                if alst == 'NOT EXPANDED':
                    upf_error(f'{blk} Not expanded in hier yaml.', f'{fkw}')

                # alias
                if not alst[0] or alst[0] == 'None':
                    upf_error(f'{blk} Not found <ALIAS> value in hier yaml. Must start upper string.', f'{fkw}')
                if re.search(r'^\d+', alst[0]):
                    upf_error(
                        f'Found {blk} <ALIAS> value begins with digital number in hier yaml. Must begin with upper string.', f'{fkw}')
                if re.search(r'[a-z]', alst[0]):
                    upf_error(f'Found {blk} <ALIAS> value includes lower string in hier yaml. Must use upper string.', f'{fkw}')
                if re.search(r'\_|\-|\$|\%|\@|\#|\&|\[|\]|\(|\)|\*|\:', alst[0]):
                    upf_error(
                        f'Found {blk} <ALIAS> value includes "_|-|$|%|@|#|&|[|]|(|)|*|:" in hier yaml. Must NOT use special string.', f'{fkw}')

                # hdlevel
                if not alst[1] or alst[1] == 'None':
                    upf_error(
                        f'{blk} Not found <HDLEVEL> value in hier yaml. Must be "blk/sys/top/crg/lib/soft" value.',
                        f'{fkw}')
                lvslst = ('top', 'sys', 'blk', 'soft', 'lib', 'crg')
                if alst[1] not in lvslst:
                    upf_error(f'Not Found {blk} <HDLEVEL> value includs "blk/sys/top/crg/lib/soft" in hier yaml.',
                                f'{fkw}')
                if curflg:
                    if crgs:
                        if blk in crgs:
                            if alst[1] not in ('crg'):
                                upf_error(f'{blk} belongs to crgs, but hdlevel NOT found in "(crg)"', f'{fkw}')
                    if ips:
                        if blk in ips:
                            if alst[1] not in ('soft', 'lib'):
                                upf_error(
                                    f'{blk} belongs to mac_inst or dig_inst, but hdlevel NOT found in "(soft, lib)"', f'{fkw}')

                    if digs:
                        if blk in digs:
                            if alst[1] not in ('soft'):
                                upf_error(f'{blk} belongs to dig_inst, but hdlevel NOT found in "(soft)"', f'{fkw}')

                    if macs:
                        if blk in macs:
                            if alst[1] not in ('lib'):
                                upf_error(f'{blk} belongs to mac_inst, but hdlevel NOT found in "(lib)"', f'{fkw}')

                    if insts:
                        if blk in insts:
                            if alst[1] not in ('top', 'sys', 'blk'):
                                upf_error(f'{blk} belongs to harden insts, but hdlevel NOT found in "(top, blk, sys)"', f'{fkw}')

                # prime_pwr
                if not alst[2] or alst[2] == 'None':
                    upf_error(f'{blk} Not found <PRIME_PWR> value in hier yaml. Must set in <pwr> and <prime_pwr>.', f'{fkw}')
                # need add mim supply check???

                # intg_nlb
                # if not alst[3] or alst[3] == 'None':
                #     upf_error(f'{blk} Not found <INTG_NLB> value in hier yaml. Must be "net/lib/bbx" value.', f'{fkw}')
                # if alst[3] not in ('net', 'lib', 'bbx'):
                #     upf_error(f'Not found {blk} <INTG_NLB> value includes "net/lib/bbx" in hier yaml.', f'{fkw}')

                # # constr_dir
                # if alst[1] in ('blk', 'sys', 'top'):
                #     if not alst[7] or alst[7] == 'None':
                #         upf_warn(f'Not found {blk} constr_dir value for project mode.', f'{fkw}')
                #     else:
                #         rcdir = self.parse_env_path(alst[7])
                #         if not os.path.exists(rcdir):
                #             upf_error(f'Not found {blk} constr_dir real path. Must be the absolute existed path.', f'{fkw}')
                #         else:
                #             if not re.search(r'\/$', alst[7]):
                #                 upf_error(f'Not found {blk} constr_dir value include "/" for the end string', f'{fkw}')

                #             if alst[7].endswith('sdc') or alst[7].endswith('sdc/'):
                #                 upf_error(f'The end strings of {blk} constr_dir value must NOT be "sdc/" or "sdc"', f'{fkw}')

                #             vardata = self.get_vars_from_vardef(blk, self._vfdir, 'vardef.json')
                #             print('__var: ', blk, self._vfdir, vardata)
                #             if re.search(r'\$env(', vardata['SDC_DIR']):
                #                 sdir = self.parse_env_path(vardata['SDC_DIR'])
                #             else:
                #                 sdir = vardata['SDC_DIR']
                #             if re.search(r'\$env(', alst[7]):
                #                 gdir = self.parse_env_path(alst[7])
                #             else:
                #                 gdir = alst[7]
                #             if sdir != gdir:
                #                 upf_warn(f'NOT consistency bet <SDC_DIR> in vardef and <constr_dir> in hier yaml. If proj mode, Need be consistent.')

                # if alst[1] == 'crg':
                #     if not alst[7] or alst[7] == 'None':
                #         upf_error(f'{blk} is crg module. Must specify constr_dir value.', f'{fkw}')
                #     else:
                #         if not os.path.exists(alst[7]):
                #             upf_error(f'Not found {blk} constr_dir real path. If proj mode, Must be the absolute existed path.', f'{fkw}')
                #         else:
                #             if not re.search(r'\/$', alst[7]):
                #                 upf_error(f'Not found {blk} constr_dir value include "/" for the end string', f'{fkw}')
                #             if alst[7].endswith('sdc') or alst[7].endswith('sdc/'):
                #                 upf_error(f'The end strings of {blk} constr_dir value must NOT be "sdc/" or "sdc"', f'{fkw}')


class UPF_DG(object):
    def __init__(self):
        self._sheets = {}
        self._hier_tree = {}
        self._vardefdata = {}
        self._vardefpcell = {}
        self._wb = {}
        self._inputs = BaseInputs()
        # self._vardef = VarDefSheet()

        self.proj_mode = False
        self._vfile_data = None
        self._pmfile_data = None
        self._objfile_data = None
        self._objfile_list = None
        self._data = None
        self._upfdir = ''
        self._mdname = ''
        self._alias = ''
        
        # self._mdname = ''
        # self._blkalias = ''
        # self._blklvl = ''

    @property
    def hier_tree(self):
        return self._hier_tree
    
    @hier_tree.setter
    def hier_tree(self, hier_tree):
        self._hier_tree = hier_tree
        # if self._mdname and self._hier_tree._blocks[self._mdname]:
        #     self._blkalias = self._hier_tree._blocks[self._mdname].alias
        #     self._blklvl = self._hier_tree._blocks[self._mdname].hdlevel
 
    def load_design_guide(self,dg_file,kwd=''):
        # self._upfdir = abspath(dirname(dirname(dg_file)))
        self._wb = openpyxl.load_workbook(dg_file)

        valdef_sheet = self._wb['VarDef']
        # start = 0
        # for i in range(1, valdef_sheet.max_row+1):
        #     if valdef_sheet.cell(i,1).value == 'Variable':
        #         start = i + 1
        #         break
        # for i in range(start, valdef_sheet.max_row+1):
        #     key = valdef_sheet.cell(row=i, column=1).value
        #     val = valdef_sheet.cell(row=i, column=2).value
        #     self._vardef[key] = val


        self._sheets = {
            'VarDef'        : VarDefSheet(self, 'VarDef'),
            'PDomain'       : PDomainSheet(self, 'PDomain'),
            'PStrategy'     : PStrategySheet(self, 'PStrategy'),     
            'PMode'         : PModeSheet(self, 'PMode'),
        }

        self._vardefdata = self._sheets['VarDef'].get_vardef_value(valdef_sheet)
        self._vardefpcell = self._sheets['VarDef'].get_table_contxt(valdef_sheet,'VarDef')

        # # read and convert excel to json data
        for sheetname,sheet in self._sheets.items():
            sheet.read_data()
        # if kwd == 'json':
        #     for sheetname,sheet in self._sheets.items():
        #         json_file = dirname(dirname(dg_file)) + '/json' + f'/{sheetname.lower()}.json'
        #         sheet.read_data()
        #         sheet.dump_json(json_file)
            

    def read_vfile(self, vfile,kwd=''):
        self._upfdir = abspath(dirname(dirname(vfile)))
        self._vfile_data = self._inputs.read_vfile(vfile)
        self._mdname = self._vfile_data['module_name']
    #    print(self._vfile_data)
        if kwd == 'json':
            self._data = self._vfile_data
            json_file = dirname(dirname(vfile)) + '/json' + f'/pvlog.json'
            self.write_json(json_file)

    def read_pmfile(self, pmfile,kwd=''):
        self._pmfile_data = self._inputs.read_pmfile(pmfile)
        #print(self._pmfile_data)
        if kwd == 'json':
            self._data = self._pmfile_data
            json_file = dirname(dirname(pmfile)) + '/json' + f'/pmcell.json'
            self.write_json(json_file)
        
    def read_objfile(self, objfile,kwd=''):
        self._objfile_data, self._objfile_list = self._inputs.read_objfile(objfile)
        #print(self._objfile_data)
        #print(self._objfile_list)
        if kwd == 'json':
            self._data = self._objfile_data
            json_file = dirname(dirname(objfile)) + '/json' + f'/pobj.json'
            self.write_json(json_file)
      

    def read_data(self):
        for sht in self._sheets.values():
            sht.read_data()       

    def update_dg(self):
        for sht in self._sheets.values():
            sht.update_sheet()

    def check_dg(self):
        for sht in self._sheets.values():
            sht.check_sheet()

    def change_dg(self,dgfile):
        for sht in self._sheets.values():
            sht.change_sheet(dgfile)


# upfdg is XupfDesignGuide object
class BaseSheet(object):
    def __init__(self, upfdg, sheetname):
        self._upfdg = upfdg
        self._sheetname = sheetname
        self._data = []
        self._vardef = {}
        self._pdnmdict = {}
    
    def get_sheet(self):
        return self._upfdg._wb[self._sheetname]

    def read_data(self):
        raise NotImplementedError(self.__class__.__name__ + ' raad_data not implemented yet')

    def write_json(self, filepath):
        os.makedirs(dirname(filepath), exist_ok=True)
        jsonstr = json.dumps(self._data, indent=4)
        with open(filepath,'w') as fw:
            print(jsonstr, file=fw)

    def find_sheet(self, sheet, skw):
        start_rowg = 1
        # TABCONST = ['PMVAR','PMCELL','PMHIER','PMDOMAIN','PMNETWORK','PMBOUNDARY','PMISO','PMLS','PMRET','PMPSW','PMRPT','PMMODE']
        TABCONST = ['PMVAR','PMCELL','PMDOMAIN','PMNETWORK','PMBOUNDARY','PMISO','PMLS','PMRET','PMPSW','PMMODE']
        for i in range(1,sheet.max_row+1):
            if skw in TABCONST and sheet.cell(i,1).value == skw:
                start_rowg = i + 1
                break  
        return  start_rowg 

    def get_vardef_value(self, sheet):

        start_rowg = self.find_sheet(sheet, 'PMVAR')
        # end_rowg = self.find_sheet(sheet, 'PMHIER')
        end_rowg = self.find_sheet(sheet, 'PMCELL')
        for i in range(start_rowg + 1, end_rowg-1):
            key = sheet.cell(row=i, column=1).value
            val = sheet.cell(row=i, column=2).value
            self._vardef[key] = val

        self._vardef['UPF_DIR'] = self._upfdg._upfdir + '/' if not self._upfdg._upfdir.endswith('/') else self._upfdg._upfdir
        self._vardef['COM_DIR'] = self._upfdg._upfdir + '/' if not self._upfdg._upfdir.endswith('/') else self._upfdg._upfdir
        self._vardef['UPF_VERSION'] = '2.1'
        self._vardef['BOUNDARY_MODE'] = 'lower'
        # self._vardef['HD_MOD_NAME'] = self._upfdg._mdname
        # self._vardef['HD_PROCESS'] = ''
        self._vardef['SS_MODE'] = 'full'
        self._vardef['SCOPE_TYPE'] = 'parent; # self'

        return self._vardef   

    def set_name_style(self, kw):
        #time_stamp = time.strftime("%Y%m%d%H%M%S", time.localtime())
        #CONST = f'Generic_Xupf_{time_stamp}'
        CONST = f'Generic_XUPF'
        return kw + '_' + CONST

    # def change_space(self, dgfile):

    #     #sheet = self.get_sheet()
    #     shtname = self._sheetname
    #     sheet = self._upfdg._wb[shtname]

    #     # Find variable start row num, below "Variable" header 
    #     start =0       
    #     df = pd.read_excel(dgfile, sheet_name=shtname, engine='openpyxl') 
    #     df.loc[len(df)] = list(df.columns)
    #     for col in df.columns:
    #         index = list(df.columns).index(col)
    #         letter = get_column_letter(index + 1)
    #         collen = df[col].apply(lambda x: len(self.max_str(str(x)).encode())).max()
    #         # sheet.column dimensions[letter).width = collen*o.9 
    #         sheet.column_dimensions[letter].width = collen * 1.05

    def max_str(self, li):
        max = 0
        max_str = ''
        try:
            for i in li.split('\n'):
                if len(i) > max:
                    max =len(i)
                    max_str=i 
            return max_str 
        except:
            return li
        


    # showErrorMessage=False,showDropDown=True
    def add_dropdown(self, sheet, options, start, end):       
        dv = DataValidation(type="list", formula1=options, showErrorMessage=False)
        sheet.add_data_validation(dv)
        if len(start) == 2 and len(end) == 2:
            for i in range(start[0], end[0] + 1):
                for j in range(start[1], end[1] + 1):
                    dv.add(sheet.cell(i,j))
        if len(start) == 1 and len(end) == 1:
            dv.add(sheet.cell(start[0],end[0]))


    def cell_style1(self, sheet, start, end):
        border=Border(left=Side(border_style='thin', color='000000'),
                      right=Side(border_style='thin', color='000000'),
                      top=Side(border_style='thin', color='000000'),
                      bottom=Side(border_style='thin', color='000000'))
        #bgfill = PatternFill(fill_type='solid', start_color='fff2cc', end_color='fff2cc') 
        #bgfill = PatternFill(fill_type = 'solid', start_color='197e00',end_color='197e00')
        bgfill = PatternFill(fill_type = 'solid', start_color='FF385724',end_color='FF333300')
        font = Font(name='等线', size=11, color='FFFFFF')
        for i in range(start[0], end[0] + 1):
            for j in range(start[1], end[1] + 1):
                sheet.cell(i,j).border=border 
                sheet.cell(i,j).fill=bgfill
                sheet.cell(i,j).font=font
                sheet.cell(i,j).alignment = Alignment(horizontal='left', vertical='center',wrapText=True) 
                sheet.cell(i,j).alignment = Alignment(horizontal='left', vertical='center',wrapText=True) 
                sheet.cell(i,j).alignment=Alignment(horizontal='left', vertical='center') 

    def cell_style2(self, sheet, start, end):
        border=Border(left=Side(border_style='thin', color='000000'),
                      right=Side(border_style='thin', color='000000'),
                      top=Side(border_style='thin', color='000000'),
                      bottom=Side(border_style='thin', color='000000'))
        #bgfill = PatternFill(fill_type='solid', start_color='fff2cc', end_color='fff2cc') 
        #bgfill = PatternFill(fill_type = 'solid', start_color='197e00',end_color='197e00')
        bgfill = PatternFill(fill_type = 'solid', start_color='FFFFFF',end_color='FFFFFF')
        #font = Font(name='等线', size=11, color='FFFFFF')
        for i in range(start[0], end[0] + 1):
            for j in range(start[1], end[1] + 1):
                sheet.cell(i,j).border=border 
                sheet.cell(i,j).fill=bgfill
                #sheet.cell(i,j).font=font
                sheet.cell(i,j).alignment = Alignment(horizontal='left', vertical='center',wrapText=True) 
                sheet.cell(i,j).alignment = Alignment(horizontal='left', vertical='center',wrapText=True) 
                sheet.cell(i,j).alignment=Alignment(horizontal='left', vertical='center',wrapText=True)

    def get_supply_infos(self):

        delkeys = ['module_name', 'ISO_CTRL', 'RET_SAVE', 'RET_RES', 'PSO_CTRL', 'PSO_ACK']
        supply_datag = self._upfdg._vfile_data
        supply_data = {}

        for ky,vl in supply_datag.items():
            if ky not in delkeys:
                supply_data[ky] = vl

        # print('supply_data: ', supply_data)
        supply_kw = []
        supply_val = []
        supply_vss = []
        supply_tmp = ''
        for key,val in supply_data.items():
            if '0v' in val or '0.0v' in val:
                supply_vss.append(key)
                upf_info(f'Ground pin is {key}')
            elif 'PSO' in val:
                supply_kw.append(key)
                for i in range(1, int(val[-1]) + 1):
                    supply_kw.append(key + f'_PSW{i}')
                supply_tmp = supply_tmp + ' ' + supply_data[key].split('PSO')[0].strip()
            else:
                supply_kw.append(key)
                supply_tmp = supply_tmp + ' ' + supply_data[key].strip()

        #print(supply_tmp.strip().split(','))
        float_list = [float(x.strip('v')) for x in supply_tmp.strip().split()]
        unique_floats = set(float_list)
        sorted_floats = sorted(unique_floats, reverse=True)
        supply_val = [str(x) + 'v' for x in sorted_floats]
        supply_val.append('off')
        supply_val.append('0v')

        return supply_kw,supply_val,supply_vss,supply_data
    

    def get_ctl_sig(self, ctsig):
        ctrl = []
        for ct in ctsig:
            if re.search(r'\[\d+:\d+\]', ct):
                sig = ct.split('[')[0].strip()
                st = int(ct.split(':')[0].strip()[-1])
                ed = int(ct.split(':')[1].strip()[0])
                for i in range(ed,st+1):
                    ctrl.append(sig + '[' + str(i) + ']')
            else:
                ctrl.append(ct)
        return ctrl

    def get_table_loc(self,sheet, shnm='') -> dict:

        #TABCONSTT = ['PMVAR','PMCELL','PMHIER','PMDOMAIN','PMNETWORK','PMBOUNDARY','PMISO','PMLS','PMPSW','PMRET','PMRPT','PMMODE']

        if not shnm:
            sheetname = self._sheetname
        else:
            sheetname = shnm

        if sheetname == 'VarDef':
            TABCONST = ['PMVAR','PMCELL']
        if sheetname == 'PDomain':
            TABCONST = ['PMDOMAIN','PMNETWORK','PMBOUNDARY']
        if sheetname == 'PStrategy':
            TABCONST = ['PMISO','PMLS','PMPSW','PMRET']
        if sheetname == 'PMode':
            TABCONST = ['PMMODE']

        # row_start max_col              
        row_start = ''
        max_row = ''
        max_col = ''
        # row_start max_col max_row
        table_row_loc = {}
        for kw in TABCONST:
            strow = self.find_sheet(sheet, kw)
            row_start = str(strow)
            for i in range(1,sheet.max_column + 1):
                if sheet.cell(strow,i).value == 'Comment':
                    max_col = str(i)
                    break

            if kw in ['PMCELL','PMBOUNDARY','PMRET','PMMODE']:
                #table_row_loc[kw] = row_start[kw] + ' ' + str(int(row_start[kw].split()[0]) + 20)
                table_row_loc[kw] = row_start + ' ' + str(sheet.max_row + 2) + ' ' + max_col
            else:
                idx = TABCONST.index(kw) + 1
                # print('dfsg: ', TABCONST,idx)
                max_row = self.find_sheet(sheet,TABCONST[idx]) - 1
                #table_row_loc[kw] = row_start[kw] + ' ' + str(int(row_start[TABCONST[idx]].split()[0]) - 2)
                table_row_loc[kw] = row_start + ' ' + str(max_row) + ' ' + max_col

        return table_row_loc
  
    def get_table_contxt(self,sheet, shnm='', tabnm=[]) -> dict:
        # row_start max_col max_row
        tab_loc = self.get_table_loc(sheet,shnm)
        print('sheet: tab_loc: ', tab_loc)

        if not shnm:
            sheetname = self._sheetname
        else:
            sheetname = shnm

        #TABCONST = ['PMVAR','PMCELL','PMHIER','PMDOMAIN','PMNETWORK','PMBOUNDARY','PMISO','PMLS','PMRET','PMPSW','PMRPT','PMMODE']
        if sheetname == 'VarDef':
            # TABCONST = ['PMVAR','PMCELL','PMHIER']
            if not tabnm:
                TABCONST = ['PMVAR','PMCELL']
            else:
                TABCONST = tabnm
        if sheetname == 'PDomain':
            if not tabnm:
                TABCONST = ['PMDOMAIN','PMNETWORK','PMBOUNDARY']
            else:
                TABCONST = tabnm
        if sheetname == 'PStrategy':
            # TABCONST = ['PMISO','PMLS','PMRET','PMPSW','PMRPT']
            if not tabnm:
                TABCONST = ['PMISO','PMLS','PMRET','PMPSW']
            else:
                TABCONST = tabnm
        if sheetname == 'PMode':
            if not tabnm:
                TABCONST = ['PMMODE']
            else:
                TABCONST = tabnm

        table_contxt = {}
        #row_contxt = {}
        if TABCONST:
            for kw in TABCONST:
                start_row = int(tab_loc[kw].split(' ')[0])
                end_row = int(tab_loc[kw].split(' ')[1])
                end_col = int(tab_loc[kw].split(' ')[2])
                if kw == 'XPMVAR':
                    for i in range(start_row,end_row+1):
                        key = sheet.cell(i+1,1).value
                        val = str(sheet.cell(i+1,2).value)
                        if key:
                            table_contxt[key] = val.strip()
                        # print('PMVARdfd: ', table_contxt)
                        # if key and val:
                        #     table_contxt[key] = val
                else:
                    table_contxt.update(self.get_row_txt(sheet,kw,start_row,end_row,end_col))

        return table_contxt

    def get_row_txt(self, sheet, kw, start_row, end_row, end_col):
        row_contxt = {}
        table_contxt = {}
        for i in range(1, end_row - start_row):
            for j in range(1, end_col + 1):
                key = sheet.cell(start_row, j).value
                val = sheet.cell(start_row + i, j).value
                val_col1 = sheet.cell(start_row + i, 1).value
                if val_col1:
                    if re.search(r'^#', val_col1.strip()):
                        continue
                if key:     key = str(key).strip()
                if val:     val = str(val).strip()
                row_contxt[key] = val
                # if key and val:
                #     row_contxt[key] = val

            all_none = all(ele is None for ele in list(row_contxt.values()))
            if not all_none and row_contxt:
                table_contxt[f'{kw}_Row{start_row + i}'] = row_contxt
            row_contxt = {}
            # for key in table_contxt.keys():
            #     if 'TMCLK' in key:
            #         print(table_contxt)

        return table_contxt


    def get_impl_obj(self, sheet, start_rowg, kwd):
        keywds = list(self._upfdg._objfile_data.keys())
        TABCONST = ['PMVAR','PMCELL','PMHIER','PMDOMAIN','PMNETWORK','PMBOUNDARY','PMISO','PMLS','PMPSW','PMRET','PMRPT','PMMODE']        
        
        supply_kw, supply_vol,supply_vss, supply_data = self.get_supply_infos()
        supply_kw = supply_kw.extend(supply_vss)
        notvss = [x for x in supply_vss if not x in ['VSS']]

        #print(keywds)

        if kwd == 'PMNETWORK':
            # SupplyPortNet	NPwellNet	InstList	MapSupplyList	Comment
            row_tmp = start_rowg
            row_tot = start_rowg + 10
            if len(supply_kw) > 8:
                sheet.insert_rows(start_rowg + 3, 8)
                row_tot += 8
                #self.cell_style2(sheet,[start_rowg + 3,1], [start_rowg + 12,5])
            if len(keywds) > 4:
                sheet.insert_rows(start_rowg + 11, 6)
                row_tot += 6
            self.cell_style2(sheet,[start_rowg + 1,1], [row_tot + 1,5])

            virpwr = [x for x in supply_kw if re.search(r'_PSW\d+',x)]
            relpwr = [x for x in supply_kw if not re.search(r'_PSW\d+',x)]
            cmt = ['PAL','PAL OUT','OUT','SNE1','SNE2','SNE3','SNE4','SNE5']
            if len(notvss) > 0:
                ncmt = notvss.extend(cmt)
            else:
                ncmt = cmt
            #print(cmt)
            self.add_dropdown(sheet, '"' + ','.join(supply_kw) + '"', [start_rowg + 1, 1], [row_tot + 1 , 1])
            self.add_dropdown(sheet, '"' + ','.join(relpwr) + '"', [start_rowg + 1, 2], [row_tot + 1 , 2])
            self.add_dropdown(sheet, '"' + ','.join(ncmt) + '"', [start_rowg + 1, 5], [row_tot + 1, 5])
            
            for i in range(1, len(supply_kw) + 1):               
                sheet.cell(start_rowg + i,1).value = supply_kw[i-1]
                if re.search(r'_PSW\d+',supply_kw[i-1]):
                    nwell = supply_kw[i-1].split('_PSW')[0].strip()
                    sheet.cell(start_rowg + i,2).value = nwell
                else:
                    sheet.cell(start_rowg + i,2).value = supply_kw[i-1]
                row_tmp += 1
            #print(row_tmp)

            for ky in keywds:
                #row_tmp += 1
                if '_conspy_insts' in ky:
                    row_tmp += 1
                    nky = ky.split('_conspy_insts')[0].strip()
                    sheet.cell(row_tmp,1).value = self._upfdg._objfile_data[nky + '_outer_spy']
                    sheet.cell(row_tmp,3).value = nky + '_conspy_insts' #self._upfdg._objfile_data[nky + '_conspy_insts']
                    #self.add_dropdown(sheet,'"' + ky + '"',)
                    sheet.cell(row_tmp,4).value = self._upfdg._objfile_data[nky + '_inner_spy']
                if '_conspy_hinsts' in ky:
                    row_tmp += 1
                    nky = ky.split('_conspy_hinsts')[0].strip()
                    sheet.cell(row_tmp,1).value = self._upfdg._objfile_data[nky + '_outer_spy']
                    #sheet.cell(row_tmp,3).value = self._upfdg._objfile_data[nky + '_conspy_hinsts']
                    sheet.cell(row_tmp,3).value = nky + '_conspy_hinsts'
                    sheet.cell(row_tmp,4).value = self._upfdg._objfile_data[nky + '_inner_spy']
            
            row_tmp = 0

        if kwd == 'PMBOUNDARY':
            bd_ele = []
            bd_exd = []
            flg_ele = 0
            flg_exd = 0
            for ky in keywds:
                if re.search(r'_spa_inport|_spa_outport|_spa_inhpin|_spa_outhpin', ky):
                    bd_ele.append(ky)
                    flg_ele = 1
                if re.search(r'_exdspa_inport|_exdspa_outport|_exdspa_inhpin|_exdspa_outhpin', ky):
                    bd_exd.append(ky)
                    flg_exd = 1
                if re.search(r'_spa_in$|_spa_out$|_spa_input|_spa_output|_spa_inpin|_spa_outpin', ky):
                    bd_ele.append(ky)
                    flg_ele = 1
                if re.search(r'_exdspa_in$|_exdspa_out$|_exdspa_input|_exdspa_output|_exdspa_inpin|_exdspa_outpin', ky):
                    bd_exd.append(ky)
                    flg_exd = 1                    
            if not flg_ele:     bd_ele = None
            if not flg_exd:     bd_exd = None

            return bd_ele,bd_exd

        if kwd == 'PMISO':
            iso_ele = []
            iso_exd = []
            iso_no = []
            flg_ele = 0
            flg_exd = 0
            flg_no = 0
            for ky in keywds:
                if re.search(r'_iso_inport|_iso_outport|_iso_inhpin|_iso_outhpin|_ctliso_inport|_ctliso_inhpin|_fdthiso_inport|_fdthiso_outport|_fdthiso_inhpin|_fdthiso_outhpin', ky):
                    iso_ele.append(ky)
                    flg_ele = 1
                if re.search(r'_exdiso_inport|_exdiso_outport|_exdiso_inhpin|_exdiso_outhpin', ky):
                    iso_exd.append(ky)
                    flg_exd = 1
                if re.search(r'_noiso_inport|_noiso_outport|_noiso_inhpin|_noiso_outhpin', ky):
                    iso_no.append(ky)
                    flg_no = 1

                if re.search(r'_iso_in$|_iso_out$|_ctliso_in$|_fdthiso_in$|_fdthiso_out$', ky):
                    iso_ele.append(ky)
                    flg_ele = 1
                if re.search(r'_exdiso_in$|_exdiso_out$', ky):
                    iso_exd.append(ky)
                    flg_exd = 1
                if re.search(r'_noiso_in$|_noiso_out$', ky):
                    iso_no.append(ky)
                    flg_no = 1

                if re.search(r'_iso_input|_iso_output|_ctliso_input|_fdthiso_input|_fdthiso_output', ky):
                    iso_ele.append(ky)
                    flg_ele = 1
                if re.search(r'_iso_inpin|_iso_outpin|_ctliso_inpin|_fdthiso_inpin|_fdthiso_outpin', ky):
                    iso_ele.append(ky)
                    flg_ele = 1                    
                if re.search(r'_exdiso_input|_exdiso_output|_exdiso_outpin', ky):
                    iso_exd.append(ky)
                    flg_exd = 1
                if re.search(r'_noiso_input|_noiso_output|_noiso_outpin', ky):
                    iso_no.append(ky)
                    flg_no = 1

            if not flg_ele:     iso_ele = None
            if not flg_exd:     iso_exd = None
            if not flg_no:      iso_no = None

            return iso_ele,iso_exd,iso_no
                

        if kwd == 'PMLS':
            ls_ele = []
            ls_exd = []
            ls_no = []
            flg_ele = 0
            flg_exd = 0
            flg_no = 0                
            for ky in keywds:
                if re.search(r'_ls_inport|_ls_outport|_ls_inhpin|_ls_outhpin|_fdthls_inport|_fdthls_outport|_fdthls_inhpin|_fdthls_outhpin', ky):
                    ls_ele.append(ky)
                    flg_ele = 1
                if re.search(r'_exdls_inport|_exdls_outport|_exdls_inhpin|_exdls_outhpin', ky):
                    ls_exd.append(ky)
                    flg_exd = 1
                if re.search(r'_nols_inport|_nols_outport|_nols_inhpin|_nols_outhpin', ky):
                    ls_no.append(ky) 
                    flg_no = 1

                if re.search(r'_ls_in$|_ls_out$|_fdthls_in$|_fdthls_out$|_ls_input|_ls_output|_fdthls_input|_fdthls_output|_ls_inpin|_ls_outpin|_fdthls_inpin|_fdthls_outpin', ky):
                    ls_ele.append(ky)
                    flg_ele = 1
                if re.search(r'_exdls_in$|_exdls_out$|_exdls_input|_exdls_output|_exdls_inpin|_exdls_outpin', ky):
                    ls_exd.append(ky)
                    flg_exd = 1
                if re.search(r'_nols_in$|_nols_out$|_nols_input|_nols_output|_nols_inpin|_nols_outpin', ky):
                    ls_no.append(ky) 
                    flg_no = 1                     
            if not flg_ele:     iso_ele = None
            if not flg_exd:     iso_exd = None
            if not flg_no:      iso_no = None

            return ls_ele,ls_exd,ls_no
        
        if kwd == 'PMPSW':
            psw_ele = []
            psw_exd = []
            flg_ele = 0
            flg_exd = 0
            for ky in keywds:
                if re.search(r'_ctlpsw_inport|_ctlpsw_inhpin', ky):
                    psw_ele.append(ky)
                    flg_ele = 1
                if re.search(r'_ackpsw_inport|_ackpsw_inhpin', ky):
                    psw_exd.append(ky)
                    flg_exd = 1

                if re.search(r'_ctlpsw_in$|_ctlpsw_input|_ctlpsw_inpin', ky):
                    psw_ele.append(ky)
                    flg_ele = 1
                if re.search(r'_ackpsw_in$|_ackpsw_input|_ackpsw_inpin', ky):
                    psw_exd.append(ky)
                    flg_exd = 1
            if not flg_ele:     psw_ele = None
            if not flg_exd:     psw_exd = None    

            return psw_ele,psw_exd

        if kwd == 'PMRET':
            ret_save = []
            ret_res = []
            ret_ele = []
            ret_exd = []
            ret_no = []
            flg_save = 0
            flg_res = 0
            flg_ele = 0
            flg_exd = 0
            flg_no = 0                
            for ky in keywds:
                if re.search(r'_saveret_inport|_saveret_inhpin', ky):
                    ret_save.append(ky)
                    flg_save = 1
                if re.search(r'_resret_inport|_resret_inhpin', ky):
                    ret_res.append(ky)
                    flg_res = 1
                if re.search(r'_ret_insts|_ret_hinsts', ky):
                    ret_ele.append(ky)
                    flg_ele = 1
                if re.search(r'_exdret_insts|_exdret_hinsts', ky):
                    ret_exd.append(ky)
                    flg_exd = 1
                if re.search(r'_noret_insts|_noret_hinsts', ky):
                    ret_no.append(ky)
                    flg_no = 1

                if re.search(r'_saveret_in$|_saveret_input|_saveret_inpin', ky):
                    ret_save.append(ky)
                    flg_save = 1
                if re.search(r'_resret_in$|_resret_input|_resret_inpin', ky):
                    ret_res.append(ky)
                    flg_res = 1


            if not flg_save:    ret_save = None
            if not flg_res:     ret_res = None 
            if not flg_ele:     ret_ele = None
            if not flg_exd:     ret_exd = None
            if not flg_no:      ret_no = None

            return ret_save,ret_res,ret_ele,ret_exd,ret_no

        if kwd == 'PMRPT':
            rpt_ele = []
            rpt_exd = []
            flg_ele = 0
            flg_exd = 0
            for ky in keywds:
                if re.search(r'_rpt_inport|_rpt_outport|_rpt_inhpin|_rpt_outhpin', ky):
                    rpt_ele.append(ky)
                    flg_ele = 1
                if re.search(r'_exdrpt_inport|_exdrpt_outport|_exdrpt_inhpin|_exdrpt_outhpin', ky):
                    rpt_exd.append(ky)
                    flg_exd = 1

            for ky in keywds:
                if re.search(r'_rpt_in$|_rpt_out$|_rpt_input|_rpt_output|_rpt_inpin|_rpt_outpin', ky):
                    rpt_ele.append(ky)
                    flg_ele = 1
                if re.search(r'_exdrpt_in$|_exdrpt_out$|_exdrpt_input|_exdrpt_output|_exdrpt_inpin|_exdrpt_outpin', ky):
                    rpt_exd.append(ky)
                    flg_exd = 1

            if not flg_ele:     rpt_ele = None
            if not flg_exd:     rpt_exd = None     

            return rpt_ele,rpt_exd
        
    def save_text(self, context,file):
        with open(file, 'w') as fw:
            fw.write(context)

    def get_rows(self,pmdata,keyrow,kwd,ckwd):
        pmdict = {}
        pmkeys = []
        pmlist = []
        tab = keyrow.split('_')[0]
        # print('pmdata: ', pmdata)

        # pmlist = [(key, val) for key, val in pmdata.items() if keyrow in key and not re.search(r'^#',val[f'{ckwd}'].strip()) and val[f'{kwd}']]
        for key, val in pmdata.items():
            if keyrow in key and not re.search(r'^#', val[f'{ckwd}'].strip()):
                if '|' in kwd:
                    kwlst = kwd.split('|')
                    kflg = False
                    for kl in kwlst:
                        if val[f'{kl}']: kflg = True
                    if kflg:
                        pmlist.append((key, val))
                else:
                    if val[f'{kwd}']:
                        pmlist.append((key, val))

        if pmlist:
            for k,v in pmlist:
                pmdict[k] = v
                pmkeys.append(k)
            pmkeys.sort(key=lambda x: int(x.split("Row")[1]))
        else:
            print(f'The table {tab} is empty.')

        # print(f'real tabale {tab}', pmkeys,pmdict)

        return pmdict, pmkeys     

    def get_pdname(self,blkalias,sheet):
        i = 0
        pddic = {}

        # print('get_pdname--sheet: ', sheet)
        contxt = self.get_table_contxt(sheet,'PDomain',['PMDOMAIN'])
        pdkeys = list(contxt.keys())
        # print('pdfsmcxc: ', contxt)
        pdict, prows = self.get_rows(contxt,'PMDOMAIN_Row','PDName','PDName')
        # print('pdict: prows', pdict, prows)
        if prows:
            for prow in prows:
                rowdict = pdict[prow]
                pdname = rowdict['PDName']
                i += 1

                prmdic = [(key,val) for key,val in rowdict.items() if val == 'PRM']
                if len(prmdic) > 1:
                    upf_error(f'{prow} has two or more primary power. Please check it.')
                    prm = 'NOPRM'
                else:
                    prm = prmdic[0][0]

                nprm = ''.join(prm.split('_'))
                self._pdnmdict[f'{pdname}'] = f'PD{i}_${{{blkalias}}}_{nprm}_{pdname}'
        else:
            upf_error(f'Can not find power domain definition. Please check it.')

        return self._pdnmdict


class BaseInputs(object):
    def __init__(self):
        self.vfile_data = {}
        self.pmfile_data = {}
        self.objfile_data = {}
               
    def read_vfile(self, vfile) -> dict:

        lines = self.read_text(vfile)
        for line in lines:
            line = line.replace('\n','').replace('\r','').replace('\t',' ').strip()
            if re.search(r'^\/\/', line):
                continue
            if re.search(r'^module', line):
                self.vfile_data['module_name']= re.split(' +',line)[1].strip().replace('(','')
            if re.search(r'\/\/#(\d+[vV]#|\d+.\d+[vV]#|\d+.\d+[vV]:)', line):
                #print(line)
                port_pwrg = re.split('//#',line)[0].strip().replace(',', '')
                port_pwrg = [x for x in re.split(' ', port_pwrg) if x]
                # port_pwr = re.split(' ', port_pwrg)[1].strip()
                port_pwr = port_pwrg[1]
                port_volx = re.split('//#',line)[1].strip() # maybe include 'PSO'
                if re.search(r'#PSO \d+#', port_volx):                    
                    #port_volg = ' PSO' + re.split(r'#PSO ',port_volx)[1].replace('#','')
                    port_volg = ' ' + re.split(r'#',port_volx)[1].strip().replace(' ','')
                    port_volx = re.split(r'#PSO ', port_volx)[0].strip()
                    #print(port_volg)
                else:
                    port_volg = ''

                if re.search(r':', port_volx):
                    port_vol = ' '.join(re.split(':',port_volx)).strip().lower().replace('#', '') + port_volg
                else:
                    port_vol = port_volx.lower().replace('#', '') + port_volg
                self.vfile_data[port_pwr] = port_vol
       
        self.vfile_data['ISO_CTRL'] = self.get_ctrl_ports(lines, 'ISO_CTRL')
        self.vfile_data['RET_SAVE'] = self.get_ctrl_ports(lines, 'RET_SAVE')
        self.vfile_data['RET_RES'] = self.get_ctrl_ports(lines, 'RET_RES')
        self.vfile_data['PSO_CTRL'] = self.get_ctrl_ports(lines, 'PSO_CTRL')
        self.vfile_data['PSO_ACK'] = self.get_ctrl_ports(lines, 'PSO_ACK')

        return self.vfile_data



    def read_pmfile(self, pmfile) -> dict:

        pm_data = self.read_yaml(pmfile)

        # ISOROW1, ELSROW1,
        isocells = self.get_pmcell_info(pm_data, 'ISOPowerMCell')
        lscells = self.get_pmcell_info(pm_data, 'LSPowerMCell')
        elscells = self.get_pmcell_info(pm_data, 'ELSPowerMCell')
        retcells = self.get_pmcell_info(pm_data, 'RETPowerMCell')
        pswcells = self.get_pmcell_info(pm_data, 'PSWPowerMCell')

        isocells.update(lscells)
        isocells.update(elscells)
        isocells.update(retcells)
        isocells.update(pswcells)
        self.pmfile_data = isocells

        return self.pmfile_data



    def read_objfile(self, objfile) -> dict:
        
        obj_lines = self.read_text(objfile)

        objfile_data = {}
        var_list = []
        var_nline = ''

        for line in obj_lines:
            if re.search(r'^#', line):
                var_nline += f'\n{line}'
                #var_list += line
            elif re.search(r'^set\s+', line):
                var_list.append(re.split(' +', line)[1].strip())
                if re.search(r'[|]',line):
                    nline = line.replace('[','\\[')
                    nline = nline.replace(']','\\]')
                else:
                    nline = line
                var_nline += f'\n{nline}'
            elif re.search(r'\w+[|\d+]',line):
                nline = line.replace('[','\\[')
                nline = nline.replace(']','\\]')
                var_nline += f'\n{nline}'              
            else:
                var_nline += f'\n{line}'
        #var_list = [re.split(' +', line)[1].strip() for line in obj_lines if re.search(r'^set\s+', line)]
        #print(f'objfile var list {var_list}')

        # tmpfile = dirname(objfile) + '/objtmp.tcl'
        tmpfile = os.path.join(os.path.dirname(objfile),'objtmp.tcl')
        # 确保目录存在
        os.makedirs(os.path.dirname(tmpfile), exist_ok=True)
        with open(tmpfile, 'w') as fw:
            fw.write(var_nline)        
        #print(var_nline)

        tcl_intp = tk.Tcl()
        # tcl_intp.eval(f'source {tmpfile}')
        #tcl_intp.eval(f'source {objfile}')
        #tcl_intp.eval(' '.join(var_nline))

        # 使用TCL的文件路径语法来source文件，确保跨平台兼容
        # 将路径转换为TCL格式
        tcl_path = tmpfile.replace(os.sep, '/')
        tcl_intp.eval(f'source [file normalize {{{tcl_path}}}]')
        
        for var in var_list:
            nvar = tcl_intp.getvar(var).strip()
            if re.search(r'\[|\]',nvar):
                nvar = nvar.replace('\\[','[')
                nvar = nvar.replace('\\]',']')
            self.objfile_data[var] = nvar

        if os.path.exists(tmpfile):
            os.system(f'rm -f {tmpfile}')

        return self.objfile_data,var_list
        


    def read_yaml(self, yaml_file):

        yaml_data = {}
        if not os.path.exists(yaml_file):
            raise FileExistsError(f'{yaml_file} does not exists')
        with open(yaml_file, 'r') as fh:
            yaml_data = yaml.load(fh, yaml.FullLoader)

        return yaml_data
        
    def get_pmcell_info(self, pmdata, kwd) -> dict:

        pm_num = len(pmdata[kwd].keys())
        kws = kwd.replace('PowerMCell', '')

        pm_data = {}
        if pm_num == 0:
            upf_warn(f'{kws} power cell not found.')
        elif pm_num == 1:
            kws = kws + 'Row1'
            pm_data[kws] = pmdata[kwd]['TypeIndex1']
        else:
            for num in range(1,pm_num + 1):
                kwn = kws + f'Row{num}'
                pm_data[kwn] = pmdata[kwd][f'TypeIndex{num}']

        return pm_data
    


    def read_text(self, file):
        if not os.path.exists(file):
            raise FileExistsError(f'{file} does not exists')
            # upf_error(f'{file} not exist. Please check it.')
            # exit(1)
        else:
            txt_list = []
            with open(file,'r') as fh:
                for line in fh.readlines():
                    #pattern = r"\[\s*(\d+)\s*:\s*(\d+)\]"
                    if line.strip() == "":
                        continue
                    # if line.strip().startswith("//"):
                    #     continue                     
                    line = re.sub(r"\[\s*(\d+)\s*:\s*(\d+)\s*\]", r"[\1:\2]", line)
                    txt_list.append(line.strip())
        
            return txt_list

    def get_ctrl_ports(self, lnlist, kwd) -> list:

        full_list = []
        stwp = ''
        stw = []

        nkw = r'//#' + f'{kwd}' + '#'
        #print(nkw)
        for line in lnlist:
            if re.search(r'^\/\/', line):
                continue
            if re.search(f'{nkw}', line):
                # kw_loc = re.split(' +',line).index(r'//#ISO_CTRL#')
                # iso_ctrl.append(re.split(' +',line)[kw_loc - 1].strip().replace(',', ' ').strip())
                #stw = re.split('//#ISO_CTRL',line)[-2].strip()
                stwp = re.split(f'{nkw}',line)[-2].strip()
                if re.search(r'\/\/$', stwp):
                    stwp.replace('//','')
                elif re.search(r'\/\/\w*\s*', stwp):
                    stwp = re.split('//', stwp)[-2].strip()

                #patn = r'\[(\d+:\d+)|(\d+\s+:\d+)|(\d+:\s+\d+)|(\s+\d+:\d+)\]'    
                if re.search(r'\[\d+:\d+\]', stwp):
                    portnum = ''.join(re.findall(r'\[\d+:\d+\]', stwp)).strip()
                    stwp = re.split(r'\[\d+:\d+\]', stwp)[-1].strip()
                    #stw = re.split('wire|logic|byte|bit|reg', stw)[-1].strip()
                    #stw = re.split('input|output', stw)[-1].strip()                 
                    stw = re.split(',', stwp)[:-1] 
                    stw = [st + portnum for st in stw]
                elif re.search(r'wire|logic|byte|bit|reg', stwp):
                    stwp = re.split('wire|logic|byte|bit|reg', stwp)[-1].strip()
                    stw = re.split(',', stwp)[:-1]
                elif re.search(r'input|output', stwp):
                    stwp = re.split('input|output', stwp)[-1].strip()
                    stw = re.split(',', stwp)[:-1].strip()

                #nested_list.append(stw)
                full_list.extend(stw)

        #return list(chain(*nested_list))
        return full_list


    # def find_start_cell_location(self,sheet,var):
    #     start = 0
    #     for i in range(1, sheet.max_row+1):
    #         if sheet.cell(i,1).value == var:
    #             start = i + 1
    #             break
    #     for i in range(start, sheet.max_row+1):
    #         key = sheet.cell(row=i, column=1).value
    #         val = sheet.cell(row=i, column=2).value
    #         self._vardef[key] = val


class VarDefSheet(BaseSheet):
    def __init__(self,*args):
        super().__init__(*args)  
        self._vardata = {}     

    def check_sheet(self):
        sheet = self.get_sheet()
        # vardata = self.get_var_cell(sheet)
        vardata = self.get_vardef_value(sheet)
        #print('vardataxxx: ',vardata)

        # module name 
        mdname = vardata['HD_MOD_NAME']
        if mdname != self._upfdg._vfile_data['module_name']:
            upf_error(f'Module name from vardef is different from the empty vfile.','chk_sht')

        # hier variable name
        # cblk = self._hiertree.get_block_by_name(mdname)
        # lvl = cblk.hdlevel
        # hierlst = [x for x in vardata.keys() if 'HIER_' in x]
        # if hierlst:
        #     for hier in hierlst:
        #         if lvl in ('blk','sys','top'):
        #             if not f'HIER_{lvl.upper()}_' in hier and hier != 'HIER_EXPD_STYLE':
        #                 upf_error(f'Naming of variable "{hier}" does not start with "HIER_{lvl.upper()}_".','chk_sht')
        #             if not vardata[hier].endswith('/') and hier != 'HIER_EXPD_STYLE':
        #                 upf_error(f'Variable "{hier}" value does not end with "/".','chk_sht')

        # check UPF_DIR/COM_DIR/DFT_DIR with the end of '/'
        # for hir in ('UPF_DIR','COM_DIR'):
        #     if not vardata[hir].endswith('/'):
        #         upf_error(f'Variable "{hir}" value does not end with "/".','chk_sht')

    def read_data(self):
        sheet = self.get_sheet()
        self._vardata = self.get_table_contxt(sheet, 'VarDef')
        # nvaldata = {}
        # nvaldata["TMVAR_Row16"] = {
        #     "Variable": "UPF_DIR",
        #     "Value": f'{self._upfdg._upfdir}',
        #     "Comment": ''
        # }
        # nvaldata["TMVAR_Row17"] = {
        #     "Variable": "COM_DIR",
        #     "Value": f'{self._upfdg._upfdir}',
        #     "Comment": ''
        # }
        # nvaldata["TMVAR_Row18"] = {
        #     "Variable": "UPF_VERSION",
        #     "Value": '2.1',
        #     "Comment": ''
        # }
        # nvaldata["TMVAR_Row19"] = {
        #     "Variable": "HD_MOD_NAME",
        #     "Value": f'{self._upfdg._mdname}',
        #     "Comment": ''
        # }
        # nvaldata["TMVAR_Row20"] = {
        #     "Variable": "HD_PROCESS",
        #     "Value": '',
        #     "Comment": ''
        # }
        # nvaldata["TMVAR_Row21"] = {
        #     "Variable": "SCOPE_TYPE",
        #     "Value": 'parent; # self',
        #     "Comment": ''
        # }
        # self._vardata.update(nvaldata)
        # print('_vardata: ', self._vardata)

    def get_var_cell(self,sht):
        start = 0
        end = 0
        vardef_data = {}
        for i in range(1, sht.max_row+1):
            if sht.cell(i,1).value == 'Variable':
                start = i + 1
            # if sht.cell(i,1).value == 'TMHIER':
            #     end = i - 1
        for i in range(start, sht.max_row+1):
            key = sht.cell(row=i, column=1).value
            val = sht.cell(row=i, column=2).value
            if key is not None:
                vardef_data[key] = val

        tmp_vardata = {}
        print(self._vardata)
        for val in list(self._vardata.values()):
            if 'Variable' in list(val.keys()):
                tmp_vardata[val['Variable']] = val['Value']
    

        for ky in list(vardef_data.keys()):
            for k,v in tmp_vardata.items():
                if k == ky:
                    upf_info(f'{k} is already defined in vardef sheet.')
                else:
                    vardef_data[k] = v

        return vardef_data

    def change_sheet(self):
        pass

    def dump_json(self,json_file):
        self._data = self._vardata
        self.write_json(json_file)



class PDomainSheet(BaseSheet):
    def __init__(self,*args):
        super().__init__(*args)
        self._pddata = {}
        #self._pdnmdic = {}
        
    def check_sheet(self):
        '''
        # only during -dg option
        # addition of supply infos in PMDOMAIN table
        # addition of dropdown infos from pmempty and pmobj file
        '''

        sheet = self.get_sheet()

        supply_kw, supply_vol, supply_vss, supply_data = self.get_supply_infos()

        # find PMDOMAIN table 
        start_rowg = self.find_sheet(sheet, 'PMDOMAIN') 


    def read_data(self):
        sheet = self.get_sheet()
        self._pddata = self.get_table_contxt(sheet,'PDomain')
        # print('_pddata: ', self._pddata)

    def change_sheet(self):
        pass

    def dump_json(self,json_file):
        self._data = self._pddata
        self.write_json(json_file)



class PStrategySheet(BaseSheet):
    def __init__(self,*args):
        super().__init__(*args)
        self._psdata = {}
        #self._pdname = {}
        self._pmels = {}
        

    def check_sheet(self):
        sheet = self.get_sheet()

        supply_kw, supply_vol, supply_vss, supply_data = self.get_supply_infos()


        # PDName	Location	SrcSupply	SinkSupply	DiffSupply	SupplyIn	EnCtrlSens	ClampVal	ApplyPorts	Elements	ExcludeList	NoISO	Comment
        # find PMISO table 
        start_rowg = self.find_sheet(sheet, 'PMISO') 
        kw_iso = ['self', 'parent','fanout']

        kw_iso = ['ELS_A', 'ELS_B','ELS_C','ELS_D', 'ELS_E', 'force', '-update']
        self.add_dropdown(sheet, '"' + ','.join(kw_iso) + '"', [start_rowg + 1, 13], [start_rowg + 5, 13])

        # PDName	Location	SrcSupply	SinkSupply	SupplyIn SupplyOut Rule	ApplyPorts	Elements	ExcludeList	NoLS	Comment
        # find PMLS table 
        start_rowg = self.find_sheet(sheet, 'PMLS') 
        kw_ls = ['self', 'parent','other','fanout','automatic']


        # PDName	SupplyIn	SupplyOut	EnCtrl	AckResp	OnState	OffState	CtrlAckSupply Comment
        # find PMPSW table 
        start_rowg = self.find_sheet(sheet, 'PMPSW')
        in_psw = []
        out_psw = []
        inpsw = []
        outpsw = []        

        pso_ctl = []
        pso_ack = []

        psoctl = []
        psoack = []
        psw_ctl,psw_ack = self.get_impl_obj(sheet, start_rowg, 'PMPSW')
        #psoctl = self._upfdg._vfile_data['PSO_CTRL']
        psoctl1 = self.get_ctl_sig(self._upfdg._vfile_data['PSO_CTRL'])
        #psoack = self._upfdg._vfile_data['PSO_ACK']
        psoack1 = self.get_ctl_sig(self._upfdg._vfile_data['PSO_ACK'])
        if psw_ctl:
            psoctl.extend(psoctl1)
            psoctl.extend(psw_ctl)
        else:
            psoctl = psoctl1
        if psw_ack:
            psoack.extend(psoack1)
            psoack.extend(psw_ack)
        else:
            psoack = psoack1

        ctlpin = []
        ackpin = []   
        psokw = re.findall(r'PSWRow\d+', ' '.join(self._upfdg._pmfile_data.keys()))    
        for i in range(1, len(psokw) + 1):
            ctlpin.append(self._upfdg._pmfile_data[f'PSWRow{i}']['CtrlPin'])
            ackpin.append(self._upfdg._pmfile_data[f'PSWRow{i}']['AckPin'])


        # PDName	SupplyIn	SaveCtrl	RestCtrl	Elements	ExcludeList	NoRET	RetRegs	Comment
        # find PMRET table 
        start_rowg = self.find_sheet(sheet, 'PMRET')
        #retsave = ' '.join(self._upfdg._vfile_data['RET_SAVE'])
        retsave1 = self.get_ctl_sig(self._upfdg._vfile_data['RET_SAVE'])
        #retres = ' '.join(self._upfdg._vfile_data['RET_RES'])
        retres1 = self.get_ctl_sig(self._upfdg._vfile_data['RET_RES'])
        # retkw = re.findall(r'RETRow\d+', ' '.join(self._upfdg._pmfile_data.keys()))
        # self._upfdg._pmfile_data[f'RETRow{i}']['SavePin'] + ' | ' + self._upfdg._pmfile_data[f'RETRow{i}']['ResPin']
        # for i in range(1, len(retkw) + 1)




    def read_data(self):
        sheet = self.get_sheet()
        self._psdata = self.get_table_contxt(sheet,'PStrategy')
        # print('_psdata: ', self._psdata)

    def dump_json(self,json_file):
        self._data = self._psdata
        self.write_json(json_file)


    def change_sheet(self):
        pass




class PModeSheet(BaseSheet):
    def __init__(self,*args):
        super().__init__(*args)
        self._pmdata = {}
        


    def check_sheet(self):
        
        sheet = self.get_sheet()

        supply_kw, supply_vol, supply_vss,supply_data = self.get_supply_infos()

        # find PMMODE table 
        start_rowg = self.find_sheet(sheet, 'PMMODE') 
        for i in range(1, len(supply_kw) + 1):
            sheet.cell(start_rowg, i + 1).value = supply_kw[i-1]
        sheet.cell(start_rowg, len(supply_kw) + 2).value = 'Comment'


    def read_data(self):
        sheet = self.get_sheet()
        self._pmdata = self.get_table_contxt(sheet,'PMode')
        # print('_pmdata: ', self._pmdata)

    def dump_json(self,json_file):
        self._data = self._pmdata
        self.write_json(json_file)

    def change_sheet(self):
        pass




###############################################################################

def modify_line_in_file(file_path, search_pattern, replacement):
    # 打开文件并逐行读取内容
    with open(file_path, 'r') as file:
        lines = file.readlines()

    # 遍历每一行并进行匹配和替换
    modified_lines = []
    for line in lines:
        if search_pattern in line:
            modified_line = line.replace(search_pattern, replacement)
            modified_lines.append(modified_line)
        else:
            modified_lines.append(line)

    # 将修改后的内容写回文件
    with open(file_path, 'w') as file:
        file.writelines(modified_lines)

def rm_exist_log(logdir):
    # LOG_DIR = os.getenv('TASK_LOGS_DIR')
    # logdir = os.path.join(LOG_DIR,taskid)
    # logdir = r'E:\stone\work\smalltool\pycharm\const0203_sysflatok_full\logs\adfeicc67ere'
    rlogfiles = os.listdir(logdir)
    for logfile in rlogfiles:
        # if logfile.endswith('.log') or logfile.endswith('.rpt'):
        if logfile in ['chk_hym.rpt','chk_sht.rpt','full_chk.rpt']:
            os.remove(os.path.join(logdir,logfile))


def printlog(context, file='upf_gen.log',logdir=''):
    # if not logdir:
    #     # LOG_DIR = os.getenv('TASK_LOGS_DIR')
    #     # logdir = os.path.join(LOG_DIR,taskid)
    #     logdir = r'E:\stone\work\smalltool\pycharm\const0203_sysflatok_full\logs\adfeicc67ere'
    #     rlogdir = os.path.join(logdir,file)
    # else:
    #     rlogdir = logdir

    if not logdir:
        LOG_DIR = os.getenv('TASK_LOGS_DIR')
        taskid = os.getenv('CURRENT_TASK_ID', 'default_task')
        if LOG_DIR:
            logdir = os.path.join(LOG_DIR, taskid)
            os.makedirs(logdir, exist_ok=True)
            rlogdir = os.path.join(logdir, file)
        else:
            # 如果环境变量不存在，使用当前目录
            rlogdir = file
    else:
        rlogdir = logdir

    if os.path.exists(rlogdir):
        with open(rlogdir, 'a') as fw:
            fw.write(context)
    else:
        with open(rlogdir, 'w') as fw:
            fw.write(context)

def movelogrpt(msgnm, flog, fdir):
    if os.path.exists(f'{msgnm}'):
        if os.path.exists(flog):
            # with open('upf_gen.log','r') as fh:
            #     for line in fh.readlines():
            #         #txt_list.append(line.strip())
            #         txt_list += f'{line.strip()} \n'
            # with open(logfile, 'a') as fw:
            #     fw.write(txt_list)
            os.system(f'rm -f {flog}')
            os.system(f'mv {msgnm} {fdir}')
        else:
            os.system(f'mv {msgnm} {fdir}')
    else:
        upf_warn(f'Can not find {msgnm}.')


#########################################################################################################################
# upf message
full_log_message_list =  []
full_rpt_message_list =  []

def upf_log(level, msg, out=sys.stdout):
    print(f'{level.upper()}: {msg}', flush=True, file=out)

def upf_info(msg, kw='log'):
    # if msg not in full_log_message_list:
    if kw == 'log':
        full_log_message_list.append(msg)
        printlog(f'UPF_INFO: {msg} \n', 'full_msg.log')
    else:
        full_rpt_message_list.append(msg)
        printlog(f'UPF_INFO: {msg} \n', 'full_chk.rpt')
    # upf_log('UPF_INFO', msg)
    if kw == 'log':
        printlog(f'UPF_INFO: {msg} \n')
    if kw == 'chk_hym':
        printlog(f'UPF_INFO: {msg} \n', 'chk_hym.rpt')
    if kw == 'chk_sht':
        printlog(f'UPF_INFO: {msg} \n', 'chk_sht.rpt')
    if kw == 'chk_dti':
        printlog(f'UPF_INFO: {msg} \n', 'chk_dti.rpt')
    if kw == 'chk_upf':
        printlog(f'UPF_INFO: {msg} \n', 'chk_upf.rpt')

def upf_warn(msg, kw='log'):
    # if msg not in full_log_message_list:
    if kw == 'log':
        full_log_message_list.append(msg)
        printlog(f'UPF_WARN: {msg} \n', 'full_msg.log')
    else:
        full_rpt_message_list.append(msg)
        printlog(f'UPF_WARN: {msg} \n', 'full_chk.rpt')
    print(f'\033[0:31mUPF_WARN\033[0m: {msg}', flush=True)
    # print(f'UPF_WARN: {msg}', flush=True)
    if kw == 'log':
        printlog(f'UPF_WARN: {msg} \n')
    if kw == 'chk_hym':
        printlog(f'UPF_WARN: {msg} \n', 'chk_hym.rpt')
    if kw == 'chk_sht':
        printlog(f'UPF_WARN: {msg} \n', 'chk_sht.rpt')
    if kw == 'chk_dti':
        printlog(f'UPF_WARN: {msg} \n', 'chk_dti.rpt')
    if kw == 'chk_upf':
        printlog(f'UPF_WARN: {msg} \n', 'chk_upf.rpt')
         

def upf_error(msg, kw='log'):
    # if msg not in full_log_message_list:
    if kw == 'log':
        full_log_message_list.append(msg)
        printlog(f'UPF_ERROR: {msg} \n', 'full_msg.log')
    else:
        full_rpt_message_list.append(msg)
        printlog(f'UPF_ERROR: {msg} \n', 'full_chk.rpt')
    print(f'\033[0:31mUPF_ERROR\033[0m: {msg}', flush=True)
    # print(f'UPF_ERROR: {msg}', flush=True)
    if kw == 'log':
        printlog(f'UPF_ERROR: {msg} \n')
    if kw == 'chk_hym':
        printlog(f'UPF_ERROR: {msg} \n', 'chk_hym.rpt')
    if kw == 'chk_sht':
        printlog(f'UPF_ERROR: {msg} \n', 'chk_sht.rpt')
    if kw == 'chk_dti':
        printlog(f'UPF_ERROR: {msg} \n', 'chk_dti.rpt')
    if kw == 'chk_upf':
        printlog(f'UPF_ERROR: {msg} \n', 'chk_upf.rpt')

def upf_fatal(msg, kw='log'):
    # if msg not in full_log_message_list:
    if kw == 'log':
        full_log_message_list.append(msg)
        printlog(f'UPF_FATAL: {msg} \n', 'full_msg.log')
    else:
        full_rpt_message_list.append(msg)
        printlog(f'UPF_FATAL: {msg} \n', 'full_chk.rpt')
    print(f'\033[0:31mUPF_FATAL\033[0m: {msg}', flush=True)
    # print(f'UPF_FATAL: {msg}', flush=True)
    if kw == 'log':
        printlog(f'UPF_FATAL: {msg} \n')
    if kw == 'chk_hym':
        printlog(f'UPF_FATAL: {msg} \n', 'chk_hym.rpt')
    if kw == 'chk_sht':
        printlog(f'UPF_FATAL: {msg} \n', 'chk_sht.rpt')
    if kw == 'chk_dti':
        printlog(f'UPF_FATAL: {msg} \n', 'chk_dti.rpt')
    if kw == 'chk_upf':
        printlog(f'UPF_FATAL: {msg} \n', 'chk_upf.rpt')
           

def upf_args():
    print(f'##++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++')
    print(f'## UPF Generation Optional Arguments Presentation:')
    print(f'## -hlp:        All of flow option usage reference.')
    print(f'## -usg:        Flow step command usage reference.')
    print(f'## -tmp:        Write out template UPFs for reference')
    print(f'## -gen_dir:    UPF generationg directory')
    print(f'## -hier_yaml:  Design hierarchy info. from user input file')
    print(f'## -setup:      Build setup directories from blocks defined in hier yaml')
    print(f'## -blocks:     Indicate only current design name for UPF generation')
    #print(f'## -flat:       Generate flatten UPF based on current design. If not flat, only generate current design level only UPF')
    print(f'## -dg:         Generate design guide file to be present UPF request format')
    #print(f'## -idg:        Incrementally update design guide file based on existed dg file and updated input files')
    print(f'## -upf:        Write out UPF files')
    print(f'## -check_hym: Check hier yaml data before generating UPF')
    print(f'## -check_only: Check input data before generating UPF')
    print(f'## -check_upf:  Check UPF consistency after generating UPF')
    print(f'## -proj:       Open project mode. Maybe need set some related project environment variables')
    print(f'##++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++')


def upf_usage():
    print(f'##++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++')
    print(f'## Before flow run, user must provide input data including hier yaml and empty vfile, dg file and crg/ip UPFs to be integrated ')
    print(f'## For hier yaml and empty vfile, must follow format of hier_pwr.yaml and pmempty.v in template/ftemp')
    print(f'## For crg/ip UPFs, must follow header format of crg.upf and userip.upf in template/ftemp')
    print(f'##++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++')
    print(f'##++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++')
    print(f'## Step I: Build the whole UPF generation directory structure.')
    print(f'## Cmd Line: xconst sdgen -gen_dir <upfdir> -hier_yaml <hier_file> -setup -blocks <blk_name> [-tmp]')
    print(f'##++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++')
    print(f'##++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++')
    print(f'## Step II: Generate initial design guide file according to current input data.')
    print(f'## Cmd Line: xconst sdgen -gen_dir <upfdir> -hier_yaml <hier_file> -dg -blocks <blk_name>')
    print(f'##++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++')
    print(f'##++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++')
    print(f'## Step III: Check and debug input data after users provide all of input files.')
    print(f'## Cmd  Line: xconst sdgen -gen_dir <upfdir> -hier_yaml <hier_file> -chk_only -blocks <blk_name>')
    print(f'##++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++')
    print(f'##++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++')
    print(f'## Step IV: Generate UPF files if all of input files are OK.')
    print(f'## Cmd  Line: xconst sdgen -gen_dir <upfdir> -hier_yaml <hier_file> -upf -blocks <blk_name> [-proj] [-flat]')
    print(f'##++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++')
    print(f'##++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++')
    print(f'## Step V: Check and debug UPF output files.')
    print(f'## Cmd  Line: xconst sdgen -gen_dir <upfdir> -hier_yaml <hier_file> -chk_upf -blocks <blk_name>')
    print(f'##++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++')


# ORG_COMDIR = os.getenv('ECS_TEMPLATES_DIR')
# 全局环境变量定义
TEMP_DIR = os.getenv('TEMP_UPLOAD_DIR')
LOG_DIR = os.getenv('TASK_LOGS_DIR')
# ORG_COMDIR = r'E:\stone\work\smalltool\pycharm\const0203_sysflatok_full\templates'
# TEMP_DIR = r'E:\stone\work\smalltool\pycharm\const0203_sysflatok_full\temp'
# LOG_DIR = r'E:\stone\work\smalltool\pycharm\const0203_sysflatok_full\logs'
# ufpadfeicc67ere


def upf_dg_chk(*arglist):
    # # 🔥 关键修复：在函数内部重新获取环境变量，确保正确传递
    # global TEMP_DIR, LOG_DIR
    # TEMP_DIR = os.getenv('TEMP_UPLOAD_DIR')
    # LOG_DIR = os.getenv('TASK_LOGS_DIR')

    # # 验证环境变量是否正确设置
    # if not TEMP_DIR:
    #     print(f"❌ 错误: TEMP_UPLOAD_DIR 环境变量未设置")
    #     sys.exit(1)
    # if not LOG_DIR:
    #     print(f"❌ 错误: TASK_LOGS_DIR 环境变量未设置")
    #     sys.exit(1)

    # print(f"✅ 环境变量检查: TEMP_DIR={TEMP_DIR}, LOG_DIR={LOG_DIR}")

    if len(arglist) == 0: arglist = ['-h']
    parser = argparse.ArgumentParser(prog='upf_dg_chk', description='design constraint excel file generation script')
    parser.add_argument('-taskid', help='Task ID name', default='ufpadfeicc67ere', required='False')
    parser.add_argument('-chk', help='Write or update design guide files', action='store_true')
    parser.add_argument('-usr', help='User permission for upf generation')

    args = parser.parse_args(args=arglist)
    taskid = args.taskid

    if args.chk:
        hier_yaml = os.path.join(TEMP_DIR, taskid, 'hier.yaml')
        pvfile = os.path.join(TEMP_DIR, taskid, 'pvlog.v')
        logdir = os.path.join(LOG_DIR,taskid)
        pmfile = os.path.join(TEMP_DIR, taskid, 'pcell.yaml')
        objfile = os.path.join(TEMP_DIR, taskid, 'pobj.tcl')

        # tfile = os.path.join(TEMP_DIR, 'tune.upf')
        # tfile = os.path.join(ORG_COMDIR, 'tune.upf')

        dg_file = os.path.join(TEMP_DIR, taskid,'pcont.xlsx')
        dg_file = re.sub(r'/+', '/', dg_file)
        # lock_file = f'{TEMP_DIR}/.~lock.pcont.xlsx#'
        # lock_file = re.sub(r'/+', '/', lock_file)

        # check hier yaml existence
        if not os.path.exists(hier_yaml):
            upf_error(f'hier yaml file not found {hier_yaml}')
            exit(1)

        # check pvfile existence
        if not os.path.exists(pvfile):
            upf_error(f'Empty vfile not found {pvfile}')
            exit(1)

        # check pobj existence
        if not os.path.exists(objfile):
            upf_error(f'object file not found {objfile}')
            exit(1)

        # check pmfile existence
        if not os.path.exists(pmfile):
            upf_error(f'pmcell file not found {pmfile}')
            exit(1)

        # check pcont existence
        if not os.path.exists(dg_file):
            upf_error(f'Design constraint file not found {dg_file}')
            exit(1)

        # check logdir existence
        if not os.path.exists(logdir):
            upf_error(f'logdir not found {logdir}')
            exit(1)

        rm_exist_log(logdir)
        # if os.path.exists(lock_file):
        #     upf_fatal('pcont.xlsx is in edit mode. Please close it')

        hier_tree = HierPwrTree(hier_yaml)
        upfdg = UPF_DG()
        upfdg.hier_tree = hier_tree

        upfdg.read_vfile(pvfile)
        upfdg.read_pmfile(pmfile)
        upfdg.read_objfile(objfile)

        if os.path.exists(dg_file):
            upfdg.load_design_guide(dg_file)
        else:
            upf_error(f'Can not find {dg_file}')
            exit(1)

        mdname = upfdg._mdname
        upfdg.hier_tree.check_hym(mdname)
        upfdg.check_dg()
        # upfdg.save_workbook(dg_file)

        print(f'Design guide file {dg_file} is verified.')



if __name__ == '__main__':
    if len(sys.argv) < 2 or (len(sys.argv) > 2 and 'upf_dg_chk' not in sys.argv[1]):
        upf_error('Missing some parameters for SDC generation')
        locals()['upf_dg_chk']('-h')
        exit(1)
    app_name = sys.argv[1]
    if app_name in locals():
        locals()[app_name](*sys.argv[2:])
    else:
        raise NameError(f'The application of DataBase generation {app_name} not found')


